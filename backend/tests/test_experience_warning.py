from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from app.experience_warning import (
    EXPERIENCE_POOL_HEADERS,
    DEFAULT_SELECTED_EXPERIENCE_FIELDS,
    WARNING_DETAIL_FIELD,
    WARNING_PARAMETER_FIELD,
    analyze_workbook_warnings,
    analyze_workbook_warnings_with_progress,
    build_warning_report_lines,
    import_experience_pool,
    write_warnings_to_workbook,
)
from app.fill_engine import FillEngine
from app.knowledge_base import KnowledgeBase


def _write_base_kb(path: Path, physical_experience: object = None) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "序号",
            "要素1",
            "要素2",
            "要素3",
            "要素4",
            "要素5",
            "单位",
            "基价",
            "备注",
            "【经验数】实物工作费调整系数",
            "【经验数解释】-实物工作费调整系数",
            "【经验数】技术工作费调整系数",
            "【经验数解释】-技术工作费调整系数",
        ]
    )
    sheet.append([1, "控制测量", "首级控制测量", "", "GPS测量C级", "中等", "个", 4274, "", physical_experience, "", None, ""])
    workbook.save(path)
    workbook.close()


def _write_source_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "表2-通用工程测量费用"
    sheet.append(["标题"])
    sheet.append(["序号", "工作任务"])
    sheet.append(
        [
            None,
            "项目",
            "内容",
            "类别",
            "比例尺",
            "单位",
            "数量",
            "实物工作费调整系数",
            "单价",
            "小计",
            "技术工作费调整系数",
            "小计",
            None,
            None,
            None,
            None,
            "【经验数】单价",
            "【DWB批注】价格or总体",
            "【经验数】实物工作费调整系数",
            "【DWB批注】-实物工作费调整系数",
            "【经验数】技术工作费调整系数",
            "【DWB批注】-技术工作费调整系数",
        ]
    )
    sheet.append(
        [
            "映射行",
            "要素1",
            "要素2",
            "要素4",
            "要素5",
            "单位",
            None,
            "实物工作费调整系数-测试",
            "单价匹配-测试",
            None,
            "技术工作费调整系数-测试",
        ]
    )
    sheet.append(
        [
            1,
            "控制测量",
            "首级控制测量",
            "GPS测量C级",
            "中等",
            "个",
            2,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            4274,
            "经验单价一",
            0.6,
            "经验实物一",
            0.22,
            "经验技术一",
        ]
    )
    sheet.append(
        [
            2,
            "控制测量",
            "首级控制测量",
            "GPS测量C级",
            "中等",
            "个",
            2,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            4300,
            "经验单价二",
            0.8,
            "经验实物二",
            0.25,
            "经验技术二",
        ]
    )
    workbook.save(path)
    workbook.close()


def _write_filled_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "表2-通用工程测量费用"
    sheet.append(["映射行", "要素1", "要素2", "要素4", "要素5", "单位", "数量", "基价", "输出-实物工作费调整系数", "输出-技术工作费调整系数"])
    sheet.append([1, "控制测量", "首级控制测量", "GPS测量C级", "中等", "个", 2, 5000, 0.7, 0.235])
    workbook.save(path)
    workbook.close()


def _pool_record(**updates: object) -> dict[str, object]:
    record = {header: "" for header in EXPERIENCE_POOL_HEADERS}
    record.update(
        {
            "来源文件": "历史控制价.xlsx",
            "来源sheet": "表2-通用工程测量费用",
            "来源行": 18,
            "要素1": "控制测量",
            "要素2": "首级控制测量",
            "要素3": "",
            "要素4": "GPS测量C级",
            "要素5": "中等",
            "单位": "个",
        }
    )
    record.update(updates)
    return record


def _write_pool_workbook(path: Path, records: list[dict[str, object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "经验池"
    sheet.append(EXPERIENCE_POOL_HEADERS)
    for record in records:
        sheet.append([record.get(header) for header in EXPERIENCE_POOL_HEADERS])
    workbook.save(path)
    workbook.close()


def _write_warning_workbook(
    path: Path,
    *,
    current_price: object,
    current_physical: object = None,
    current_technical: object = None,
    row_values: dict[str, object] | None = None,
    headers: list[str] | None = None,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "表2-通用工程测量费用"
    header_row = headers or [
        "映射行",
        "要素1",
        "要素2",
        "要素3",
        "要素4",
        "要素5",
        "单位",
        "数量",
        "基价",
        "输出-实物工作费调整系数",
        "输出-技术工作费调整系数",
    ]
    sheet.append(header_row)
    values = {
        "映射行": 1,
        "要素1": "控制测量",
        "要素2": "首级控制测量",
        "要素3": "",
        "要素4": "GPS测量C级",
        "要素5": "中等",
        "单位": "个",
        "数量": 2,
        "基价": current_price,
        "输出-实物工作费调整系数": current_physical,
        "输出-技术工作费调整系数": current_technical,
    }
    if row_values:
        values.update(row_values)
    sheet.append([values.get(header) for header in header_row])
    workbook.save(path)
    workbook.close()


def test_import_experience_pool_writes_independent_pool_file(tmp_path):
    source = tmp_path / "source.xlsx"
    pool = tmp_path / "pool.xlsx"
    _write_source_workbook(source)

    summary = import_experience_pool(source, pool, selected_fields=DEFAULT_SELECTED_EXPERIENCE_FIELDS)

    assert pool.exists()
    assert summary["imported_rows"] == 2
    workbook = load_workbook(pool, read_only=True, data_only=True)
    try:
        rows = list(workbook.active.iter_rows(values_only=True))
    finally:
        workbook.close()
    assert list(rows[0]) == EXPERIENCE_POOL_HEADERS
    record = dict(zip(rows[0], rows[1]))
    assert [record[field] for field in ["要素1", "要素2", "要素3", "要素4", "要素5", "单位"]] == [
        "控制测量",
        "首级控制测量",
        None,
        "GPS测量C级",
        "中等",
        "个",
    ]
    assert record["基价"] == 4274
    assert record["实物工作费调整系数"] == 0.6
    assert record["技术工作费调整系数"] == 0.22
    assert record["原表备注1"] == "经验单价一"
    assert record["原表备注2"] == "经验实物一"
    assert record["原表备注3"] == "经验技术一"


def test_import_experience_pool_uses_template_workbook_styles(tmp_path):
    source = tmp_path / "source.xlsx"
    pool = tmp_path / "pool.xlsx"
    template = tmp_path / "template.xlsx"
    _write_source_workbook(source)

    template_workbook = Workbook()
    template_sheet = template_workbook.active
    template_sheet.title = "经验池"
    template_sheet.append(EXPERIENCE_POOL_HEADERS)
    template_sheet.append([""] * len(EXPERIENCE_POOL_HEADERS))
    template_sheet["A1"].fill = PatternFill(fill_type="solid", fgColor="112233")
    template_sheet["A2"].fill = PatternFill(fill_type="solid", fgColor="445566")
    template_sheet.column_dimensions["A"].width = 28
    template_sheet.row_dimensions[2].height = 24
    template_sheet.freeze_panes = "A2"
    template_workbook.save(template)
    template_workbook.close()

    import_experience_pool(
        source,
        pool,
        selected_fields=[DEFAULT_SELECTED_EXPERIENCE_FIELDS[0]],
        template_path=template,
    )

    workbook = load_workbook(pool)
    try:
        sheet = workbook.active
        assert sheet["A1"].fill.fgColor.rgb == "00112233"
        assert sheet["A2"].fill.fgColor.rgb == "00445566"
        assert sheet.column_dimensions["A"].width == 28
        assert sheet.row_dimensions[2].height == 24
        assert sheet.freeze_panes == "A2"
    finally:
        workbook.close()


def test_import_experience_pool_restyles_existing_pool_from_template(tmp_path):
    source = tmp_path / "source.xlsx"
    pool = tmp_path / "pool.xlsx"
    template = tmp_path / "template.xlsx"
    _write_source_workbook(source)

    pool_workbook = Workbook()
    pool_sheet = pool_workbook.active
    pool_sheet.append(EXPERIENCE_POOL_HEADERS)
    pool_sheet.append(["旧数据"] + [""] * (len(EXPERIENCE_POOL_HEADERS) - 1))
    pool_workbook.save(pool)
    pool_workbook.close()

    template_workbook = Workbook()
    template_sheet = template_workbook.active
    template_sheet.append(EXPERIENCE_POOL_HEADERS)
    template_sheet.append([""] * len(EXPERIENCE_POOL_HEADERS))
    template_sheet["A1"].fill = PatternFill(fill_type="solid", fgColor="ABCDEF")
    template_sheet["A2"].fill = PatternFill(fill_type="solid", fgColor="FEDCBA")
    template_sheet.column_dimensions["A"].width = 31
    template_workbook.save(template)
    template_workbook.close()

    import_experience_pool(
        source,
        pool,
        selected_fields=[DEFAULT_SELECTED_EXPERIENCE_FIELDS[0]],
        template_path=template,
    )

    workbook = load_workbook(pool)
    try:
        sheet = workbook.active
        assert sheet["A1"].fill.fgColor.rgb == "00ABCDEF"
        assert sheet["A2"].fill.fgColor.rgb == "00FEDCBA"
        assert sheet.column_dimensions["A"].width == 31
    finally:
        workbook.close()


def test_import_experience_pool_with_column_mapping_preserves_source_column_names(tmp_path):
    source = tmp_path / "messy-source.xlsx"
    pool = tmp_path / "pool.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "乱版控制价"
    sheet.append(["项目抬头"])
    sheet.append(
        [
            "序号",
            "项目名称",
            "作业内容",
            "类别名称",
            "比例尺名称",
            "复杂程度名称",
            "计量单位",
            "工程数量",
            "综合单价",
            "实物经验",
            "技术经验",
            "备注A",
            "备注B",
            "备注C",
            "海域水深",
        ]
    )
    sheet.append(
        [
            "",
            "要素1",
            "要素2",
            "要素3",
            "要素4",
            "要素5",
            "单位",
            "工程量",
            "基价",
            "实物工作费调整系数",
            "技术工作费调整系数",
            "原表备注1",
            "原表备注2",
            "原表备注3",
            "其他参数1",
        ]
    )
    sheet.append(
        [
            1,
            "控制测量",
            "首级控制测量",
            "GPS",
            "GPS测量C级",
            "中等",
            "个",
            8,
            4274,
            0.6,
            0.22,
            "价格来自批注",
            "实物系数批注",
            "技术系数批注",
            "浅水",
        ]
    )
    workbook.save(source)
    workbook.close()

    summary = import_experience_pool(
        source,
        pool,
        sheet_configs=[
            {
                "sheet_name": "乱版控制价",
                "enabled": True,
                "header_row": 3,
                "column_mapping": {
                    "要素1": "B",
                    "要素2": "C",
                    "要素3": "D",
                    "要素4": "E",
                    "要素5": "F",
                    "单位": "G",
                    "工程量": "H",
                    "基价": "I",
                    "实物工作费调整系数": "J",
                    "技术工作费调整系数": "K",
                    "原表备注1": "L",
                    "原表备注2": "M",
                    "原表备注3": "N",
                    "其他参数1": "O",
                },
            }
        ],
    )

    assert summary["imported_rows"] == 1
    workbook = load_workbook(pool, read_only=True, data_only=True)
    try:
        rows = list(workbook.active.iter_rows(values_only=True))
    finally:
        workbook.close()
    record = dict(zip(rows[0], rows[1]))
    assert record["要素1列名"] == "项目名称"
    assert record["要素1"] == "控制测量"
    assert record["单位列名"] == "计量单位"
    assert record["基价列名"] == "综合单价"
    assert record["工程量列名"] == "工程数量"
    assert record["工程量"] == 8
    assert record["其他参数1列名称"] == "海域水深"
    assert record["其他参数1读取数值"] == "浅水"
    assert record["原表备注1"] == "价格来自批注"


def test_import_experience_pool_uses_template_headers_without_modifying_template(tmp_path):
    source = tmp_path / "source.xlsx"
    pool = tmp_path / "pool.xlsx"
    template = tmp_path / "【经验池】【模板勿动】-管勘智算.xlsx"
    _write_source_workbook(source)

    template_workbook = Workbook()
    template_sheet = template_workbook.active
    template_sheet.title = "经验池"
    template_sheet.append([*EXPERIENCE_POOL_HEADERS, "模板扩展列"])
    template_sheet.append(["模板说明行，不应被导入函数修改"])
    template_workbook.save(template)
    template_workbook.close()

    summary = import_experience_pool(
        source,
        pool,
        selected_fields=DEFAULT_SELECTED_EXPERIENCE_FIELDS,
        template_path=template,
    )

    assert summary["imported_rows"] == 2
    output = load_workbook(pool, read_only=True, data_only=True)
    try:
        rows = list(output.active.iter_rows(values_only=True))
    finally:
        output.close()
    assert rows[0][-1] == "模板扩展列"
    assert rows[1][-1] is None

    template_after = load_workbook(template, read_only=True, data_only=True)
    try:
        assert template_after.active.max_row == 2
        assert template_after.active["A2"].value == "模板说明行，不应被导入函数修改"
    finally:
        template_after.close()


def test_warning_analysis_lists_multiple_experience_values_for_same_row(tmp_path):
    source = tmp_path / "source.xlsx"
    pool = tmp_path / "pool.xlsx"
    filled = tmp_path / "filled.xlsx"
    _write_source_workbook(source)
    _write_filled_workbook(filled)
    import_experience_pool(source, pool)

    result = analyze_workbook_warnings(filled, pool)

    assert result["summary"]["warning_rows"] == 1
    price_warning = next(item for item in result["warnings"] if item["metric"] == "基价")
    assert price_warning["experience_values"] == [4274, 4300]
    assert price_warning["match_mode"] == "字段完全匹配"
    assert price_warning["match_mode_detail"] == "字段完全匹配"
    assert price_warning["experience_average"] == 4287
    assert price_warning["deviation_percent"] == 16.631677
    assert price_warning["severity"] == "low"
    assert price_warning["severity_label"] == "低风险"
    assert "相对经验池平均值 4287 偏离 16.6317%" in price_warning["suggested_action"]
    assert "控制测量" in price_warning["row_key"]
    assert result["summary"]["candidate_rows"] == 1
    assert result["summary"]["checked_rows"] == 1
    assert result["summary"]["no_comparable_rows"] == 0
    assert result["summary"]["high_rows"] == 0
    assert result["summary"]["low_rows"] == 1
    assert result["summary"]["match_mode_counts"]["字段完全匹配"] == 1
    assert result["summary"]["summary_text"] == "经验池预警：输入候选 1 行，可比选 1 行，未找到同类 0 行，发现 1 条预警，其中高风险 0 条、低风险 1 条；匹配模式：字段完全匹配 1 行"
    assert result["summary"]["metric_counts"]["基价"] == 1


def test_warning_report_lines_include_summary_and_key_items(tmp_path):
    source = tmp_path / "source.xlsx"
    pool = tmp_path / "pool.xlsx"
    filled = tmp_path / "filled.xlsx"
    _write_source_workbook(source)
    _write_filled_workbook(filled)
    import_experience_pool(source, pool)
    result = analyze_workbook_warnings(filled, pool)

    lines = build_warning_report_lines(result["summary"], result["warnings"], limit=3)

    assert lines[0] == "经验池预警：输入候选 1 行，可比选 1 行，未找到同类 0 行，发现 1 条预警，其中高风险 0 条、低风险 1 条；匹配模式：字段完全匹配 1 行"
    assert any("表2-通用工程测量费用 第 2 行：字段完全匹配；基价 当前值 5000" in line for line in lines)
    assert any("字段完全匹配" in line for line in lines)
    assert any("经验池平均值 4287" in line for line in lines)
    assert any("实际偏离率 16.631677%" in line for line in lines)
    assert any("说明：经验池仅用于预警和比选" in line for line in lines)


def test_write_warnings_to_workbook_adds_parameter_and_detail_columns(tmp_path):
    source = tmp_path / "source.xlsx"
    pool = tmp_path / "pool.xlsx"
    filled = tmp_path / "filled.xlsx"
    _write_source_workbook(source)
    _write_filled_workbook(filled)
    import_experience_pool(source, pool)
    result = analyze_workbook_warnings(filled, pool)

    summary = write_warnings_to_workbook(filled, result["row_results"])

    assert summary["written_rows"] == 1
    workbook = load_workbook(filled, data_only=True)
    try:
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        parameter_index = headers.index(WARNING_PARAMETER_FIELD) + 1
        detail_index = headers.index(WARNING_DETAIL_FIELD) + 1
        assert sheet.cell(row=2, column=parameter_index).value == "基价"
        detail = sheet.cell(row=2, column=detail_index).value
        assert "匹配模式：字段完全匹配" in detail
        assert "经验池平均值 4287" in detail
        assert "经验范围 4274~4300" in detail
        assert "实际偏离率 16.631677%" in detail
        assert "低风险阈值 5%" in detail
        assert "高风险阈值 20%" in detail
        assert "source.xlsx / 表2-通用工程测量费用 第5行 的 基价 为 4274" in detail
    finally:
        workbook.close()


def test_write_warnings_to_workbook_writes_no_warning_rows(tmp_path):
    source = tmp_path / "source.xlsx"
    pool = tmp_path / "pool.xlsx"
    filled = tmp_path / "filled-no-warning.xlsx"
    _write_source_workbook(source)
    import_experience_pool(source, pool)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "表2-通用工程测量费用"
    sheet.append(["映射行", "要素1", "要素2", "要素4", "要素5", "单位", "数量", "基价", "输出-实物工作费调整系数", "输出-技术工作费调整系数"])
    sheet.append([1, "控制测量", "首级控制测量", "GPS测量C级", "中等", "个", 2, 4274, 0.7, 0.235])
    workbook.save(filled)
    workbook.close()

    result = analyze_workbook_warnings(filled, pool)
    summary = write_warnings_to_workbook(filled, result["row_results"])

    assert result["summary"]["warning_rows"] == 0
    assert summary["written_rows"] == 1
    workbook = load_workbook(filled, data_only=True)
    try:
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        parameter_index = headers.index(WARNING_PARAMETER_FIELD) + 1
        detail_index = headers.index(WARNING_DETAIL_FIELD) + 1
        assert sheet.cell(row=2, column=parameter_index).value == "无预警"
        detail = sheet.cell(row=2, column=detail_index).value
        assert "已比对参数：" in detail
        assert "基价（平均值 4287；实际偏离率 0.303242%" in detail
        assert "匹配模式：字段完全匹配" in detail
        assert "当前阈值：低风险 5%；高风险 20%" in detail
        assert "结论：未超过阈值，故无预警。" in detail
    finally:
        workbook.close()


def test_warning_analysis_uses_ordered_match_but_not_prefix_match(tmp_path):
    filled = tmp_path / "filled-ordered.xlsx"
    pool = tmp_path / "pool-ordered.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "表3-地质测绘"
    sheet.append(["要素1", "要素2", "要素3", "要素4", "要素5", "单位", "数量", "基价"])
    sheet.append(["岩土工程勘察", "地质测绘", "", "比例-1:500", "中等", "km2", 3, 2])
    workbook.save(filled)
    workbook.close()

    _write_pool_workbook(
        pool,
        [
            {
                **_pool_record(
                    来源sheet="表3-地质测绘",
                    要素1="岩土工程勘察",
                    要素2="地质测绘",
                    要素3="比例-1:500",
                    要素4="",
                    要素5="中等",
                    单位="km2",
                    基价=1,
                )
            }
        ],
    )

    result = analyze_workbook_warnings(filled, pool)

    assert result["summary"]["candidate_rows"] == 1
    assert result["summary"]["checked_rows"] == 1
    assert result["summary"]["no_comparable_rows"] == 0
    assert result["summary"]["match_mode_counts"]["非空要素顺序匹配"] == 1
    warning = result["warnings"][0]
    assert warning["match_mode"] == "非空要素顺序匹配"
    assert warning["match_mode_detail"] == "非空要素顺序匹配"
    assert "前缀匹配" not in warning["warning_detail"]


def test_warning_analysis_keeps_no_warning_when_deviation_within_low_threshold(tmp_path):
    source = tmp_path / "source.xlsx"
    pool = tmp_path / "pool.xlsx"
    filled = tmp_path / "filled-low-threshold.xlsx"
    _write_source_workbook(source)
    import_experience_pool(source, pool)
    _write_warning_workbook(filled, current_price=4400, current_physical=0.7, current_technical=0.235)

    result = analyze_workbook_warnings(filled, pool)

    assert result["summary"]["warning_rows"] == 0
    assert result["summary"]["checked_rows"] == 1
    assert len(result["row_results"]) == 1
    assert result["row_results"][0]["warning_parameter"] == "无预警"
    assert "结论：未超过阈值，故无预警。" in result["row_results"][0]["warning_detail"]


def test_warning_analysis_marks_low_risk_when_deviation_between_thresholds(tmp_path):
    source = tmp_path / "source.xlsx"
    pool = tmp_path / "pool.xlsx"
    filled = tmp_path / "filled-low-risk.xlsx"
    _write_source_workbook(source)
    import_experience_pool(source, pool)
    _write_warning_workbook(filled, current_price=5000, current_physical=0.7, current_technical=0.235)

    result = analyze_workbook_warnings(filled, pool)

    assert result["summary"]["warning_rows"] == 1
    assert result["summary"]["high_rows"] == 0
    assert result["summary"]["low_rows"] == 1
    assert result["warnings"][0]["severity"] == "low"


def test_warning_analysis_marks_high_risk_when_deviation_exceeds_high_threshold(tmp_path):
    pool = tmp_path / "pool-high-risk.xlsx"
    filled = tmp_path / "filled-high-risk.xlsx"
    _write_pool_workbook(pool, [_pool_record(基价=100)])
    _write_warning_workbook(filled, current_price=130, current_physical=None, current_technical=None)

    result = analyze_workbook_warnings(filled, pool)

    assert result["summary"]["warning_rows"] == 1
    assert result["summary"]["high_rows"] == 1
    assert result["warnings"][0]["severity"] == "high"
    assert result["warnings"][0]["deviation_percent"] == 30


def test_warning_analysis_keeps_no_warning_when_average_is_zero_and_current_is_zero(tmp_path):
    pool = tmp_path / "pool-zero.xlsx"
    filled = tmp_path / "filled-zero.xlsx"
    _write_pool_workbook(pool, [_pool_record(基价=0)])
    _write_warning_workbook(filled, current_price=0, current_physical=None, current_technical=None)

    result = analyze_workbook_warnings(filled, pool)

    assert result["summary"]["warning_rows"] == 0
    assert result["row_results"][0]["warning_parameter"] == "无预警"
    assert "基价（平均值 0；实际偏离率 0%" in result["row_results"][0]["warning_detail"]


def test_warning_analysis_keeps_no_warning_when_average_is_zero_and_current_is_nonzero(tmp_path):
    pool = tmp_path / "pool-zero-nonzero.xlsx"
    filled = tmp_path / "filled-zero-nonzero.xlsx"
    _write_pool_workbook(pool, [_pool_record(基价=0)])
    _write_warning_workbook(filled, current_price=15, current_physical=None, current_technical=None)

    result = analyze_workbook_warnings(filled, pool)

    assert result["summary"]["warning_rows"] == 0
    assert result["row_results"][0]["warning_parameter"] == "无预警"
    assert "基价（平均值 0；实际偏离率 0%" in result["row_results"][0]["warning_detail"]


def test_warning_analysis_only_checks_rows_with_filter_value(tmp_path):
    pool = tmp_path / "pool-filter-enter.xlsx"
    filled = tmp_path / "filled-filter-enter.xlsx"
    _write_pool_workbook(pool, [_pool_record(基价=100)])
    _write_warning_workbook(filled, current_price=130, current_physical=None, current_technical=None, row_values={"数量": 5})

    result = analyze_workbook_warnings(filled, pool)

    assert result["summary"]["candidate_rows"] == 1
    assert result["summary"]["checked_rows"] == 1
    assert result["summary"]["warning_rows"] == 1


def test_warning_analysis_skips_rows_when_filter_value_is_blank(tmp_path):
    pool = tmp_path / "pool-filter-blank.xlsx"
    filled = tmp_path / "filled-filter-blank.xlsx"
    _write_pool_workbook(pool, [_pool_record(基价=100)])
    _write_warning_workbook(filled, current_price=130, current_physical=None, current_technical=None, row_values={"数量": "   "})

    result = analyze_workbook_warnings(filled, pool)

    assert result["summary"]["candidate_rows"] == 0
    assert result["summary"]["checked_rows"] == 0
    assert result["summary"]["warning_rows"] == 0
    assert result["row_results"] == []


def test_warning_analysis_skips_rows_when_filter_value_is_zero(tmp_path):
    pool = tmp_path / "pool-filter-zero.xlsx"
    filled = tmp_path / "filled-filter-zero.xlsx"
    _write_pool_workbook(pool, [_pool_record(基价=100)])
    _write_warning_workbook(filled, current_price=130, current_physical=None, current_technical=None, row_values={"数量": 0})

    result = analyze_workbook_warnings(filled, pool)

    assert result["summary"]["candidate_rows"] == 0
    assert result["summary"]["checked_rows"] == 0
    assert result["summary"]["warning_rows"] == 0


def test_warning_analysis_raises_when_filter_field_is_missing(tmp_path):
    pool = tmp_path / "pool-filter-missing.xlsx"
    filled = tmp_path / "filled-filter-missing.xlsx"
    _write_pool_workbook(pool, [_pool_record(基价=100)])
    _write_warning_workbook(
        filled,
        current_price=130,
        current_physical=None,
        current_technical=None,
        headers=["映射行", "要素1", "要素2", "要素3", "要素4", "要素5", "单位", "基价"],
    )

    try:
        analyze_workbook_warnings(filled, pool)
    except ValueError as exc:
        assert str(exc) == "预警 sheet 表2-通用工程测量费用 未映射过滤字段：数量"
    else:
        raise AssertionError("预期缺少过滤字段时抛出 ValueError")


def test_warning_analysis_inherits_vertical_merged_cell_values_for_all_rows(tmp_path):
    filled = tmp_path / "filled-merged-warning.xlsx"
    pool = tmp_path / "pool-merged-warning.xlsx"

    filled_workbook = Workbook()
    filled_sheet = filled_workbook.active
    filled_sheet.title = "表2-通用工程测量费用"
    filled_sheet.append(["映射行", "要素1", "要素2", "要素3", "要素4", "要素5", "单位", "数量", "基价"])
    filled_sheet.append([1, "控制测量", "首级控制测量", "", "GPS测量C级", "中等", "个", 6, 5000])
    filled_sheet.append([2, None, "首级控制测量", "", "GPS测量C级", "中等", "个", None, 5000])
    filled_sheet.merge_cells("H2:H3")
    filled_sheet.merge_cells("B2:B3")
    filled_workbook.save(filled)
    filled_workbook.close()

    pool_workbook = Workbook()
    pool_sheet = pool_workbook.active
    pool_sheet.title = "经验池"
    pool_sheet.append(EXPERIENCE_POOL_HEADERS)
    record = {header: "" for header in EXPERIENCE_POOL_HEADERS}
    record.update(
        {
            "来源文件": "历史控制价.xlsx",
            "来源sheet": "表2-通用工程测量费用",
            "来源行": 9,
            "要素1": "控制测量",
            "要素2": "首级控制测量",
            "要素3": "",
            "要素4": "GPS测量C级",
            "要素5": "中等",
            "单位": "个",
            "基价": 4274,
        }
    )
    pool_sheet.append([record.get(header) for header in EXPERIENCE_POOL_HEADERS])
    pool_workbook.save(pool)
    pool_workbook.close()

    progress_events: list[dict[str, object]] = []
    result = analyze_workbook_warnings_with_progress(
        filled,
        pool,
        progress_callback=lambda payload: progress_events.append(dict(payload)),
    )

    assert result["summary"]["candidate_rows"] == 2
    assert result["summary"]["checked_rows"] == 2
    assert result["summary"]["warning_rows"] == 2
    assert len(result["row_results"]) == 2
    assert [row["excel_row"] for row in result["row_results"]] == [2, 3]
    assert all(row["warning_parameter"] == "基价" for row in result["row_results"])
    assert progress_events[0]["status"] == "running"
    assert progress_events[-1]["status"] == "completed"
    assert progress_events[-1]["processed_rows"] == 2
    assert progress_events[-1]["total_rows"] == 2


def test_experience_pool_does_not_feed_second_layer_matching(tmp_path):
    kb_path = tmp_path / "kb.xlsx"
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    _write_base_kb(kb_path, physical_experience=None)

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["要素1", "要素2", "要素4", "要素5", "单位", "基价", "输出-实物工作费调整系数"])
    sheet.append(["控制测量", "首级控制测量", "GPS测量C级", "中等", "个", None, None])
    workbook.save(input_path)
    workbook.close()

    summary = FillEngine(KnowledgeBase.from_excel(kb_path)).fill_workbook(input_path, output_path)

    assert summary.physical_experience_rows == 0
    output = load_workbook(output_path, data_only=True)
    try:
        assert output.active["G2"].value == "待复核"
    finally:
        output.close()
