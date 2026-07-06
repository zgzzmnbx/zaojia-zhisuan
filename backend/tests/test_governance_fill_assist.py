from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

import app.main as main_module
from app.experience_warning import EXPERIENCE_POOL_HEADERS
from app.main import app
from app.schemas import FillSummary, ReviewRow


def _write_state(job_dir: Path, output_name: str, summary: FillSummary) -> None:
    (job_dir / main_module.PROCESS_STATE_FILENAME).write_text(
        main_module.json.dumps(
            {
                "input_filename": "input.xlsx",
                "input_excel": "",
                "output_excel": output_name,
                "output_report": "report.docx",
                "summary": summary.to_dict(),
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )


def _minimal_summary() -> FillSummary:
    return FillSummary(
        total_data_rows=1,
        price_column="基价",
        filled_rows=0,
        matched_rows=0,
        unchanged_rows=0,
        review_rows=1,
        conflict_rows=0,
        output_excel="output.xlsx",
        output_report="report.docx",
        report_text="",
        table_preview={
            "sheet_name": "表2",
            "header_row": 1,
            "headers": ["要素1", "要素2", "要素3", "要素4", "要素5", "单位", "基价"],
            "rows": [["控制测量", "GPS测量E级", "", "中等", "", "点", "待复核"]],
            "row_numbers": [2],
        },
        review_details=[
            ReviewRow(
                excel_row=2,
                status="unmatched",
                message="要素4开始没有匹配",
                values={"要素1": "控制测量", "要素2": "GPS测量E级", "要素4": "中等", "单位": "点"},
            )
        ],
    )


def test_experience_pool_governance_reports_quality_issues(tmp_path, monkeypatch):
    pool_path = tmp_path / "experience.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "经验池"
    sheet.append(["来源文件", "来源sheet", "来源行", "要素1", "要素2", "要素3", "要素4", "要素5", "单位", "基价", "实物工作费调整系数", "技术工作费调整系数"])
    sheet.append(["a.xlsx", "表2", 5, "控制测量", "GPS", "", "中等", "", "点", 100, 1, 0.22])
    sheet.append(["b.xlsx", "表2", 6, "控制测量", "GPS", "", "中等", "", "点", 120, 1, 0.22])
    sheet.append(["c.xlsx", "表2", 7, "控制测量", "GPS", "", "中等", "", "", "bad", "", 0.22])
    workbook.save(pool_path)
    workbook.close()
    monkeypatch.setattr(main_module, "RUNTIME_DIR", tmp_path / "runtime")
    monkeypatch.setattr(main_module, "DEFAULT_EXPERIENCE_POOL_PATH", pool_path)
    monkeypatch.setattr(main_module, "LEGACY_EXPERIENCE_POOL_PATH", tmp_path / "missing.xlsx")

    response = TestClient(app).get("/api/quality/experience-pool")

    assert response.status_code == 200
    payload = response.json()
    assert payload["summary"]["issue_count"] >= 3
    assert payload["summary"]["categories"]["duplicate_record"] == 1
    assert payload["summary"]["categories"]["empty_field"] >= 1
    assert Path(payload["report_path"]).exists()


def test_risk_summary_combines_review_rows_and_warning_details(tmp_path, monkeypatch):
    job_dir = tmp_path / "job-risk"
    job_dir.mkdir()
    summary = _minimal_summary()
    summary.warning_details = [
        {
            "sheet_name": "表2",
            "excel_row": 2,
            "metric": "基价",
            "severity": "high",
            "severity_label": "高风险",
            "message": "高于经验池均值 25%",
            "experience_average": 100,
            "deviation_percent": 25,
            "sample_count": 5,
        }
    ]
    _write_state(job_dir, "output.xlsx", summary)
    monkeypatch.setattr(main_module, "RUNTIME_DIR", tmp_path)

    response = TestClient(app).get("/api/risk/summary", params={"job_id": "job-risk"})

    assert response.status_code == 200
    payload = response.json()
    assert payload["summary"]["total"] == 2
    assert payload["summary"]["type_counts"]["待复核"] == 1
    assert payload["summary"]["type_counts"]["经验池偏离"] == 1


def test_fill_assist_candidates_and_confirm_write_trace(tmp_path, monkeypatch):
    kb_path = tmp_path / "kb.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["要素1", "要素2", "要素3", "要素4", "要素5", "单位", "基价"])
    sheet.append(["控制测量", "GPS测量E级", "", "简单", "", "点", 3000])
    workbook.save(kb_path)
    workbook.close()

    job_dir = tmp_path / "job-fill"
    job_dir.mkdir()
    output_path = job_dir / "output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "表2"
    sheet.append(["要素1", "要素2", "要素3", "要素4", "要素5", "单位", "基价", "匹配状态", "匹配说明"])
    sheet.append(["控制测量", "GPS测量E级", "", "中等", "", "点", "待复核", "待复核", "要素4开始没有匹配"])
    workbook.save(output_path)
    workbook.close()
    _write_state(job_dir, output_path.name, _minimal_summary())
    monkeypatch.setattr(main_module, "RUNTIME_DIR", tmp_path)
    monkeypatch.setattr(main_module, "DEFAULT_KB_PATH", kb_path)
    monkeypatch.setattr(main_module, "DEFAULT_EXPERIENCE_POOL_PATH", tmp_path / "missing-pool.xlsx")
    monkeypatch.setattr(main_module, "LEGACY_EXPERIENCE_POOL_PATH", tmp_path / "missing-legacy.xlsx")

    client = TestClient(app)
    response = client.post(
        "/api/fill-assist/candidates",
        json={"job_id": "job-fill", "sheet_name": "表2", "row_number": 2, "target_header": "基价"},
    )
    assert response.status_code == 200
    payload = response.json()
    assert payload["context"]["target_column"] == 7
    assert payload["candidates"][0]["source"] == "knowledge_similar"
    assert payload["candidates"][0]["value"] == 3000
    assert payload["trace"][0]["kind"] == "匹配过程"
    assert payload["trace"][0]["source"] == "输出 Excel 匹配说明列"

    confirm_response = client.post(
        "/api/fill-assist/confirm",
        json={
            "job_id": "job-fill",
            "sheet_name": "表2",
            "row_number": 2,
            "column_number": 7,
            "candidate": payload["candidates"][0],
            "note": "人工确认采用相似知识库记录",
        },
    )
    assert confirm_response.status_code == 200
    edited = load_workbook(output_path)
    try:
        cell = edited["表2"]["G2"]
        assert cell.value == 3000
        assert "辅助填价人工确认" in cell.comment.text
        assert "二维知识库第 2 行" in cell.comment.text
    finally:
        edited.close()
    logs = main_module.json.loads((job_dir / main_module.MANUAL_EDIT_LOG_FILENAME).read_text(encoding="utf-8"))
    assert logs[-1]["source"] == "fill-assist"
    assert logs[-1]["candidate"]["source"] == "knowledge_similar"


def test_fill_assist_candidates_include_experience_pool_source(tmp_path, monkeypatch):
    kb_path = tmp_path / "kb.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["要素1", "要素2", "要素3", "要素4", "要素5", "单位", "基价"])
    sheet.append(["控制测量", "GPS测量E级", "", "简单", "", "点", 3000])
    workbook.save(kb_path)
    workbook.close()

    pool_path = tmp_path / "experience-pool.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(EXPERIENCE_POOL_HEADERS)
    for index, value in enumerate([3200, 3400, 3600], start=2):
        record = {header: "" for header in EXPERIENCE_POOL_HEADERS}
        record.update(
            {
                "来源文件": "历史控制价.xlsx",
                "来源sheet": "表2",
                "来源行": index,
                "要素1": "控制测量",
                "要素2": "GPS测量E级",
                "要素4": "中等",
                "单位": "点",
                "基价": value,
                "工程量": 1,
            }
        )
        sheet.append([record.get(header) for header in EXPERIENCE_POOL_HEADERS])
    workbook.save(pool_path)
    workbook.close()

    job_dir = tmp_path / "job-fill-experience"
    job_dir.mkdir()
    output_path = job_dir / "output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "表2"
    sheet.append(["要素1", "要素2", "要素3", "要素4", "要素5", "单位", "基价", "匹配状态", "匹配说明"])
    sheet.append(["控制测量", "GPS测量E级", "", "中等", "", "点", "待复核", "待复核", "要素4开始没有匹配"])
    workbook.save(output_path)
    workbook.close()

    _write_state(job_dir, output_path.name, _minimal_summary())
    monkeypatch.setattr(main_module, "RUNTIME_DIR", tmp_path)
    monkeypatch.setattr(main_module, "DEFAULT_KB_PATH", kb_path)
    monkeypatch.setattr(main_module, "DEFAULT_EXPERIENCE_POOL_PATH", pool_path)
    monkeypatch.setattr(main_module, "LEGACY_EXPERIENCE_POOL_PATH", tmp_path / "missing-legacy.xlsx")

    response = TestClient(app).post(
        "/api/fill-assist/candidates",
        json={"job_id": "job-fill-experience", "sheet_name": "表2", "row_number": 2, "target_header": "基价"},
    )

    assert response.status_code == 200
    candidates = response.json()["candidates"]
    experience_candidates = [candidate for candidate in candidates if candidate["source"] == "experience_pool"]
    assert experience_candidates
    assert experience_candidates[0]["source_label"] == "经验池同类均值"
    assert experience_candidates[0]["value"] == 3400
    assert experience_candidates[0]["sample_count"] == 3
    assert candidates[0]["source"] == "experience_pool"
    assert candidates[0]["similarity"] == 100


def test_process_can_defer_matching_until_batch_match(tmp_path, monkeypatch):
    kb_path = tmp_path / "kb.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["要素1", "要素2", "要素3", "要素4", "要素5", "单位", "基价"])
    sheet.append(["控制测量", "GPS测量E级", "", "中等", "", "点", 3203])
    workbook.save(kb_path)
    workbook.close()

    input_path = tmp_path / "input.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "表2"
    sheet.append(["要素1", "要素2", "要素3", "要素4", "要素5", "单位", "基价"])
    sheet.append(["控制测量", "GPS测量E级", "", "中等", "", "点", ""])
    workbook.save(input_path)
    workbook.close()

    monkeypatch.setattr(main_module, "RUNTIME_DIR", tmp_path / "runtime")
    monkeypatch.setattr(main_module, "DEFAULT_KB_PATH", kb_path)
    client = TestClient(app)

    with input_path.open("rb") as handle:
        prepare_response = client.post(
            "/api/process",
            files={"file": ("input.xlsx", handle, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={
                "column_mapping": main_module.json.dumps(
                    {
                        "要素1": "A",
                        "要素2": "B",
                        "要素3": "C",
                        "要素4": "D",
                        "要素5": "E",
                        "单位": "F",
                        "输出-价格列": "G",
                    },
                    ensure_ascii=False,
                ),
                "only_match_rows_with_value": "false",
                "defer_matching": "true",
            },
        )

    assert prepare_response.status_code == 200
    prepared = prepare_response.json()
    assert prepared["summary"]["matching_status"] == "pending"
    assert prepared["summary"]["filled_rows"] == 0
    assert prepared["downloads"]["excel"] == ""

    refresh_response = client.post(
        "/api/preview/refresh",
        json={
            "job_id": prepared["job_id"],
            "header_rows": {"表2": 1},
        },
    )
    assert refresh_response.status_code == 200

    batch_response = client.post("/api/process/batch-match", json={"job_id": prepared["job_id"]})

    assert batch_response.status_code == 200
    matched = batch_response.json()
    assert matched["summary"]["matching_status"] == "completed"
    assert matched["summary"]["filled_rows"] == 1
    output_path = main_module.RUNTIME_DIR / prepared["job_id"] / matched["summary"]["output_excel"]
    edited = load_workbook(output_path, data_only=True)
    try:
        assert edited["表2"]["G2"].value == 3203
    finally:
        edited.close()


def test_batch_match_preserves_current_output_values_after_workload_like_prefill(tmp_path, monkeypatch):
    kb_path = tmp_path / "kb.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["要素1", "要素2", "要素3", "要素4", "要素5", "单位", "基价", "实物工作费调整系数", "技术工作费调整系数"])
    sheet.append(["控制测量", "GPS测量E级", "", "中等", "", "点", 3203, 0.6, 0.22])
    workbook.save(kb_path)
    workbook.close()

    input_path = tmp_path / "input.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "表2"
    sheet.append(["要素1", "要素2", "要素3", "要素4", "要素5", "单位", "基价", "数量"])
    sheet.append(["控制测量", "GPS测量E级", "", "中等", "", "点", "", ""])
    workbook.save(input_path)
    workbook.close()

    monkeypatch.setattr(main_module, "RUNTIME_DIR", tmp_path / "runtime")
    monkeypatch.setattr(main_module, "DEFAULT_KB_PATH", kb_path)
    client = TestClient(app)

    with input_path.open("rb") as handle:
        prepare_response = client.post(
            "/api/process",
            files={"file": ("input.xlsx", handle, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={
                "column_mapping": main_module.json.dumps(
                    {
                        "要素1": "A",
                        "要素2": "B",
                        "要素3": "C",
                        "要素4": "D",
                        "要素5": "E",
                        "单位": "F",
                        "输出-价格列": "G",
                    },
                    ensure_ascii=False,
                ),
                "only_match_rows_with_value": "false",
                "defer_matching": "true",
            },
        )

    assert prepare_response.status_code == 200
    prepared = prepare_response.json()
    output_path = main_module.RUNTIME_DIR / prepared["job_id"] / prepared["summary"]["output_excel"]
    output_book = load_workbook(output_path)
    try:
        output_book["表2"]["H2"].value = 26
        output_book.save(output_path)
    finally:
        output_book.close()

    batch_response = client.post("/api/process/batch-match", json={"job_id": prepared["job_id"]})

    assert batch_response.status_code == 200
    matched = batch_response.json()
    assert matched["summary"]["matching_status"] == "completed"
    edited = load_workbook(output_path, data_only=True)
    try:
        assert edited["表2"]["G2"].value == 3203
        assert edited["表2"]["H2"].value == 26
    finally:
        edited.close()
