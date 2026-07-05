import csv
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from app import adjustment_rules
from app.fill_engine import FillEngine
from app.knowledge_base import KnowledgeBase
from app.normalization import normalize_key_part


ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = ROOT / "03-知识库-二维数据库制作"
KB_PATH = DATA_DIR / "【数据库】【导入】.xlsx"
RULE_DIR = ROOT / "backend" / "app" / "rules"
ADJUSTMENT_RULE_FILES = [
    "physical_factor_rules",
    "physical_factor_overrides",
    "technical_fee_rules",
]


def find_data_file(*tokens: str, exclude: tuple[str, ...] = (), required: bool = True) -> Path | None:
    for path in DATA_DIR.glob("*.xlsx"):
        if path.name.startswith("~$"):
            continue
        if all(token in path.name for token in tokens) and not any(token in path.name for token in exclude):
            return path
    if not required:
        return None
    raise FileNotFoundError(f"未找到测试数据文件：{tokens}")


INPUT_PATH = find_data_file("输入100", "空单价100", exclude=("答案",), required=False)
ANSWER_PATH = find_data_file("输入100-答案", "空单价100", required=False)
PROJECT_EXAMPLE_PATH = find_data_file("项目例子", "控制价")


def read_csv_rule_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), [dict(row) for row in reader if row.get("rule_id")]


def read_xlsx_rule_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    headers = [str(value or "").strip() for value in rows[0]]
    data: list[dict[str, str]] = []
    for values in rows[1:]:
        row = {
            header: "" if value is None else str(value).strip()
            for header, value in zip(headers, values)
        }
        if row.get("rule_id"):
            data.append(row)
    return headers, data


def write_rule_csv(path: Path, headers: list[str], rows: list[list[str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)


def write_rule_xlsx(path: Path, headers: list[str], rows: list[list[str]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "rules"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def test_normalize_key_part_handles_parentheses_and_scale_prefix():
    assert normalize_key_part("三角（边）一级") == normalize_key_part("三角(边)一级")
    assert normalize_key_part("Ａ：Ｂ，Ｃ／Ｄ") == normalize_key_part("A:B,C/D")
    assert normalize_key_part("1:500") == normalize_key_part("比例-1:500")
    assert normalize_key_part("  GPS测量D级\r\n") == "gps测量d级"
    assert normalize_key_part("岩\u2028土\u200b工程 勘察") == normalize_key_part("岩土工程勘察")


def test_knowledge_base_finds_known_price():
    kb = KnowledgeBase.from_excel(KB_PATH)

    result = kb.lookup(
        {
            "要素1": "岩土工程勘察",
            "要素2": "地质测绘",
            "要素3": "",
            "要素4": "比例-1:500",
            "要素5": "复杂",
            "单位": "km2",
        }
    )

    assert result.status == "matched"
    assert result.price == 17213


def test_knowledge_base_reads_first_sheet_columns_a_to_h_only(tmp_path):
    kb_path = tmp_path / "kb-a-to-h.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "第一个sheet"
    sheet.append(["序号", "要素1", "要素2", "要素3", "要素4", "要素5", "单位", "基价", "基价"])
    sheet.append([1, "测试工程", "测试项目", "", "", "", "项", 123, 999])
    other = workbook.create_sheet("第二个sheet")
    other.append(["序号", "要素1", "要素2", "要素3", "要素4", "要素5", "单位", "基价"])
    other.append([1, "测试工程", "测试项目", "", "", "", "项", 456])
    workbook.save(kb_path)

    kb = KnowledgeBase.from_excel(kb_path)
    result = kb.lookup(
        {
            "要素1": "测试工程",
            "要素2": "测试项目",
            "要素3": "",
            "要素4": "",
            "要素5": "",
            "单位": "项",
        }
    )

    assert result.status == "matched"
    assert result.price == 123


def test_knowledge_base_reads_experience_adjustment_fields(tmp_path):
    kb_path = tmp_path / "kb-experience.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "序号",
        "要素1",
        "要素2",
        "要素3",
        "要素4",
        "要素5",
        "单位",
        "基价",
        "备注",
        "【经验数】实物工作费调整系数",
        "【经验数解释】-实物工作费调整系数",
        "【经验数】技术工作费调整系数",
        "【经验数解释】-技术工作费调整系数",
    ])
    sheet.append([1, "经验工程", "经验项目", "", "", "", "项", 123, "", 1.35, "实物经验说明", 0.66, "技术经验说明"])
    workbook.save(kb_path)

    kb = KnowledgeBase.from_excel(kb_path)
    result = kb.lookup(
        {
            "要素1": "经验工程",
            "要素2": "经验项目",
            "要素3": "",
            "要素4": "",
            "要素5": "",
            "单位": "项",
        }
    )

    assert result.status == "matched"
    candidate = result.candidates[0]
    assert candidate["【经验数】实物工作费调整系数"] == 1.35
    assert candidate["【经验数解释】-实物工作费调整系数"] == "实物经验说明"
    assert candidate["【经验数】技术工作费调整系数"] == 0.66
    assert candidate["【经验数解释】-技术工作费调整系数"] == "技术经验说明"


def test_knowledge_base_ignores_spacing_and_cell_line_breaks_when_matching():
    kb = KnowledgeBase(
        [
            {
                "要素1": "岩土工程勘察",
                "要素2": "地质测绘",
                "要素3": "",
                "要素4": "比例-1:500",
                "要素5": "复杂",
                "单位": "km2",
                "基价": 17213,
                "_excel_row": 2,
            }
        ]
    )

    result = kb.lookup(
        {
            "要素1": "岩 土\r\n工程 勘察",
            "要素2": "地质\u2028测绘",
            "要素3": "",
            "要素4": "１：５００",
            "要素5": "复\u200b杂",
            "单位": "k m 2",
        }
    )

    assert result.status == "matched"
    assert result.price == 17213


def test_adjustment_rule_excel_workbooks_exist_and_match_csv_fields():
    for stem in ADJUSTMENT_RULE_FILES:
        csv_headers, csv_rows = read_csv_rule_rows(RULE_DIR / f"{stem}.csv")
        xlsx_headers, xlsx_rows = read_xlsx_rule_rows(RULE_DIR / f"{stem}.xlsx")

        assert xlsx_headers == csv_headers
        assert [row["rule_id"] for row in xlsx_rows] == [row["rule_id"] for row in csv_rows]
        assert xlsx_rows == csv_rows


def test_adjustment_engine_prefers_excel_rules_over_csv(monkeypatch, tmp_path):
    physical_headers = [
        "rule_id",
        "sheet_tokens",
        "trigger_keywords",
        "coefficient",
        "source_type",
        "confidence",
        "basis",
        "review_required",
        "note",
    ]
    technical_headers = [
        "rule_id",
        "sheet_tokens",
        "business_keywords",
        "category_keywords",
        "coefficient",
        "source_type",
        "confidence",
        "basis",
        "formula_effective",
        "review_required",
        "note",
    ]
    write_rule_csv(
        tmp_path / "physical_factor_rules.csv",
        physical_headers,
        [["P-XLSX-PRIORITY", "表4", "泥浆护壁", "1.5", "CSV", "低", "csv", "false", "CSV 规则"]],
    )
    write_rule_xlsx(
        tmp_path / "physical_factor_rules.xlsx",
        physical_headers,
        [["P-XLSX-PRIORITY", "表4", "泥浆护壁", "1.8", "Excel", "高", "xlsx", "false", "Excel 规则"]],
    )
    write_rule_csv(tmp_path / "physical_factor_overrides.csv", physical_headers, [])
    write_rule_csv(
        tmp_path / "technical_fee_rules.csv",
        technical_headers,
        [["T-XLSX-TECH", "表4", "岩土工程勘察", "甲级", "1.2", "CSV", "高", "csv", "true", "false", "技术规则"]],
    )
    write_rule_xlsx(
        tmp_path / "technical_fee_rules.xlsx",
        technical_headers,
        [["T-XLSX-TECH", "表4", "岩土工程勘察", "甲级", "1.3", "Excel", "高", "xlsx", "true", "false", "Excel 技术规则"]],
    )
    monkeypatch.setattr(adjustment_rules, "RULE_DIR", tmp_path)

    engine = adjustment_rules.AdjustmentEngine.from_default_rules()
    result = engine.evaluate(
        "表4-通用工程勘察费用",
        {"要素1": "岩土工程勘察", "要素2": "钻探", "要素3": "", "要素4": "", "要素5": "", "单位": "m"},
        ["泥浆护壁", "甲级"],
    )

    assert result.physical.value == "待复核"
    assert result.physical.status == "review"
    assert result.technical.value == 1.3
    assert "Excel 技术规则" in result.technical.message


def test_technical_fee_first_layer_rules_from_standard_workbook():
    engine = adjustment_rules.AdjustmentEngine.from_default_rules()

    table2_aerial = engine.evaluate(
        "表2-通用工程测量费用",
        {"要素1": "线路航测", "要素2": "", "要素3": "", "要素4": "", "要素5": "", "单位": "km"},
        [],
    )
    assert table2_aerial.technical.value == 0
    assert table2_aerial.technical.rules[0].rule_id == "T-T2-AERIAL-LINE-000"

    table2_default = engine.evaluate(
        "表2-通用工程测量费用",
        {"要素1": "地面测量", "要素2": "控制测量", "要素3": "", "要素4": "", "要素5": "", "单位": "点"},
        [],
    )
    assert table2_default.technical.value == 0.22
    assert table2_default.technical.rules[0].rule_id == "T-T2-DEFAULT-022"

    table4_point = engine.evaluate(
        "表4-通用工程勘察费用",
        {"要素1": "勘探点测放", "要素2": "", "要素3": "", "要素4": "", "要素5": "", "单位": "点"},
        [],
    )
    assert table4_point.technical.value == 0
    assert table4_point.technical.status == "matched"

    table4_high_density = engine.evaluate(
        "表4-通用工程勘察费用",
        {"要素1": "工程物探", "要素2": "高密度电法经验", "要素3": "", "要素4": "", "要素5": "", "单位": "km"},
        ["物探"],
    )
    assert table4_high_density.technical.value == 1
    assert table4_high_density.technical.rules[0].rule_id == "T-T4-GEOPHYSICS-HIGH-DENSITY-100"


def test_technical_fee_high_priority_rule_wins_within_first_layer(monkeypatch, tmp_path):
    physical_headers = [
        "rule_id",
        "sheet_tokens",
        "trigger_keywords",
        "coefficient",
        "source_type",
        "confidence",
        "basis",
        "review_required",
        "note",
    ]
    technical_headers = [
        "rule_id",
        "sheet_tokens",
        "business_keywords",
        "category_keywords",
        "coefficient",
        "source_type",
        "confidence",
        "basis",
        "formula_effective",
        "review_required",
        "note",
        "priority",
    ]
    write_rule_csv(tmp_path / "physical_factor_rules.csv", physical_headers, [])
    write_rule_csv(tmp_path / "physical_factor_overrides.csv", physical_headers, [])
    write_rule_csv(
        tmp_path / "technical_fee_rules.csv",
        technical_headers,
        [
            [
                "T-NORMAL-FIRST",
                "表4",
                "工程物探",
                "物探",
                "0.22",
                "第一层标准规则",
                "高",
                "test",
                "true",
                "false",
                "普通规则",
                "",
            ],
            [
                "T-HIGH-SECOND",
                "表4",
                "工程物探",
                "高密度电法经验",
                "1.0",
                "第一层标准规则",
                "高",
                "test",
                "true",
                "false",
                "高优先级规则",
                "高",
            ],
        ],
    )
    monkeypatch.setattr(adjustment_rules, "RULE_DIR", tmp_path)

    engine = adjustment_rules.AdjustmentEngine.from_default_rules()
    result = engine.evaluate(
        "表4-通用工程勘察费用",
        {"要素1": "工程物探", "要素2": "高密度电法经验", "要素3": "", "要素4": "", "要素5": "", "单位": "km"},
        ["物探"],
    )

    assert result.technical.value == 1
    assert result.technical.rules[0].rule_id == "T-HIGH-SECOND"


def test_adjustment_engine_falls_back_to_csv_when_excel_is_broken(monkeypatch, tmp_path):
    physical_headers = [
        "rule_id",
        "sheet_tokens",
        "trigger_keywords",
        "coefficient",
        "source_type",
        "confidence",
        "basis",
        "review_required",
        "note",
    ]
    technical_headers = [
        "rule_id",
        "sheet_tokens",
        "business_keywords",
        "category_keywords",
        "coefficient",
        "source_type",
        "confidence",
        "basis",
        "formula_effective",
        "review_required",
        "note",
    ]
    write_rule_csv(
        tmp_path / "physical_factor_rules.csv",
        physical_headers,
        [["P-CSV-FALLBACK", "表4", "泥浆护壁", "1.6", "CSV", "高", "csv", "false", "CSV 回退规则"]],
    )
    (tmp_path / "physical_factor_rules.xlsx").write_text("not a valid xlsx", encoding="utf-8")
    write_rule_csv(tmp_path / "physical_factor_overrides.csv", physical_headers, [])
    write_rule_csv(
        tmp_path / "technical_fee_rules.csv",
        technical_headers,
        [["T-CSV-TECH", "表4", "岩土工程勘察", "甲级", "1.2", "CSV", "高", "csv", "true", "false", "技术规则"]],
    )
    (tmp_path / "technical_fee_rules.xlsx").write_text("not a valid xlsx", encoding="utf-8")
    monkeypatch.setattr(adjustment_rules, "RULE_DIR", tmp_path)

    engine = adjustment_rules.AdjustmentEngine.from_default_rules()
    result = engine.evaluate(
        "表4-通用工程勘察费用",
        {"要素1": "岩土工程勘察", "要素2": "钻探", "要素3": "", "要素4": "", "要素5": "", "单位": "m"},
        ["泥浆护壁", "甲级"],
    )

    assert result.physical.value == "待复核"
    assert result.physical.status == "review"
    assert result.technical.value == 1.2
    assert "技术规则" in result.technical.message


def test_fill_engine_matches_answer_workbook(tmp_path):
    if INPUT_PATH is None or ANSWER_PATH is None:
        pytest.skip("缺少 100 行空单价输入表或答案表")
    output_path = tmp_path / "filled.xlsx"
    kb = KnowledgeBase.from_excel(KB_PATH)
    summary = FillEngine(kb).fill_workbook(INPUT_PATH, output_path)

    assert summary.total_data_rows == 100
    assert summary.filled_rows == 100
    assert summary.review_rows == 0
    assert summary.report_text == "输入100行，匹配100行。"
    price_index = summary.table_preview["headers"].index("基价测试列")
    status_index = summary.table_preview["headers"].index("匹配状态")
    assert summary.table_preview["rows"][0][price_index] == 17213
    assert summary.table_preview["rows"][0][status_index] == "已匹配"

    generated = load_workbook(output_path, data_only=True)
    answer = load_workbook(ANSWER_PATH, data_only=True)
    generated_ws = generated.active
    answer_ws = answer.active

    # Generated workbook keeps the input columns and fills the existing price column.
    generated_prices = [generated_ws.cell(row=i, column=7).value for i in range(2, 102)]
    answer_prices = [answer_ws.cell(row=i, column=9).value for i in range(2, 102)]

    assert [str(v) for v in generated_prices] == [str(v) for v in answer_prices]


def test_fill_engine_uses_user_column_mapping_for_shuffled_input(tmp_path):
    input_path = tmp_path / "shuffled.xlsx"
    output_path = tmp_path / "filled-shuffled.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["备注", "复杂程度", "价格输入", "专业", "单位列", "比例尺", "工作项", "空列"])
    sheet.append(["样例", "复杂", "空单价", "岩土工程勘察", "km2", "比例-1:500", "地质测绘", ""])
    sheet.append(["样例", "不存在", "空单价", "岩土工程勘察", "km2", "比例-1:500", "地质测绘", ""])
    workbook.save(input_path)

    kb = KnowledgeBase.from_excel(KB_PATH)
    summary = FillEngine(kb).fill_workbook(
        input_path,
        output_path,
        column_mapping={
            "要素1": "专业",
            "要素2": "工作项",
            "要素3": "空列",
            "要素4": "比例尺",
            "要素5": "复杂程度",
            "单位": "单位列",
            "价格列": "价格输入",
        },
    )

    assert summary.total_data_rows == 2
    assert summary.filled_rows == 1
    assert summary.review_rows == 1
    assert summary.price_column == "价格输入"

    generated = load_workbook(output_path)
    ws = generated.active
    headers = [cell.value for cell in ws[1]]

    assert ws.cell(row=2, column=3).value == 17213
    assert ws.cell(row=3, column=3).value == "待复核"
    assert ws.cell(row=2, column=3).fill.fill_type == "solid"
    assert ws.cell(row=2, column=3).fill.fgColor.rgb == "00C6EFCE"
    assert ws.cell(row=2, column=3).font.color.rgb == "00006100"
    assert ws.cell(row=3, column=3).fill.fill_type == "solid"
    assert ws.cell(row=3, column=3).fill.fgColor.rgb == "00FFC7CE"
    assert ws.cell(row=3, column=3).font.color.rgb == "009C0006"
    assert "匹配状态" in headers
    assert "候选数量" in headers
    assert "匹配说明" in headers

    status_column = headers.index("匹配状态") + 1
    candidate_column = headers.index("候选数量") + 1
    note_column = headers.index("匹配说明") + 1

    assert ws.cell(row=2, column=status_column).value == "已匹配"
    assert ws.cell(row=2, column=candidate_column).value == 1
    assert "匹配知识库第" in ws.cell(row=2, column=note_column).value
    assert ws.cell(row=3, column=status_column).value == "待复核"
    assert "没有匹配" in ws.cell(row=3, column=note_column).value


def test_fill_engine_uses_header_row_when_mapping_line_is_not_first_row(tmp_path):
    input_path = tmp_path / "header-row.xlsx"
    output_path = tmp_path / "filled-header-row.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["项目名称", "测试"])
    sheet.append(["说明", "非表头"])
    sheet.append(["价格输入", "专业", "工作项", "比例尺", "复杂程度", "单位列", "空列"])
    sheet.append(["空单价", "岩土工程勘察", "地质测绘", "比例-1:500", "复杂", "km2", ""])
    workbook.save(input_path)

    kb = KnowledgeBase.from_excel(KB_PATH)
    summary = FillEngine(kb).fill_workbook(
        input_path,
        output_path,
        column_mapping={
            "要素1": "B",
            "要素2": "C",
            "要素3": "G",
            "要素4": "D",
            "要素5": "E",
            "单位": "F",
            "价格列": "A",
        },
        header_row=3,
    )

    assert summary.total_data_rows == 1
    assert summary.filled_rows == 1
    assert summary.table_preview["headers"][0] == "价格输入"
    assert summary.table_preview["rows"][0][0] == 17213

    ws = load_workbook(output_path, data_only=True).active
    headers = [cell.value for cell in ws[3]]
    assert "匹配状态" in headers
    assert ws.cell(row=4, column=1).value == 17213
    assert ws.cell(row=4, column=headers.index("匹配状态") + 1).value == "已匹配"


def test_match_report_starts_at_column_r(tmp_path):
    input_path = tmp_path / "report-start.xlsx"
    output_path = tmp_path / "filled-report-start.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["价格输入", "要素1", "要素2", "要素3", "要素4", "要素5", "单位"])
    sheet.append(["空单价", "岩土工程勘察", "地质测绘", "", "比例-1:500", "复杂", "km2"])
    workbook.save(input_path)

    kb = KnowledgeBase.from_excel(KB_PATH)
    FillEngine(kb).fill_workbook(input_path, output_path)

    ws = load_workbook(output_path, data_only=True).active
    assert ws.cell(row=1, column=18).value == "匹配状态"
    assert ws.cell(row=1, column=19).value == "候选数量"
    assert ws.cell(row=1, column=20).value == "匹配说明"
    assert ws.cell(row=1, column=21).value == "匹配说明-实物工作费调整系数"
    assert ws.cell(row=1, column=22).value == "匹配说明-技术工作费调整系数"


def test_match_report_uses_reserved_position_before_column_r(tmp_path):
    input_path = tmp_path / "report-reserved.xlsx"
    output_path = tmp_path / "filled-report-reserved.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["价格输入", "要素1", "要素2", "要素3", "要素4", "要素5", "单位", "匹配报告预留位置"])
    sheet.append(["空单价", "岩土工程勘察", "地质测绘", "", "比例-1:500", "复杂", "km2", ""])
    workbook.save(input_path)

    kb = KnowledgeBase.from_excel(KB_PATH)
    FillEngine(kb).fill_workbook(input_path, output_path)

    ws = load_workbook(output_path, data_only=True).active
    assert ws.cell(row=1, column=9).value == "匹配状态"
    assert ws.cell(row=1, column=10).value == "候选数量"
    assert ws.cell(row=1, column=11).value == "匹配说明"
    assert ws.cell(row=1, column=12).value == "匹配说明-实物工作费调整系数"
    assert ws.cell(row=1, column=13).value == "匹配说明-技术工作费调整系数"
    assert ws.cell(row=1, column=18).value is None


def test_fill_engine_writes_output_adjustment_columns(tmp_path):
    input_path = tmp_path / "adjustment-columns.xlsx"
    output_path = tmp_path / "filled-adjustment-columns.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "表2-通用工程测量费用"
    sheet.append([
        "单价匹配-测试",
        "要素1",
        "要素2",
        "要素3",
        "要素4",
        "要素5",
        "单位",
        "实物工作费调整系数",
        "技术工作费调整系数",
    ])
    sheet.append(["空单价", "工程测量", "控制测量", "", "", "", "点", "", ""])
    workbook.save(input_path)

    kb = KnowledgeBase([])
    summary = FillEngine(kb).fill_workbook(
        input_path,
        output_path,
        column_mapping={
            "输出-价格列": "A",
            "要素1": "B",
            "要素2": "C",
            "要素3": "D",
            "要素4": "E",
            "要素5": "F",
            "单位": "G",
            "输出-实物工作费调整系数": "H",
            "输出-技术工作费调整系数": "I",
        },
    )

    assert summary.price_column == "A"
    ws = load_workbook(output_path, data_only=True).active
    assert ws.cell(row=2, column=8).value == "待复核"
    assert ws.cell(row=2, column=9).value == 0.22
    assert ws.cell(row=2, column=8).fill.fill_type == "solid"
    assert ws.cell(row=2, column=8).fill.fgColor.rgb == "00FFC7CE"
    assert ws.cell(row=2, column=9).fill.fill_type == "solid"
    assert ws.cell(row=2, column=9).fill.fgColor.rgb == "00C6EFCE"
    assert summary.physical_matched_rows == 0
    assert summary.physical_experience_rows == 0
    assert summary.physical_review_rows == 1
    assert summary.technical_matched_rows == 1
    assert summary.technical_experience_rows == 0
    assert summary.technical_review_rows == 0


def test_adjustment_cells_use_green_yellow_red_layer_colors(tmp_path):
    input_path = tmp_path / "adjustment-layer-colors.xlsx"
    output_path = tmp_path / "filled-adjustment-layer-colors.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "表4-通用工程勘察费用"
    sheet.append([
        "单价匹配-测试",
        "要素1",
        "要素2",
        "要素3",
        "要素4",
        "要素5",
        "单位",
        "备注",
        "技术工作类别",
        "实物工作费调整系数",
        "技术工作费调整系数",
    ])
    sheet.append(["空单价", "岩土工程勘察", "钻探", "", "", "", "m", "泥浆护壁", "甲级", "", ""])
    workbook.save(input_path)

    FillEngine(KnowledgeBase([])).fill_workbook(
        input_path,
        output_path,
        column_mapping={
            "输出-价格列": "A",
            "要素1": "B",
            "要素2": "C",
            "要素3": "D",
            "要素4": "E",
            "要素5": "F",
            "单位": "G",
            "输出-实物工作费调整系数": "J",
            "输出-技术工作费调整系数": "K",
        },
    )

    ws = load_workbook(output_path, data_only=True).active
    assert ws.cell(row=2, column=1).value == "待复核"
    assert ws.cell(row=2, column=1).fill.fgColor.rgb == "00FFC7CE"
    assert ws.cell(row=2, column=10).value == "待复核"
    assert ws.cell(row=2, column=10).fill.fgColor.rgb == "00FFC7CE"
    assert ws.cell(row=2, column=11).value == 1.2
    assert ws.cell(row=2, column=11).fill.fgColor.rgb == "00C6EFCE"


def test_fill_engine_uses_second_layer_experience_when_first_layer_misses(tmp_path):
    input_path = tmp_path / "adjustment-experience.xlsx"
    output_path = tmp_path / "filled-adjustment-experience.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "表4-通用工程勘察费用"
    sheet.append([
        "单价匹配-测试",
        "要素1",
        "要素2",
        "要素3",
        "要素4",
        "要素5",
        "单位",
        "实物工作费调整系数",
        "技术工作费调整系数",
    ])
    sheet.append(["空单价", "经验工程", "经验项目", "", "", "", "项", "", ""])
    workbook.save(input_path)

    kb = KnowledgeBase(
        [
            {
                "要素1": "经验工程",
                "要素2": "经验项目",
                "要素3": "",
                "要素4": "",
                "要素5": "",
                "单位": "项",
                "基价": 123,
                "【经验数】实物工作费调整系数": 1.35,
                "【经验数解释】-实物工作费调整系数": "实物经验说明",
                "【经验数】技术工作费调整系数": 0.66,
                "【经验数解释】-技术工作费调整系数": "技术经验说明",
                "_excel_row": 2,
            }
        ]
    )
    summary = FillEngine(kb).fill_workbook(
        input_path,
        output_path,
        column_mapping={
            "输出-价格列": "A",
            "要素1": "B",
            "要素2": "C",
            "要素3": "D",
            "要素4": "E",
            "要素5": "F",
            "单位": "G",
            "输出-实物工作费调整系数": "H",
            "输出-技术工作费调整系数": "I",
        },
    )

    ws = load_workbook(output_path, data_only=True).active
    assert ws.cell(row=2, column=1).value == 123
    assert ws.cell(row=2, column=8).value == 1.35
    assert ws.cell(row=2, column=9).value == 0.66
    assert ws.cell(row=2, column=8).fill.fill_type == "solid"
    assert ws.cell(row=2, column=8).fill.fgColor.rgb == "00FFF2CC"
    assert ws.cell(row=2, column=9).fill.fill_type == "solid"
    assert ws.cell(row=2, column=9).fill.fgColor.rgb == "00FFF2CC"

    headers = [cell.value for cell in ws[1]]
    physical_note = ws.cell(row=2, column=headers.index("匹配说明-实物工作费调整系数") + 1).value
    technical_note = ws.cell(row=2, column=headers.index("匹配说明-技术工作费调整系数") + 1).value
    assert "第二层经验提示层" in physical_note
    assert "实物经验说明" in physical_note
    assert "第二层经验提示层" in technical_note
    assert "技术经验说明" in technical_note
    assert summary.physical_matched_rows == 0
    assert summary.physical_experience_rows == 1
    assert summary.physical_review_rows == 0
    assert summary.technical_matched_rows == 0
    assert summary.technical_experience_rows == 1
    assert summary.technical_review_rows == 0


def test_fill_engine_uses_physical_experience_but_does_not_override_technical_first_layer(tmp_path):
    input_path = tmp_path / "adjustment-experience-no-override.xlsx"
    output_path = tmp_path / "filled-adjustment-experience-no-override.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "表4-通用工程勘察费用"
    sheet.append([
        "单价匹配-测试",
        "要素1",
        "要素2",
        "要素3",
        "要素4",
        "要素5",
        "单位",
        "备注",
        "技术工作类别",
        "实物工作费调整系数",
        "技术工作费调整系数",
    ])
    sheet.append(["空单价", "岩土工程勘察", "钻探", "", "", "", "m", "泥浆护壁", "甲级", "", ""])
    workbook.save(input_path)

    kb = KnowledgeBase(
        [
            {
                "要素1": "岩土工程勘察",
                "要素2": "钻探",
                "要素3": "",
                "要素4": "",
                "要素5": "",
                "单位": "m",
                "基价": 456,
                "【经验数】实物工作费调整系数": 9.99,
                "【经验数解释】-实物工作费调整系数": "实物第二层经验",
                "【经验数】技术工作费调整系数": 8.88,
                "【经验数解释】-技术工作费调整系数": "不应覆盖技术",
                "_excel_row": 2,
            }
        ]
    )
    FillEngine(kb).fill_workbook(
        input_path,
        output_path,
        column_mapping={
            "输出-价格列": "A",
            "要素1": "B",
            "要素2": "C",
            "要素3": "D",
            "要素4": "E",
            "要素5": "F",
            "单位": "G",
            "输出-实物工作费调整系数": "J",
            "输出-技术工作费调整系数": "K",
        },
    )

    ws = load_workbook(output_path, data_only=True).active
    assert ws.cell(row=2, column=10).value == 9.99
    assert ws.cell(row=2, column=11).value == 1.2
    assert ws.cell(row=2, column=10).fill.fgColor.rgb == "00FFF2CC"
    assert ws.cell(row=2, column=11).fill.fgColor.rgb == "00C6EFCE"


def test_adjustment_engine_writes_standard_physical_and_technical_coefficients(tmp_path):
    input_path = tmp_path / "adjustment-standard.xlsx"
    output_path = tmp_path / "filled-adjustment-standard.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "表4-通用工程勘察费用"
    sheet.append([
        "单价匹配-测试",
        "要素1",
        "要素2",
        "要素3",
        "要素4",
        "要素5",
        "单位",
        "备注",
        "技术工作类别",
        "实物工作费调整系数",
        "技术工作费调整系数",
    ])
    sheet.append(["空单价", "岩土工程勘察", "钻探", "", "", "", "m", "泥浆护壁", "甲级", "", ""])
    sheet.append(["空单价", "岩土工程勘察", "钻探", "", "", "", "m", "泥浆护壁、夜间作业", "乙级", "", ""])
    workbook.save(input_path)

    FillEngine(KnowledgeBase([])).fill_workbook(
        input_path,
        output_path,
        column_mapping={
            "输出-价格列": "A",
            "要素1": "B",
            "要素2": "C",
            "要素3": "D",
            "要素4": "E",
            "要素5": "F",
            "单位": "G",
            "输出-实物工作费调整系数": "J",
            "输出-技术工作费调整系数": "K",
        },
    )

    ws = load_workbook(output_path, data_only=True).active
    assert ws.cell(row=2, column=10).value == "待复核"
    assert ws.cell(row=2, column=11).value == 1.2
    assert ws.cell(row=3, column=10).value == "待复核"
    assert ws.cell(row=3, column=11).value == 1
    assert ws.cell(row=2, column=10).fill.fgColor.rgb == "00FFC7CE"
    assert ws.cell(row=2, column=11).fill.fgColor.rgb == "00C6EFCE"
    assert ws.cell(row=3, column=10).fill.fgColor.rgb == "00FFC7CE"
    assert ws.cell(row=3, column=11).fill.fgColor.rgb == "00C6EFCE"
    headers = [cell.value for cell in ws[1]]
    physical_note_col = headers.index("匹配说明-实物工作费调整系数") + 1
    technical_note_col = headers.index("匹配说明-技术工作费调整系数") + 1
    row2_technical_note = ws.cell(row=2, column=technical_note_col).value
    row2_physical_note = ws.cell(row=2, column=physical_note_col).value

    assert "第一层规则暂未启用" in row2_physical_note
    assert "T-T4-GEO-A" in row2_technical_note
    assert "甲级" in row2_technical_note
    assert "系数 1.2" in row2_technical_note


def test_adjustment_engine_writes_table3_geological_mapping_coefficients(tmp_path):
    input_path = tmp_path / "adjustment-table3.xlsx"
    output_path = tmp_path / "filled-adjustment-table3.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "表3-地质测绘"
    sheet.append([
        "单价匹配-测试",
        "要素1",
        "要素2",
        "要素3",
        "要素4",
        "要素5",
        "单位",
        "备注",
        "技术工作类别",
        "实物工作费调整系数",
        "技术工作费调整系数",
    ])
    sheet.append(["空单价", "岩土工程勘察", "地质测绘", "", "比例-1:500", "复杂", "km2", "带状工程地质测绘", "甲级", "", ""])
    workbook.save(input_path)

    FillEngine(KnowledgeBase([])).fill_workbook(
        input_path,
        output_path,
        column_mapping={
            "输出-价格列": "A",
            "要素1": "B",
            "要素2": "C",
            "要素3": "D",
            "要素4": "E",
            "要素5": "F",
            "单位": "G",
            "输出-实物工作费调整系数": "J",
            "输出-技术工作费调整系数": "K",
        },
    )

    ws = load_workbook(output_path, data_only=True).active
    assert ws.cell(row=2, column=10).value == "待复核"
    assert ws.cell(row=2, column=11).value == 1.2
    assert ws.cell(row=2, column=10).fill.fgColor.rgb == "00FFC7CE"
    assert ws.cell(row=2, column=11).fill.fgColor.rgb == "00C6EFCE"


def test_adjustment_engine_uses_table2_standard_physical_factor_from_manual(tmp_path):
    input_path = tmp_path / "adjustment-table2-manual.xlsx"
    output_path = tmp_path / "filled-adjustment-table2-manual.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "表2-通用工程测量费用"
    sheet.append([
        "单价匹配-测试",
        "要素1",
        "要素2",
        "要素3",
        "要素4",
        "要素5",
        "单位",
        "备注",
        "实物工作费调整系数",
        "技术工作费调整系数",
    ])
    sheet.append(["空单价", "工程测量", "地形测量", "", "", "", "km2", "数字化测绘", "", ""])
    workbook.save(input_path)

    FillEngine(KnowledgeBase([])).fill_workbook(
        input_path,
        output_path,
        column_mapping={
            "输出-价格列": "A",
            "要素1": "B",
            "要素2": "C",
            "要素3": "D",
            "要素4": "E",
            "要素5": "F",
            "单位": "G",
            "输出-实物工作费调整系数": "I",
            "输出-技术工作费调整系数": "J",
        },
    )

    ws = load_workbook(output_path, data_only=True).active
    assert ws.cell(row=2, column=9).value == "待复核"
    assert ws.cell(row=2, column=10).value == 0.22
    assert ws.cell(row=2, column=9).fill.fgColor.rgb == "00FFC7CE"
    assert ws.cell(row=2, column=10).fill.fgColor.rgb == "00C6EFCE"
    headers = [cell.value for cell in ws[1]]
    physical_note = ws.cell(row=2, column=headers.index("匹配说明-实物工作费调整系数") + 1).value
    assert "第一层规则暂未启用" in physical_note


def test_adjustment_engine_uses_2009_survey_deduction_with_standard_source(tmp_path):
    input_path = tmp_path / "adjustment-2009-survey.xlsx"
    output_path = tmp_path / "filled-adjustment-2009-survey.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "表2-通用工程测量费用"
    sheet.append([
        "单价匹配-测试",
        "要素1",
        "要素2",
        "要素3",
        "要素4",
        "要素5",
        "单位",
        "备注",
        "实物工作费调整系数",
        "技术工作费调整系数",
    ])
    sheet.append(["空单价", "工程测量", "像控点连测", "", "", "", "幅", "航线网布点核减15%", "", ""])
    workbook.save(input_path)

    FillEngine(KnowledgeBase([])).fill_workbook(
        input_path,
        output_path,
        column_mapping={
            "输出-价格列": "A",
            "要素1": "B",
            "要素2": "C",
            "要素3": "D",
            "要素4": "E",
            "要素5": "F",
            "单位": "G",
            "输出-实物工作费调整系数": "I",
            "输出-技术工作费调整系数": "J",
        },
    )

    ws = load_workbook(output_path, data_only=True).active
    assert ws.cell(row=2, column=9).value == "待复核"
    assert ws.cell(row=2, column=9).fill.fgColor.rgb == "00FFC7CE"
    headers = [cell.value for cell in ws[1]]
    physical_note = ws.cell(row=2, column=headers.index("匹配说明-实物工作费调整系数") + 1).value
    assert "第一层规则暂未启用" in physical_note


def test_adjustment_engine_combines_general_temperature_and_altitude_factors(tmp_path):
    input_path = tmp_path / "adjustment-general-natural.xlsx"
    output_path = tmp_path / "filled-adjustment-general-natural.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "表4-通用工程勘察费用"
    sheet.append([
        "单价匹配-测试",
        "要素1",
        "要素2",
        "要素3",
        "要素4",
        "要素5",
        "单位",
        "备注",
        "技术工作类别",
        "实物工作费调整系数",
        "技术工作费调整系数",
    ])
    sheet.append(["空单价", "岩土工程勘察", "钻探", "", "", "", "m", "气温>=35℃，海拔3001~3500m", "甲级", "", ""])
    workbook.save(input_path)

    FillEngine(KnowledgeBase([])).fill_workbook(
        input_path,
        output_path,
        column_mapping={
            "输出-价格列": "A",
            "要素1": "B",
            "要素2": "C",
            "要素3": "D",
            "要素4": "E",
            "要素5": "F",
            "单位": "G",
            "输出-实物工作费调整系数": "J",
            "输出-技术工作费调整系数": "K",
        },
    )

    ws = load_workbook(output_path, data_only=True).active
    assert ws.cell(row=2, column=10).value == "待复核"
    assert ws.cell(row=2, column=11).value == 1.2
    assert ws.cell(row=2, column=10).fill.fgColor.rgb == "00FFC7CE"
    assert ws.cell(row=2, column=11).fill.fgColor.rgb == "00C6EFCE"
    headers = [cell.value for cell in ws[1]]
    physical_note = ws.cell(row=2, column=headers.index("匹配说明-实物工作费调整系数") + 1).value
    assert "第一层规则暂未启用" in physical_note


def test_adjustment_engine_writes_first_layer_zero_technical_fee_for_survey_point(tmp_path):
    input_path = tmp_path / "adjustment-review.xlsx"
    output_path = tmp_path / "filled-adjustment-review.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "表4-通用工程勘察费用"
    sheet.append([
        "单价匹配-测试",
        "要素1",
        "要素2",
        "要素3",
        "要素4",
        "要素5",
        "单位",
        "备注",
        "实物工作费调整系数",
        "技术工作费调整系数",
    ])
    sheet.append(["空单价", "岩土工程勘察", "勘探点测放", "", "", "", "点", "", "", ""])
    workbook.save(input_path)

    FillEngine(KnowledgeBase([])).fill_workbook(
        input_path,
        output_path,
        column_mapping={
            "输出-价格列": "A",
            "要素1": "B",
            "要素2": "C",
            "要素3": "D",
            "要素4": "E",
            "要素5": "F",
            "单位": "G",
            "输出-实物工作费调整系数": "I",
            "输出-技术工作费调整系数": "J",
        },
    )

    ws = load_workbook(output_path, data_only=True).active
    assert ws.cell(row=2, column=9).value == "待复核"
    assert ws.cell(row=2, column=10).value == 0
    assert ws.cell(row=2, column=9).fill.fgColor.rgb == "00FFC7CE"
    assert ws.cell(row=2, column=10).fill.fgColor.rgb == "00C6EFCE"
    headers = [cell.value for cell in ws[1]]
    technical_note_col = headers.index("匹配说明-技术工作费调整系数") + 1
    technical_note = ws.cell(row=2, column=technical_note_col).value
    assert "输出系数 0" in technical_note
    assert "勘探点测放" in technical_note
    assert "T-T4-POINT-000" in technical_note


def test_fill_engine_processes_selected_candidate_sheets(tmp_path):
    input_path = tmp_path / "multi-sheet.xlsx"
    output_path = tmp_path / "filled-multi-sheet.xlsx"

    workbook = Workbook()
    ws1 = workbook.active
    ws1.title = "表2 测量"
    ws1.append(["单价匹配-测试", "要素1", "要素2", "要素4", "要素5", "单位"])
    ws1.append(["空单价", "岩土工程勘察", "地质测绘", "比例-1:500", "复杂", "km2"])
    ws2 = workbook.create_sheet("表3 跳过")
    ws2.append(["单价匹配-测试", "要素1", "要素2", "要素4", "要素5", "单位"])
    ws2.append(["空单价", "岩土工程勘察", "地质测绘", "比例-1:500", "复杂", "km2"])
    workbook.save(input_path)

    kb = KnowledgeBase.from_excel(KB_PATH)
    summary = FillEngine(kb).fill_workbook(
        input_path,
        output_path,
        sheet_configs=[
            {
                "sheet_name": "表2 测量",
                "enabled": True,
                "header_row": 1,
                "column_mapping": {
                    "输出-价格列": "A",
                    "要素1": "B",
                    "要素2": "C",
                    "要素3": "空元素列",
                    "要素4": "D",
                    "要素5": "E",
                    "单位": "F",
                },
            },
            {
                "sheet_name": "表3 跳过",
                "enabled": False,
                "header_row": 1,
                "column_mapping": {
                    "输出-价格列": "A",
                    "要素1": "B",
                    "要素2": "C",
                    "要素3": "空元素列",
                    "要素4": "D",
                    "要素5": "E",
                    "单位": "F",
                },
            },
        ],
    )

    assert summary.total_data_rows == 1
    assert summary.filled_rows == 1
    workbook = load_workbook(output_path, data_only=True)
    assert workbook["表2 测量"].cell(row=2, column=1).value == 17213
    assert workbook["表3 跳过"].cell(row=2, column=1).value == "空单价"


def test_table_preview_includes_enabled_sheets_and_fifty_rows(tmp_path):
    input_path = tmp_path / "preview-multi-sheet.xlsx"
    output_path = tmp_path / "filled-preview-multi-sheet.xlsx"

    workbook = Workbook()
    for index, sheet_name in enumerate(["表2 预览", "表3 预览"]):
        sheet = workbook.active if index == 0 else workbook.create_sheet(sheet_name)
        sheet.title = sheet_name
        sheet.append(["价格输入", "要素1", "要素2", "要素3", "要素4", "要素5", "单位"])
        for row_index in range(60):
            sheet.append(["空单价", "测试工程", f"测试项目{row_index}", "", "", "", "项"])
    workbook.save(input_path)

    kb = KnowledgeBase(
        [
            {
                "要素1": "测试工程",
                "要素2": f"测试项目{row_index}",
                "要素3": "",
                "要素4": "",
                "要素5": "",
                "单位": "项",
                "基价": row_index,
                "_excel_row": row_index + 2,
            }
            for row_index in range(60)
        ]
    )
    summary = FillEngine(kb).fill_workbook(
        input_path,
        output_path,
        sheet_configs=[
            {
                "sheet_name": "表2 预览",
                "enabled": True,
                "header_row": 1,
                "column_mapping": {
                    "输出-价格列": "A",
                    "要素1": "B",
                    "要素2": "C",
                    "要素3": "D",
                    "要素4": "E",
                    "要素5": "F",
                    "单位": "G",
                },
            },
            {
                "sheet_name": "表3 预览",
                "enabled": True,
                "header_row": 1,
                "column_mapping": {
                    "输出-价格列": "A",
                    "要素1": "B",
                    "要素2": "C",
                    "要素3": "D",
                    "要素4": "E",
                    "要素5": "F",
                    "单位": "G",
                },
            },
        ],
    )

    assert [sheet["sheet_name"] for sheet in summary.table_preview["sheets"]] == ["表2 预览", "表3 预览"]
    assert len(summary.table_preview["sheets"][0]["rows"]) == 50
    assert len(summary.table_preview["sheets"][1]["rows"]) == 50
    assert summary.table_preview["rows"] == summary.table_preview["sheets"][0]["rows"]


def test_formula_cached_value_is_used_for_matching_project_example_f51(tmp_path):
    output_path = tmp_path / "filled-formula-f51.xlsx"

    kb = KnowledgeBase.from_excel(KB_PATH)
    summary = FillEngine(kb).fill_workbook(
        PROJECT_EXAMPLE_PATH,
        output_path,
        sheet_configs=[
            {
                    "sheet_name": "表4-通用工程勘察费用",
                    "enabled": True,
                    "header_row": 4,
                    "column_mapping": {
                        "输出-价格列": "AC",
                        "要素1": "B",
                        "要素2": "C",
                        "要素3": "D",
                        "要素4": "E",
                        "要素5": "F",
                    "单位": "G",
                },
            }
        ],
    )

    assert summary.matched_rows >= 158
    workbook = load_workbook(output_path, data_only=True)
    ws = workbook["表4-通用工程勘察费用"]
    assert ws["AC51"].value != "待复核"
    headers = [cell.value for cell in ws[4]]
    assert ws.cell(row=51, column=headers.index("匹配状态") + 1).value == "已匹配"


def test_technical_fee_uses_project_example_technical_column_for_hydro_row90(tmp_path):
    output_path = tmp_path / "filled-hydro-row90.xlsx"

    kb = KnowledgeBase.from_excel(KB_PATH)
    FillEngine(kb).fill_workbook(
        PROJECT_EXAMPLE_PATH,
        output_path,
        sheet_configs=[
            {
                "sheet_name": "表4-通用工程勘察费用",
                "enabled": True,
                "header_row": 4,
                "column_mapping": {
                    "输出-价格列": "AC",
                    "要素1": "B",
                    "要素2": "C",
                    "要素3": "D",
                    "要素4": "E",
                    "要素5": "F",
                    "单位": "G",
                    "输出-技术工作费调整系数": "AE",
                },
            }
        ],
    )

    workbook = load_workbook(output_path, data_only=True)
    ws = workbook["表4-通用工程勘察费用"]
    merged_value_map = FillEngine._build_merged_value_map(ws)
    assert FillEngine._read_mapped_value(ws, 90, 2, merged_value_map) == "水文地质勘察"
    assert FillEngine._read_mapped_value(ws, 90, 5, merged_value_map) == "利用钻孔抽水"
    assert FillEngine._read_mapped_value(ws, 90, 15, merged_value_map) == "复杂"
    assert ws["AE90"].value == 0.33


def test_input_chart_scale_placeholder_is_treated_as_blank(tmp_path):
    input_path = tmp_path / "chart-scale-placeholder.xlsx"
    output_path = tmp_path / "filled-chart-scale-placeholder.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["价格输入", "要素1", "要素2", "要素3", "要素4", "要素5", "单位"])
    sheet.append(["空单价", "测试工程", "测试项目", "成图比例", "", "简单", "项"])
    workbook.save(input_path)

    kb = KnowledgeBase(
        [
            {
                "要素1": "测试工程",
                "要素2": "测试项目",
                "要素3": "",
                "要素4": "",
                "要素5": "简单",
                "单位": "项",
                "基价": 321,
                "_excel_row": 2,
            }
        ]
    )
    summary = FillEngine(kb).fill_workbook(input_path, output_path)

    assert summary.filled_rows == 1
    ws = load_workbook(output_path, data_only=True).active
    assert ws.cell(row=2, column=1).value == 321


def test_match_report_can_be_disabled(tmp_path):
    input_path = tmp_path / "report-disabled.xlsx"
    output_path = tmp_path / "filled-report-disabled.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["价格输入", "要素1", "要素2", "要素3", "要素4", "要素5", "单位"])
    sheet.append(["空单价", "岩土工程勘察", "地质测绘", "", "比例-1:500", "复杂", "km2"])
    workbook.save(input_path)

    kb = KnowledgeBase.from_excel(KB_PATH)
    FillEngine(kb).fill_workbook(input_path, output_path, output_match_report=False)

    ws = load_workbook(output_path, data_only=True).active
    assert ws.cell(row=1, column=18).value is None


def test_fill_engine_treats_empty_element_column_as_blank_value(tmp_path):
    input_path = tmp_path / "empty-element.xlsx"
    output_path = tmp_path / "filled-empty-element.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["价格输入", "专业", "工作项", "比例尺", "复杂程度", "单位列"])
    sheet.append(["空单价", "岩土工程勘察", "地质测绘", "比例-1:500", "复杂", "km2"])
    workbook.save(input_path)

    kb = KnowledgeBase.from_excel(KB_PATH)
    summary = FillEngine(kb).fill_workbook(
        input_path,
        output_path,
        column_mapping={
            "要素1": "B",
            "要素2": "C",
            "要素3": "空元素列",
            "要素4": "D",
            "要素5": "E",
            "单位": "F",
            "价格列": "A",
        },
    )

    assert summary.filled_rows == 1
    assert summary.review_rows == 0
    assert summary.table_preview["rows"][0][0] == 17213


def test_fill_engine_reads_merged_cell_values_for_all_rows(tmp_path):
    input_path = tmp_path / "merged-values.xlsx"
    output_path = tmp_path / "filled-merged-values.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["价格输入", "专业", "工作项", "比例尺", "复杂程度", "单位列", "空列"])
    sheet.append(["空单价", "岩土工程勘察", "地质测绘", "比例-1:500", "复杂", "km2", ""])
    sheet.append(["空单价", None, "地质测绘", "比例-1:500", "复杂", "km2", ""])
    sheet.merge_cells("B2:B3")
    workbook.save(input_path)

    kb = KnowledgeBase.from_excel(KB_PATH)
    summary = FillEngine(kb).fill_workbook(
        input_path,
        output_path,
        column_mapping={
            "要素1": "B",
            "要素2": "C",
            "要素3": "空元素列",
            "要素4": "D",
            "要素5": "E",
            "单位": "F",
            "价格列": "A",
        },
    )

    assert summary.filled_rows == 2
    assert summary.review_rows == 0

    ws = load_workbook(output_path, data_only=True).active
    assert ws.cell(row=2, column=1).value == 17213
    assert ws.cell(row=3, column=1).value == 17213


def test_horizontal_merged_cells_keep_only_first_column_value_by_default(tmp_path):
    input_path = tmp_path / "horizontal-merged-values.xlsx"
    output_path = tmp_path / "filled-horizontal-merged-values.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["价格输入", "要素1", "要素2", "要素3", "要素4", "要素5", "单位"])
    sheet.append(["空单价", "测试工程", "横向值", None, "", "", "项"])
    sheet.merge_cells("C2:D2")
    workbook.save(input_path)

    kb = KnowledgeBase(
        [
            {
                "要素1": "测试工程",
                "要素2": "横向值",
                "要素3": "",
                "要素4": "",
                "要素5": "",
                "单位": "项",
                "基价": 88,
                "_excel_row": 2,
            }
        ]
    )
    summary = FillEngine(kb).fill_workbook(input_path, output_path)

    assert summary.filled_rows == 1
    assert summary.review_rows == 0
    ws = load_workbook(output_path, data_only=True).active
    assert ws.cell(row=2, column=1).value == 88


def test_vertical_merged_cell_inheritance_can_be_disabled(tmp_path):
    input_path = tmp_path / "vertical-merged-disabled.xlsx"
    output_path = tmp_path / "filled-vertical-merged-disabled.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["价格输入", "要素1", "要素2", "要素3", "要素4", "要素5", "单位"])
    sheet.append(["空单价", "测试工程", "测试项目", "", "", "", "项"])
    sheet.append(["空单价", "测试工程", None, "", "", "", "项"])
    sheet.merge_cells("C2:C3")
    workbook.save(input_path)

    kb = KnowledgeBase(
        [
            {
                "要素1": "测试工程",
                "要素2": "测试项目",
                "要素3": "",
                "要素4": "",
                "要素5": "",
                "单位": "项",
                "基价": 66,
                "_excel_row": 2,
            }
        ]
    )
    summary = FillEngine(kb).fill_workbook(
        input_path,
        output_path,
        merge_vertical_cells=False,
    )

    assert summary.filled_rows == 1
    assert summary.review_rows == 1
    ws = load_workbook(output_path, data_only=True).active
    assert ws.cell(row=2, column=1).value == 66
    assert ws.cell(row=3, column=1).value == "待复核"


def test_rectangular_merged_cells_only_inherit_down_first_column_by_default(tmp_path):
    input_path = tmp_path / "rectangular-merged-values.xlsx"
    output_path = tmp_path / "filled-rectangular-merged-values.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["价格输入", "要素1", "要素2", "要素3", "要素4", "要素5", "单位"])
    sheet.append(["空单价", "矩形值", None, None, "", "", "项"])
    sheet.append(["空单价", None, None, None, "", "", "项"])
    sheet.merge_cells("B2:D3")
    workbook.save(input_path)

    kb = KnowledgeBase(
        [
            {
                "要素1": "矩形值",
                "要素2": "",
                "要素3": "",
                "要素4": "",
                "要素5": "",
                "单位": "项",
                "基价": 77,
                "_excel_row": 2,
            }
        ]
    )
    summary = FillEngine(kb).fill_workbook(input_path, output_path)

    assert summary.filled_rows == 2
    assert summary.review_rows == 0
    ws = load_workbook(output_path, data_only=True).active
    assert ws.cell(row=2, column=1).value == 77
    assert ws.cell(row=3, column=1).value == 77


def test_ordered_non_empty_elements_match_when_empty_element_moves(tmp_path):
    input_path = tmp_path / "ordered-elements.xlsx"
    output_path = tmp_path / "filled-ordered-elements.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["价格输入", "要素1", "要素2", "要素3", "要素4", "要素5", "单位"])
    sheet.append(["空单价", "岩土工程勘察", "地质测绘", "比例-1:500", "复杂", "", "km2"])
    workbook.save(input_path)

    kb = KnowledgeBase.from_excel(KB_PATH)
    summary = FillEngine(kb).fill_workbook(input_path, output_path)

    assert summary.filled_rows == 1
    assert summary.review_rows == 0
    assert summary.table_preview["rows"][0][0] == 17213


def test_level_prefix_alias_matches_knowledge_base_value(tmp_path):
    kb = KnowledgeBase(
        [
            {
                "要素1": "测试工程",
                "要素2": "级别-Ⅰ类",
                "要素3": "",
                "要素4": "",
                "要素5": "",
                "单位": "项",
                "基价": 99,
                "_excel_row": 12,
            }
        ]
    )

    result = kb.lookup(
        {
            "要素1": "测试工程",
            "要素2": "Ⅰ类",
            "要素3": "",
            "要素4": "",
            "要素5": "",
            "单位": "项",
        }
    )

    assert result.status == "matched"
    assert result.price == 99


def test_fill_engine_skips_blank_subtotal_and_total_rows_without_report(tmp_path):
    input_path = tmp_path / "ignored-rows.xlsx"
    output_path = tmp_path / "filled-ignored-rows.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["价格输入", "要素1", "要素2", "要素3", "要素4", "要素5", "单位"])
    sheet.append(["空单价", "", "地质测绘", "", "比例-1:500", "复杂", "km2"])
    sheet.append(["空单价", "小计", "地质测绘", "", "比例-1:500", "复杂", "km2"])
    sheet.append(["空单价", "合计", "地质测绘", "", "比例-1:500", "复杂", "km2"])
    sheet.append(["空单价", "岩土工程勘察", "地质测绘", "", "比例-1:500", "复杂", "km2"])
    workbook.save(input_path)

    kb = KnowledgeBase.from_excel(KB_PATH)
    summary = FillEngine(kb).fill_workbook(input_path, output_path)

    assert summary.total_data_rows == 1
    assert summary.filled_rows == 1
    assert summary.review_rows == 0
    assert summary.report_text == "输入1行，匹配1行。"
    assert len(summary.price_logs) == 1

    ws = load_workbook(output_path, data_only=True).active
    assert ws.cell(row=2, column=1).value == "空单价"
    assert ws.cell(row=3, column=1).value == "空单价"
    assert ws.cell(row=4, column=1).value == "空单价"
    assert ws.cell(row=5, column=1).value == 17213
    assert ws.cell(row=2, column=18).value is None
    assert ws.cell(row=3, column=18).value is None
    assert ws.cell(row=4, column=18).value is None
    assert ws.cell(row=5, column=18).value == "已匹配"


def test_fill_engine_skips_rows_without_selected_filter_value(tmp_path):
    input_path = tmp_path / "quantity-filter.xlsx"
    output_path = tmp_path / "filled-quantity-filter.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["价格输入", "要素1", "要素2", "要素3", "要素4", "要素5", "单位", "数量"])
    sheet.append(["空单价", "岩土工程勘察", "地质测绘", "", "比例-1:500", "复杂", "km2", ""])
    sheet.append(["空单价", "岩土工程勘察", "地质测绘", "", "比例-1:500", "复杂", "km2", 0])
    sheet.append(["空单价", "岩土工程勘察", "地质测绘", "", "比例-1:500", "复杂", "km2", "=0"])
    sheet.append(["空单价", "岩土工程勘察", "地质测绘", "", "比例-1:500", "复杂", "km2", 3])
    workbook.save(input_path)

    kb = KnowledgeBase.from_excel(KB_PATH)
    summary = FillEngine(kb).fill_workbook(
        input_path,
        output_path,
        only_match_rows_with_value=True,
        match_value_filter_field="数量",
    )

    assert summary.total_data_rows == 1
    assert summary.filled_rows == 1
    assert summary.review_rows == 0
    assert summary.report_text == "输入1行，匹配1行。"

    ws = load_workbook(output_path, data_only=True).active
    assert ws.cell(row=2, column=1).value == "空单价"
    assert ws.cell(row=3, column=1).value == "空单价"
    assert ws.cell(row=4, column=1).value == "空单价"
    assert ws.cell(row=5, column=1).value == 17213
    assert ws.cell(row=2, column=18).value is None
    assert ws.cell(row=3, column=18).value is None
    assert ws.cell(row=4, column=18).value is None
    assert ws.cell(row=5, column=18).value == "已匹配"


def test_fill_engine_requires_selected_filter_column_when_enabled(tmp_path):
    input_path = tmp_path / "missing-quantity.xlsx"
    output_path = tmp_path / "filled-missing-quantity.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["价格输入", "要素1", "要素2", "要素3", "要素4", "要素5", "单位"])
    sheet.append(["空单价", "岩土工程勘察", "地质测绘", "", "比例-1:500", "复杂", "km2"])
    workbook.save(input_path)

    kb = KnowledgeBase.from_excel(KB_PATH)
    with pytest.raises(ValueError, match="未找到指定列：数量"):
        FillEngine(kb).fill_workbook(
            input_path,
            output_path,
            only_match_rows_with_value=True,
            match_value_filter_field="数量",
        )


def test_match_report_marks_not_found_red_and_review_yellow(tmp_path):
    input_path = tmp_path / "report-colors.xlsx"
    output_path = tmp_path / "filled-report-colors.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["价格输入", "要素1", "要素2", "要素3", "要素4", "要素5", "单位"])
    sheet.append(["空单价", "没有这项", "", "", "", "", "项"])
    sheet.append(["空单价", "冲突工程", "冲突项目", "", "", "", "项"])
    workbook.save(input_path)

    kb = KnowledgeBase(
        [
            {
                "要素1": "冲突工程",
                "要素2": "冲突项目",
                "要素3": "",
                "要素4": "",
                "要素5": "",
                "单位": "项",
                "基价": 100,
                "_excel_row": 2,
            },
            {
                "要素1": "冲突工程",
                "要素2": "冲突项目",
                "要素3": "",
                "要素4": "",
                "要素5": "",
                "单位": "项",
                "基价": 200,
                "_excel_row": 3,
            },
        ]
    )
    summary = FillEngine(kb).fill_workbook(input_path, output_path)

    assert summary.review_rows == 2
    assert summary.conflict_rows == 1
    ws = load_workbook(output_path).active
    headers = [cell.value for cell in ws[1]]
    status_column = headers.index("匹配状态") + 1
    note_column = headers.index("匹配说明") + 1

    assert ws.cell(row=2, column=status_column).value == "待复核"
    assert ws.cell(row=2, column=status_column).fill.fgColor.rgb == "00FFC7CE"
    assert ws.cell(row=2, column=note_column).fill.fgColor.rgb == "00FFC7CE"
    assert ws.cell(row=3, column=status_column).value == "待复核"
    assert ws.cell(row=3, column=status_column).fill.fgColor.rgb == "00FFF2CC"
    assert ws.cell(row=3, column=note_column).fill.fgColor.rgb == "00FFF2CC"
