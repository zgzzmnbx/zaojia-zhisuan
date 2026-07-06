from __future__ import annotations

from pathlib import Path
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook

from app import workload_term_rules as workload_term_rules_module
from app.workload_capture import (
    DEFAULT_SELECTED_WORKLOAD_FIELDS,
    SOURCE_DIAGNOSTIC_FIELD,
    TARGET_QUANTITY_FIELD,
    WRITE_MODE_CONSERVATIVE,
    WRITE_MODE_OVERWRITE,
    capture_workload,
    suggest_workload_column_mapping,
)


def _write_source(path: Path, duplicate: bool = False) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "委托方工作量"
    sheet.append(["项目", "内容", "类别", "比例尺", "单位", "数量", "实物工作费调整系数", "技术工作费调整系数", "备注"])
    sheet.append(["控制测量", "首级控制测量", "GPS测量C级", "中等", "个", 26, 0.6, 0.22, "起算点"])
    if duplicate:
        sheet.append(["控制测量", "首级控制测量", "GPS测量C级", "中等", "个", 30, 0.6, 0.22, "重复"])
    else:
        sheet.append(["地形测量", "站场地形图", "复杂", "1:500", "km2", 0.5, 1.5, 0.22, "未匹配"])
    workbook.save(path)
    workbook.close()


def _write_target(path: Path, duplicate: bool = False) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "表2 通用工程测量费用"
    sheet.append(["要素1", "要素2", "要素3", "要素4", "单位", TARGET_QUANTITY_FIELD, "实物工作费调整系数(信息抓取)", "技术工作费调整系数(信息抓取)", "委托方备注(信息抓取)", "抓取日志"])
    sheet.append(["控制测量", "首级控制测量", "GPS测量C级", "中等", "个", None, None, None, None, None])
    if duplicate:
        sheet.append(["控制测量", "首级控制测量", "GPS测量C级", "中等", "个", None, None, None, None, None])
    workbook.save(path)
    workbook.close()


def _write_target_unit_mismatch(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "表2 通用工程测量费用"
    sheet.append(["要素1", "要素2", "要素3", "要素4", "单位", TARGET_QUANTITY_FIELD, "抓取日志"])
    sheet.append(["控制测量", "首级控制测量", "GPS测量C级", "中等", "个", None, None])
    workbook.save(path)
    workbook.close()


def _source_configs() -> list[dict[str, object]]:
    return [
        {
            "sheet_name": "委托方工作量",
            "enabled": True,
            "header_row": 1,
            "column_mapping": {
                "要素1": "A",
                "要素2": "B",
                "要素3": "C",
                "要素4": "D",
                "单位": "E",
                "数量": "F",
                "实物工作费调整系数": "G",
                "技术工作费调整系数": "H",
                "委托方备注": "I",
            },
        }
    ]


def _target_configs() -> list[dict[str, object]]:
    return [
        {
            "sheet_name": "表2 通用工程测量费用",
            "enabled": True,
            "header_row": 1,
            "column_mapping": {
                "要素1": "A",
                "要素2": "B",
                "要素3": "C",
                "要素4": "D",
                "单位": "E",
                "数量(信息抓取)": "F",
                "实物工作费调整系数(信息抓取)": "G",
                "技术工作费调整系数(信息抓取)": "H",
                "委托方备注(信息抓取)": "I",
                "抓取日志": "J",
            },
        }
    ]


def _write_term_rules(path: Path, *, enable_unit_alias: bool = True, enable_all_field_alias: bool = True) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    unit_sheet = workbook.create_sheet("单位严格等价")
    unit_sheet.append(["单位严格等价"])
    unit_sheet.append(["说明", "测试"])
    unit_sheet.append(["推荐口径", "测试"])
    unit_sheet.append([])
    unit_sheet.append([])
    unit_sheet.append(["启用", "原值", "归并值", "是否双向", "生效模块", "匹配强度", "建议层级", "备注"])
    unit_sheet.append(["是" if enable_unit_alias else "否", "个", "点", "是", "全部", "严格等价", "正式匹配", "测试"])

    exact_sheet = workbook.create_sheet("要素严格等价")
    exact_sheet.append(["要素严格等价"])
    exact_sheet.append(["说明", "测试"])
    exact_sheet.append(["建议口径", "测试"])
    exact_sheet.append([])
    exact_sheet.append([])
    exact_sheet.append(["启用", "字段名", "原值", "归并值", "是否双向", "生效模块", "匹配强度", "建议层级", "备注"])
    exact_sheet.append(["是" if enable_all_field_alias else "否", None, "首级控制测量", "控制测量", "是", "全部", "严格等价", "正式匹配", "测试"])
    exact_sheet.append(["是", None, None, None, None, None, None, None, "不完整规则行应忽略"])

    weak_sheet = workbook.create_sheet("要素弱等价")
    weak_sheet.append(["要素弱等价"])
    weak_sheet.append(["说明", "测试"])
    weak_sheet.append(["推荐口径", "测试"])
    weak_sheet.append([])
    weak_sheet.append([])
    weak_sheet.append(["启用", "字段名", "上位值", "下位值", "匹配方向", "生效模块", "建议层级", "是否允许自动写值", "备注", "维护人说明"])

    workbook.create_sheet("使用说明")
    workbook.create_sheet("模块作用域说明")
    workbook.save(path)
    workbook.close()


def _write_formula_cached_value(path: Path, cell_ref: str, formula: str, cached_value: str) -> None:
    with ZipFile(path, "r") as source_zip:
        files = {name: source_zip.read(name) for name in source_zip.namelist()}
    sheet_xml = files["xl/worksheets/sheet1.xml"].decode("utf-8")
    original = f'<c r="{cell_ref}"><f>{formula}</f><v></v></c>'
    replacement = f'<c r="{cell_ref}"><f>{formula}</f><v>{cached_value}</v></c>'
    if original not in sheet_xml:
        raise AssertionError(f"未找到公式单元格 {cell_ref}")
    files["xl/worksheets/sheet1.xml"] = sheet_xml.replace(original, replacement, 1).encode("utf-8")
    with ZipFile(path, "w") as target_zip:
        for name, content in files.items():
            target_zip.writestr(name, content)


def test_capture_workload_fills_target_and_marks_source(tmp_path):
    source = tmp_path / "source.xlsx"
    target = tmp_path / "target.xlsx"
    output_source = tmp_path / "source-out.xlsx"
    output_target = tmp_path / "target-out.xlsx"
    _write_source(source)
    _write_target(target)

    summary = capture_workload(
        source,
        target,
        output_source,
        output_target,
        _source_configs(),
        _target_configs(),
        selected_fields=DEFAULT_SELECTED_WORKLOAD_FIELDS,
    )

    assert summary["filled_rows"] == 1
    assert summary["unmatched_source_rows"] == 1
    output = load_workbook(output_target, data_only=True)
    try:
        sheet = output["表2 通用工程测量费用"]
        assert sheet["F2"].value == 26
        assert sheet["G2"].value == 0.6
        assert sheet["H2"].value == 0.22
        assert sheet["I2"].value == "起算点"
        assert "抓取成功" in sheet["J2"].value
        assert "字段完全匹配" in sheet["J2"].value
        assert sheet["F2"].fill.fgColor.rgb == "00D9EAD3"
    finally:
        output.close()

    marked_source = load_workbook(output_source, data_only=True)
    try:
        sheet = marked_source["委托方工作量"]
        assert "抓取成功" in sheet["J2"].value
        assert "字段完全匹配" in sheet["J2"].value
        assert sheet["K1"].value == SOURCE_DIAGNOSTIC_FIELD
        assert "识别结果" in sheet["K2"].value
        assert "最可能目标" in sheet["K2"].value
        assert "未抓取成功" in sheet["J3"].value
        assert "识别结果" in sheet["K3"].value
        assert sheet["F3"].fill.fgColor.rgb == "00F4CCCC"
    finally:
        marked_source.close()


def test_capture_workload_conservative_mode_keeps_existing_target_value(tmp_path):
    source = tmp_path / "source.xlsx"
    target = tmp_path / "target.xlsx"
    output_source = tmp_path / "source-out.xlsx"
    output_target = tmp_path / "target-out.xlsx"
    _write_source(source)
    _write_target(target)
    workbook = load_workbook(target)
    try:
        workbook["表2 通用工程测量费用"]["F2"].value = 99
        workbook.save(target)
    finally:
        workbook.close()

    summary = capture_workload(
        source,
        target,
        output_source,
        output_target,
        _source_configs(),
        _target_configs(),
        selected_fields=[TARGET_QUANTITY_FIELD],
        write_mode=WRITE_MODE_CONSERVATIVE,
    )

    assert summary["filled_rows"] == 0
    assert summary["overwritten_rows"] == 0
    assert summary["skipped_existing_rows"] == 1
    assert summary["skipped_existing_cells"] == 1
    output = load_workbook(output_target, data_only=True)
    try:
        sheet = output["表2 通用工程测量费用"]
        assert sheet["F2"].value == 99
        assert "已匹配但未写入" in sheet["J2"].value
    finally:
        output.close()


def test_capture_workload_overwrite_mode_replaces_existing_target_value(tmp_path):
    source = tmp_path / "source.xlsx"
    target = tmp_path / "target.xlsx"
    output_source = tmp_path / "source-out.xlsx"
    output_target = tmp_path / "target-out.xlsx"
    _write_source(source)
    _write_target(target)
    workbook = load_workbook(target)
    try:
        workbook["表2 通用工程测量费用"]["F2"].value = 99
        workbook.save(target)
    finally:
        workbook.close()

    summary = capture_workload(
        source,
        target,
        output_source,
        output_target,
        _source_configs(),
        _target_configs(),
        selected_fields=[TARGET_QUANTITY_FIELD],
        write_mode=WRITE_MODE_OVERWRITE,
    )

    assert summary["filled_rows"] == 0
    assert summary["overwritten_rows"] == 1
    assert summary["overwritten_cells"] == 1
    output = load_workbook(output_target, data_only=True)
    try:
        sheet = output["表2 通用工程测量费用"]
        assert sheet["F2"].value == 26
        assert "覆盖1项" in sheet["J2"].value
    finally:
        output.close()


def test_capture_workload_warns_when_one_source_matches_multiple_targets(tmp_path):
    source = tmp_path / "source.xlsx"
    target = tmp_path / "target.xlsx"
    output_source = tmp_path / "source-out.xlsx"
    output_target = tmp_path / "target-out.xlsx"
    _write_source(source)
    _write_target(target, duplicate=True)

    summary = capture_workload(
        source,
        target,
        output_source,
        output_target,
        _source_configs(),
        _target_configs(),
    )

    assert summary["filled_rows"] == 0
    assert summary["duplicate_warning_rows"] == 2
    assert len(summary["issue_log_preview"]) == 2
    assert all("一对多预警" in row["message"] for row in summary["issue_log_preview"])
    output = load_workbook(output_target, data_only=True)
    try:
        sheet = output["表2 通用工程测量费用"]
        assert sheet["F2"].value is None
        assert "一对多预警" in sheet["J2"].value
        assert "一对多预警" in sheet["J3"].value
        assert sheet["F2"].fill.fgColor.rgb == "00FFF2CC"
        assert sheet["J2"].fill.fgColor.rgb == "00FFF2CC"
    finally:
        output.close()

    marked_source = load_workbook(output_source, data_only=True)
    try:
        sheet = marked_source["委托方工作量"]
        assert "一对多预警" in sheet["J2"].value
        assert sheet["F2"].fill.fgColor.rgb == "00FFF2CC"
        assert sheet["K2"].fill.fgColor.rgb == "00FFF2CC"
    finally:
        marked_source.close()


def test_suggest_workload_column_mapping_for_source_headers():
    headers = ["序号", "项目", "内容", "类别", "比例尺", "单位", "数量", "调整系数", "技术工作费调整系数", "备注"]

    mapping = suggest_workload_column_mapping(headers, "source")

    assert mapping["要素1"] == "B"
    assert mapping["要素2"] == "C"
    assert mapping["要素3"] == "D"
    assert mapping["要素4"] == "E"
    assert mapping["单位"] == "F"
    assert mapping["数量"] == "G"
    assert mapping["实物工作费调整系数"] == "H"
    assert mapping["技术工作费调整系数"] == "I"


def test_suggest_workload_column_mapping_uses_adjacent_fallback_by_default():
    headers = ["项目", "未命名列", "类别", "比例尺", "单位", "数量"]

    mapping = suggest_workload_column_mapping(headers, "source")

    assert mapping["要素1"] == "A"
    assert mapping["要素2"] == "B"
    assert mapping["要素3"] == "C"
    assert mapping["要素4"] == "D"
    assert mapping["单位"] == "E"
    assert mapping["数量"] == "F"


def test_suggest_workload_column_mapping_uses_element_sequence_before_default_aliases():
    headers = ["项目", "第一后续列", "第二后续列", "第三后续列", "第四后续列", "单位", "数量"]

    mapping = suggest_workload_column_mapping(headers, "source")

    assert mapping["要素1"] == "A"
    assert mapping["要素2"] == "B"
    assert mapping["要素3"] == "C"
    assert mapping["要素4"] == "D"
    assert mapping["要素5"] == "E"
    assert mapping["单位"] == "F"
    assert mapping["数量"] == "G"


def test_suggest_workload_column_mapping_stops_element_sequence_at_unit_column():
    headers = ["项目", "内容", "类别", "单位", "数量"]

    mapping = suggest_workload_column_mapping(headers, "source")

    assert mapping["要素1"] == "A"
    assert mapping["要素2"] == "B"
    assert mapping["要素3"] == "C"
    assert mapping["要素4"] == ""
    assert mapping["要素5"] == ""
    assert mapping["单位"] == "D"
    assert mapping["数量"] == "E"


def test_suggest_workload_column_mapping_can_disable_adjacent_fallback():
    headers = ["项目", "未命名列", "类别", "比例尺", "单位", "数量"]

    mapping = suggest_workload_column_mapping(
        headers,
        "source",
        adjacent_fallback_enabled=False,
        element_sequence_enabled=False,
    )

    assert mapping["要素1"] == "A"
    assert mapping["要素2"] == ""
    assert mapping["要素3"] == "C"


def test_suggest_workload_column_mapping_can_disable_element_sequence_for_source():
    headers = ["项目", "第一后续列", "第二后续列", "第三后续列", "第四后续列", "单位"]

    mapping = suggest_workload_column_mapping(
        headers,
        "source",
        adjacent_fallback_enabled=False,
        element_sequence_enabled=False,
    )

    assert mapping["要素1"] == "A"
    assert mapping["要素2"] == ""
    assert mapping["要素3"] == ""
    assert mapping["要素4"] == ""
    assert mapping["要素5"] == ""


def test_suggest_workload_column_mapping_uses_adjacent_fallback_for_target():
    headers = ["要素1", "未命名列", "要素3", "要素4", "单位", "数量(信息抓取)"]

    mapping = suggest_workload_column_mapping(headers, "target")

    assert mapping["要素1"] == "A"
    assert mapping["要素2"] == "B"
    assert mapping["要素3"] == "C"
    assert mapping["要素4"] == "D"
    assert mapping["单位"] == "E"
    assert mapping["数量(信息抓取)"] == "F"


def test_suggest_workload_column_mapping_does_not_use_element_sequence_for_target_by_default():
    headers = ["要素1", "第一后续列", "第二后续列", "第三后续列", "第四后续列", "单位"]

    mapping = suggest_workload_column_mapping(headers, "target", adjacent_fallback_enabled=False)

    assert mapping["要素1"] == "A"
    assert mapping["要素2"] == ""
    assert mapping["要素3"] == ""
    assert mapping["要素4"] == ""
    assert mapping["要素5"] == ""


def test_suggest_workload_column_mapping_can_enable_element_sequence_for_target():
    headers = ["要素1", "第一后续列", "第二后续列", "第三后续列", "第四后续列", "单位"]

    mapping = suggest_workload_column_mapping(
        headers,
        "target",
        adjacent_fallback_enabled=False,
        element_sequence_enabled=True,
    )

    assert mapping["要素1"] == "A"
    assert mapping["要素2"] == "B"
    assert mapping["要素3"] == "C"
    assert mapping["要素4"] == "D"
    assert mapping["要素5"] == "E"


def test_capture_workload_uses_ordered_match_mode_b(tmp_path):
    source = tmp_path / "source-ordered.xlsx"
    target = tmp_path / "target-ordered.xlsx"
    output_source = tmp_path / "source-ordered-out.xlsx"
    output_target = tmp_path / "target-ordered-out.xlsx"

    source_book = Workbook()
    source_sheet = source_book.active
    source_sheet.title = "委托方工作量"
    source_sheet.append(["项目", "内容", "类别", "比例尺", "单位", "数量", "备注"])
    source_sheet.append(["控制测量", None, "GPS测量C级", "中等", "个", 26, "顺序匹配"])
    source_book.save(source)
    source_book.close()

    target_book = Workbook()
    target_sheet = target_book.active
    target_sheet.title = "表2 通用工程测量费用"
    target_sheet.append(["要素1", "要素2", "要素3", "要素4", "单位", TARGET_QUANTITY_FIELD, "抓取日志"])
    target_sheet.append(["控制测量", "GPS测量C级", None, "中等", "个", None, None])
    target_book.save(target)
    target_book.close()

    summary = capture_workload(
        source,
        target,
        output_source,
        output_target,
        _source_configs(),
        [
            {
                "sheet_name": "表2 通用工程测量费用",
                "enabled": True,
                "header_row": 1,
                "column_mapping": {
                    "要素1": "A",
                    "要素2": "B",
                    "要素3": "C",
                    "要素4": "D",
                    "单位": "E",
                    TARGET_QUANTITY_FIELD: "F",
                    "抓取日志": "G",
                },
            }
        ],
        selected_fields=[TARGET_QUANTITY_FIELD],
    )

    assert summary["filled_rows"] == 1
    assert "非空要素顺序匹配" in summary["log_preview"][0]["message"]
    output = load_workbook(output_target, data_only=True)
    try:
        sheet = output["表2 通用工程测量费用"]
        assert sheet["F2"].value == 26
        assert "非空要素顺序匹配" in sheet["G2"].value
    finally:
        output.close()


def test_capture_workload_filters_rows_by_selected_non_empty_field(tmp_path):
    source = tmp_path / "source-filter.xlsx"
    target = tmp_path / "target-filter.xlsx"
    output_source = tmp_path / "source-filter-out.xlsx"
    output_target = tmp_path / "target-filter-out.xlsx"

    source_book = Workbook()
    source_sheet = source_book.active
    source_sheet.title = "委托方工作量"
    source_sheet.append(["项目", "内容", "类别", "比例尺", "单位", "数量", "实物工作费调整系数"])
    source_sheet.append(["控制测量", "平面控制", "一等", "1:500", "km", 12, 1.1])
    source_sheet.append(["控制测量", "导线测量", "二等", "1:500", "km", None, 1.3])
    source_book.save(source)
    source_book.close()

    target_book = Workbook()
    target_sheet = target_book.active
    target_sheet.title = "表2 通用工程测量费用"
    target_sheet.append(["要素1", "要素2", "要素3", "要素4", "单位", TARGET_QUANTITY_FIELD, "抓取日志"])
    target_sheet.append(["控制测量", "平面控制", "一等", "1:500", "km", None, None])
    target_sheet.append(["控制测量", "导线测量", "二等", "1:500", "km", None, None])
    target_book.save(target)
    target_book.close()

    summary = capture_workload(
        source,
        target,
        output_source,
        output_target,
        _source_configs(),
        [
            {
                "sheet_name": "表2 通用工程测量费用",
                "enabled": True,
                "header_row": 1,
                "column_mapping": {
                    "要素1": "A",
                    "要素2": "B",
                    "要素3": "C",
                    "要素4": "D",
                    "单位": "E",
                    TARGET_QUANTITY_FIELD: "F",
                    "抓取日志": "G",
                },
            }
        ],
        selected_fields=[TARGET_QUANTITY_FIELD],
        filter_non_empty_field="数量",
    )

    assert summary["source_rows"] == 1
    assert summary["filled_rows"] == 1
    assert len(summary["issue_log_preview"]) == 1
    assert summary["issue_log_preview"][0]["excel_row"] == 3
    assert "模式A+B均未命中" in summary["issue_log_preview"][0]["message"]
    output = load_workbook(output_target, data_only=True)
    try:
        sheet = output["表2 通用工程测量费用"]
        assert sheet["F2"].value == 12
        assert sheet["F3"].value is None
        assert "模式A+B均未命中" in sheet["G3"].value
    finally:
        output.close()

    marked_source = load_workbook(output_source, data_only=True)
    try:
        sheet = marked_source["委托方工作量"]
        assert "抓取成功" in sheet["H2"].value
        assert sheet["H3"].value in (None, "")
        assert sheet["I3"].value in (None, "")
    finally:
        marked_source.close()


def test_capture_workload_filters_out_zero_and_formula_zero_rows(tmp_path, monkeypatch):
    monkeypatch.setattr(
        workload_term_rules_module,
        "DEFAULT_WORKLOAD_TERM_RULES_PATH",
        tmp_path / "missing-term-rules.xlsx",
    )

    source = tmp_path / "source-filter-zero.xlsx"
    target = tmp_path / "target-filter-zero.xlsx"
    output_source = tmp_path / "source-filter-zero-out.xlsx"
    output_target = tmp_path / "target-filter-zero-out.xlsx"

    source_book = Workbook()
    source_sheet = source_book.active
    source_sheet.title = "委托方工作量"
    source_sheet.append(["项目", "内容", "类别", "比例尺", "单位", "数量"])
    source_sheet.append(["控制测量", "平面控制", "一等", "1:500", "km", 12])
    source_sheet.append(["控制测量", "导线测量", "二等", "1:500", "km", 0])
    source_sheet.append(["控制测量", "水准测量", "三等", "1:500", "km", "=0"])
    source_book.save(source)
    source_book.close()
    _write_formula_cached_value(source, "F4", "0", "0")

    target_book = Workbook()
    target_sheet = target_book.active
    target_sheet.title = "表2 通用工程测量费用"
    target_sheet.append(["要素1", "要素2", "要素3", "要素4", "单位", TARGET_QUANTITY_FIELD, "抓取日志"])
    target_sheet.append(["控制测量", "平面控制", "一等", "1:500", "km", None, None])
    target_sheet.append(["控制测量", "导线测量", "二等", "1:500", "km", None, None])
    target_sheet.append(["控制测量", "水准测量", "三等", "1:500", "km", None, None])
    target_book.save(target)
    target_book.close()

    summary = capture_workload(
        source,
        target,
        output_source,
        output_target,
        _source_configs(),
        [
            {
                "sheet_name": "表2 通用工程测量费用",
                "enabled": True,
                "header_row": 1,
                "column_mapping": {
                    "要素1": "A",
                    "要素2": "B",
                    "要素3": "C",
                    "要素4": "D",
                    "单位": "E",
                    TARGET_QUANTITY_FIELD: "F",
                    "抓取日志": "G",
                },
            }
        ],
        selected_fields=[TARGET_QUANTITY_FIELD],
        filter_non_empty_field="数量",
    )

    assert summary["source_rows"] == 1
    assert summary["filled_rows"] == 1

    marked_source = load_workbook(output_source, data_only=True)
    try:
        sheet = marked_source["委托方工作量"]
        assert "抓取成功" in sheet["G2"].value
        assert sheet["G3"].value in (None, "")
        assert sheet["H3"].value in (None, "")
        assert sheet["G4"].value in (None, "")
        assert sheet["H4"].value in (None, "")
    finally:
        marked_source.close()


def test_capture_workload_source_diagnostic_identifies_unit_mismatch(tmp_path, monkeypatch):
    source = tmp_path / "source-unit.xlsx"
    target = tmp_path / "target-unit.xlsx"
    output_source = tmp_path / "source-unit-out.xlsx"
    output_target = tmp_path / "target-unit-out.xlsx"

    source_book = Workbook()
    source_sheet = source_book.active
    source_sheet.title = "委托方工作量"
    source_sheet.append(["项目", "内容", "类别", "比例尺", "单位", "数量"])
    source_sheet.append(["控制测量", "首级控制测量", "GPS测量C级", "中等", "点", 26])
    source_book.save(source)
    source_book.close()

    _write_target_unit_mismatch(target)
    monkeypatch.setattr(
        workload_term_rules_module,
        "DEFAULT_WORKLOAD_TERM_RULES_PATH",
        tmp_path / "missing-term-rules.xlsx",
    )

    capture_workload(
        source,
        target,
        output_source,
        output_target,
        _source_configs(),
        [
            {
                "sheet_name": "表2 通用工程测量费用",
                "enabled": True,
                "header_row": 1,
                "column_mapping": {
                    "要素1": "A",
                    "要素2": "B",
                    "要素3": "C",
                    "要素4": "D",
                    "单位": "E",
                    TARGET_QUANTITY_FIELD: "F",
                    "抓取日志": "G",
                },
            }
        ],
        selected_fields=[TARGET_QUANTITY_FIELD],
    )

    marked_source = load_workbook(output_source, data_only=True)
    try:
        sheet = marked_source["委托方工作量"]
        assert sheet["G1"].value == "抓取日志"
        assert sheet["H1"].value == SOURCE_DIAGNOSTIC_FIELD
        assert "最可能目标：表2 通用工程测量费用 第 2 行" in sheet["H2"].value
        assert "单位不一致" in sheet["H2"].value
        assert "源单位=点" in sheet["H2"].value
        assert "目标单位=个" in sheet["H2"].value
    finally:
        marked_source.close()


def test_capture_workload_uses_term_rules_for_unit_and_all_element_fields(tmp_path, monkeypatch):
    rules_path = tmp_path / "term-rules.xlsx"
    _write_term_rules(rules_path)
    monkeypatch.setattr(workload_term_rules_module, "DEFAULT_WORKLOAD_TERM_RULES_PATH", rules_path)

    source = tmp_path / "source-term.xlsx"
    target = tmp_path / "target-term.xlsx"
    output_source = tmp_path / "source-term-out.xlsx"
    output_target = tmp_path / "target-term-out.xlsx"

    source_book = Workbook()
    source_sheet = source_book.active
    source_sheet.title = "委托方工作量"
    source_sheet.append(["项目", "内容", "类别", "比例尺", "单位", "数量", "备注"])
    source_sheet.append(["控制测量", "首级控制测量", "GPS测量C级", "中等", "点", 26, "术语归并"])
    source_book.save(source)
    source_book.close()

    target_book = Workbook()
    target_sheet = target_book.active
    target_sheet.title = "表2 通用工程测量费用"
    target_sheet.append(["要素1", "要素2", "要素3", "要素4", "单位", TARGET_QUANTITY_FIELD, "抓取日志"])
    target_sheet.append(["控制测量", "控制测量", "GPS测量C级", "中等", "个", None, None])
    target_book.save(target)
    target_book.close()

    summary = capture_workload(
        source,
        target,
        output_source,
        output_target,
        _source_configs(),
        [
            {
                "sheet_name": "表2 通用工程测量费用",
                "enabled": True,
                "header_row": 1,
                "column_mapping": {
                    "要素1": "A",
                    "要素2": "B",
                    "要素3": "C",
                    "要素4": "D",
                    "单位": "E",
                    TARGET_QUANTITY_FIELD: "F",
                    "抓取日志": "G",
                },
            }
        ],
        selected_fields=[TARGET_QUANTITY_FIELD],
    )

    assert summary["filled_rows"] == 1
    output = load_workbook(output_target, data_only=True)
    try:
        sheet = output["表2 通用工程测量费用"]
        assert sheet["F2"].value == 26
        assert "字段完全匹配" in sheet["G2"].value
    finally:
        output.close()


def test_capture_workload_writes_formula_cached_values_instead_of_formula_text(tmp_path, monkeypatch):
    monkeypatch.setattr(
        workload_term_rules_module,
        "DEFAULT_WORKLOAD_TERM_RULES_PATH",
        tmp_path / "missing-term-rules.xlsx",
    )

    source = tmp_path / "source-formula.xlsx"
    target = tmp_path / "target-formula.xlsx"
    output_source = tmp_path / "source-formula-out.xlsx"
    output_target = tmp_path / "target-formula-out.xlsx"

    source_book = Workbook()
    source_sheet = source_book.active
    source_sheet.title = "委托方工作量"
    source_sheet.append(["项目", "内容", "类别", "比例尺", "单位", "数量", "实物工作费调整系数"])
    source_sheet.append(["控制测量", "首级控制测量", "GPS测量C级", "中等", "个", "=6+6", "=1.1"])
    source_book.save(source)
    source_book.close()
    _write_formula_cached_value(source, "F2", "6+6", "12")
    _write_formula_cached_value(source, "G2", "1.1", "1.1")

    target_book = Workbook()
    target_sheet = target_book.active
    target_sheet.title = "表2 通用工程测量费用"
    target_sheet.append(["要素1", "要素2", "要素3", "要素4", "单位", TARGET_QUANTITY_FIELD, "实物工作费调整系数(信息抓取)", "抓取日志"])
    target_sheet.append(["控制测量", "首级控制测量", "GPS测量C级", "中等", "个", None, None, None])
    target_book.save(target)
    target_book.close()

    summary = capture_workload(
        source,
        target,
        output_source,
        output_target,
        _source_configs(),
        [
            {
                "sheet_name": "表2 通用工程测量费用",
                "enabled": True,
                "header_row": 1,
                "column_mapping": {
                    "要素1": "A",
                    "要素2": "B",
                    "要素3": "C",
                    "要素4": "D",
                    "单位": "E",
                    TARGET_QUANTITY_FIELD: "F",
                    "实物工作费调整系数(信息抓取)": "G",
                    "抓取日志": "H",
                },
            }
        ],
        selected_fields=[TARGET_QUANTITY_FIELD, "实物工作费调整系数(信息抓取)"],
    )

    assert summary["filled_rows"] == 1
    output = load_workbook(output_target, data_only=False)
    try:
        sheet = output["表2 通用工程测量费用"]
        assert sheet["F2"].value == 12
        assert sheet["G2"].value == 1.1
        assert sheet["F2"].data_type != "f"
        assert sheet["G2"].data_type != "f"
    finally:
        output.close()


def test_capture_workload_optional_target_fields_can_be_left_unused(tmp_path, monkeypatch):
    monkeypatch.setattr(
        workload_term_rules_module,
        "DEFAULT_WORKLOAD_TERM_RULES_PATH",
        tmp_path / "missing-term-rules.xlsx",
    )

    source = tmp_path / "source-optional.xlsx"
    target = tmp_path / "target-optional.xlsx"
    output_source = tmp_path / "source-optional-out.xlsx"
    output_target = tmp_path / "target-optional-out.xlsx"
    _write_source(source)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "表2 通用工程测量费用"
    sheet.append(["要素1", "要素2", "要素3", "要素4", "单位", TARGET_QUANTITY_FIELD, "抓取日志"])
    sheet.append(["控制测量", "首级控制测量", "GPS测量C级", "中等", "个", None, None])
    workbook.save(target)
    workbook.close()

    summary = capture_workload(
        source,
        target,
        output_source,
        output_target,
        _source_configs(),
        [
            {
                "sheet_name": "表2 通用工程测量费用",
                "enabled": True,
                "header_row": 1,
                "column_mapping": {
                    "要素1": "A",
                    "要素2": "B",
                    "要素3": "C",
                    "要素4": "D",
                    "单位": "E",
                    TARGET_QUANTITY_FIELD: "F",
                    "抓取日志": "G",
                    "实物工作费调整系数(信息抓取)": "",
                    "技术工作费调整系数(信息抓取)": "",
                    "委托方备注(信息抓取)": "",
                },
            }
        ],
        selected_fields=DEFAULT_SELECTED_WORKLOAD_FIELDS,
    )

    assert summary["filled_rows"] == 1
    output = load_workbook(output_target, data_only=True)
    try:
        sheet = output["表2 通用工程测量费用"]
        headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
        assert TARGET_QUANTITY_FIELD in headers
        assert "实物工作费调整系数(信息抓取)" not in headers
        assert "技术工作费调整系数(信息抓取)" not in headers
        assert "委托方备注(信息抓取)" not in headers
        assert sheet["F2"].value == 26
    finally:
        output.close()
