from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string

from .adjustment_rules import AdjustmentEngine, AdjustmentEvaluation, AdjustmentResult
from .knowledge_base import (
    EXPERIENCE_PHYSICAL_FIELD,
    EXPERIENCE_PHYSICAL_NOTE_FIELD,
    EXPERIENCE_TECHNICAL_FIELD,
    EXPERIENCE_TECHNICAL_NOTE_FIELD,
    KnowledgeBase,
)
from .normalization import is_blank_price, normalize_key_part
from .schemas import FIELD_COLUMNS, FillSummary, ReviewRow

PHYSICAL_ADJUSTMENT_REPORT_COLUMN = "匹配说明-实物工作费调整系数"
TECHNICAL_ADJUSTMENT_REPORT_COLUMN = "匹配说明-技术工作费调整系数"
DIAGNOSTIC_COLUMNS = [
    "匹配状态",
    "候选数量",
    "匹配说明",
    PHYSICAL_ADJUSTMENT_REPORT_COLUMN,
    TECHNICAL_ADJUSTMENT_REPORT_COLUMN,
]
EMPTY_ELEMENT_COLUMN = "空元素列"
REPORT_START_COLUMN = 18
PRICE_COLUMN_FIELDS = ("输出-价格列", "价格列")
PHYSICAL_ADJUSTMENT_FIELD = "输出-实物工作费调整系数"
TECHNICAL_ADJUSTMENT_FIELD = "输出-技术工作费调整系数"
TECHNICAL_CONTEXT_COLUMNS = ("要素-技术", "要素技术", "技术类别", "技术工作类别")
PREVIEW_MAPPING_HEADER_MARKERS = {"映射行", "要素1", "要素2", "要素3", "要素4", "要素5", "单位"}
VALUE_FILTER_ALIASES = {
    "数量": ("数量", "工程量", "工程数量", "工程量合计"),
}
MATCHED_PRICE_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
REVIEW_PRICE_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
EXPERIENCE_ADJUSTMENT_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
MATCHED_PRICE_FONT_COLOR = "006100"
REVIEW_PRICE_FONT_COLOR = "9C0006"


class FillEngine:
    def __init__(self, knowledge_base: KnowledgeBase, adjustment_engine: AdjustmentEngine | None = None) -> None:
        self.knowledge_base = knowledge_base
        self.adjustment_engine = adjustment_engine or AdjustmentEngine.from_default_rules()

    def fill_workbook(
        self,
        input_path: str | Path,
        output_path: str | Path,
        column_mapping: dict[str, str] | None = None,
        header_row: int = 1,
        output_match_report: bool = True,
        merge_vertical_cells: bool = True,
        merge_horizontal_cells: bool = True,
        only_match_rows_with_value: bool = False,
        match_value_filter_field: str = "数量",
        sheet_configs: list[dict[str, Any]] | None = None,
    ) -> FillSummary:
        workbook = load_workbook(input_path)
        value_workbook = load_workbook(input_path, data_only=True)
        configs = sheet_configs or [
            {
                "sheet_name": workbook.active.title,
                "enabled": True,
                "header_row": header_row,
                "column_mapping": column_mapping,
                "output_match_report": output_match_report,
                "merge_vertical_cells": merge_vertical_cells,
                "merge_horizontal_cells": merge_horizontal_cells,
                "only_match_rows_with_value": only_match_rows_with_value,
                "match_value_filter_field": match_value_filter_field,
            }
        ]

        total_data_rows = 0
        filled_rows = 0
        matched_rows = 0
        unchanged_rows = 0
        review_details: list[ReviewRow] = []
        price_logs: list[str] = []
        conflict_rows = 0
        physical_matched_rows = 0
        physical_experience_rows = 0
        physical_review_rows = 0
        technical_matched_rows = 0
        technical_experience_rows = 0
        technical_review_rows = 0
        price_column_name = ""
        preview_header_rows = {
            str(config.get("sheet_name") or workbook.active.title): int(config.get("header_row") or header_row)
            for config in configs
        }

        for config in configs:
            if config.get("enabled") is False:
                continue
            sheet_name = str(config.get("sheet_name") or workbook.active.title)
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"输入表不存在候选 sheet：{sheet_name}")
            sheet = workbook[sheet_name]
            value_sheet = value_workbook[sheet_name]
            current_header_row = int(config.get("header_row") or header_row)
            current_mapping = config.get("column_mapping")
            current_output_match_report = bool(config.get("output_match_report", output_match_report))
            current_merge_vertical_cells = bool(config.get("merge_vertical_cells", merge_vertical_cells))
            current_merge_horizontal_cells = bool(config.get("merge_horizontal_cells", merge_horizontal_cells))
            current_only_match_rows_with_value = bool(config.get("only_match_rows_with_value", only_match_rows_with_value))
            current_match_value_filter_field = str(config.get("match_value_filter_field") or match_value_filter_field)
            sheet_summary = self._fill_sheet(
                sheet,
                value_sheet,
                column_mapping=current_mapping,
                header_row=current_header_row,
                output_match_report=current_output_match_report,
                merge_vertical_cells=current_merge_vertical_cells,
                merge_horizontal_cells=current_merge_horizontal_cells,
                only_match_rows_with_value=current_only_match_rows_with_value,
                match_value_filter_field=current_match_value_filter_field,
            )
            if not price_column_name:
                price_column_name = sheet_summary.price_column
            total_data_rows += sheet_summary.total_data_rows
            filled_rows += sheet_summary.filled_rows
            matched_rows += sheet_summary.matched_rows
            unchanged_rows += sheet_summary.unchanged_rows
            conflict_rows += sheet_summary.conflict_rows
            review_details.extend(sheet_summary.review_details)
            price_logs.extend([f"{sheet.title}：{line}" for line in sheet_summary.price_logs])
            physical_matched_rows += sheet_summary.physical_matched_rows
            physical_experience_rows += sheet_summary.physical_experience_rows
            physical_review_rows += sheet_summary.physical_review_rows
            technical_matched_rows += sheet_summary.technical_matched_rows
            technical_experience_rows += sheet_summary.technical_experience_rows
            technical_review_rows += sheet_summary.technical_review_rows

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        value_workbook.close()

        report_text = f"输入{total_data_rows}行，匹配{matched_rows}行。"
        table_preview = self._build_multi_sheet_table_preview(
            [
                (workbook[sheet_name], preview_header_rows.get(sheet_name, 1))
                for sheet_name in workbook.sheetnames
            ],
            max_rows=50,
        )
        return FillSummary(
            total_data_rows=total_data_rows,
            price_column=price_column_name,
            filled_rows=filled_rows,
            matched_rows=matched_rows,
            unchanged_rows=unchanged_rows,
            review_rows=len(review_details),
            conflict_rows=conflict_rows,
            output_excel=output_path.name,
            output_report="",
            report_text=report_text,
            table_preview=table_preview,
            review_details=review_details,
            price_logs=price_logs,
            physical_matched_rows=physical_matched_rows,
            physical_experience_rows=physical_experience_rows,
            physical_review_rows=physical_review_rows,
            technical_matched_rows=technical_matched_rows,
            technical_experience_rows=technical_experience_rows,
            technical_review_rows=technical_review_rows,
        )

    def _fill_sheet(
        self,
        sheet: Any,
        value_sheet: Any,
        column_mapping: dict[str, str] | None = None,
        header_row: int = 1,
        output_match_report: bool = True,
        merge_vertical_cells: bool = True,
        merge_horizontal_cells: bool = True,
        only_match_rows_with_value: bool = False,
        match_value_filter_field: str = "数量",
    ) -> FillSummary:
        headers = [cell.value for cell in value_sheet[header_row]]
        header_map = {str(name).strip(): idx for idx, name in enumerate(headers, start=1) if name}

        field_map = self._resolve_field_map(header_map, column_mapping)
        missing = [name for name in FIELD_COLUMNS if name not in field_map]
        if missing:
            raise ValueError(f"输入表缺少必要列：{', '.join(missing)}")

        price_column_name, price_column_index = self._find_price_column(header_map, headers, column_mapping)
        physical_adjustment_index = self._find_optional_output_column(
            header_map,
            headers,
            column_mapping,
            PHYSICAL_ADJUSTMENT_FIELD,
            "实物工作费调整系数",
        )
        technical_adjustment_index = self._find_optional_output_column(
            header_map,
            headers,
            column_mapping,
            TECHNICAL_ADJUSTMENT_FIELD,
            "技术工作费调整系数",
        )
        diagnostic_map = (
            self._ensure_diagnostic_columns(sheet, header_map, header_row)
            if output_match_report
            else {}
        )
        total_data_rows = 0
        filled_rows = 0
        matched_rows = 0
        unchanged_rows = 0
        review_details: list[ReviewRow] = []
        price_logs: list[str] = []
        physical_matched_rows = 0
        physical_experience_rows = 0
        physical_review_rows = 0
        technical_matched_rows = 0
        technical_experience_rows = 0
        technical_review_rows = 0
        merged_value_map = self._build_merged_value_map(
            value_sheet,
            merge_vertical_cells=merge_vertical_cells,
            merge_horizontal_cells=merge_horizontal_cells,
        )
        filter_column_index = (
            self._find_value_filter_column(headers, match_value_filter_field)
            if only_match_rows_with_value
            else None
        )
        if only_match_rows_with_value and filter_column_index is None:
            raise ValueError(f"{sheet.title} 未找到指定列：{match_value_filter_field}")

        for excel_row in range(header_row + 1, sheet.max_row + 1):
            raw_price = sheet.cell(row=excel_row, column=price_column_index).value
            current_price = (
                None
                if isinstance(raw_price, str) and raw_price.lstrip().startswith("=")
                else value_sheet.cell(row=excel_row, column=price_column_index).value
            )
            values = {
                name: self._read_mapped_value(value_sheet, excel_row, field_map[name], merged_value_map)
                for name in FIELD_COLUMNS
            }
            for context_name in TECHNICAL_CONTEXT_COLUMNS:
                if context_name in header_map:
                    values[context_name] = self._read_mapped_value(
                        value_sheet,
                        excel_row,
                        header_map[context_name],
                        merged_value_map,
                    )
            if self._is_ignored_row(values.get("要素1")):
                continue
            if filter_column_index is not None:
                filter_value = self._read_mapped_value(value_sheet, excel_row, filter_column_index, merged_value_map)
                if not self._has_value_for_matching_filter(filter_value):
                    continue
            total_data_rows += 1
            row_values = self._read_row_values(value_sheet, excel_row, merged_value_map)
            adjustment = self.adjustment_engine.evaluate(sheet.title, values, row_values)
            result = self.knowledge_base.lookup(values)
            adjustment = self._apply_second_layer_experience(adjustment, result)
            if physical_adjustment_index:
                physical_cell = sheet.cell(row=excel_row, column=physical_adjustment_index)
                physical_cell.value = adjustment.physical.value
                self._mark_adjustment_cell(physical_cell, adjustment.physical.status)
                physical_matched_rows, physical_experience_rows, physical_review_rows = self._count_adjustment_status(
                    adjustment.physical.status,
                    physical_matched_rows,
                    physical_experience_rows,
                    physical_review_rows,
                )
            if technical_adjustment_index:
                technical_cell = sheet.cell(row=excel_row, column=technical_adjustment_index)
                technical_cell.value = adjustment.technical.value
                self._mark_adjustment_cell(technical_cell, adjustment.technical.status)
                technical_matched_rows, technical_experience_rows, technical_review_rows = self._count_adjustment_status(
                    adjustment.technical.status,
                    technical_matched_rows,
                    technical_experience_rows,
                    technical_review_rows,
                )
            if not is_blank_price(current_price):
                unchanged_rows += 1
                price_logs.append(f"第 {excel_row} 行：原有价格保留，未执行匹配。")
                self._write_diagnostics(
                    sheet,
                    excel_row,
                    diagnostic_map,
                    status="原有价格",
                    candidate_count="",
                    note="价格列已有值，未改动",
                    physical_adjustment_note=adjustment.physical.message,
                    technical_adjustment_note=adjustment.technical.message,
                )
                continue

            candidate_count = len(result.candidates)
            if result.status == "matched":
                price_cell = sheet.cell(row=excel_row, column=price_column_index)
                price_cell.value = result.price
                self._mark_price_cell(price_cell, matched=True)
                filled_rows += 1
                matched_rows += 1
                price_logs.append(f"第 {excel_row} 行：已匹配，候选数量 {candidate_count}，{result.message}。")
                self._write_diagnostics(
                    sheet,
                    excel_row,
                    diagnostic_map,
                    status="已匹配",
                    candidate_count=candidate_count,
                    note=result.message,
                    physical_adjustment_note=adjustment.physical.message,
                    technical_adjustment_note=adjustment.technical.message,
                )
                continue

            price_cell = sheet.cell(row=excel_row, column=price_column_index)
            price_cell.value = "待复核"
            self._mark_price_cell(price_cell, matched=False)
            review_note = self._build_review_note(result.message, result.candidates)
            price_logs.append(f"第 {excel_row} 行：待复核，候选数量 {candidate_count}，{review_note}。")
            self._write_diagnostics(
                sheet,
                excel_row,
                diagnostic_map,
                status="待复核",
                candidate_count=candidate_count,
                note=review_note,
                physical_adjustment_note=adjustment.physical.message,
                technical_adjustment_note=adjustment.technical.message,
                severity="unmatched" if result.status == "not_found" else "review",
            )
            review_details.append(
                ReviewRow(
                    excel_row=excel_row,
                    status=result.status,
                    message=result.message,
                    values=values,
                )
            )

        conflict_rows = sum(1 for row in review_details if row.status == "conflict")
        report_text = f"输入{total_data_rows}行，匹配{matched_rows}行。"
        table_preview = self._build_table_preview(sheet, max_rows=50, header_row=header_row)
        return FillSummary(
            total_data_rows=total_data_rows,
            price_column=price_column_name,
            filled_rows=filled_rows,
            matched_rows=matched_rows,
            unchanged_rows=unchanged_rows,
            review_rows=len(review_details),
            conflict_rows=conflict_rows,
            output_excel="",
            output_report="",
            report_text=report_text,
            table_preview=table_preview,
            review_details=review_details,
            price_logs=price_logs,
            physical_matched_rows=physical_matched_rows,
            physical_experience_rows=physical_experience_rows,
            physical_review_rows=physical_review_rows,
            technical_matched_rows=technical_matched_rows,
            technical_experience_rows=technical_experience_rows,
            technical_review_rows=technical_review_rows,
        )

    @classmethod
    def _resolve_field_map(
        cls,
        header_map: dict[str, int],
        column_mapping: dict[str, str] | None,
    ) -> dict[str, int]:
        resolved: dict[str, int] = {}
        for name in FIELD_COLUMNS:
            if column_mapping and column_mapping.get(name):
                if cls._is_empty_element_reference(column_mapping[name]):
                    resolved[name] = 0
                else:
                    resolved[name] = cls._resolve_column_reference(column_mapping[name], header_map)
            elif name in header_map:
                resolved[name] = header_map[name]
            elif name in {"要素2", "要素3", "要素4", "要素5"}:
                resolved[name] = 0
        return resolved

    @classmethod
    def _find_price_column(
        cls,
        header_map: dict[str, int],
        headers: list[Any],
        column_mapping: dict[str, str] | None = None,
    ) -> tuple[str, int]:
        if column_mapping:
            for field in PRICE_COLUMN_FIELDS:
                if column_mapping.get(field):
                    column_name = column_mapping[field]
                    return column_name, cls._resolve_column_reference(column_name, header_map)

        preferred = ["单价匹配-测试", "基价测试列", "基价", "单价", "价格"]
        for name in preferred:
            if name in header_map:
                return name, header_map[name]
        for idx, value in enumerate(headers, start=1):
            text = str(value or "").replace(" ", "")
            if any(token in text for token in ["基价", "单价", "价格"]):
                return str(value), idx
        raise ValueError("输入表未找到基价/单价/价格列")

    @classmethod
    def _find_optional_output_column(
        cls,
        header_map: dict[str, int],
        headers: list[Any],
        column_mapping: dict[str, str] | None,
        mapping_field: str,
        header_token: str,
    ) -> int | None:
        if column_mapping and column_mapping.get(mapping_field):
            return cls._resolve_column_reference(column_mapping[mapping_field], header_map)
        compact_token = header_token.replace(" ", "")
        for idx, value in enumerate(headers, start=1):
            compact = str(value or "").replace(" ", "")
            if compact_token in compact:
                return idx
        return None

    @classmethod
    def _find_value_filter_column(cls, headers: list[Any], field: str) -> int | None:
        aliases = VALUE_FILTER_ALIASES.get(str(field).strip(), (str(field).strip(),))
        normalized_aliases = {cls._normalize_header_token(alias) for alias in aliases if str(alias).strip()}
        if not normalized_aliases:
            return None

        normalized_headers = [
            (idx, cls._normalize_header_token(value))
            for idx, value in enumerate(headers, start=1)
        ]
        for idx, header in normalized_headers:
            if header in normalized_aliases:
                return idx
        for idx, header in normalized_headers:
            if header and any(alias in header for alias in normalized_aliases):
                return idx
        return None

    @staticmethod
    def _normalize_header_token(value: Any) -> str:
        return str(value or "").replace(" ", "").replace("\n", "").replace("\r", "").strip()

    @staticmethod
    def _has_value_for_matching_filter(value: Any) -> bool:
        if value is None or isinstance(value, bool):
            return False
        if isinstance(value, (int, float)):
            return abs(float(value)) > 1e-9
        text = str(value).strip()
        if not text:
            return False
        numeric_text = text.replace(",", "").replace("，", "")
        try:
            return abs(float(numeric_text)) > 1e-9
        except ValueError:
            return True

    @staticmethod
    def _resolve_column_reference(reference: str, header_map: dict[str, int]) -> int:
        text = str(reference).strip()
        if text in header_map:
            return header_map[text]
        try:
            return column_index_from_string(text.upper())
        except ValueError as exc:
            raise ValueError(f"输入表未找到映射列：{reference}") from exc

    @staticmethod
    def _is_empty_element_reference(reference: str) -> bool:
        return str(reference).strip() == EMPTY_ELEMENT_COLUMN

    @staticmethod
    def _build_merged_value_map(
        sheet: Any,
        merge_vertical_cells: bool = True,
        merge_horizontal_cells: bool = True,
    ) -> dict[tuple[int, int], Any]:
        merged_value_map: dict[tuple[int, int], Any] = {}
        for merged_range in sheet.merged_cells.ranges:
            value = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for column in range(merged_range.min_col, merged_range.max_col + 1):
                    if row == merged_range.min_row and column == merged_range.min_col:
                        merged_value_map[(row, column)] = value
                    elif column == merged_range.min_col and merge_vertical_cells:
                        merged_value_map[(row, column)] = value
                    elif column > merged_range.min_col and not merge_horizontal_cells:
                        if row == merged_range.min_row or merge_vertical_cells:
                            merged_value_map[(row, column)] = value
        return merged_value_map

    @staticmethod
    def _read_mapped_value(
        sheet: Any,
        excel_row: int,
        column_index: int,
        merged_value_map: dict[tuple[int, int], Any],
    ) -> Any:
        if column_index == 0:
            return ""
        value = sheet.cell(row=excel_row, column=column_index).value
        if value is not None:
            return FillEngine._clean_input_value(value)
        return FillEngine._clean_input_value(merged_value_map.get((excel_row, column_index)))

    @staticmethod
    def _read_row_values(
        sheet: Any,
        excel_row: int,
        merged_value_map: dict[tuple[int, int], Any],
    ) -> list[Any]:
        return [
            FillEngine._read_mapped_value(sheet, excel_row, column, merged_value_map)
            for column in range(1, sheet.max_column + 1)
        ]

    @staticmethod
    def _clean_input_value(value: Any) -> Any:
        if isinstance(value, str) and value.strip() == "成图比例":
            return ""
        return value

    @staticmethod
    def _is_ignored_row(first_element: Any) -> bool:
        return normalize_key_part(first_element) in {"", "小计", "合计"}

    @classmethod
    def _apply_second_layer_experience(
        cls,
        adjustment: AdjustmentEvaluation,
        result: Any,
    ) -> AdjustmentEvaluation:
        if result.status != "matched" or not result.candidates:
            return adjustment

        candidate = result.candidates[0]
        physical = adjustment.physical
        technical = adjustment.technical
        if physical.status in {"default", "review"} and not physical.rules:
            physical = cls._build_experience_adjustment(
                candidate=candidate,
                value_field=EXPERIENCE_PHYSICAL_FIELD,
                note_field=EXPERIENCE_PHYSICAL_NOTE_FIELD,
                label="实物工作费调整系数",
            ) or physical
        if technical.status == "review" and not technical.rules:
            technical = cls._build_experience_adjustment(
                candidate=candidate,
                value_field=EXPERIENCE_TECHNICAL_FIELD,
                note_field=EXPERIENCE_TECHNICAL_NOTE_FIELD,
                label="技术工作费调整系数",
            ) or technical
        return AdjustmentEvaluation(physical=physical, technical=technical)

    @staticmethod
    def _build_experience_adjustment(
        candidate: dict[str, Any],
        value_field: str,
        note_field: str,
        label: str,
    ) -> AdjustmentResult | None:
        value = FillEngine._normalize_experience_number(candidate.get(value_field))
        if value is None:
            return None
        note = str(candidate.get(note_field) or "").strip()
        message = (
            f"第二层经验提示层：第一层未命中{label}标准规则，"
            f"按单价知识库经验数 {_format_adjustment_number(value)} 填写。"
        )
        if note:
            message += f"经验说明：{note}。"
        return AdjustmentResult(value=value, status="experience", message=message)

    @staticmethod
    def _normalize_experience_number(value: Any) -> int | float | None:
        if value is None:
            return None
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return FillEngine._clean_adjustment_number(float(value))
        text = str(value).strip()
        if text in {"", "/", "待复核"} or "不需要经验数" in text:
            return None
        try:
            return FillEngine._clean_adjustment_number(float(text))
        except ValueError:
            return None

    @staticmethod
    def _clean_adjustment_number(value: float) -> int | float:
        rounded = round(value, 6)
        return int(rounded) if rounded.is_integer() else rounded

    @staticmethod
    def _ensure_diagnostic_columns(sheet: Any, header_map: dict[str, int], header_row: int) -> dict[str, int]:
        diagnostic_map: dict[str, int] = {}
        existing = {name: header_map[name] for name in DIAGNOSTIC_COLUMNS if name in header_map}
        if len(existing) == len(DIAGNOSTIC_COLUMNS):
            return existing
        if existing:
            diagnostic_map.update(existing)
            next_column = max(existing.values()) + 1
            for name in DIAGNOSTIC_COLUMNS:
                if name in diagnostic_map:
                    continue
                while sheet.cell(row=header_row, column=next_column).value:
                    next_column += 1
                sheet.cell(row=header_row, column=next_column).value = name
                diagnostic_map[name] = next_column
                next_column += 1
            return diagnostic_map
        next_column = header_map.get("匹配报告预留位置", REPORT_START_COLUMN)
        if "匹配报告预留位置" in header_map:
            next_column += 1
        else:
            while any(sheet.cell(row=header_row, column=next_column + offset).value for offset in range(len(DIAGNOSTIC_COLUMNS))):
                next_column += 1
        for name in DIAGNOSTIC_COLUMNS:
            sheet.cell(row=header_row, column=next_column).value = name
            diagnostic_map[name] = next_column
            next_column += 1
        return diagnostic_map

    @staticmethod
    def _write_diagnostics(
        sheet: Any,
        excel_row: int,
        diagnostic_map: dict[str, int],
        status: str,
        candidate_count: int | str,
        note: str,
        physical_adjustment_note: str = "",
        technical_adjustment_note: str = "",
        severity: str = "",
    ) -> None:
        if not diagnostic_map:
            return
        sheet.cell(row=excel_row, column=diagnostic_map["匹配状态"]).value = status
        sheet.cell(row=excel_row, column=diagnostic_map["候选数量"]).value = candidate_count
        sheet.cell(row=excel_row, column=diagnostic_map["匹配说明"]).value = note
        sheet.cell(row=excel_row, column=diagnostic_map[PHYSICAL_ADJUSTMENT_REPORT_COLUMN]).value = physical_adjustment_note
        sheet.cell(row=excel_row, column=diagnostic_map[TECHNICAL_ADJUSTMENT_REPORT_COLUMN]).value = technical_adjustment_note
        if severity:
            fill = REVIEW_PRICE_FILL if severity == "unmatched" else EXPERIENCE_ADJUSTMENT_FILL
            for column in diagnostic_map.values():
                sheet.cell(row=excel_row, column=column).fill = fill
            return
        for report_column in (PHYSICAL_ADJUSTMENT_REPORT_COLUMN, TECHNICAL_ADJUSTMENT_REPORT_COLUMN):
            cell = sheet.cell(row=excel_row, column=diagnostic_map[report_column])
            if "待复核" in str(cell.value or ""):
                cell.fill = EXPERIENCE_ADJUSTMENT_FILL

    @staticmethod
    def _mark_price_cell(cell: Any, matched: bool) -> None:
        font = copy(cell.font)
        if matched:
            cell.fill = MATCHED_PRICE_FILL
            font.color = MATCHED_PRICE_FONT_COLOR
        else:
            cell.fill = REVIEW_PRICE_FILL
            font.color = REVIEW_PRICE_FONT_COLOR
        cell.font = font

    @staticmethod
    def _mark_adjustment_cell(cell: Any, status: str) -> None:
        if status == "matched":
            cell.fill = MATCHED_PRICE_FILL
        elif status == "experience":
            cell.fill = EXPERIENCE_ADJUSTMENT_FILL
        else:
            cell.fill = REVIEW_PRICE_FILL

    @staticmethod
    def _count_adjustment_status(
        status: str,
        matched_rows: int,
        experience_rows: int,
        review_rows: int,
    ) -> tuple[int, int, int]:
        if status == "matched":
            return matched_rows + 1, experience_rows, review_rows
        if status == "experience":
            return matched_rows, experience_rows + 1, review_rows
        return matched_rows, experience_rows, review_rows + 1

    @staticmethod
    def _build_review_note(message: str, candidates: list[dict[str, Any]]) -> str:
        if not candidates:
            return f"{message}；请检查输入列映射、要素内容或单位是否与知识库一致"
        prices = sorted({str(candidate.get("基价", "")) for candidate in candidates})
        return f"{message}；候选基价：{'、'.join(prices)}"

    @staticmethod
    def _build_table_preview(sheet: Any, max_rows: int, header_row: int = 1) -> dict[str, Any]:
        headers = [cell.value if cell.value is not None else "" for cell in sheet[header_row]]
        headers = FillEngine.preview_display_headers(sheet, header_row, headers)
        rows: list[list[Any]] = []
        for excel_row in range(header_row + 1, min(sheet.max_row, header_row + max_rows) + 1):
            rows.append(
                [
                    sheet.cell(row=excel_row, column=column).value
                    if sheet.cell(row=excel_row, column=column).value is not None
                    else ""
                    for column in range(1, sheet.max_column + 1)
                ]
            )
        return {"sheet_name": sheet.title, "header_row": header_row, "headers": headers, "rows": rows}

    @classmethod
    def preview_display_headers(
        cls,
        sheet: Any,
        header_row: int,
        raw_headers: list[Any],
        column_count: int | None = None,
    ) -> list[Any]:
        headers = [value if value is not None else "" for value in raw_headers]
        if not cls._is_mapping_preview_header(headers) or header_row <= 1:
            return headers

        column_total = column_count or len(headers)
        visual_merged_values = cls._build_merged_value_map(
            sheet,
            merge_vertical_cells=True,
            merge_horizontal_cells=False,
        )
        source_rows: list[list[Any]] = []
        for row_index in range(max(1, header_row - 2), header_row):
            row_values = [
                sheet.cell(row=row_index, column=column_index).value
                if sheet.cell(row=row_index, column=column_index).value is not None
                else visual_merged_values.get((row_index, column_index), "")
                for column_index in range(1, column_total + 1)
            ]
            if sum(1 for value in row_values if str(value or "").strip()) <= 1 and row_index < header_row - 1:
                continue
            source_rows.append(row_values)

        if not source_rows:
            return headers

        display_headers: list[Any] = []
        for column_index in range(column_total):
            parts: list[str] = []
            for row_values in source_rows:
                text = str(row_values[column_index] if column_index < len(row_values) else "").strip()
                if text and text not in parts:
                    parts.append(text)
            raw_text = str(headers[column_index] if column_index < len(headers) else "").strip()
            if parts:
                display_headers.append(" / ".join(parts))
            elif raw_text and raw_text not in PREVIEW_MAPPING_HEADER_MARKERS:
                display_headers.append(raw_text)
            else:
                display_headers.append(f"列{column_index + 1}")
        return display_headers

    @staticmethod
    def _is_mapping_preview_header(headers: list[Any]) -> bool:
        values = {str(value or "").strip() for value in headers if str(value or "").strip()}
        return "映射行" in values and "要素1" in values

    @classmethod
    def _build_multi_sheet_table_preview(
        cls,
        sheets: list[tuple[Any, int]],
        max_rows: int,
    ) -> dict[str, Any]:
        previews = [
            cls._build_table_preview(sheet, max_rows=max_rows, header_row=header_row)
            for sheet, header_row in sheets
        ]
        if not previews:
            return {"sheet_name": "", "headers": [], "rows": [], "sheets": []}
        first = previews[0]
        return {
            "sheet_name": first["sheet_name"],
            "headers": first["headers"],
            "rows": first["rows"],
            "sheets": previews,
        }


def _format_adjustment_number(value: int | float) -> str:
    return str(int(value)) if isinstance(value, float) and value.is_integer() else str(value)
