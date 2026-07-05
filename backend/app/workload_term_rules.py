from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .normalization import normalize_key_part
from .paths import DEFAULT_WORKLOAD_TERM_RULES_PATH
from .schemas import FIELD_COLUMNS


ELEMENT_FIELDS = FIELD_COLUMNS[:5]
WORKLOAD_SCOPE_MARKERS = {
    "",
    "*",
    "all",
    "全部",
    "工作量抓取",
    "原始工作量抓取",
    "工作量信息抓取",
    "数量抓取",
    "数量信息抓取",
}
TRUE_MARKERS = {"1", "true", "yes", "y", "是"}
DIAGNOSTIC_ONLY_MARKERS = {"仅诊断提示", "仅候选比选"}


@dataclass(frozen=True)
class WorkloadTermRules:
    unit_aliases: dict[str, str]
    element_aliases: dict[str, dict[str, str]]
    source_path: Path | None = None

    @classmethod
    def load(cls, path: str | Path | None = None) -> "WorkloadTermRules":
        rule_path = Path(path) if path else DEFAULT_WORKLOAD_TERM_RULES_PATH
        empty = cls(
            unit_aliases={},
            element_aliases={field: {} for field in ELEMENT_FIELDS},
            source_path=rule_path if rule_path.exists() else None,
        )
        if not rule_path.exists():
            return empty

        workbook = load_workbook(rule_path, read_only=True, data_only=True)
        try:
            unit_aliases: dict[str, str] = {}
            element_aliases = {field: {} for field in ELEMENT_FIELDS}

            unit_sheet = workbook["单位严格等价"] if "单位严格等价" in workbook.sheetnames else None
            if unit_sheet is not None:
                header_row, header_map = _find_header_row(unit_sheet, {"启用", "原值", "归并值"})
                if header_row:
                    for row in range(header_row + 1, unit_sheet.max_row + 1):
                        record = _read_row(unit_sheet, row, header_map)
                        if not _is_enabled_row(record):
                            continue
                        if not _scope_matches(record.get("生效模块")):
                            continue
                        if not _is_formal_matching_rule(
                            record.get("匹配强度"),
                            record.get("建议层级"),
                            record.get("是否允许自动写值"),
                        ):
                            continue
                        original = normalize_key_part(record.get("原值"))
                        merged = normalize_key_part(record.get("归并值"))
                        if not original or not merged:
                            continue
                        unit_aliases[original] = merged

            exact_sheet = workbook["要素严格等价"] if "要素严格等价" in workbook.sheetnames else None
            if exact_sheet is not None:
                header_row, header_map = _find_header_row(exact_sheet, {"启用", "原值", "归并值"})
                if header_row:
                    for row in range(header_row + 1, exact_sheet.max_row + 1):
                        record = _read_row(exact_sheet, row, header_map)
                        if not _is_enabled_row(record):
                            continue
                        if not _scope_matches(record.get("生效模块")):
                            continue
                        if not _is_formal_matching_rule(
                            record.get("匹配强度"),
                            record.get("建议层级"),
                            record.get("是否允许自动写值"),
                        ):
                            continue
                        original = normalize_key_part(record.get("原值"))
                        merged = normalize_key_part(record.get("归并值"))
                        if not original or not merged:
                            continue
                        for field in _resolve_target_fields(record.get("字段名")):
                            element_aliases[field][original] = merged

            weak_sheet = workbook["要素弱等价"] if "要素弱等价" in workbook.sheetnames else None
            if weak_sheet is not None:
                header_row, header_map = _find_header_row(weak_sheet, {"启用", "上位值", "下位值"})
                if header_row:
                    for row in range(header_row + 1, weak_sheet.max_row + 1):
                        record = _read_row(weak_sheet, row, header_map)
                        if not _is_enabled_row(record):
                            continue
                        if not _scope_matches(record.get("生效模块")):
                            continue
                        if not _is_formal_matching_rule(
                            record.get("匹配强度"),
                            record.get("建议层级"),
                            record.get("是否允许自动写值"),
                        ):
                            continue
                        upper = normalize_key_part(record.get("上位值"))
                        lower = normalize_key_part(record.get("下位值"))
                        if not upper or not lower:
                            continue
                        direction = _clean_text(record.get("匹配方向"))
                        merged = upper
                        if direction == "下位可匹配上位":
                            merged = lower
                        for field in _resolve_target_fields(record.get("字段名")):
                            element_aliases[field][upper] = merged
                            element_aliases[field][lower] = merged

            return cls(unit_aliases=unit_aliases, element_aliases=element_aliases, source_path=rule_path)
        finally:
            workbook.close()

    def make_key(self, row: dict[str, Any]) -> tuple[str, ...]:
        return tuple(self.normalize_field(field, row.get(field)) for field in FIELD_COLUMNS)

    def make_ordered_key(self, row: dict[str, Any]) -> tuple[str, tuple[str, ...]]:
        unit = self.normalize_field("单位", row.get("单位"))
        elements = tuple(
            self.canonical_element(field, row.get(field))
            for field in ELEMENT_FIELDS
            if self.canonical_element(field, row.get(field))
        )
        return unit, elements

    def normalize_field(self, field: str, value: Any) -> str:
        normalized = normalize_key_part(value)
        if not normalized:
            return ""
        if field == "单位":
            return _resolve_alias(normalized, self.unit_aliases)
        if field in ELEMENT_FIELDS:
            return _resolve_alias(normalized, self.element_aliases.get(field, {}))
        return normalized

    def canonical_element(self, field: str, value: Any) -> str:
        normalized = self.normalize_field(field, value)
        return _strip_known_prefix(normalized)


def _find_header_row(sheet: Any, required_headers: set[str]) -> tuple[int, dict[str, int]]:
    scan_limit = min(sheet.max_row, 20)
    for row in range(1, scan_limit + 1):
        header_map: dict[str, int] = {}
        for column in range(1, sheet.max_column + 1):
            header = _clean_text(sheet.cell(row=row, column=column).value)
            if header and header not in header_map:
                header_map[header] = column
        if required_headers.issubset(header_map):
            return row, header_map
    return 0, {}


def _read_row(sheet: Any, row: int, header_map: dict[str, int]) -> dict[str, Any]:
    return {
        header: sheet.cell(row=row, column=column).value
        for header, column in header_map.items()
    }


def _is_enabled_row(record: dict[str, Any]) -> bool:
    return _clean_text(record.get("启用")).lower() in TRUE_MARKERS


def _scope_matches(raw_scope: Any) -> bool:
    scope = _clean_text(raw_scope)
    if not scope:
        return True
    parts = {
        part.strip()
        for raw_part in scope.replace("；", ",").replace("、", ",").replace("/", ",").split(",")
        for part in [raw_part]
        if part.strip()
    }
    if not parts:
        return True
    return any(part.lower() in WORKLOAD_SCOPE_MARKERS for part in parts)


def _is_formal_matching_rule(raw_strength: Any, raw_level: Any, raw_allow_auto: Any) -> bool:
    strength = _clean_text(raw_strength)
    level = _clean_text(raw_level)
    allow_auto = _clean_text(raw_allow_auto).lower() in TRUE_MARKERS
    if strength in DIAGNOSTIC_ONLY_MARKERS or level in DIAGNOSTIC_ONLY_MARKERS:
        return False
    if raw_allow_auto is not None and _clean_text(raw_allow_auto) and not allow_auto:
        return False
    if not strength and not level and raw_allow_auto is None:
        return True
    return strength in {"", "严格等价"} or level in {"", "正式匹配"}


def _resolve_target_fields(raw_field: Any) -> list[str]:
    field_name = _clean_text(raw_field)
    if not field_name:
        return ELEMENT_FIELDS[:]
    return [field_name] if field_name in ELEMENT_FIELDS else []


def _clean_text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _resolve_alias(value: str, aliases: dict[str, str]) -> str:
    current = value
    seen: set[str] = set()
    while current and current not in seen and current in aliases:
        seen.add(current)
        next_value = aliases[current]
        if not next_value or next_value == current:
            break
        current = next_value
    return current


def _strip_known_prefix(value: str) -> str:
    if value.startswith("级别-"):
        return value.removeprefix("级别-")
    if value.startswith("比例-"):
        return value.removeprefix("比例-")
    return value
