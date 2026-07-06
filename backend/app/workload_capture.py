from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter

from .fill_engine import FillEngine
from .normalization import normalize_key_part
from .schemas import FIELD_COLUMNS
from .workload_term_rules import WorkloadTermRules

SOURCE_QUANTITY_FIELD = "数量"
SOURCE_PHYSICAL_FIELD = "实物工作费调整系数"
SOURCE_TECHNICAL_FIELD = "技术工作费调整系数"
SOURCE_REMARK_FIELD = "委托方备注"

TARGET_QUANTITY_FIELD = "数量(信息抓取)"
TARGET_PHYSICAL_FIELD = "实物工作费调整系数(信息抓取)"
TARGET_TECHNICAL_FIELD = "技术工作费调整系数(信息抓取)"
TARGET_REMARK_FIELD = "委托方备注(信息抓取)"
CAPTURE_LOG_FIELD = "抓取日志"
SOURCE_DIAGNOSTIC_FIELD = "抓取诊断"

SOURCE_CAPTURE_FIELDS = [
    SOURCE_QUANTITY_FIELD,
    SOURCE_PHYSICAL_FIELD,
    SOURCE_TECHNICAL_FIELD,
    SOURCE_REMARK_FIELD,
]
TARGET_CAPTURE_FIELDS = [
    TARGET_QUANTITY_FIELD,
    TARGET_PHYSICAL_FIELD,
    TARGET_TECHNICAL_FIELD,
    TARGET_REMARK_FIELD,
]
TARGET_OPTIONAL_FIELDS = [
    TARGET_PHYSICAL_FIELD,
    TARGET_TECHNICAL_FIELD,
    TARGET_REMARK_FIELD,
]
DEFAULT_SELECTED_WORKLOAD_FIELDS = TARGET_CAPTURE_FIELDS[:]
SOURCE_MAPPING_FIELDS = [*FIELD_COLUMNS, *SOURCE_CAPTURE_FIELDS]
TARGET_MAPPING_FIELDS = [*FIELD_COLUMNS, *TARGET_CAPTURE_FIELDS, CAPTURE_LOG_FIELD]
WORKLOAD_FIELD_PREFERENCE_FIELDS = SOURCE_MAPPING_FIELDS[:]
WORKLOAD_TARGET_FIELD_PREFERENCE_FIELDS = TARGET_MAPPING_FIELDS[:]
DEFAULT_WORKLOAD_FILTER_FIELD = SOURCE_QUANTITY_FIELD

SOURCE_TO_TARGET_FIELD = {
    TARGET_QUANTITY_FIELD: SOURCE_QUANTITY_FIELD,
    TARGET_PHYSICAL_FIELD: SOURCE_PHYSICAL_FIELD,
    TARGET_TECHNICAL_FIELD: SOURCE_TECHNICAL_FIELD,
    TARGET_REMARK_FIELD: SOURCE_REMARK_FIELD,
}

MATCH_MODE_EXACT = "字段完全匹配"
MATCH_MODE_ORDERED = "非空要素顺序匹配"
WRITE_MODE_CONSERVATIVE = "conservative"
WRITE_MODE_OVERWRITE = "overwrite"

MATCH_FILL = PatternFill(fill_type="solid", fgColor="D9EAD3")
WARNING_DUPLICATE_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
WARNING_DUPLICATE_FONT = Font(color="7F6000")
WARNING_FILL = PatternFill(fill_type="solid", fgColor="F4CCCC")
WARNING_FONT = Font(color="9C0006")


@dataclass
class SourceRecord:
    sheet_name: str
    excel_row: int
    key_values: dict[str, Any]
    values: dict[str, Any]
    value_columns: dict[str, int]
    log_column: int
    diagnostic_column: int
    exact_key: tuple[str, ...]
    ordered_key: tuple[str, tuple[str, ...]]


@dataclass
class TargetRecord:
    sheet_name: str
    excel_row: int
    key_values: dict[str, Any]
    output_columns: dict[str, int]
    log_column: int
    exact_key: tuple[str, ...]
    ordered_key: tuple[str, tuple[str, ...]]
    message: str = ""
    status: str = "skipped"
    matched_source: SourceRecord | None = None
    match_mode: str = ""


@dataclass
class WorkloadCaptureResult:
    source_rows: list[SourceRecord] = field(default_factory=list)
    target_rows: list[TargetRecord] = field(default_factory=list)
    filled_rows: int = 0
    overwritten_rows: int = 0
    skipped_existing_rows: int = 0
    written_cells: int = 0
    overwritten_cells: int = 0
    skipped_existing_cells: int = 0
    warning_rows: int = 0
    unmatched_source_rows: int = 0
    duplicate_warning_rows: int = 0


@dataclass
class WorkloadMatch:
    match_mode: str
    sources: list[SourceRecord]
    targets: list[TargetRecord]


def _is_duplicate_warning(match: WorkloadMatch) -> bool:
    return len(match.sources) > 1 or len(match.targets) > 1


def _warning_style(duplicate: bool) -> tuple[PatternFill, Font]:
    if duplicate:
        return WARNING_DUPLICATE_FILL, WARNING_DUPLICATE_FONT
    return WARNING_FILL, WARNING_FONT


def capture_workload(
    workload_path: str | Path,
    target_path: str | Path,
    output_workload_path: str | Path,
    output_target_path: str | Path,
    source_sheet_configs: list[dict[str, Any]],
    target_sheet_configs: list[dict[str, Any]],
    selected_fields: list[str] | None = None,
    filter_non_empty_field: str | None = DEFAULT_WORKLOAD_FILTER_FIELD,
    write_mode: str = WRITE_MODE_OVERWRITE,
) -> dict[str, Any]:
    selected = [field for field in (selected_fields or DEFAULT_SELECTED_WORKLOAD_FIELDS) if field in TARGET_CAPTURE_FIELDS]
    if not selected:
        raise ValueError("至少选择一个工作量抓取字段")
    if write_mode not in {WRITE_MODE_CONSERVATIVE, WRITE_MODE_OVERWRITE}:
        raise ValueError("工作量抓取写入模式只能是 conservative 或 overwrite")

    workload_path = Path(workload_path)
    target_path = Path(target_path)
    output_workload_path = Path(output_workload_path)
    output_target_path = Path(output_target_path)
    output_workload_path.parent.mkdir(parents=True, exist_ok=True)
    output_target_path.parent.mkdir(parents=True, exist_ok=True)

    source_workbook = load_workbook(workload_path)
    source_value_workbook = load_workbook(workload_path, data_only=True)
    target_workbook = load_workbook(target_path)
    term_rules = WorkloadTermRules.load()
    try:
        result = WorkloadCaptureResult()
        source_records = _collect_source_records(
            source_workbook,
            source_value_workbook,
            source_sheet_configs,
            selected,
            filter_non_empty_field=filter_non_empty_field,
            term_rules=term_rules,
        )
        target_records = _collect_target_records(target_workbook, target_sheet_configs, selected, term_rules=term_rules)
        result.source_rows = source_records
        result.target_rows = target_records

        source_exact_index = _index_records(source_records, "exact_key")
        source_ordered_index = _index_records(source_records, "ordered_key")
        target_exact_index = _index_records(target_records, "exact_key")
        target_ordered_index = _index_records(target_records, "ordered_key")

        for target in target_records:
            match = _lookup_target_match(
                target,
                source_exact_index,
                source_ordered_index,
                target_exact_index,
                target_ordered_index,
            )
            if len(match.sources) == 1 and len(match.targets) == 1:
                source = match.sources[0]
                write_result = _write_target_success(
                    target_workbook[target.sheet_name],
                    target,
                    source,
                    selected,
                    match.match_mode,
                    write_mode=write_mode,
                )
                written = int(write_result["written_fields"])
                overwritten = int(write_result["overwritten_fields"])
                skipped_existing = int(write_result["skipped_existing_fields"])
                result.written_cells += written
                result.overwritten_cells += overwritten
                result.skipped_existing_cells += skipped_existing
                if written:
                    result.filled_rows += 1
                if overwritten:
                    result.overwritten_rows += 1
                if skipped_existing and not written and not overwritten:
                    result.skipped_existing_rows += 1
                target.status = "matched" if written or overwritten else "skipped_existing"
                target.matched_source = source
                target.match_mode = match.match_mode
                if written or overwritten:
                    target.message = (
                        f"抓取成功：{match.match_mode}，匹配工作量表 {source.sheet_name} 第 {source.excel_row} 行；"
                        f"写入{written}项，覆盖{overwritten}项。"
                    )
                else:
                    target.message = (
                        f"已匹配但未写入：保守模式下目标行已有值，匹配工作量表 {source.sheet_name} 第 {source.excel_row} 行。"
                    )
            else:
                message = _target_warning_message(target, match)
                _write_target_warning(
                    target_workbook[target.sheet_name],
                    target,
                    selected,
                    message,
                    duplicate=_is_duplicate_warning(match),
                )
                target.status = "warning"
                target.match_mode = match.match_mode
                target.message = message
                result.warning_rows += 1
                if _is_duplicate_warning(match):
                    result.duplicate_warning_rows += 1

        for source in source_records:
            match = _lookup_source_match(
                source,
                source_exact_index,
                source_ordered_index,
                target_exact_index,
                target_ordered_index,
            )
            sheet = source_workbook[source.sheet_name]
            diagnosis = _build_source_diagnosis(source, match, target_records)
            if len(match.sources) == 1 and len(match.targets) == 1:
                target = match.targets[0]
                _write_source_log(
                    sheet,
                    source,
                    f"抓取成功：{match.match_mode}，写入控制价计算表 {target.sheet_name} 第 {target.excel_row} 行。",
                    diagnosis,
                    matched=True,
                )
            else:
                message = _source_warning_message(source, match)
                _write_source_log(
                    sheet,
                    source,
                    message,
                    diagnosis,
                    matched=False,
                    duplicate=_is_duplicate_warning(match),
                )
                result.unmatched_source_rows += 1

        source_workbook.save(output_workload_path)
        target_workbook.save(output_target_path)
    finally:
        source_workbook.close()
        source_value_workbook.close()
        target_workbook.close()

    return {
        "source_file": workload_path.name,
        "target_file": target_path.name,
        "output_workload": output_workload_path.name,
        "output_target": output_target_path.name,
        "selected_fields": selected,
        "source_rows": len(result.source_rows),
        "target_rows": len(result.target_rows),
        "filled_rows": result.filled_rows,
        "overwritten_rows": result.overwritten_rows,
        "skipped_existing_rows": result.skipped_existing_rows,
        "written_cells": result.written_cells,
        "overwritten_cells": result.overwritten_cells,
        "skipped_existing_cells": result.skipped_existing_cells,
        "warning_rows": result.warning_rows,
        "unmatched_source_rows": result.unmatched_source_rows,
        "duplicate_warning_rows": result.duplicate_warning_rows,
        "write_mode": write_mode,
        "issue_log_preview": [
            {
                "sheet_name": row.sheet_name,
                "excel_row": row.excel_row,
                "status": row.status,
                "message": row.message,
            }
            for row in result.target_rows
            if _is_target_issue_log(row)
        ][:20],
        "log_preview": [
            {
                "sheet_name": row.sheet_name,
                "excel_row": row.excel_row,
                "status": row.status,
                "message": row.message,
            }
            for row in result.target_rows[:20]
        ],
    }


def _is_target_issue_log(row: TargetRecord) -> bool:
    if row.status == "warning":
        return True
    return "一对多" in row.message or "未抓取" in row.message or "未匹配" in row.message


def default_workload_field_preferences() -> dict[str, list[str]]:
    return {
        "要素1": ["要素1", "项目", "工作任务", "项目名称", "专业"],
        "要素2": ["要素2", "内容", "工作内容", "作业内容"],
        "要素3": ["要素3", "类别", "类别名称"],
        "要素4": ["要素4", "比例尺", "成图比例", "规格"],
        "要素5": ["要素5", "级别", "复杂程度", "等级"],
        "单位": ["单位", "计量单位"],
        SOURCE_QUANTITY_FIELD: ["数量", "工程量合计", "工程量"],
        SOURCE_PHYSICAL_FIELD: ["实物工作费调整系数", "实物系数", "调整系数", "附加系数"],
        SOURCE_TECHNICAL_FIELD: ["技术工作费调整系数", "技术系数"],
        SOURCE_REMARK_FIELD: ["委托方备注", "备注", "说明"],
    }


def default_workload_target_field_preferences() -> dict[str, list[str]]:
    return {
        "要素1": ["要素1", "项目", "工作任务", "项目名称", "专业"],
        "要素2": ["要素2", "内容", "工作内容", "作业内容"],
        "要素3": ["要素3", "类别", "类别名称"],
        "要素4": ["要素4", "比例尺", "成图比例", "规格"],
        "要素5": ["要素5", "级别", "复杂程度", "等级"],
        "单位": ["单位", "计量单位"],
        TARGET_QUANTITY_FIELD: [TARGET_QUANTITY_FIELD, "数量", "工程量"],
        TARGET_PHYSICAL_FIELD: [TARGET_PHYSICAL_FIELD, "实物工作费调整系数"],
        TARGET_TECHNICAL_FIELD: [TARGET_TECHNICAL_FIELD, "技术工作费调整系数"],
        TARGET_REMARK_FIELD: [TARGET_REMARK_FIELD, "委托方备注", "备注"],
        CAPTURE_LOG_FIELD: [CAPTURE_LOG_FIELD],
    }


def suggest_workload_column_mapping(
    headers: list[str],
    role: str,
    preferences: dict[str, list[str]] | None = None,
    adjacent_fallback_enabled: bool = True,
    element_sequence_enabled: bool | None = None,
) -> dict[str, str]:
    if element_sequence_enabled is None:
        element_sequence_enabled = role == "source"
    if role == "source":
        defaults = default_workload_field_preferences()
        mapping: dict[str, str] = {field: "" for field in SOURCE_MAPPING_FIELDS}
        mapping["要素1"] = _find_preferred_column(headers, preferences.get("要素1", []) if preferences else [], defaults.get("要素1", []))
        if element_sequence_enabled:
            _apply_element_sequence_fallback(mapping, headers)
        for field in SOURCE_MAPPING_FIELDS:
            if mapping.get(field):
                continue
            mapping[field] = _find_preferred_column(headers, preferences.get(field, []) if preferences else [], defaults.get(field, []))
        if adjacent_fallback_enabled:
            _apply_adjacent_column_fallback(mapping, SOURCE_MAPPING_FIELDS, headers)
        return mapping

    mapping = {field: "" for field in TARGET_MAPPING_FIELDS}
    defaults = default_workload_target_field_preferences()
    mapping["要素1"] = _find_preferred_column(headers, preferences.get("要素1", []) if preferences else [], defaults.get("要素1", []))
    if element_sequence_enabled:
        _apply_element_sequence_fallback(mapping, headers)
    for field in TARGET_MAPPING_FIELDS:
        if mapping.get(field):
            continue
        mapping[field] = _find_preferred_column(headers, preferences.get(field, []) if preferences else [], defaults.get(field, []))
    if adjacent_fallback_enabled:
        _apply_adjacent_column_fallback(mapping, TARGET_MAPPING_FIELDS, headers)
    return mapping


def _apply_element_sequence_fallback(mapping: dict[str, str], headers: list[str]) -> None:
    element1_column = mapping.get("要素1")
    if not element1_column:
        return
    try:
        element1_index = column_index_from_string(element1_column)
    except ValueError:
        return
    for offset, field in enumerate(["要素2", "要素3", "要素4", "要素5"], start=1):
        column_index = element1_index + offset
        if column_index > len(headers):
            break
        if _is_unit_header(headers[column_index - 1]):
            break
        mapping[field] = get_column_letter(column_index)


def _is_unit_header(value: object) -> bool:
    normalized = normalize_key_part(value)
    return normalized in {"单位", "计量单位"}


def _apply_adjacent_column_fallback(mapping: dict[str, str], ordered_fields: list[str], headers: list[str]) -> None:
    max_column = len(headers)
    for index, field in enumerate(ordered_fields):
        if mapping.get(field) or index == 0:
            continue
        previous_column = mapping.get(ordered_fields[index - 1])
        if not previous_column:
            continue
        try:
            next_column_index = column_index_from_string(previous_column) + 1
        except ValueError:
            continue
        if 1 <= next_column_index <= max_column:
            if field in {"要素2", "要素3", "要素4", "要素5"} and _is_unit_header(headers[next_column_index - 1]):
                continue
            mapping[field] = get_column_letter(next_column_index)


def _collect_source_records(
    workbook: Any,
    value_workbook: Any,
    configs: list[dict[str, Any]],
    selected_target_fields: list[str],
    filter_non_empty_field: str | None = DEFAULT_WORKLOAD_FILTER_FIELD,
    term_rules: WorkloadTermRules | None = None,
) -> list[SourceRecord]:
    rules = term_rules or WorkloadTermRules.load()
    selected_source_fields = [SOURCE_TO_TARGET_FIELD[field] for field in selected_target_fields]
    records: list[SourceRecord] = []
    for config in configs:
        if not config.get("enabled", True):
            continue
        sheet_name = str(config.get("sheet_name") or "").strip()
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        value_sheet = value_workbook[sheet_name] if sheet_name in value_workbook.sheetnames else sheet
        header_row = max(1, int(config.get("header_row") or 1))
        mapping = _clean_mapping(config.get("column_mapping"))
        merged_value_map = FillEngine._build_merged_value_map(sheet)
        value_merged_map = FillEngine._build_merged_value_map(value_sheet)
        log_column, diagnostic_column = _ensure_source_output_columns(sheet, header_row)
        filter_column = mapping.get(filter_non_empty_field) if filter_non_empty_field else None
        if filter_non_empty_field and not filter_column:
            raise ValueError(f"工作量表 sheet {sheet_name} 未映射过滤字段：{filter_non_empty_field}")
        for excel_row in range(header_row + 1, sheet.max_row + 1):
            key_values = _read_key_values(sheet, excel_row, mapping, merged_value_map)
            values = {
                field: _read_mapped_value(value_sheet, excel_row, mapping.get(field), value_merged_map)
                for field in selected_source_fields
            }
            if not _has_key_content(key_values):
                continue
            if filter_non_empty_field:
                filter_value = _read_mapped_value(value_sheet, excel_row, filter_column, value_merged_map)
                if not _has_filter_value(filter_value):
                    continue
            if not _has_any_capture_value(values):
                continue
            records.append(
                SourceRecord(
                    sheet_name=sheet_name,
                    excel_row=excel_row,
                    key_values=key_values,
                    values=values,
                    value_columns={
                        field: column_index_from_string(mapping[field])
                        for field in selected_source_fields
                        if mapping.get(field)
                    },
                    log_column=log_column,
                    diagnostic_column=diagnostic_column,
                    exact_key=rules.make_key(key_values),
                    ordered_key=rules.make_ordered_key(key_values),
                )
            )
    return records


def _collect_target_records(
    workbook: Any,
    configs: list[dict[str, Any]],
    selected_fields: list[str],
    term_rules: WorkloadTermRules | None = None,
) -> list[TargetRecord]:
    rules = term_rules or WorkloadTermRules.load()
    records: list[TargetRecord] = []
    for config in configs:
        if not config.get("enabled", True):
            continue
        sheet_name = str(config.get("sheet_name") or "").strip()
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        header_row = max(1, int(config.get("header_row") or 1))
        mapping = _clean_mapping(config.get("column_mapping"))
        merged_value_map = FillEngine._build_merged_value_map(sheet)
        output_columns = _ensure_target_columns(sheet, header_row, mapping, selected_fields)
        log_column = output_columns[CAPTURE_LOG_FIELD]
        for excel_row in range(header_row + 1, sheet.max_row + 1):
            key_values = _read_key_values(sheet, excel_row, mapping, merged_value_map)
            if not _has_key_content(key_values):
                continue
            records.append(
                TargetRecord(
                    sheet_name=sheet_name,
                    excel_row=excel_row,
                    key_values=key_values,
                    output_columns=output_columns,
                    log_column=log_column,
                    exact_key=rules.make_key(key_values),
                    ordered_key=rules.make_ordered_key(key_values),
                )
            )
    return records


def _lookup_target_match(
    target: TargetRecord,
    source_exact_index: dict[tuple[str, ...], list[SourceRecord]],
    source_ordered_index: dict[tuple[str, tuple[str, ...]], list[SourceRecord]],
    target_exact_index: dict[tuple[str, ...], list[TargetRecord]],
    target_ordered_index: dict[tuple[str, tuple[str, ...]], list[TargetRecord]],
) -> WorkloadMatch:
    exact_sources = source_exact_index.get(target.exact_key, [])
    if exact_sources:
        return WorkloadMatch(
            match_mode=MATCH_MODE_EXACT,
            sources=exact_sources,
            targets=target_exact_index.get(target.exact_key, []),
        )
    ordered_sources = source_ordered_index.get(target.ordered_key, [])
    if ordered_sources:
        return WorkloadMatch(
            match_mode=MATCH_MODE_ORDERED,
            sources=ordered_sources,
            targets=target_ordered_index.get(target.ordered_key, []),
        )
    return WorkloadMatch(match_mode="", sources=[], targets=[])


def _lookup_source_match(
    source: SourceRecord,
    source_exact_index: dict[tuple[str, ...], list[SourceRecord]],
    source_ordered_index: dict[tuple[str, tuple[str, ...]], list[SourceRecord]],
    target_exact_index: dict[tuple[str, ...], list[TargetRecord]],
    target_ordered_index: dict[tuple[str, tuple[str, ...]], list[TargetRecord]],
) -> WorkloadMatch:
    exact_targets = target_exact_index.get(source.exact_key, [])
    if exact_targets:
        return WorkloadMatch(
            match_mode=MATCH_MODE_EXACT,
            sources=source_exact_index.get(source.exact_key, []),
            targets=exact_targets,
        )
    ordered_targets = target_ordered_index.get(source.ordered_key, [])
    if ordered_targets:
        return WorkloadMatch(
            match_mode=MATCH_MODE_ORDERED,
            sources=source_ordered_index.get(source.ordered_key, []),
            targets=ordered_targets,
        )
    return WorkloadMatch(match_mode="", sources=[], targets=[])


def _read_key_values(
    sheet: Any,
    excel_row: int,
    mapping: dict[str, str],
    merged_value_map: dict[tuple[int, int], Any],
) -> dict[str, Any]:
    return {
        field: _read_mapped_value(sheet, excel_row, mapping.get(field), merged_value_map)
        for field in FIELD_COLUMNS
    }


def _read_mapped_value(
    sheet: Any,
    excel_row: int,
    column: str | None,
    merged_value_map: dict[tuple[int, int], Any],
) -> Any:
    if not column:
        return None
    try:
        return FillEngine._read_mapped_value(sheet, excel_row, column_index_from_string(column), merged_value_map)
    except ValueError:
        return None


def _ensure_target_columns(
    sheet: Any,
    header_row: int,
    mapping: dict[str, str],
    selected_fields: list[str],
) -> dict[str, int]:
    columns: dict[str, int] = {}
    next_column = _last_used_column(sheet, header_row) + 1
    for field in [*selected_fields, CAPTURE_LOG_FIELD]:
        column_letter = mapping.get(field)
        if column_letter:
            try:
                columns[field] = column_index_from_string(column_letter)
                continue
            except ValueError:
                pass
        if field in TARGET_OPTIONAL_FIELDS:
            continue
        while sheet.cell(row=header_row, column=next_column).value:
            next_column += 1
        sheet.cell(row=header_row, column=next_column).value = field
        columns[field] = next_column
        next_column += 1
    return columns


def _ensure_source_output_columns(sheet: Any, header_row: int) -> tuple[int, int]:
    log_column = 0
    diagnostic_column = 0
    for column in range(1, _last_used_column(sheet, header_row) + 1):
        header = str(sheet.cell(row=header_row, column=column).value or "").strip()
        if header == CAPTURE_LOG_FIELD:
            log_column = column
        elif header == SOURCE_DIAGNOSTIC_FIELD:
            diagnostic_column = column
    next_column = _last_used_column(sheet, header_row) + 1
    if not log_column:
        log_column = next_column
        sheet.cell(row=header_row, column=log_column).value = CAPTURE_LOG_FIELD
        next_column += 1
    if not diagnostic_column:
        preferred_column = log_column + 1
        if preferred_column > sheet.max_column or not sheet.cell(row=header_row, column=preferred_column).value:
            diagnostic_column = preferred_column
        else:
            diagnostic_column = next_column
        sheet.cell(row=header_row, column=diagnostic_column).value = SOURCE_DIAGNOSTIC_FIELD
    return log_column, diagnostic_column


def _last_used_column(sheet: Any, header_row: int) -> int:
    max_column = 1
    max_row = min(sheet.max_row, max(header_row + 30, 30))
    max_scan_column = min(sheet.max_column, 300)
    for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_scan_column):
        for cell in row:
            if cell.value not in (None, ""):
                max_column = max(max_column, cell.column)
    return max_column


def _write_target_success(
    sheet: Any,
    target: TargetRecord,
    source: SourceRecord,
    selected_fields: list[str],
    match_mode: str,
    *,
    write_mode: str,
) -> dict[str, int]:
    written_fields = 0
    overwritten_fields = 0
    skipped_existing_fields = 0
    for target_field in selected_fields:
        column = target.output_columns.get(target_field)
        if not column:
            continue
        source_field = SOURCE_TO_TARGET_FIELD[target_field]
        value = source.values.get(source_field)
        if not _has_value(value):
            continue
        cell = sheet.cell(row=target.excel_row, column=column)
        had_value = _has_value(cell.value)
        if had_value and write_mode == WRITE_MODE_CONSERVATIVE:
            skipped_existing_fields += 1
            continue
        cell.value = value
        cell.fill = MATCH_FILL
        if had_value:
            overwritten_fields += 1
        else:
            written_fields += 1
    log_cell = sheet.cell(row=target.excel_row, column=target.log_column)
    if written_fields or overwritten_fields:
        log_cell.value = (
            f"抓取成功：{match_mode}，匹配 {source.sheet_name} 第 {source.excel_row} 行；"
            f"写入{written_fields}项，覆盖{overwritten_fields}项。"
        )
        log_cell.fill = MATCH_FILL
    else:
        log_cell.value = f"已匹配但未写入：保守模式下目标行已有值，匹配 {source.sheet_name} 第 {source.excel_row} 行。"
        log_cell.fill = WARNING_DUPLICATE_FILL
        log_cell.font = WARNING_DUPLICATE_FONT
    return {
        "written_fields": written_fields,
        "overwritten_fields": overwritten_fields,
        "skipped_existing_fields": skipped_existing_fields,
    }


def _write_target_warning(
    sheet: Any,
    target: TargetRecord,
    selected_fields: list[str],
    message: str,
    *,
    duplicate: bool,
) -> None:
    fill, font = _warning_style(duplicate)
    for field in selected_fields:
        column = target.output_columns.get(field)
        if not column:
            continue
        cell = sheet.cell(row=target.excel_row, column=column)
        cell.fill = fill
        cell.font = font
    log_cell = sheet.cell(row=target.excel_row, column=target.log_column)
    log_cell.value = message
    log_cell.fill = fill
    log_cell.font = font


def _write_source_log(
    sheet: Any,
    source: SourceRecord,
    message: str,
    diagnosis: str,
    matched: bool,
    duplicate: bool = False,
) -> None:
    fill = MATCH_FILL
    font = None
    if not matched:
        fill, font = _warning_style(duplicate)
    for column in source.value_columns.values():
        sheet.cell(row=source.excel_row, column=column).fill = fill
        if font is not None:
            sheet.cell(row=source.excel_row, column=column).font = font
    log_cell = sheet.cell(row=source.excel_row, column=source.log_column)
    log_cell.value = message
    log_cell.fill = fill
    if font is not None:
        log_cell.font = font
    diagnostic_cell = sheet.cell(row=source.excel_row, column=source.diagnostic_column)
    diagnostic_cell.value = diagnosis
    diagnostic_cell.fill = fill
    if font is not None:
        diagnostic_cell.font = font


def _target_warning_message(target: TargetRecord, match: WorkloadMatch) -> str:
    if not match.sources:
        return "未抓取：工作量表中未找到要素1-5和单位匹配的记录（模式A+B均未命中）。"
    if len(match.sources) > 1:
        rows = "、".join(f"{source.sheet_name}第{source.excel_row}行" for source in match.sources[:5])
        return f"一对多预警：{match.match_mode}下工作量表匹配到 {len(match.sources)} 行（{rows}），未自动填写。"
    if len(match.targets) > 1:
        rows = "、".join(f"{row.sheet_name}第{row.excel_row}行" for row in match.targets[:5])
        return f"一对多预警：{match.match_mode}下同一工作量记录对应控制价计算表 {len(match.targets)} 行（{rows}），未自动填写。"
    return "未抓取：匹配关系异常。"


def _source_warning_message(source: SourceRecord, match: WorkloadMatch) -> str:
    if not match.targets:
        return "未抓取成功：控制价计算表中未找到要素1-5和单位匹配的行（模式A+B均未命中）。"
    if len(match.sources) > 1:
        return f"一对多预警：{match.match_mode}下同一匹配键在工作量表中出现 {len(match.sources)} 行，未自动抓取。"
    if len(match.targets) > 1:
        rows = "、".join(f"{target.sheet_name}第{target.excel_row}行" for target in match.targets[:5])
        return f"一对多预警：{match.match_mode}下对应控制价计算表 {len(match.targets)} 行（{rows}），未自动抓取。"
    return "未抓取成功：匹配关系异常。"


def _build_source_diagnosis(
    source: SourceRecord,
    match: WorkloadMatch,
    all_targets: list[TargetRecord],
) -> str:
    lines = [f"识别结果：{_describe_key_values(source.key_values)}"]
    if len(match.sources) == 1 and len(match.targets) == 1:
        target = match.targets[0]
        lines.append(f"最可能目标：{_describe_target(target)}")
        lines.append(f"结论：已抓取成功（{match.match_mode}）。")
        return "\n".join(lines)

    target, reason = _diagnose_target_gap(source, all_targets)
    if target is not None:
        lines.append(f"最可能目标：{_describe_target(target)}")
    else:
        lines.append("最可能目标：未找到明显接近的目标行。")
    lines.append(f"卡点：{reason}")
    return "\n".join(lines)


def _diagnose_target_gap(
    source: SourceRecord,
    targets: list[TargetRecord],
) -> tuple[TargetRecord | None, str]:
    if not targets:
        return None, "当前未启用任何目标 sheet。"

    same_order_different_unit = [
        target
        for target in targets
        if target.ordered_key[1] == source.ordered_key[1]
        and target.exact_key[-1] != source.exact_key[-1]
    ]
    if same_order_different_unit:
        target = _choose_best_target(source, same_order_different_unit)
        return (
            target,
            f"单位不一致：非空要素顺序已经对上，但源单位={_display_value(source.key_values.get('单位'))}，"
            f"目标单位={_display_value(target.key_values.get('单位'))}。",
        )

    same_unit_same_a1 = [
        target
        for target in targets
        if target.exact_key[-1] == source.exact_key[-1]
        and target.exact_key[0] == source.exact_key[0]
    ]
    if same_unit_same_a1:
        target = _choose_best_target(source, same_unit_same_a1)
        return target, _sequence_mismatch_reason(source, target, "单位一致、要素1一致，但要素顺序不一致")

    same_unit = [
        target
        for target in targets
        if target.exact_key[-1] == source.exact_key[-1]
    ]
    if same_unit:
        target = _choose_best_target(source, same_unit)
        return (
            target,
            f"单位一致，但要素1不一致：源要素1={_display_value(source.key_values.get('要素1'))}，"
            f"目标要素1={_display_value(target.key_values.get('要素1'))}。",
        )

    same_a1 = [
        target
        for target in targets
        if target.exact_key[0] == source.exact_key[0]
    ]
    if same_a1:
        target = _choose_best_target(source, same_a1)
        return (
            target,
            f"要素1接近，但单位不一致：源单位={_display_value(source.key_values.get('单位'))}，"
            f"目标单位={_display_value(target.key_values.get('单位'))}。",
        )

    target = _choose_best_target(source, targets)
    if target is None:
        return None, "未找到可比较的目标行。"
    return target, "未找到明显同类：当前最接近的目标行在单位和要素顺序上都没有对齐。"


def _sequence_mismatch_reason(source: SourceRecord, target: TargetRecord, prefix: str) -> str:
    source_sequence = " -> ".join(_display_sequence(source.key_values))
    target_sequence = " -> ".join(_display_sequence(target.key_values))
    return (
        f"{prefix}。源非空序列=[{source_sequence}]；目标非空序列=[{target_sequence}]。"
    )


def _choose_best_target(source: SourceRecord, targets: list[TargetRecord]) -> TargetRecord | None:
    if not targets:
        return None
    return max(
        targets,
        key=lambda target: (
            _field_match_count(source, target),
            _ordered_prefix_length(source.ordered_key[1], target.ordered_key[1]),
            _ordered_common_count(source.ordered_key[1], target.ordered_key[1]),
            int(source.exact_key[0] == target.exact_key[0]),
            int(source.exact_key[-1] == target.exact_key[-1]),
            -target.excel_row,
        ),
    )


def _field_match_count(source: SourceRecord, target: TargetRecord) -> int:
    return sum(1 for index, value in enumerate(source.exact_key) if value and value == target.exact_key[index])


def _ordered_prefix_length(source_ordered: tuple[str, ...], target_ordered: tuple[str, ...]) -> int:
    prefix = 0
    for source_part, target_part in zip(source_ordered, target_ordered):
        if source_part != target_part:
            break
        prefix += 1
    return prefix


def _ordered_common_count(source_ordered: tuple[str, ...], target_ordered: tuple[str, ...]) -> int:
    return len(set(source_ordered).intersection(target_ordered))


def _describe_target(target: TargetRecord) -> str:
    return f"{target.sheet_name} 第 {target.excel_row} 行：{_describe_key_values(target.key_values)}"


def _describe_key_values(values: dict[str, Any]) -> str:
    fields = [*FIELD_COLUMNS]
    return "；".join(f"{field}={_display_value(values.get(field))}" for field in fields)


def _display_sequence(values: dict[str, Any]) -> list[str]:
    sequence: list[str] = []
    for field in FIELD_COLUMNS[:5]:
        value = values.get(field)
        if normalize_key_part(value):
            sequence.append(_display_value(value))
    return sequence


def _display_value(value: Any) -> str:
    if value is None:
        return "空"
    text = str(value).replace("\n", "").strip()
    return text if text else "空"


def _index_records(
    records: list[SourceRecord] | list[TargetRecord],
    key_name: str,
) -> dict[Any, list[Any]]:
    index: dict[Any, list[Any]] = {}
    for record in records:
        index.setdefault(getattr(record, key_name), []).append(record)
    return index


def _clean_mapping(raw_mapping: Any) -> dict[str, str]:
    return {
        str(key): str(value).strip()
        for key, value in dict(raw_mapping or {}).items()
        if value is not None and str(value).strip()
    }


def _has_key_content(row: dict[str, Any]) -> bool:
    return bool(normalize_key_part(row.get("要素1"))) and bool(normalize_key_part(row.get("单位")))


def _has_value(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def _has_filter_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    text = str(value).strip()
    if text == "":
        return False
    try:
        return float(text.replace(",", "")) != 0
    except ValueError:
        return True


def _has_any_capture_value(values: dict[str, Any]) -> bool:
    return any(_has_value(value) for value in values.values())


def _find_preferred_column(headers: list[str], preferred_aliases: list[str], default_aliases: list[str]) -> str:
    found = _find_column_letter(headers, preferred_aliases)
    if found:
        return found
    for alias in preferred_aliases:
        if alias in default_aliases:
            continue
        found = _find_column_by_token(headers, alias)
        if found:
            return found
    found = _find_column_letter(headers, default_aliases)
    if found:
        return found
    for alias in default_aliases:
        found = _find_column_by_token(headers, alias)
        if found:
            return found
    return ""


def _find_column_letter(headers: list[str], names: list[str]) -> str:
    compact_names = [name.replace(" ", "") for name in names]
    for index, header in enumerate(headers, start=1):
        header_text = str(header or "").strip()
        compact_header = header_text.replace(" ", "")
        if header_text in names or compact_header in compact_names:
            return get_column_letter(index)
    for index, header in enumerate(headers, start=1):
        compact_header = str(header or "").replace(" ", "")
        if any(name and name in compact_header for name in compact_names):
            return get_column_letter(index)
    return ""


def _find_column_by_token(headers: list[str], token: str) -> str:
    compact_token = token.replace(" ", "")
    for index, header in enumerate(headers, start=1):
        if compact_token in str(header or "").replace(" ", ""):
            return get_column_letter(index)
    return ""
