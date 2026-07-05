from __future__ import annotations

from collections import defaultdict
from copy import copy
from datetime import datetime
from pathlib import Path
from statistics import mean
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter

from .fill_engine import FillEngine
from .knowledge_base import KnowledgeBase
from .normalization import normalize_key_part
from .schemas import FIELD_COLUMNS

PRICE_METRIC = "基价"
PHYSICAL_METRIC = "实物工作费调整系数"
TECHNICAL_METRIC = "技术工作费调整系数"
DEFAULT_SELECTED_EXPERIENCE_FIELDS = [PRICE_METRIC, PHYSICAL_METRIC, TECHNICAL_METRIC]
WARNING_PARAMETER_FIELD = "预警参数"
WARNING_DETAIL_FIELD = "预警细节"
WARNING_OUTPUT_FIELDS = [WARNING_PARAMETER_FIELD, WARNING_DETAIL_FIELD]
WARNING_HIGH_FILL = PatternFill(fill_type="solid", fgColor="F4CCCC")
WARNING_MEDIUM_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
WARNING_FONT = Font(color="9C0006")
WARNING_NONE_FILL = PatternFill(fill_type="solid", fgColor="E8F5E9")
WARNING_NONE_FONT = Font(color="166534")
MATCH_MODE_EXACT = "字段完全匹配"
MATCH_MODE_ORDERED = "非空要素顺序匹配"
MAX_CONSECUTIVE_BLANK_FILTER_ROWS = 200
NO_WARNING_PARAMETER_TEXT = "无预警"
DEFAULT_LOW_RISK_WARNING_PERCENT = 5.0
DEFAULT_HIGH_RISK_WARNING_PERCENT = 20.0
DEFAULT_WARNING_FILTER_FIELD = "数量"
WARNING_FILTER_FIELDS = [DEFAULT_WARNING_FILTER_FIELD]
WARNING_FILTER_FIELD_ALIASES = {
    DEFAULT_WARNING_FILTER_FIELD: ["数量", "工程量", "工程数量", "工程量合计"],
}

LEGACY_EXPERIENCE_POOL_HEADERS = [
    "来源文件",
    "要素1",
    "要素2",
    "要素3",
    "要素4",
    "要素5",
    "单位",
    PRICE_METRIC,
    "基价说明",
    PHYSICAL_METRIC,
    "实物说明",
    TECHNICAL_METRIC,
    "技术说明",
    "来源sheet",
    "来源行",
    "导入时间",
]

EXPERIENCE_MAPPING_FIELDS = [
    *FIELD_COLUMNS,
    PRICE_METRIC,
    "工程量",
    PHYSICAL_METRIC,
    TECHNICAL_METRIC,
    "其他参数1",
    "其他参数2",
    "原表备注1",
    "原表备注2",
    "原表备注3",
]

EXPERIENCE_POOL_HEADERS = [
    "来源文件",
    "来源sheet",
    "来源行",
    "导入时间",
    "导入批次",
    "来源表头行",
    "要素1列名",
    "要素1",
    "要素2列名",
    "要素2",
    "要素3列名",
    "要素3",
    "要素4列名",
    "要素4",
    "要素5列名",
    "要素5",
    "单位列名",
    "单位",
    "基价列名",
    PRICE_METRIC,
    "工程量列名",
    "工程量",
    "实物工作费调整系数列名",
    PHYSICAL_METRIC,
    "技术工作费调整系数列名",
    TECHNICAL_METRIC,
    "其他参数1列名称",
    "其他参数1读取数值",
    "其他参数2列名称",
    "其他参数2读取数值",
    "原表备注1列名",
    "原表备注1",
    "原表备注2列名",
    "原表备注2",
    "原表备注3列名",
    "原表备注3",
    "数据签名",
]

SOURCE_SHEET_LAYOUTS = {
    "表2": {
        "key_columns": {
            "要素1": "B",
            "要素2": "C",
            "要素3": None,
            "要素4": "D",
            "要素5": "E",
            "单位": "F",
        },
        "experience_columns": {
            PRICE_METRIC: ("Q", "R"),
            PHYSICAL_METRIC: ("S", "T"),
            TECHNICAL_METRIC: ("U", "V"),
        },
    },
    "表3": {
        "key_columns": {
            "要素1": "B",
            "要素2": "C",
            "要素3": "D",
            "要素4": "E",
            "要素5": "F",
            "单位": "G",
        },
        "experience_columns": {
            PRICE_METRIC: ("AL", "AM"),
            PHYSICAL_METRIC: ("AN", "AO"),
            TECHNICAL_METRIC: ("AP", "AQ"),
        },
    },
    "表4": {
        "key_columns": {
            "要素1": "B",
            "要素2": "C",
            "要素3": "D",
            "要素4": "E",
            "要素5": "F",
            "单位": "G",
        },
        "experience_columns": {
            PRICE_METRIC: ("V", "W"),
            PHYSICAL_METRIC: ("X", "Y"),
            TECHNICAL_METRIC: ("Z", "AA"),
        },
    },
}


def import_experience_pool(
    source_path: str | Path,
    pool_path: str | Path,
    selected_fields: list[str] | None = None,
    sheet_configs: list[dict[str, Any]] | None = None,
    template_path: str | Path | None = None,
    filter_non_empty_field: str | None = None,
) -> dict[str, Any]:
    selected = set(selected_fields or DEFAULT_SELECTED_EXPERIENCE_FIELDS)
    source_path = Path(source_path)
    pool_path = Path(pool_path)
    pool_path.parent.mkdir(parents=True, exist_ok=True)
    template = Path(template_path) if template_path else None
    pool_headers = _pool_headers_from_template(template)
    pool_workbook, pool_sheet = _open_or_create_pool(pool_path, pool_headers, template)
    source_workbook = load_workbook(source_path, data_only=True)
    imported_rows = 0
    skipped_rows = 0
    try:
        imported_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        import_batch = datetime.now().strftime("%Y%m%d%H%M%S")
        if sheet_configs:
            imported_rows, skipped_rows = _import_mapped_sheets(
                source_workbook,
                pool_sheet,
                source_path.name,
                selected,
                sheet_configs,
                imported_at,
                import_batch,
                pool_headers,
                filter_non_empty_field=filter_non_empty_field,
            )
            pool_workbook.save(pool_path)
            return {
                "pool_path": str(pool_path),
                "source_file": source_path.name,
                "imported_rows": imported_rows,
                "skipped_rows": skipped_rows,
                "selected_fields": sorted(selected),
            }
        for sheet_name in source_workbook.sheetnames:
            layout = _layout_for_sheet(sheet_name)
            if not layout:
                continue
            sheet = source_workbook[sheet_name]
            merged_value_map = FillEngine._build_merged_value_map(sheet)
            source_column_names = _legacy_source_column_names(sheet, layout)
            for excel_row in range(5, sheet.max_row + 1):
                key_values = _read_key_values(sheet, excel_row, layout["key_columns"], merged_value_map)
                if not _has_key_content(key_values):
                    skipped_rows += 1
                    continue
                experience = _read_experience_values(
                    sheet,
                    excel_row,
                    layout["experience_columns"],
                    selected,
                    merged_value_map,
                )
                if not any(_normalize_number(value) is not None for value, _note in experience.values()):
                    skipped_rows += 1
                    continue
                _append_pool_row(
                    pool_sheet,
                    _pool_row(
                        source_path.name,
                        key_values,
                        experience,
                        sheet_name,
                        excel_row,
                        imported_at,
                        source_column_names=source_column_names,
                        remarks={
                            "原表备注1": experience.get(PRICE_METRIC, (None, None))[1],
                            "原表备注2": experience.get(PHYSICAL_METRIC, (None, None))[1],
                            "原表备注3": experience.get(TECHNICAL_METRIC, (None, None))[1],
                        },
                        header_row=4,
                        import_batch=import_batch,
                        pool_headers=pool_headers,
                    ),
                )
                imported_rows += 1
        pool_workbook.save(pool_path)
    finally:
        pool_workbook.close()
        source_workbook.close()
    return {
        "pool_path": str(pool_path),
        "source_file": source_path.name,
        "imported_rows": imported_rows,
        "skipped_rows": skipped_rows,
        "selected_fields": sorted(selected),
    }


def analyze_workbook_warnings(workbook_path: str | Path, pool_path: str | Path) -> dict[str, Any]:
    return analyze_workbook_warnings_with_progress(workbook_path, pool_path)


def analyze_workbook_warnings_with_progress(
    workbook_path: str | Path,
    pool_path: str | Path,
    progress_callback: Any | None = None,
    low_risk_warning_ratio: float = DEFAULT_LOW_RISK_WARNING_PERCENT / 100,
    high_risk_warning_ratio: float = DEFAULT_HIGH_RISK_WARNING_PERCENT / 100,
    only_check_rows_with_value: bool = True,
    value_filter_field: str = DEFAULT_WARNING_FILTER_FIELD,
) -> dict[str, Any]:
    workbook_path = Path(workbook_path)
    pool_path = Path(pool_path)
    if not pool_path.exists():
        return {
            "summary": {
                "pool_enabled": False,
                "checked_rows": 0,
                "warning_rows": 0,
                "high_rows": 0,
                "low_rows": 0,
                "medium_rows": 0,
                "metric_counts": {},
                "summary_text": "经验池未启用：尚未导入经验池，本次未执行经验值预警。",
            },
            "warnings": [],
        }

    pool_matcher = _load_pool_matcher(pool_path)
    workbook = load_workbook(workbook_path, data_only=True)
    total_candidate_rows = _count_warning_candidate_rows(
        workbook,
        only_check_rows_with_value=only_check_rows_with_value,
        value_filter_field=value_filter_field,
    )
    warnings: list[dict[str, Any]] = []
    row_results: list[dict[str, Any]] = []
    candidate_rows = 0
    checked_rows = 0
    no_comparable_rows = 0
    match_mode_counts: dict[str, int] = {}
    low_risk_warning_percent = _ratio_to_percent(low_risk_warning_ratio)
    high_risk_warning_percent = _ratio_to_percent(high_risk_warning_ratio)
    if progress_callback:
        progress_callback(
            {
                "status": "running",
                "processed_rows": 0,
                "total_rows": total_candidate_rows,
                "matched_rows": 0,
                "warning_rows": 0,
            }
        )
    try:
        for sheet in workbook.worksheets:
            merged_value_map = FillEngine._build_merged_value_map(sheet)
            header_row = _detect_header_row(sheet, merged_value_map)
            headers = _headers_at(sheet, header_row, merged_value_map)
            header_map = _header_map(headers)
            if not all(field in header_map for field in ("要素1", "单位")):
                continue
            metric_columns = _metric_columns(headers)
            if not metric_columns:
                continue
            filter_column_index = _warning_filter_column_index(headers, value_filter_field) if only_check_rows_with_value else None
            if only_check_rows_with_value and not filter_column_index:
                raise ValueError(f"预警 sheet {sheet.title} 未映射过滤字段：{value_filter_field}")
            for excel_row in range(header_row + 1, sheet.max_row + 1):
                row_values = _row_values(sheet, excel_row, len(headers), merged_value_map)
                key_values = {
                    field: _value_from_row(row_values, header_map.get(field))
                    for field in FIELD_COLUMNS
                }
                if not _has_key_content(key_values):
                    continue
                if only_check_rows_with_value:
                    filter_value = _value_from_row(row_values, filter_column_index)
                    if not _has_warning_filter_value(filter_value):
                        continue
                candidate_rows += 1
                match_mode, match_mode_detail, records = pool_matcher.lookup(key_values)
                if not records:
                    no_comparable_rows += 1
                    continue
                checked_rows += 1
                if match_mode:
                    match_mode_counts[match_mode] = match_mode_counts.get(match_mode, 0) + 1
                compared_metrics: list[dict[str, Any]] = []
                unavailable_metrics: list[dict[str, Any]] = []
                observed_metrics: list[str] = []
                row_warning_items: list[dict[str, Any]] = []
                for metric, column_index in metric_columns.items():
                    current_value = _normalize_number(_value_from_row(row_values, column_index))
                    if current_value is None:
                        continue
                    observed_metrics.append(metric)
                    experience_values = _unique_numbers(record.get(metric) for record in records)
                    if not experience_values:
                        unavailable_metrics.append(
                            {
                                "metric": metric,
                                "reason": "该参数无可比经验值",
                            }
                        )
                        continue
                    metric_result = _metric_warning_result(
                        metric,
                        current_value,
                        experience_values,
                        low_risk_warning_ratio,
                        high_risk_warning_ratio,
                    )
                    compared_metrics.append(metric_result)
                    if metric_result["severity"] == "none":
                        continue
                    warning = {
                        "sheet_name": sheet.title,
                        "excel_row": excel_row,
                        **metric_result,
                        "row_key": _format_row_key(key_values),
                        "match_mode": match_mode,
                        "match_mode_detail": match_mode_detail,
                        "low_risk_threshold_percent": low_risk_warning_percent,
                        "high_risk_threshold_percent": high_risk_warning_percent,
                        "source_rows": [
                            {
                                "source_file": record.get("来源文件", ""),
                                "source_sheet": record.get("来源sheet", ""),
                                "source_row": record.get("来源行", ""),
                                "value": record.get(metric),
                                "note": _metric_note(record, metric),
                            }
                            for record in records
                            if _normalize_number(record.get(metric)) is not None
                        ],
                    }
                    warnings.append(warning)
                    row_warning_items.append(warning)
                    warnings[-1]["warning_parameter"] = _warning_parameter_text(warnings[-1])
                    warnings[-1]["warning_detail"] = _warning_detail_text(warnings[-1])
                if observed_metrics:
                    row_result: dict[str, Any] = {
                        "sheet_name": sheet.title,
                        "excel_row": excel_row,
                        "row_key": _format_row_key(key_values),
                        "match_mode": match_mode,
                        "match_mode_detail": match_mode_detail,
                        "observed_metrics": observed_metrics,
                        "compared_metrics": compared_metrics,
                        "unavailable_metrics": unavailable_metrics,
                        "low_risk_threshold_percent": low_risk_warning_percent,
                        "high_risk_threshold_percent": high_risk_warning_percent,
                    }
                    if row_warning_items:
                        row_result["warning_parameter"] = _row_warning_parameters(row_warning_items)
                        row_result["warning_detail"] = _row_warning_details(row_warning_items)
                        row_result["severity"] = _row_warning_severity(row_warning_items)
                    else:
                        row_result["warning_parameter"] = NO_WARNING_PARAMETER_TEXT
                        row_result["warning_detail"] = _no_warning_detail_text(
                            match_mode_detail,
                            compared_metrics,
                            unavailable_metrics,
                            low_risk_warning_percent,
                            high_risk_warning_percent,
                        )
                        row_result["severity"] = "none"
                    row_results.append(row_result)
                if progress_callback:
                    progress_callback(
                        {
                            "status": "running",
                            "processed_rows": candidate_rows,
                            "total_rows": total_candidate_rows,
                            "matched_rows": checked_rows,
                            "warning_rows": len(warnings),
                        }
                    )
    finally:
        workbook.close()
    if progress_callback:
        progress_callback(
            {
                "status": "completed",
                "processed_rows": candidate_rows,
                "total_rows": total_candidate_rows,
                "matched_rows": checked_rows,
                "warning_rows": len(warnings),
            }
        )

    high_rows = sum(1 for item in warnings if item["severity"] == "high")
    low_rows = sum(1 for item in warnings if item["severity"] == "low")
    metric_counts = _metric_counts(warnings)
    summary = {
        "pool_enabled": True,
        "total_candidate_rows": total_candidate_rows,
        "candidate_rows": candidate_rows,
        "checked_rows": checked_rows,
        "no_comparable_rows": no_comparable_rows,
        "warning_rows": len(warnings),
        "high_rows": high_rows,
        "low_rows": low_rows,
        "medium_rows": low_rows,
        "metric_counts": metric_counts,
        "match_mode_counts": match_mode_counts,
        "low_risk_threshold_percent": low_risk_warning_percent,
        "high_risk_threshold_percent": high_risk_warning_percent,
    }
    summary["summary_text"] = _summary_text(summary)
    return {
        "summary": summary,
        "warnings": warnings,
        "row_results": row_results,
    }


def write_warnings_to_workbook(workbook_path: str | Path, warning_rows: list[dict[str, Any]]) -> dict[str, Any]:
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        return {"written_rows": 0, "warning_columns": WARNING_OUTPUT_FIELDS}
    workbook = load_workbook(workbook_path)
    grouped = _group_warnings_by_row(warning_rows)
    written_rows = 0
    try:
        for sheet in workbook.worksheets:
            merged_value_map = FillEngine._build_merged_value_map(sheet)
            header_row = _detect_header_row(sheet, merged_value_map)
            headers = _headers_at(sheet, header_row, merged_value_map)
            if not headers:
                continue
            output_columns = _ensure_warning_columns(sheet, header_row)
            sheet_warnings = grouped.get(sheet.title, {})
            for excel_row, warning_row in sheet_warnings.items():
                if excel_row <= header_row:
                    continue
                parameter_cell = sheet.cell(row=excel_row, column=output_columns[WARNING_PARAMETER_FIELD])
                detail_cell = sheet.cell(row=excel_row, column=output_columns[WARNING_DETAIL_FIELD])
                parameter_cell.value = warning_row.get("warning_parameter") or ""
                detail_cell.value = warning_row.get("warning_detail") or ""
                severity = str(warning_row.get("severity") or "")
                if severity == "high":
                    fill = WARNING_HIGH_FILL
                    font = WARNING_FONT
                elif severity == "low":
                    fill = WARNING_MEDIUM_FILL
                    font = WARNING_FONT
                else:
                    fill = WARNING_NONE_FILL
                    font = WARNING_NONE_FONT
                for cell in (parameter_cell, detail_cell):
                    cell.fill = fill
                    cell.font = font
                written_rows += 1
        if written_rows or grouped:
            workbook.save(workbook_path)
    finally:
        workbook.close()
    return {"written_rows": written_rows, "warning_columns": WARNING_OUTPUT_FIELDS}


def build_warning_report_lines(
    warning_summary: dict[str, Any],
    warning_details: list[dict[str, Any]],
    limit: int = 8,
) -> list[str]:
    if not warning_summary:
        return ["经验池预警：未执行经验池预警分析。"]
    if warning_summary.get("executed") is False:
        return [str(warning_summary.get("summary_text") or "经验池预警尚未执行：点击运行后再生成预警明细。")]
    if not warning_summary.get("pool_enabled"):
        return [str(warning_summary.get("summary_text") or "经验池未启用：尚未导入经验池，本次未执行经验值预警。")]

    lines = [str(warning_summary.get("summary_text") or _summary_text(warning_summary))]
    if not warning_details:
        lines.append("预警明细：当前输出数字未触发经验池预警。")
    else:
        lines.append("预警明细：以下列出重点预警，经验值仅用于比选和人工复核。")
        for warning in warning_details[:limit]:
            match_mode_detail = str(warning.get("match_mode_detail") or warning.get("match_mode") or "").strip()
            lines.append(
                f"{warning.get('severity_label', _severity_label(str(warning.get('severity', ''))))}："
                f"{warning.get('sheet_name', '')} 第 {warning.get('excel_row', '')} 行："
                f"{match_mode_detail + '；' if match_mode_detail else ''}"
                f"{warning.get('metric', '')} 当前值 {warning.get('current_value', '')}；"
                f"经验池平均值 {warning.get('experience_average', '')}；"
                f"经验范围 {warning.get('experience_range_text', '')}；"
                f"实际偏离率 {warning.get('deviation_percent', '')}%；"
                f"{warning.get('suggested_action') or warning.get('message', '')}"
            )
        if len(warning_details) > limit:
            lines.append(f"其余 {len(warning_details) - limit} 条预警请查看前端预览区或输出 JSON 明细。")
    lines.append("说明：经验池仅用于预警和比选，不参与基价匹配、系数回填或第二层经验提示。")
    return lines


def _group_warnings_by_row(warning_rows: list[dict[str, Any]]) -> dict[str, dict[int, dict[str, Any]]]:
    grouped: dict[str, dict[int, dict[str, Any]]] = {}
    for warning in warning_rows:
        sheet_name = str(warning.get("sheet_name") or "").strip()
        try:
            excel_row = int(warning.get("excel_row") or 0)
        except (TypeError, ValueError):
            continue
        if not sheet_name or excel_row <= 0:
            continue
        grouped.setdefault(sheet_name, {})[excel_row] = warning
    return grouped


def _ensure_warning_columns(sheet: Any, header_row: int) -> dict[str, int]:
    header_map = _header_map(_headers_at(sheet, header_row))
    columns: dict[str, int] = {}
    for field in WARNING_OUTPUT_FIELDS:
        if field in header_map:
            columns[field] = header_map[field]
    next_column = _last_used_column(sheet, header_row) + 1
    for field in WARNING_OUTPUT_FIELDS:
        if field in columns:
            continue
        while sheet.cell(row=header_row, column=next_column).value:
            next_column += 1
        sheet.cell(row=header_row, column=next_column).value = field
        columns[field] = next_column
        next_column += 1
    return columns


def _last_used_column(sheet: Any, header_row: int) -> int:
    max_row = min(sheet.max_row, max(header_row + 30, 30))
    last_column = 0
    for row in sheet.iter_rows(min_row=1, max_row=max_row):
        for cell in row:
            if cell.value not in (None, ""):
                last_column = max(last_column, cell.column)
    return last_column or sheet.max_column


def _row_warning_parameters(warnings: list[dict[str, Any]]) -> str:
    metrics = []
    for warning in warnings:
        metric = str(warning.get("metric") or "").strip()
        if metric and metric not in metrics:
            metrics.append(metric)
    return "；".join(metrics)


def _row_warning_details(warnings: list[dict[str, Any]]) -> str:
    return "；".join(
        str(warning.get("warning_detail") or _warning_detail_text(warning)).strip()
        for warning in warnings
        if str(warning.get("warning_detail") or _warning_detail_text(warning)).strip()
    )


def _row_warning_severity(warnings: list[dict[str, Any]]) -> str:
    if any(str(warning.get("severity") or "") == "high" for warning in warnings):
        return "high"
    if any(str(warning.get("severity") or "") == "low" for warning in warnings):
        return "low"
    return "none"


def _warning_parameter_text(warning: dict[str, Any]) -> str:
    return str(warning.get("metric") or "").strip()


def _warning_detail_text(warning: dict[str, Any]) -> str:
    metric = str(warning.get("metric") or "").strip()
    current_value = warning.get("current_value", "")
    severity_label = warning.get("severity_label") or _severity_label(str(warning.get("severity") or ""))
    match_mode_detail = str(warning.get("match_mode_detail") or warning.get("match_mode") or "").strip()
    message = str(warning.get("message") or "").strip()
    source_text = _source_comparison_text(warning.get("source_rows") or [], metric)
    suggested_action = str(warning.get("suggested_action") or "").strip()
    average_value = warning.get("experience_average", "")
    experience_min = warning.get("experience_min", "")
    experience_max = warning.get("experience_max", "")
    deviation_percent = warning.get("deviation_percent", "")
    low_threshold_percent = warning.get("low_risk_threshold_percent", "")
    high_threshold_percent = warning.get("high_risk_threshold_percent", "")
    parts = [f"{severity_label}：{message}" if message else f"{severity_label}：{metric} 当前值 {current_value} 触发经验池预警。"]
    parts.append(f"经验池平均值 {average_value}")
    parts.append(f"经验范围 {experience_min}~{experience_max}")
    parts.append(f"实际偏离率 {deviation_percent}%")
    parts.append(f"低风险阈值 {low_threshold_percent}%")
    parts.append(f"高风险阈值 {high_threshold_percent}%")
    if match_mode_detail:
        parts.append(f"匹配模式：{match_mode_detail}")
    if source_text:
        parts.append(f"来源：{source_text.removeprefix('对比其他项目发现：').removesuffix('。')}")
    if suggested_action and suggested_action != message:
        parts.append(suggested_action)
    return "；".join(part for part in parts if part)


def _no_warning_detail_text(
    match_mode_detail: str | None,
    compared_metrics: list[dict[str, Any]],
    unavailable_metrics: list[dict[str, Any]],
    low_threshold_percent: float,
    high_threshold_percent: float,
) -> str:
    compared_parts = []
    for item in compared_metrics:
        metric = str(item.get("metric") or "").strip()
        if not metric:
            continue
        compared_parts.append(
            f"{metric}（平均值 {item.get('experience_average')}；实际偏离率 {item.get('deviation_percent')}%；"
            f"低风险阈值 {low_threshold_percent:g}% ；高风险阈值 {high_threshold_percent:g}%）"
        )
    unavailable_parts = [
        f"{str(item.get('metric') or '').strip()}（{str(item.get('reason') or '').strip()}）"
        for item in unavailable_metrics
        if str(item.get("metric") or "").strip()
    ]
    parts = [f"已比对参数：{'；'.join(compared_parts) if compared_parts else '无'}"]
    if unavailable_parts:
        parts.append(f"无可比经验值：{'；'.join(unavailable_parts)}")
    if match_mode_detail:
        parts.append(f"匹配模式：{match_mode_detail}")
    parts.append(f"当前阈值：低风险 {low_threshold_percent:g}%；高风险 {high_threshold_percent:g}%")
    parts.append("结论：未超过阈值，故无预警。")
    return "；".join(parts)


def _source_comparison_text(source_rows: list[dict[str, Any]], metric: str, limit: int = 3) -> str:
    comparisons = []
    for source in source_rows:
        value = source.get("value")
        if _normalize_number(value) is None:
            continue
        source_file = str(source.get("source_file") or "其他项目").strip()
        source_sheet = str(source.get("source_sheet") or "").strip()
        source_row = str(source.get("source_row") or "").strip()
        location = source_file
        if source_sheet:
            location += f" / {source_sheet}"
        if source_row:
            location += f" 第{source_row}行"
        note = str(source.get("note") or "").strip()
        text = f"{location} 的 {metric} 为 {value}"
        if note:
            text += f"（{note}）"
        comparisons.append(text)
        if len(comparisons) >= limit:
            break
    if not comparisons:
        return ""
    suffix = "等" if len(source_rows) > limit else ""
    return f"对比其他项目发现：{'；'.join(comparisons)}{suffix}。"


def _pool_headers_from_template(template_path: Path | None) -> list[str]:
    if not template_path or not template_path.exists():
        return EXPERIENCE_POOL_HEADERS
    workbook = load_workbook(template_path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        headers = [str(value).strip() if value is not None else "" for value in _headers_at(sheet, 1)]
    finally:
        workbook.close()
    _validate_pool_headers(headers, template_path)
    return headers


def _validate_pool_headers(headers: list[str], source: Path) -> None:
    required_headers = [
        "来源文件",
        "来源sheet",
        "来源行",
        "导入时间",
        "要素1",
        "单位",
        PRICE_METRIC,
        PHYSICAL_METRIC,
        TECHNICAL_METRIC,
    ]
    missing = [header for header in required_headers if header not in headers]
    if missing:
        raise ValueError(f"经验池模板缺少必要字段：{source}，缺少 {', '.join(missing)}")


def _open_or_create_pool(pool_path: Path, pool_headers: list[str] | None = None, template_path: Path | None = None) -> tuple[Any, Any]:
    pool_headers = pool_headers or EXPERIENCE_POOL_HEADERS
    if pool_path.exists():
        workbook = load_workbook(pool_path)
        sheet = workbook.worksheets[0]
        headers = _headers_at(sheet, 1)
        if headers == LEGACY_EXPERIENCE_POOL_HEADERS:
            sheet = _migrate_legacy_pool_sheet(workbook, sheet, pool_headers)
        elif headers != pool_headers:
            raise ValueError(f"经验池字段不符合当前版本要求：{pool_path}")
        if template_path and template_path.exists():
            _apply_pool_template_styles(sheet, template_path, pool_headers)
        return workbook, sheet
    if template_path and template_path.exists():
        workbook = load_workbook(template_path)
        sheet = workbook.worksheets[0]
        headers = [str(value).strip() if value is not None else "" for value in _headers_at(sheet, 1)]
        _validate_pool_headers(headers, template_path)
        if sheet.max_row >= 2:
            _cache_pool_template_data_row_style(sheet, 2, len(headers))
        if sheet.max_row > 1:
            sheet.delete_rows(2, sheet.max_row - 1)
        return workbook, sheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "经验池"
    sheet.append(pool_headers)
    return workbook, sheet


def _cache_pool_template_data_row_style(sheet: Any, row_index: int, max_column: int) -> None:
    styles = []
    for column in range(1, max_column + 1):
        cell = sheet.cell(row=row_index, column=column)
        styles.append(
            {
                "style": copy(cell._style) if cell.has_style else None,
                "font": copy(cell.font),
                "fill": copy(cell.fill),
                "border": copy(cell.border),
                "alignment": copy(cell.alignment),
                "protection": copy(cell.protection),
                "number_format": cell.number_format,
            }
        )
    sheet._pool_template_data_row_styles = styles
    sheet._pool_template_data_row_height = sheet.row_dimensions[row_index].height


def _apply_pool_template_styles(sheet: Any, template_path: Path, pool_headers: list[str]) -> None:
    template_workbook = load_workbook(template_path)
    try:
        template_sheet = template_workbook.worksheets[0]
        headers = [str(value).strip() if value is not None else "" for value in _headers_at(template_sheet, 1)]
        if headers != pool_headers:
            return
        max_column = len(pool_headers)
        sheet.freeze_panes = template_sheet.freeze_panes
        sheet.sheet_view.showGridLines = template_sheet.sheet_view.showGridLines
        for column in range(1, max_column + 1):
            letter = get_column_letter(column)
            source_dimension = template_sheet.column_dimensions[letter]
            target_dimension = sheet.column_dimensions[letter]
            target_dimension.width = source_dimension.width
            target_dimension.hidden = source_dimension.hidden
            _copy_cell_style(template_sheet.cell(row=1, column=column), sheet.cell(row=1, column=column))
        if template_sheet.max_row >= 2:
            sheet.row_dimensions[1].height = template_sheet.row_dimensions[1].height
            sheet.row_dimensions[2].height = template_sheet.row_dimensions[2].height
            for row in range(2, sheet.max_row + 1):
                if template_sheet.row_dimensions[2].height is not None:
                    sheet.row_dimensions[row].height = template_sheet.row_dimensions[2].height
                for column in range(1, max_column + 1):
                    _copy_cell_style(template_sheet.cell(row=2, column=column), sheet.cell(row=row, column=column))
    finally:
        template_workbook.close()


def _copy_cell_style(source: Any, target: Any) -> None:
    if source.has_style:
        target._style = copy(source._style)
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)
    target.number_format = source.number_format


def _append_pool_row(sheet: Any, values: list[Any]) -> None:
    target_row = sheet.max_row + 1
    row_styles = getattr(sheet, "_pool_template_data_row_styles", None)
    row_height = getattr(sheet, "_pool_template_data_row_height", None)
    for column, value in enumerate(values, start=1):
        cell = sheet.cell(row=target_row, column=column, value=value)
        if not row_styles or column > len(row_styles):
            continue
        style = row_styles[column - 1]
        if style.get("style") is not None:
            cell._style = copy(style["style"])
        cell.font = copy(style["font"])
        cell.fill = copy(style["fill"])
        cell.border = copy(style["border"])
        cell.alignment = copy(style["alignment"])
        cell.protection = copy(style["protection"])
        cell.number_format = style["number_format"]
    if row_height is not None:
        sheet.row_dimensions[target_row].height = row_height


def _migrate_legacy_pool_sheet(workbook: Any, sheet: Any, pool_headers: list[str]) -> Any:
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    records = []
    for row in rows:
        records.append(
            {
                header: row[index] if index < len(row) else None
                for index, header in enumerate(LEGACY_EXPERIENCE_POOL_HEADERS)
            }
        )
    workbook.remove(sheet)
    new_sheet = workbook.create_sheet("经验池", 0)
    new_sheet.append(EXPERIENCE_POOL_HEADERS)
    for record in records:
        key_values = {field: record.get(field) for field in FIELD_COLUMNS}
        experience = {
            PRICE_METRIC: (record.get(PRICE_METRIC), record.get("基价说明")),
            PHYSICAL_METRIC: (record.get(PHYSICAL_METRIC), record.get("实物说明")),
            TECHNICAL_METRIC: (record.get(TECHNICAL_METRIC), record.get("技术说明")),
        }
        _append_pool_row(
            new_sheet,
            _pool_row(
                str(record.get("来源文件") or ""),
                key_values,
                experience,
                str(record.get("来源sheet") or ""),
                int(record.get("来源行") or 0),
                str(record.get("导入时间") or ""),
                remarks={
                    "原表备注1": record.get("基价说明"),
                    "原表备注2": record.get("实物说明"),
                    "原表备注3": record.get("技术说明"),
                },
                pool_headers=pool_headers,
            ),
        )
    return new_sheet


def _layout_for_sheet(sheet_name: str) -> dict[str, Any] | None:
    for token, layout in SOURCE_SHEET_LAYOUTS.items():
        if token in sheet_name:
            return layout
    return None


def _read_key_values(
    sheet: Any,
    excel_row: int,
    key_columns: dict[str, str | None],
    merged_value_map: dict[tuple[int, int], Any],
) -> dict[str, Any]:
    return {
        field: _read_cell(sheet, excel_row, column, merged_value_map) if column else None
        for field, column in key_columns.items()
    }


def _read_experience_values(
    sheet: Any,
    excel_row: int,
    columns: dict[str, tuple[str, str]],
    selected: set[str],
    merged_value_map: dict[tuple[int, int], Any],
) -> dict[str, tuple[Any, Any]]:
    experience: dict[str, tuple[Any, Any]] = {}
    for metric, (value_column, note_column) in columns.items():
        if metric not in selected:
            experience[metric] = (None, None)
            continue
        experience[metric] = (
            _read_cell(sheet, excel_row, value_column, merged_value_map),
            _read_cell(sheet, excel_row, note_column, merged_value_map),
        )
    return experience


def _import_mapped_sheets(
    source_workbook: Any,
    pool_sheet: Any,
    source_file: str,
    selected: set[str],
    sheet_configs: list[dict[str, Any]],
    imported_at: str,
    import_batch: str,
    pool_headers: list[str],
    filter_non_empty_field: str | None = None,
) -> tuple[int, int]:
    imported_rows = 0
    skipped_rows = 0
    for config in sheet_configs:
        if not config.get("enabled", True):
            continue
        sheet_name = str(config.get("sheet_name") or "").strip()
        if not sheet_name or sheet_name not in source_workbook.sheetnames:
            skipped_rows += 1
            continue
        sheet = source_workbook[sheet_name]
        header_row = max(1, int(config.get("header_row") or 1))
        column_mapping = {
            str(key): str(value).strip()
            for key, value in dict(config.get("column_mapping") or {}).items()
            if value is not None
        }
        if filter_non_empty_field and not column_mapping.get(filter_non_empty_field):
            raise ValueError(f"已启用只导入有值行，但 {sheet_name} 未映射过滤字段：{filter_non_empty_field}")
        merged_value_map = FillEngine._build_merged_value_map(sheet)
        source_column_names = {
            field: _mapped_source_column_name(sheet, header_row, column_mapping.get(field), field)
            for field in EXPERIENCE_MAPPING_FIELDS
        }
        consecutive_blank_filter_rows = 0
        for excel_row in range(header_row + 1, sheet.max_row + 1):
            if filter_non_empty_field:
                filter_value = _read_mapped_column(sheet, excel_row, column_mapping.get(filter_non_empty_field), merged_value_map)
                if _is_blank_value(filter_value):
                    skipped_rows += 1
                    consecutive_blank_filter_rows += 1
                    if consecutive_blank_filter_rows >= MAX_CONSECUTIVE_BLANK_FILTER_ROWS:
                        break
                    continue
                consecutive_blank_filter_rows = 0
            key_values = {
                field: _read_mapped_column(sheet, excel_row, column_mapping.get(field), merged_value_map)
                for field in FIELD_COLUMNS
            }
            if not _has_key_content(key_values):
                skipped_rows += 1
                continue
            experience = {
                metric: (
                    _read_mapped_column(sheet, excel_row, column_mapping.get(metric), merged_value_map)
                    if metric in selected
                    else None,
                    None,
                )
                for metric in DEFAULT_SELECTED_EXPERIENCE_FIELDS
            }
            if not any(_normalize_number(value) is not None for value, _note in experience.values()):
                skipped_rows += 1
                continue
            _append_pool_row(
                pool_sheet,
                _pool_row(
                    source_file,
                    key_values,
                    experience,
                    sheet_name,
                    excel_row,
                    imported_at,
                    source_column_names=source_column_names,
                    quantity=_read_mapped_column(sheet, excel_row, column_mapping.get("工程量"), merged_value_map),
                    extra_params={
                        "其他参数1": _read_mapped_column(sheet, excel_row, column_mapping.get("其他参数1"), merged_value_map),
                        "其他参数2": _read_mapped_column(sheet, excel_row, column_mapping.get("其他参数2"), merged_value_map),
                    },
                    remarks={
                        "原表备注1": _read_mapped_column(sheet, excel_row, column_mapping.get("原表备注1"), merged_value_map),
                        "原表备注2": _read_mapped_column(sheet, excel_row, column_mapping.get("原表备注2"), merged_value_map),
                        "原表备注3": _read_mapped_column(sheet, excel_row, column_mapping.get("原表备注3"), merged_value_map),
                    },
                    header_row=header_row,
                    import_batch=import_batch,
                    pool_headers=pool_headers,
                ),
            )
            imported_rows += 1
    return imported_rows, skipped_rows


def _is_blank_value(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _read_mapped_column(
    sheet: Any,
    excel_row: int,
    column: str | None,
    merged_value_map: dict[tuple[int, int], Any],
) -> Any:
    if not column or column == "空元素列":
        return None
    try:
        column_index = column_index_from_string(column)
    except ValueError:
        return None
    return FillEngine._read_mapped_value(sheet, excel_row, column_index, merged_value_map)


def _read_cell(
    sheet: Any,
    excel_row: int,
    column: str,
    merged_value_map: dict[tuple[int, int], Any],
) -> Any:
    return FillEngine._read_mapped_value(
        sheet,
        excel_row,
        column_index_from_string(column),
        merged_value_map,
    )


def _legacy_source_column_names(sheet: Any, layout: dict[str, Any]) -> dict[str, str]:
    names: dict[str, str] = {}
    for field, column in layout["key_columns"].items():
        names[field] = _mapped_source_column_name(sheet, 4, column, field) if column else ""
    for metric, (value_column, _note_column) in layout["experience_columns"].items():
        names[metric] = _mapped_source_column_name(sheet, 4, value_column, metric)
    return names


def _mapped_source_column_name(sheet: Any, header_row: int, column: str | None, fallback: str) -> str:
    if not column or column == "空元素列":
        return ""
    try:
        column_index = column_index_from_string(column)
    except ValueError:
        return ""
    candidates = []
    if header_row > 1:
        candidates.append(sheet.cell(row=header_row - 1, column=column_index).value)
    candidates.append(sheet.cell(row=header_row, column=column_index).value)
    for value in candidates:
        text = str(value or "").strip()
        if text and text != fallback:
            return text
    return str(candidates[-1] or fallback).strip()


def _pool_row(
    source_file: str,
    key_values: dict[str, Any],
    experience: dict[str, tuple[Any, Any]],
    sheet_name: str,
    excel_row: int,
    imported_at: str,
    source_column_names: dict[str, str] | None = None,
    quantity: Any = None,
    extra_params: dict[str, Any] | None = None,
    remarks: dict[str, Any] | None = None,
    header_row: int | None = None,
    import_batch: str | None = None,
    pool_headers: list[str] | None = None,
) -> list[Any]:
    source_column_names = source_column_names or {}
    extra_params = extra_params or {}
    remarks = remarks or {}
    price, price_note = experience.get(PRICE_METRIC, (None, None))
    physical, physical_note = experience.get(PHYSICAL_METRIC, (None, None))
    technical, technical_note = experience.get(TECHNICAL_METRIC, (None, None))
    remark1 = remarks.get("原表备注1", price_note)
    remark2 = remarks.get("原表备注2", physical_note)
    remark3 = remarks.get("原表备注3", technical_note)
    signature = _data_signature(
        source_file,
        sheet_name,
        excel_row,
        key_values,
        {
            PRICE_METRIC: price,
            "工程量": quantity,
            PHYSICAL_METRIC: physical,
            TECHNICAL_METRIC: technical,
        },
    )
    record = {
        "来源文件": source_file,
        "来源sheet": sheet_name,
        "来源行": excel_row,
        "导入时间": imported_at,
        "导入批次": import_batch or "",
        "来源表头行": header_row or "",
        "要素1列名": source_column_names.get("要素1", ""),
        "要素1": key_values.get("要素1"),
        "要素2列名": source_column_names.get("要素2", ""),
        "要素2": key_values.get("要素2"),
        "要素3列名": source_column_names.get("要素3", ""),
        "要素3": key_values.get("要素3"),
        "要素4列名": source_column_names.get("要素4", ""),
        "要素4": key_values.get("要素4"),
        "要素5列名": source_column_names.get("要素5", ""),
        "要素5": key_values.get("要素5"),
        "单位列名": source_column_names.get("单位", ""),
        "单位": key_values.get("单位"),
        "基价列名": source_column_names.get(PRICE_METRIC, ""),
        PRICE_METRIC: _normalize_number(price),
        "工程量列名": source_column_names.get("工程量", ""),
        "工程量": _normalize_number(quantity),
        "实物工作费调整系数列名": source_column_names.get(PHYSICAL_METRIC, ""),
        PHYSICAL_METRIC: _normalize_number(physical),
        "技术工作费调整系数列名": source_column_names.get(TECHNICAL_METRIC, ""),
        TECHNICAL_METRIC: _normalize_number(technical),
        "其他参数1列名称": source_column_names.get("其他参数1", ""),
        "其他参数1读取数值": extra_params.get("其他参数1"),
        "其他参数2列名称": source_column_names.get("其他参数2", ""),
        "其他参数2读取数值": extra_params.get("其他参数2"),
        "原表备注1列名": source_column_names.get("原表备注1", "基价说明" if price_note else ""),
        "原表备注1": remark1,
        "原表备注2列名": source_column_names.get("原表备注2", "实物说明" if physical_note else ""),
        "原表备注2": remark2,
        "原表备注3列名": source_column_names.get("原表备注3", "技术说明" if technical_note else ""),
        "原表备注3": remark3,
        "数据签名": signature,
    }
    return [record.get(header) for header in (pool_headers or EXPERIENCE_POOL_HEADERS)]


def _data_signature(
    source_file: str,
    sheet_name: str,
    excel_row: int,
    key_values: dict[str, Any],
    metric_values: dict[str, Any],
) -> str:
    parts = [source_file, sheet_name, str(excel_row)]
    parts.extend(str(key_values.get(field) or "").strip() for field in FIELD_COLUMNS)
    parts.extend(str(metric_values.get(field) or "").strip() for field in [PRICE_METRIC, "工程量", PHYSICAL_METRIC, TECHNICAL_METRIC])
    return "|".join(parts)


class ExperiencePoolMatcher:
    def __init__(self, records: list[dict[str, Any]]) -> None:
        self._exact_index: dict[tuple[str, ...], list[dict[str, Any]]] = defaultdict(list)
        self._ordered_index: dict[tuple[str, tuple[str, ...]], list[dict[str, Any]]] = defaultdict(list)
        for record in records:
            self._exact_index[KnowledgeBase.make_key(record)].append(record)
            ordered_key = KnowledgeBase.make_ordered_key(record)
            self._ordered_index[ordered_key].append(record)

    def lookup(self, row: dict[str, Any]) -> tuple[str | None, str | None, list[dict[str, Any]]]:
        exact_candidates = self._exact_index.get(KnowledgeBase.make_key(row), [])
        if exact_candidates:
            return MATCH_MODE_EXACT, MATCH_MODE_EXACT, exact_candidates

        ordered_key = KnowledgeBase.make_ordered_key(row)
        ordered_candidates = self._ordered_index.get(ordered_key, [])
        if ordered_candidates:
            return MATCH_MODE_ORDERED, MATCH_MODE_ORDERED, ordered_candidates
        return None, None, []


def _load_pool_matcher(pool_path: Path) -> ExperiencePoolMatcher:
    workbook = load_workbook(pool_path, data_only=True, read_only=True)
    records: list[dict[str, Any]] = []
    try:
        sheet = workbook.worksheets[0]
        headers = _headers_at(sheet, 1)
        header_map = _header_map(headers)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            record = {
                header: row[index - 1] if index - 1 < len(row) else None
                for header, index in header_map.items()
            }
            if not _has_key_content(record):
                continue
            records.append(record)
    finally:
        workbook.close()
    return ExperiencePoolMatcher(records)


def _count_warning_candidate_rows(
    workbook: Any,
    *,
    only_check_rows_with_value: bool = True,
    value_filter_field: str = DEFAULT_WARNING_FILTER_FIELD,
) -> int:
    total = 0
    for sheet in workbook.worksheets:
        merged_value_map = FillEngine._build_merged_value_map(sheet)
        header_row = _detect_header_row(sheet, merged_value_map)
        headers = _headers_at(sheet, header_row, merged_value_map)
        header_map = _header_map(headers)
        if not all(field in header_map for field in ("要素1", "单位")):
            continue
        metric_columns = _metric_columns(headers)
        if not metric_columns:
            continue
        filter_column_index = _warning_filter_column_index(headers, value_filter_field) if only_check_rows_with_value else None
        if only_check_rows_with_value and not filter_column_index:
            raise ValueError(f"预警 sheet {sheet.title} 未映射过滤字段：{value_filter_field}")
        for excel_row in range(header_row + 1, sheet.max_row + 1):
            row_values = _row_values(sheet, excel_row, len(headers), merged_value_map)
            key_values = {
                field: _value_from_row(row_values, header_map.get(field))
                for field in FIELD_COLUMNS
            }
            if not _has_key_content(key_values):
                continue
            if only_check_rows_with_value:
                filter_value = _value_from_row(row_values, filter_column_index)
                if not _has_warning_filter_value(filter_value):
                    continue
            total += 1
    return total


def _detect_header_row(sheet: Any, merged_value_map: dict[tuple[int, int], Any] | None = None) -> int:
    merged_value_map = merged_value_map or {}
    for row_index in range(1, min(sheet.max_row, 8) + 1):
        values = _headers_at(sheet, row_index, merged_value_map)
        if any(str(value).strip() == "要素1" for value in values if value is not None):
            return row_index
    return 1


def _headers_at(sheet: Any, row_index: int, merged_value_map: dict[tuple[int, int], Any] | None = None) -> list[Any]:
    merged_value_map = merged_value_map or {}
    return [
        FillEngine._read_mapped_value(sheet, row_index, column_index, merged_value_map) or ""
        for column_index in range(1, sheet.max_column + 1)
    ]


def _header_map(headers: list[Any]) -> dict[str, int]:
    return {
        str(value).strip(): index
        for index, value in enumerate(headers, start=1)
        if str(value or "").strip()
    }


def _row_values(
    sheet: Any,
    excel_row: int,
    max_column: int,
    merged_value_map: dict[tuple[int, int], Any] | None = None,
) -> list[Any]:
    merged_value_map = merged_value_map or {}
    return [
        FillEngine._read_mapped_value(sheet, excel_row, column_index, merged_value_map)
        for column_index in range(1, max_column + 1)
    ]


def _value_from_row(row: list[Any], column_index: int | None) -> Any:
    if not column_index or column_index - 1 >= len(row):
        return None
    return row[column_index - 1]


def _metric_columns(headers: list[Any]) -> dict[str, int]:
    columns: dict[str, int] = {}
    for index, raw_header in enumerate(headers, start=1):
        header = str(raw_header or "").replace(" ", "")
        if not header or "匹配说明" in header:
            continue
        if PRICE_METRIC not in columns and any(token in header for token in ["基价", "单价", "价格"]):
            columns[PRICE_METRIC] = index
        if PHYSICAL_METRIC not in columns and "实物工作费调整系数" in header:
            columns[PHYSICAL_METRIC] = index
        if TECHNICAL_METRIC not in columns and "技术工作费调整系数" in header:
            columns[TECHNICAL_METRIC] = index
    return columns


def _warning_filter_column_index(headers: list[Any], field: str) -> int | None:
    aliases = WARNING_FILTER_FIELD_ALIASES.get(field, [field])
    normalized_aliases = [_compact_header(alias) for alias in aliases if alias]
    for index, raw_header in enumerate(headers, start=1):
        if _compact_header(raw_header) in normalized_aliases:
            return index
    return None


def _has_key_content(row: dict[str, Any]) -> bool:
    return bool(normalize_key_part(row.get("要素1"))) and bool(normalize_key_part(row.get("单位")))


def _has_warning_filter_value(value: Any) -> bool:
    if value is None or isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return abs(float(value)) > 0.000001
    text = str(value).strip()
    if not text:
        return False
    normalized = _normalize_number(text)
    if normalized is not None:
        return abs(float(normalized)) > 0.000001
    return True


def _compact_header(value: Any) -> str:
    return str(value or "").strip().replace(" ", "")


def _normalize_number(value: Any) -> int | float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return _clean_number(float(value))
    text = str(value).strip().replace(",", "")
    if text in {"", "/", "待复核", "空单价"}:
        return None
    try:
        return _clean_number(float(text))
    except ValueError:
        return None


def _clean_number(value: float) -> int | float:
    rounded = round(value, 6)
    return int(rounded) if rounded.is_integer() else rounded


def _unique_numbers(values: Any) -> list[int | float]:
    numbers = [_normalize_number(value) for value in values]
    return [number for number in numbers if number is not None]


def _metric_warning_result(
    metric: str,
    current: int | float,
    values: list[int | float],
    low_risk_warning_ratio: float,
    high_risk_warning_ratio: float,
) -> dict[str, Any]:
    average = _clean_number(float(mean(values)))
    experience_min = min(values)
    experience_max = max(values)
    deviation_ratio = _deviation_ratio(current, average)
    deviation_percent = _ratio_to_percent(deviation_ratio)
    severity = _warning_severity(deviation_ratio, average, low_risk_warning_ratio, high_risk_warning_ratio)
    return {
        "metric": metric,
        "current_value": current,
        "experience_values": values,
        "experience_average": average,
        "experience_min": experience_min,
        "experience_max": experience_max,
        "experience_values_text": "、".join(str(value) for value in values),
        "experience_range_text": f"{experience_min}~{experience_max}",
        "sample_count": len(values),
        "deviation_ratio": deviation_ratio,
        "deviation_percent": deviation_percent,
        "severity": severity,
        "severity_label": _severity_label(severity),
        "message": _warning_message(metric, current, average, experience_min, experience_max, deviation_percent, severity),
        "suggested_action": _suggested_action(metric, current, average, experience_min, experience_max, deviation_percent, severity),
    }


def _deviation_ratio(current: int | float, average: int | float) -> float:
    if abs(float(average)) < 0.000001:
        return 0.0
    return abs(float(current) - float(average)) / abs(float(average))


def _warning_severity(
    deviation_ratio: float,
    average: int | float,
    low_risk_warning_ratio: float,
    high_risk_warning_ratio: float,
) -> str:
    if abs(float(average)) < 0.000001:
        return "none"
    if deviation_ratio <= low_risk_warning_ratio:
        return "none"
    if deviation_ratio <= high_risk_warning_ratio:
        return "low"
    return "high"


def _warning_message(
    metric: str,
    current: int | float,
    average: int | float,
    experience_min: int | float,
    experience_max: int | float,
    deviation_percent: float,
    severity: str,
) -> str:
    if severity == "high":
        return (
            f"{metric} 当前值 {current}；经验池平均值 {average}；经验范围 {experience_min}~{experience_max}；"
            f"实际偏离率 {deviation_percent:g}%；已触发高风险预警。"
        )
    if severity == "low":
        return (
            f"{metric} 当前值 {current}；经验池平均值 {average}；经验范围 {experience_min}~{experience_max}；"
            f"实际偏离率 {deviation_percent:g}%；已触发低风险预警。"
        )
    return (
        f"{metric} 当前值 {current}；经验池平均值 {average}；经验范围 {experience_min}~{experience_max}；"
        f"实际偏离率 {deviation_percent:g}%；未超过阈值。"
    )


def _suggested_action(
    metric: str,
    current: int | float,
    average: int | float,
    experience_min: int | float,
    experience_max: int | float,
    deviation_percent: float,
    severity: str,
) -> str:
    if severity == "high":
        return (
            f"建议复核：{metric} 当前值 {current} 相对经验池平均值 {average} 偏离 {deviation_percent:g}%，"
            f"请核对输入要素、单位、来源依据和人工填值。"
        )
    if severity == "low":
        return (
            f"建议复核：{metric} 当前值 {current} 相对经验池平均值 {average} 偏离 {deviation_percent:g}%，"
            f"建议结合项目条件与历史范围 {experience_min}~{experience_max} 复核。"
        )
    return "未超过阈值，当前参数无需触发预警。"


def _ratio_to_percent(value: float) -> float:
    return _clean_number(float(value) * 100)


def _severity_label(severity: str) -> str:
    if severity == "high":
        return "高风险"
    if severity == "low":
        return "低风险"
    return "无预警"


def _metric_counts(warnings: list[dict[str, Any]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for warning in warnings:
        metric = str(warning.get("metric", ""))
        if metric:
            counts[metric] = counts.get(metric, 0) + 1
    return counts


def _summary_text(summary: dict[str, Any]) -> str:
    if not summary.get("pool_enabled"):
        return "经验池未启用：尚未导入经验池，本次未执行经验值预警。"
    candidate_rows = int(summary.get("candidate_rows") or 0)
    warning_rows = int(summary.get("warning_rows") or 0)
    checked_rows = int(summary.get("checked_rows") or 0)
    no_comparable_rows = int(summary.get("no_comparable_rows") or 0)
    match_mode_counts = dict(summary.get("match_mode_counts") or {})
    mode_text = "；".join(f"{mode} {count} 行" for mode, count in match_mode_counts.items() if count)
    coverage_text = (
        f"输入候选 {candidate_rows} 行，可比选 {checked_rows} 行，未找到同类 {no_comparable_rows} 行"
        if candidate_rows
        else f"可比选 {checked_rows} 行，未找到同类 {no_comparable_rows} 行"
    )
    if checked_rows == 0:
        return f"经验池预警：{coverage_text}，本次未找到可比选同类记录。"
    if warning_rows == 0:
        if mode_text:
            return f"经验池预警：{coverage_text}，各参数平均值偏离率均未超过阈值；匹配模式：{mode_text}"
        return f"经验池预警：{coverage_text}，各参数平均值偏离率均未超过阈值。"
    high_rows = int(summary.get("high_rows") or 0)
    low_rows = int(summary.get("low_rows") or summary.get("medium_rows") or 0)
    if mode_text:
        return (
            f"经验池预警：{coverage_text}，发现 {warning_rows} 条预警，"
            f"其中高风险 {high_rows} 条、低风险 {low_rows} 条；匹配模式：{mode_text}"
        )
    return (
        f"经验池预警：{coverage_text}，发现 {warning_rows} 条预警，"
        f"其中高风险 {high_rows} 条、低风险 {low_rows} 条。"
    )


def _format_row_key(row: dict[str, Any]) -> str:
    parts = [str(row.get(field) or "").strip() for field in FIELD_COLUMNS]
    return " / ".join(part for part in parts if part)


def _metric_note(record: dict[str, Any], metric: str) -> Any:
    if metric == PRICE_METRIC:
        return record.get("基价说明") or _joined_remarks(record)
    if metric == PHYSICAL_METRIC:
        return record.get("实物说明") or _joined_remarks(record)
    if metric == TECHNICAL_METRIC:
        return record.get("技术说明") or _joined_remarks(record)
    return ""


def _joined_remarks(record: dict[str, Any]) -> str:
    remarks = [
        str(record.get(field) or "").strip()
        for field in ["原表备注1", "原表备注2", "原表备注3"]
        if str(record.get(field) or "").strip()
    ]
    return "；".join(remarks)
