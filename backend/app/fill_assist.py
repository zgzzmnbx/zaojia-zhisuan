from __future__ import annotations

from pathlib import Path
from statistics import mean
from typing import Any

from openpyxl import load_workbook

from .experience_warning import (
    ExperiencePoolMatcher,
    MATCH_MODE_EXACT,
    MATCH_MODE_ORDERED,
    PRICE_METRIC,
    _has_key_content,
    _header_map,
    _headers_at,
)
from .fill_engine import FillEngine
from .knowledge_base import ELEMENT_COLUMNS, KnowledgeBase
from .normalization import normalize_key_part, normalize_price
from .schemas import FIELD_COLUMNS


def build_fill_assist_context(output_excel: Path, sheet_name: str, row_number: int, target_header: str = "") -> dict[str, Any]:
    workbook = load_workbook(output_excel, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"输出 Excel 不存在 sheet：{sheet_name}")
        sheet = workbook[sheet_name]
        merged_value_map = FillEngine._build_merged_value_map(sheet)
        header_row = _detect_header_row(sheet, merged_value_map)
        headers = _headers_at(sheet, header_row, merged_value_map)
        header_map = _header_map(headers)
        values = {
            header: merged_value_map.get((row_number, column), sheet.cell(row=row_number, column=column).value)
            for header, column in header_map.items()
            if header
        }
        row = {field: values.get(field) for field in FIELD_COLUMNS}
        target_column = _target_column(header_map, target_header)
        return {
            "sheet_name": sheet_name,
            "excel_row": row_number,
            "header_row": header_row,
            "target_header": headers[target_column - 1] if target_column and target_column <= len(headers) else target_header,
            "target_column": target_column,
            "current_value": values.get(headers[target_column - 1]) if target_column and target_column <= len(headers) else None,
            "row": row,
            "diagnostics": {
                "匹配状态": values.get("匹配状态") or values.get("输出-匹配状态") or "",
                "候选数量": values.get("候选数量") or values.get("输出-候选数量") or "",
                "匹配说明": values.get("匹配说明") or values.get("输出-匹配说明") or "",
                "预警参数": values.get("预警参数") or "",
                "预警细节": values.get("预警细节") or "",
            },
        }
    finally:
        workbook.close()


def build_fill_assist_candidates(
    row: dict[str, Any],
    *,
    knowledge_base: KnowledgeBase,
    pool_path: Path | None = None,
    limit: int = 8,
) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    candidates.extend(_knowledge_similar_candidates(row, knowledge_base))
    if pool_path and pool_path.exists():
        candidates.extend(_experience_pool_candidates(row, pool_path))
    candidates = _mark_conflicts(candidates)
    candidates.sort(key=_candidate_sort_key)
    return candidates[:limit]


def _knowledge_similar_candidates(row: dict[str, Any], knowledge_base: KnowledgeBase) -> list[dict[str, Any]]:
    unit = normalize_key_part(row.get("单位"))
    if not unit:
        return []
    input_elements = [KnowledgeBase.canonical_key_part(row.get(field)) for field in ELEMENT_COLUMNS]
    input_non_empty = [value for value in input_elements if value]
    candidates: list[dict[str, Any]] = []
    for item in knowledge_base.rows:
        if normalize_key_part(item.get("单位")) != unit:
            continue
        candidate_elements = [KnowledgeBase.canonical_key_part(item.get(field)) for field in ELEMENT_COLUMNS]
        matches = sum(1 for left, right in zip(input_elements, candidate_elements) if left and right and left == right)
        total = max(len(input_non_empty), 1)
        if matches == 0:
            continue
        differences = _element_differences(row, item)
        if len(differences) > 2 and matches < 3:
            continue
        similarity = round(matches / total, 4)
        confidence = "high" if len(differences) <= 1 and similarity >= 0.8 else "medium" if similarity >= 0.5 else "low"
        candidates.append(
            {
                "id": f"kb-{item.get('_excel_row')}",
                "source": "knowledge_similar",
                "source_label": "知识库-相似",
                "value": item.get("基价"),
                "metric": PRICE_METRIC,
                "confidence": confidence,
                "confidence_label": _confidence_label(confidence),
                "similarity": round(similarity * 100, 2),
                "source_row": item.get("_excel_row"),
                "reason": f"单位一致，{matches}/{total} 个非空要素一致。",
                "risk_tips": _difference_risk_tips(differences),
                "basis": f"二维知识库第 {item.get('_excel_row')} 行",
                "candidate_key": {field: item.get(field) for field in FIELD_COLUMNS},
            }
        )
    return candidates


def _experience_pool_candidates(row: dict[str, Any], pool_path: Path) -> list[dict[str, Any]]:
    records = _read_pool_records(pool_path)
    matcher = ExperiencePoolMatcher(records)
    match_mode, match_detail, matched = matcher.lookup(row)
    if not matched:
        return []
    numbers = [_number(record.get(PRICE_METRIC)) for record in matched]
    values = [number for number in numbers if number is not None]
    if not values:
        return []
    avg = mean(values)
    min_value = min(values)
    max_value = max(values)
    spread = 0 if avg == 0 else (max_value - min_value) / abs(avg)
    confidence = "high" if len(values) >= 5 and spread <= 0.1 else "medium" if len(values) >= 3 else "low"
    return [
        {
            "id": "experience-average",
            "source": "experience_pool",
            "source_label": "经验池同类均值",
            "value": _clean_number(avg),
            "metric": PRICE_METRIC,
            "confidence": confidence,
            "confidence_label": _confidence_label(confidence),
            "similarity": 100 if match_mode == MATCH_MODE_EXACT else 80 if match_mode == MATCH_MODE_ORDERED else 0,
            "sample_count": len(values),
            "reason": f"{match_detail or match_mode}，同类样本 {len(values)} 条。",
            "risk_tips": _experience_risk_tips(len(values), spread),
            "basis": f"经验池 {pool_path.name}",
            "source_rows": [
                {
                    "source_file": record.get("来源文件", ""),
                    "source_sheet": record.get("来源sheet", ""),
                    "source_row": record.get("来源行", ""),
                    "pool_row": record.get("_excel_row", ""),
                }
                for record in matched[:5]
            ],
            "experience_min": _clean_number(min_value),
            "experience_max": _clean_number(max_value),
        }
    ]


def _read_pool_records(pool_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(pool_path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        headers = _headers_at(sheet, 1)
        header_map = _header_map(headers)
        records: list[dict[str, Any]] = []
        for excel_row, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            record = {
                header: row[index - 1] if index - 1 < len(row) else None
                for header, index in header_map.items()
            }
            if not _has_key_content(record):
                continue
            record["_excel_row"] = excel_row
            records.append(record)
        return records
    finally:
        workbook.close()


def _detect_header_row(sheet: Any, merged_value_map: dict[tuple[int, int], Any]) -> int:
    for row_index in range(1, min(sheet.max_row, 12) + 1):
        headers = _headers_at(sheet, row_index, merged_value_map)
        compact = {str(header).strip() for header in headers if str(header or "").strip()}
        if "要素1" in compact and "单位" in compact:
            return row_index
    return 1


def _target_column(header_map: dict[str, int], target_header: str) -> int | None:
    compact_target = _compact(target_header)
    for header, column in header_map.items():
        compact_header = _compact(header)
        if compact_target and compact_header == compact_target:
            return column
    for header, column in header_map.items():
        compact_header = _compact(header)
        if compact_header in {"基价", "单价"} or "基价" in compact_header or "单价" in compact_header:
            return column
    return None


def _element_differences(row: dict[str, Any], candidate: dict[str, Any]) -> list[str]:
    differences: list[str] = []
    for field in ELEMENT_COLUMNS:
        left = KnowledgeBase.canonical_key_part(row.get(field))
        right = KnowledgeBase.canonical_key_part(candidate.get(field))
        if not left and not right:
            continue
        if left != right:
            differences.append(f"{field}：本行“{row.get(field) or ''}”，候选“{candidate.get(field) or ''}”")
    return differences


def _difference_risk_tips(differences: list[str]) -> list[str]:
    if not differences:
        return ["要素与单位均一致，但主匹配未采用时仍需核对候选冲突原因。"]
    return [f"存在要素差异：{item}" for item in differences[:3]]


def _experience_risk_tips(sample_count: int, spread: float) -> list[str]:
    tips: list[str] = []
    if sample_count < 3:
        tips.append(f"仅 {sample_count} 条历史样本，均值代表性有限。")
    if spread > 0.2:
        tips.append(f"样本离散度 {round(spread * 100, 2)}%，均值可能失真。")
    return tips or ["经验池同类样本稳定，可作为人工复核参考。"]


def _mark_conflicts(candidates: list[dict[str, Any]]) -> list[dict[str, Any]]:
    high_values = {str(candidate.get("value")) for candidate in candidates if candidate.get("confidence") == "high"}
    if len(high_values) <= 1:
        return candidates
    for candidate in candidates:
        if candidate.get("confidence") == "high":
            candidate.setdefault("risk_tips", []).append("存在多个高可信候选但数值不同，请核对要素口径。")
    return candidates


def _candidate_sort_key(candidate: dict[str, Any]) -> tuple[float, int, int, int]:
    source_rank = {"knowledge_similar": 0, "experience_pool": 1}.get(str(candidate.get("source")), 9)
    confidence_rank = {"high": 0, "medium": 1, "low": 2}.get(str(candidate.get("confidence")), 3)
    similarity = float(candidate.get("similarity") or 0)
    sample_count = int(candidate.get("sample_count") or 0)
    return (-similarity, source_rank, confidence_rank, -sample_count)


def _confidence_label(confidence: str) -> str:
    return {"high": "推荐候选", "medium": "可参考", "low": "谨慎参考"}.get(confidence, "可参考")


def _compact(value: Any) -> str:
    return str(value or "").replace(" ", "").strip().lower()


def _number(value: Any) -> int | float | None:
    parsed = normalize_price(value)
    if isinstance(parsed, (int, float)) and not isinstance(parsed, bool):
        return parsed
    return None


def _clean_number(value: float) -> int | float:
    return int(value) if float(value).is_integer() else round(value, 4)
