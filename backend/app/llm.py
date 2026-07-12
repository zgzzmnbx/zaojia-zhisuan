from __future__ import annotations

import json
import os
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

from openpyxl import load_workbook

from .paths import PROJECT_ROOT


PROVIDER_DEEPSEEK = "deepseek"
PROVIDER_SILICONFLOW = "siliconflow"
DEFAULT_PROVIDER = PROVIDER_DEEPSEEK
DEFAULT_MODEL = "deepseek-v4-flash"
DEFAULT_BASE_URL = "https://api.deepseek.com"
DEFAULT_TEMPERATURE = 0.2
DEFAULT_MAX_TOKENS = 1800


@dataclass(frozen=True)
class LlmConfig:
    provider: str = DEFAULT_PROVIDER
    model: str = DEFAULT_MODEL
    base_url: str = DEFAULT_BASE_URL

    def resolved_api_key(self) -> str:
        key_name = self.api_key_env_name()
        return os.getenv(key_name, "").strip() or _read_local_api_key(key_name)

    def api_key_env_name(self) -> str:
        if self.provider == PROVIDER_SILICONFLOW:
            return "SILICONFLOW_API_KEY"
        return "DEEPSEEK_API_KEY"

    def chat_completions_url(self) -> str:
        clean_url = self.base_url.strip().rstrip("/")
        if clean_url.endswith("/chat/completions"):
            return clean_url
        return f"{clean_url}/chat/completions"


def _read_local_api_key(key_name: str) -> str:
    env_path = PROJECT_ROOT / ".env.local"
    if not env_path.exists():
        return ""
    try:
        for line in env_path.read_text(encoding="utf-8").splitlines():
            clean = line.strip()
            if not clean or clean.startswith("#") or "=" not in clean:
                continue
            name, value = clean.split("=", 1)
            if name.strip() == key_name:
                return value.strip().strip('"').strip("'")
    except OSError:
        return ""
    return ""


def call_chat_completion(config: LlmConfig, messages: list[dict[str, str]]) -> str:
    api_key = config.resolved_api_key()
    if not api_key:
        raise ValueError(f"请先设置后端环境变量 {config.api_key_env_name()}")

    payload = json.dumps(
        {
            "model": config.model,
            "messages": messages,
            "temperature": DEFAULT_TEMPERATURE,
            "max_tokens": DEFAULT_MAX_TOKENS,
        },
        ensure_ascii=False,
    ).encode("utf-8")
    request = Request(
        config.chat_completions_url(),
        data=payload,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )

    try:
        with urlopen(request, timeout=90) as response:
            data = json.loads(response.read().decode("utf-8"))
    except HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")
        raise RuntimeError(f"大模型接口返回错误：HTTP {exc.code} {detail}") from exc
    except URLError as exc:
        raise RuntimeError(f"大模型接口连接失败：{exc.reason}") from exc

    try:
        return str(data["choices"][0]["message"]["content"]).strip()
    except (KeyError, IndexError, TypeError) as exc:
        raise RuntimeError("大模型接口返回格式异常") from exc


def build_risk_prompt(report_markdown: str, excel_path: Path, knowledge_evidence: str = "") -> list[dict[str, str]]:
    summary = report_markdown.split("## 价格识别日志", 1)[0].strip()
    fee_summary = read_fee_summary(excel_path)
    evidence_text = knowledge_evidence.strip() or "未检索到可用知识库依据。"
    user_content = "\n\n".join(
        [
            "请基于造价智算项目背景、处理摘要、费用汇总和知识库依据，输出可直接写入报告“五、其他需要注意的事项”的风险审查意见。",
            (
                "项目背景：本项目是工程造价辅助智能体，当前核心场景覆盖长输管道勘察测量最高投标限价编制；基价来自结构化计价库，"
                "实物工作费调整系数和技术工作费调整系数采用第一层标准规则、第二层经验提示、待复核三层架构。"
            ),
            (
                "要求：语言正式、稳重、可直接放进报告；只输出风险审查正文；"
                "按“一、本次处理结论”“二、主要风险概览”“三、重点复核建议”“四、规则依据提示”“五、后续处理建议”组织；"
                "可以结合知识库依据说明风险原因，但不要逐行复述匹配情况；"
                "不要裁决最终单价或系数；不要编造未提供的标准条文；证据不足时用“建议复核”表述。"
            ),
            "【处理摘要】",
            summary[:3000],
            "【费用汇总】",
            fee_summary,
            "【知识库依据】",
            evidence_text[:5000],
        ]
    )
    return [
        {
            "role": "system",
            "content": (
                "你是国家管网工程造价与招标控制价复核助手，擅长识别价格测算、规则依据和资料一致性风险。"
                "你只能基于用户提供的处理摘要、费用汇总和知识库依据写风险审查意见。"
                "不得反向修改或裁决结构化规则引擎输出的基价、单价、实物工作费调整系数和技术工作费调整系数。"
            ),
        },
        {"role": "user", "content": user_content},
    ]


def read_fee_summary(path: Path) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if "费用汇总" not in workbook.sheetnames:
            return "未发现费用汇总 sheet。"
        sheet = workbook["费用汇总"]
        rows: list[str] = []
        for row_index in range(1, min(sheet.max_row, 20) + 1):
            values = [sheet.cell(row=row_index, column=column).value for column in range(1, min(sheet.max_column, 4) + 1)]
            text = " | ".join("" if value is None else str(value) for value in values).strip()
            if text:
                rows.append(text)
        return "\n".join(rows)
    finally:
        workbook.close()


def read_excel_preview(path: Path, max_rows: int = 20, max_columns: int = 12) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows: list[str] = []
    try:
        for row in sheet.iter_rows(max_row=min(sheet.max_row, max_rows), max_col=min(sheet.max_column, max_columns), values_only=True):
            rows.append(" | ".join("" if value is None else str(value) for value in row))
    finally:
        workbook.close()
    return "\n".join(rows)
