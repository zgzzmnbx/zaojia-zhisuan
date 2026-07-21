from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .normalization import normalize_key_part

RULE_DIR = Path(__file__).resolve().parent / "rules"
TECHNICAL_CONTEXT_FIELDS = ("要素-技术", "要素技术", "技术类别", "技术工作类别")
CONTROLLED_TECHNICAL_CATEGORY_TOKENS = {
    "甲级",
    "乙级",
    "丙级",
    "简单",
    "中等",
    "复杂",
    "物探",
    "试验",
}


@dataclass(frozen=True)
class CoefficientRule:
    rule_id: str
    target: str
    sheet_tokens: tuple[str, ...]
    trigger_keywords: tuple[str, ...]
    business_keywords: tuple[str, ...]
    category_keywords: tuple[str, ...]
    coefficient: str
    source_type: str
    confidence: str
    basis: str
    formula_effective: bool
    review_required: bool
    note: str
    priority: str = ""


@dataclass(frozen=True)
class AdjustmentResult:
    value: int | float | str
    status: str
    message: str
    rules: tuple[CoefficientRule, ...] = ()


@dataclass(frozen=True)
class AdjustmentEvaluation:
    physical: AdjustmentResult
    technical: AdjustmentResult


class AdjustmentEngine:
    def __init__(self, physical_rules: list[CoefficientRule], technical_rules: list[CoefficientRule]) -> None:
        self.physical_rules = physical_rules
        self.technical_rules = sorted(
            technical_rules,
            key=lambda rule: _priority_rank(rule.priority),
            reverse=True,
        )

    @classmethod
    def from_default_rules(cls) -> "AdjustmentEngine":
        physical_rules: list[CoefficientRule] = []
        technical_rules = _load_rules(RULE_DIR / "technical_fee_rules.csv", target="technical")
        return cls(physical_rules=physical_rules, technical_rules=technical_rules)

    @classmethod
    def from_rule_assets(
        cls,
        *,
        physical_rules_path: str | Path | None,
        technical_rules_path: str | Path,
    ) -> "AdjustmentEngine":
        physical_rules = (
            _load_explicit_rule_asset(Path(physical_rules_path), target="physical")
            if physical_rules_path
            else []
        )
        technical_rules = _load_explicit_rule_asset(Path(technical_rules_path), target="technical")
        return cls(physical_rules=physical_rules, technical_rules=technical_rules)

    def evaluate(
        self,
        sheet_name: str,
        mapped_values: dict[str, Any],
        row_values: list[Any],
    ) -> AdjustmentEvaluation:
        row_text = _compact_text([sheet_name, *mapped_values.values(), *row_values])
        technical_context_text = _compact_text(
            [
                mapped_values.get(field_name)
                for field_name in TECHNICAL_CONTEXT_FIELDS
                if field_name in mapped_values
            ]
        )
        sheet_text = normalize_key_part(sheet_name)
        return AdjustmentEvaluation(
            physical=self._evaluate_physical(sheet_text, row_text),
            technical=self._evaluate_technical(sheet_text, row_text, technical_context_text),
        )

    def _evaluate_physical(self, sheet_text: str, row_text: str) -> AdjustmentResult:
        matched = [
            rule
            for rule in self.physical_rules
            if _rule_matches(rule, sheet_text, row_text) and not rule.review_required
        ]
        if not matched:
            return AdjustmentResult(
                "待复核",
                "review",
                "待复核：实物工作费调整系数第一层规则暂未启用，且第二层经验提示未命中。",
            )

        coefficients = [_parse_coefficient(rule.coefficient) for rule in matched]
        if len(coefficients) == 1:
            value = coefficients[0]
            message = (
                f"命中 1 条实物系数规则：{_format_rule_detail(matched[0])}；"
                f"综合系数={_format_number(value)}。"
            )
        else:
            value = 1 + sum(coefficient - 1 for coefficient in coefficients)
            formula_parts = " + ".join(f"({_format_number(coefficient)}-1)" for coefficient in coefficients)
            message = (
                f"命中 {len(matched)} 条实物系数规则：{_format_rules(matched)}；"
                f"按总则组合公式 1 + Σ(k-1) = 1 + {formula_parts} = {_format_number(value)}。"
            )
        return AdjustmentResult(
            _clean_number(value),
            "matched",
            message,
            tuple(matched),
        )

    def _evaluate_technical(
        self,
        sheet_text: str,
        row_text: str,
        technical_context_text: str,
    ) -> AdjustmentResult:
        for rule in self.technical_rules:
            if _technical_rule_matches(rule, sheet_text, row_text, technical_context_text):
                if rule.review_required or rule.coefficient == "待复核":
                    message = (
                        f"待复核：命中规则 {_format_rule_detail(rule)}；"
                        "该规则标记为需要人工复核，输出待复核。"
                    )
                    return AdjustmentResult("待复核", "review", message, (rule,))
                value = _parse_coefficient(rule.coefficient)
                message = (
                    f"命中技术费规则：{_format_rule_detail(rule)}；"
                    f"输出系数 {_format_number(value)}。"
                )
                return AdjustmentResult(_clean_number(value), "matched", message, (rule,))
        return AdjustmentResult(
            "待复核",
            "review",
            "待复核：未命中技术工作费调整系数规则；请检查 sheet、业务大类、技术工作类别或规则库。",
        )


def _load_rules(path: Path, target: str) -> list[CoefficientRule]:
    xlsx_path = path.with_suffix(".xlsx")
    if xlsx_path.exists():
        try:
            return _load_rules_from_xlsx(xlsx_path, target)
        except Exception:
            pass
    return _load_rules_from_csv(path, target)


def _load_explicit_rule_asset(path: Path, target: str) -> list[CoefficientRule]:
    if path.suffix.lower() == ".xlsx":
        return _load_rules_from_xlsx(path, target)
    if path.suffix.lower() == ".csv":
        return _load_rules_from_csv(path, target)
    raise ValueError(f"不支持的规则资产格式：{path.suffix}")


def _load_rules_from_csv(path: Path, target: str) -> list[CoefficientRule]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return _rows_to_rules(reader, target)


def _load_rules_from_xlsx(path: Path, target: str) -> list[CoefficientRule]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows)]
        _validate_rule_headers(headers, target, path)
        dict_rows = (
            {
                header: _cell_to_text(value)
                for header, value in zip(headers, values)
            }
            for values in rows
        )
        return _rows_to_rules(dict_rows, target)
    finally:
        workbook.close()


def _rows_to_rules(rows: Any, target: str) -> list[CoefficientRule]:
    return [
        CoefficientRule(
            rule_id=str(row.get("rule_id") or "").strip(),
            target=target,
            sheet_tokens=_split_tokens(row.get("sheet_tokens")),
            trigger_keywords=_split_tokens(row.get("trigger_keywords")),
            business_keywords=_split_tokens(row.get("business_keywords")),
            category_keywords=_split_tokens(row.get("category_keywords")),
            coefficient=str(row.get("coefficient") or "").strip(),
            source_type=str(row.get("source_type") or "").strip(),
            confidence=str(row.get("confidence") or "").strip(),
            basis=str(row.get("basis") or "").strip(),
            formula_effective=_as_bool(row.get("formula_effective"), default=True),
            review_required=_as_bool(row.get("review_required"), default=False),
            note=str(row.get("note") or "").strip(),
            priority=str(row.get("priority") or "").strip(),
        )
        for row in rows
        if row.get("rule_id")
    ]


def _validate_rule_headers(headers: list[str], target: str, path: Path) -> None:
    required = {
        "rule_id",
        "sheet_tokens",
        "coefficient",
        "source_type",
        "confidence",
        "basis",
        "review_required",
        "note",
    }
    if target == "physical":
        required.add("trigger_keywords")
    if target == "technical":
        required.update({"business_keywords", "category_keywords", "formula_effective"})
    missing = sorted(required.difference(headers))
    if missing:
        raise ValueError(f"Excel 规则表缺少字段：{path} -> {', '.join(missing)}")


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _split_tokens(value: Any) -> tuple[str, ...]:
    return tuple(normalize_key_part(part) for part in str(value or "").split(";") if normalize_key_part(part))


def _compact_text(values: list[Any]) -> str:
    return "|".join(normalize_key_part(value) for value in values if normalize_key_part(value))


def _rule_matches(rule: CoefficientRule, sheet_text: str, row_text: str) -> bool:
    if rule.sheet_tokens and not any(token in sheet_text for token in rule.sheet_tokens):
        return False
    for token_group in [rule.business_keywords, rule.category_keywords, rule.trigger_keywords]:
        if token_group and not any(token in row_text for token in token_group):
            return False
    return True


def _technical_rule_matches(
    rule: CoefficientRule,
    sheet_text: str,
    row_text: str,
    technical_context_text: str,
) -> bool:
    if rule.sheet_tokens and not any(token in sheet_text for token in rule.sheet_tokens):
        return False
    if rule.business_keywords and not any(token in row_text for token in rule.business_keywords):
        return False
    if not rule.category_keywords:
        return True
    category_text = (
        technical_context_text
        if technical_context_text and _uses_controlled_technical_categories(rule.category_keywords)
        else row_text
    )
    return any(token in category_text for token in rule.category_keywords)


def _uses_controlled_technical_categories(tokens: tuple[str, ...]) -> bool:
    return all(token in CONTROLLED_TECHNICAL_CATEGORY_TOKENS for token in tokens)


def _parse_coefficient(value: str) -> float:
    return float(value)


def _priority_rank(value: str) -> int:
    normalized = normalize_key_part(value)
    if normalized in {"高", "high"}:
        return 100
    if normalized in {"中", "medium"}:
        return 50
    if normalized in {"低", "low"}:
        return 10
    return 0


def _clean_number(value: float) -> int | float:
    rounded = round(value, 6)
    return int(rounded) if rounded.is_integer() else rounded


def _format_rules(rules: list[CoefficientRule]) -> str:
    return "、".join(_format_rule_detail(rule) for rule in rules)


def _format_rule_detail(rule: CoefficientRule) -> str:
    coefficient = rule.coefficient
    parts = [f"{rule.rule_id}（{rule.note or '未填写规则说明'}"]
    if coefficient:
        parts.append(f"系数 {coefficient}")
    if rule.source_type:
        parts.append(f"来源：{rule.source_type}")
    if rule.basis:
        parts.append(f"依据：{rule.basis}")
    if rule.target == "technical":
        parts.append(f"是否参与公式：{'是' if rule.formula_effective else '否'}")
    return "，".join(parts) + "）"


def _format_number(value: float) -> str:
    cleaned = _clean_number(value)
    return str(cleaned)


def _as_bool(value: Any, default: bool) -> bool:
    text = str(value or "").strip().lower()
    if not text:
        return default
    return text in {"1", "true", "yes", "y", "是"}
