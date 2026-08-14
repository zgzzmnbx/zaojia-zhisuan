from __future__ import annotations

import json
import hashlib
import os
import re
import sqlite3
from threading import Lock
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any
from uuid import uuid4

import httpx
from fastapi import Body, FastAPI, File, Form, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from .fill_engine import (
    EMPTY_ELEMENT_COLUMN,
    PHYSICAL_ADJUSTMENT_FIELD,
    PRICE_COLUMN_FIELDS,
    TECHNICAL_ADJUSTMENT_FIELD,
    FillEngine,
)
from .adjustment_rules import AdjustmentEngine
from . import feishu_webhook
from . import feishu_app_bot
from . import external_task_dispatch
from .knowledge_base import KnowledgeBase
from .knowledge_demo_answers import get_demo_answer, get_row_demo_answer
from .knowledge_qa import (
    ASSISTANT_TABLE_FORMAT_RULE,
    HybridRetrievalError,
    build_knowledge_answer_prompt,
    ensure_knowledge_answer,
    hybrid_search_knowledge,
    is_knowledge_question,
    knowledge_retrieval_capabilities,
    load_knowledge_retrieval_config,
    load_or_build_hybrid_index,
    prepend_ranked_candidate_recommendation,
    search_knowledge,
    strip_force_knowledge_prefix,
)
from .knowledge_libraries import (
    KnowledgeLibrarySelection,
    PROFESSIONAL_KNOWLEDGE_LIBRARY_ID,
    knowledge_library_catalog,
    parse_requested_library_ids,
    resolve_knowledge_library_selection,
)
from .knowledge_memory import (
    DEFAULT_AUTO_APPROVE_KNOWLEDGE_TYPES,
    KNOWLEDGE_MEMORY_TYPES,
    KnowledgeMemoryConflict,
    KnowledgeMemoryError,
    KnowledgeMemoryNotFound,
    KnowledgeMemoryPermissionError,
    KnowledgeMemoryStore,
    classify_knowledge_type,
    normalize_project_key,
    search_confirmed_project_memory,
)
from .trusted_experience import TrustedExperienceError, TrustedExperienceStore
from .experience_warning import (
    DEFAULT_SELECTED_EXPERIENCE_FIELDS,
    DEFAULT_HIGH_RISK_WARNING_PERCENT,
    DEFAULT_LOW_RISK_WARNING_PERCENT,
    DEFAULT_WARNING_FILTER_FIELD,
    EXPERIENCE_MAPPING_FIELDS,
    PHYSICAL_METRIC,
    PRICE_METRIC,
    TECHNICAL_METRIC,
    WARNING_FILTER_FIELDS,
    WARNING_OUTPUT_FIELDS,
    _has_warning_filter_value,
    _warning_filter_column_index,
    analyze_workbook_warnings_with_progress,
    import_experience_pool,
    write_warnings_to_workbook,
)
from .experience_governance import build_experience_pool_governance_report, write_governance_markdown
from .fill_assist import build_fill_assist_candidates, build_fill_assist_context
from .risk_items import build_standard_trace, build_structured_risk_items, summarize_risk_items
from .workload_capture import (
    DEFAULT_SELECTED_WORKLOAD_FIELDS,
    DEFAULT_WORKLOAD_FILTER_FIELD,
    SOURCE_MAPPING_FIELDS,
    SOURCE_QUANTITY_FIELD,
    TARGET_MAPPING_FIELDS,
    WRITE_MODE_CONSERVATIVE,
    WRITE_MODE_OVERWRITE,
    WORKLOAD_FIELD_PREFERENCE_FIELDS,
    WORKLOAD_TARGET_FIELD_PREFERENCE_FIELDS,
    capture_workload,
    default_workload_field_preferences,
    default_workload_target_field_preferences,
    suggest_workload_column_mapping,
)
from .llm import (
    DEFAULT_BASE_URL,
    DEFAULT_MODEL,
    DEFAULT_PROVIDER,
    KNOWLEDGE_QA_TEMPERATURE,
    LlmConfig,
    build_risk_prompt,
    call_chat_completion,
)
from .llm_usage import LlmUsageError, LlmUsageLedger
from .schemas import FIELD_COLUMNS, FillSummary, ReviewRow
from .excel_recalc import recalculate_workbook
from .formula_resolver import WorkbookFormulaResolver
from .paths import (
    BUSINESS_SKILLS_DIR,
    DEFAULT_BUSINESS_TASK_DB_PATH,
    DEFAULT_EXPERIENCE_FIELD_PREFERENCES_PATH,
    DEFAULT_EXPERIENCE_POOL_PATH,
    DEFAULT_EXPERIENCE_POOL_TEMPLATE_PATH,
    DEFAULT_EXPERIENCE_WARNING_SETTINGS_PATH,
    DEFAULT_INPUT_FIELD_PREFERENCES_PATH,
    DEFAULT_KB_PATH,
    DEFAULT_KNOWLEDGE_MEMORY_DB_PATH,
    DEFAULT_LLM_USAGE_DB_PATH,
    DEFAULT_PREVIEW_COLUMN_PREFERENCES_PATH,
    DEFAULT_PROJECT_LEDGER_DB_PATH,
    DEFAULT_REPORT_TEMPLATE_PATH,
    DEFAULT_UI_PREFERENCES_PATH,
    DEFAULT_WORKLOAD_FIELD_PREFERENCES_PATH,
    DEFAULT_WORKLOAD_TARGET_FIELD_PREFERENCES_PATH,
    LEGACY_EXPERIENCE_POOL_PATH,
    PROJECT_DEFAULT_SETTINGS_PATH,
    PROJECT_ROOT,
    RUNTIME_DIR,
)
from .project_dashboard import backfill_project_ledger
from .business_tasks import (
    BusinessTaskConflict,
    BusinessTaskError,
    BusinessTaskNotFound,
    BusinessTaskStore,
    now_iso,
)
from .digital_employee_evidence import build_onboarding_evidence
from .settlement_audit_api import router as settlement_audit_router
from .project_ledger import (
    ProjectArtifactNotFoundError,
    ProjectLedger,
    ProjectLedgerError,
    ProjectNotFoundError,
)
from .professional_skills import (
    MAX_RECOMMENDATION_FILE_BYTES,
    ProfessionalSkillError,
    ProfessionalSkillRegistry,
    SkillRuntimeContext,
)
from .report import append_risk_report, ensure_risk_report_evidence, write_report


APP_VERSION = "v5.23.3"
# `/api/health.version` 是旧版运行器的兼容字段；当前发布版本通过
# `release_version` 返回，避免旧客户端在小版本升级时误判服务不可用。
HEALTH_API_COMPAT_VERSION = "v5.19.4"
OUTPUT_FILE_PREFIX = "【输出】"
TEMP_FILE_PREFIX = "【临时】"
PROCESS_STATE_FILENAME = "process-state.json"
MANUAL_EDIT_LOG_FILENAME = "preview-manual-edits.json"
MANUAL_EDIT_FILL = PatternFill(fill_type="solid", fgColor="DDEBFF")
MANUAL_EDIT_COMMENT_AUTHOR = "造价智算"
MANUAL_EDIT_NUMERIC_HEADER_TOKENS = (
    "金额",
    "费用",
    "数量",
    "工程量",
    "基价",
    "单价",
    "合价",
    "系数",
    "税",
    "小计",
    "合计",
)
MANUAL_EDIT_READONLY_HEADERS = {
    "匹配状态",
    "候选数量",
    "匹配说明",
    "预警参数",
    "预警细节",
    "输出-匹配状态",
    "输出-候选数量",
    "输出-匹配说明",
}
DEFAULT_PREVIEW_CELL_MAX_DISPLAY_CHARS = 8
MIN_PREVIEW_COLUMN_WIDTH_PX = 72
MAX_PREVIEW_COLUMN_WIDTH_PX = 420
DEFAULT_CORE_PREVIEW_LABELS = [
    "要素1",
    "要素2",
    "要素3",
    "要素4",
    "要素5",
    "单位",
    "单价",
    "实物工作费调整系数",
    "技术工作费调整系数",
    "预警参数",
    "预警细节",
]
RISK_REPORT_KNOWLEDGE_QUERIES = [
    "第二层经验提示是什么意思？",
    "待复核是什么原因？",
    "经验池预警偏离率和阈值怎么判断？",
    "基价 单价 字段完全匹配 非空要素顺序匹配",
    "实物工作费调整系数第一层标准规则第二层经验提示",
    "技术工作费调整系数0.22依据",
    "附加调整系数为什么不能连乘？",
]
DEMO_SAMPLE_TOKENS = ("输入100", "空单价100")
WARNING_PROGRESS_DEFAULT = {
    "status": "idle",
    "processed_rows": 0,
    "total_rows": 0,
    "matched_rows": 0,
    "warning_rows": 0,
}
WARNING_PROGRESS: dict[str, dict[str, object]] = {}
WARNING_PROGRESS_LOCK = Lock()
LLM_USAGE_BACKFILL_LOCK = Lock()
LLM_USAGE_BACKFILLED_DATABASES: set[str] = set()
PROFESSIONAL_SKILL_REGISTRY = ProfessionalSkillRegistry(
    PROJECT_ROOT,
    BUSINESS_SKILLS_DIR,
    PROJECT_DEFAULT_SETTINGS_PATH,
)
INPUT_FIELD_PREFERENCE_FIELDS = [
    *FIELD_COLUMNS,
    "输出-价格列",
    PHYSICAL_ADJUSTMENT_FIELD,
    TECHNICAL_ADJUSTMENT_FIELD,
]

app = FastAPI(title="造价智算 API", version=APP_VERSION)
FRONTEND_ORIGINS = [
    "http://127.0.0.1:5173",
    "http://localhost:5173",
    "http://127.0.0.1:5174",
    "http://localhost:5174",
]
app.add_middleware(
    CORSMiddleware,
    allow_origins=FRONTEND_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
app.include_router(settlement_audit_router)


@app.get("/api/health")
def health() -> dict[str, str]:
    return {
        "status": "ok",
        "service": "guankanzhisuan",
        "version": HEALTH_API_COMPAT_VERSION,
        "release_version": APP_VERSION,
    }


def _project_filters(
    *,
    date_from: str,
    date_to: str,
    skill_id: str,
    status: str,
    source_type: str,
    keyword: str,
    risk: str,
    quality: str,
    lifecycle_stage: str,
) -> dict[str, str]:
    return {
        "date_from": date_from.strip(),
        "date_to": date_to.strip(),
        "skill_id": skill_id.strip(),
        "status": status.strip(),
        "source_type": source_type.strip(),
        "keyword": keyword.strip(),
        "risk": risk.strip(),
        "quality": quality.strip(),
        "lifecycle_stage": lifecycle_stage.strip(),
    }


@app.get("/api/projects/dashboard")
def get_projects_dashboard(
    date_from: str = Query(default=""),
    date_to: str = Query(default=""),
    skill_id: str = Query(default=""),
    status: str = Query(default=""),
    source_type: str = Query(default=""),
    keyword: str = Query(default=""),
    risk: str = Query(default=""),
    quality: str = Query(default=""),
    lifecycle_stage: str = Query(default=""),
    compare: bool = Query(default=False),
) -> dict[str, object]:
    try:
        payload = _project_ledger().dashboard(
            **_project_filters(
                date_from=date_from,
                date_to=date_to,
                skill_id=skill_id,
                status=status,
                source_type=source_type,
                keyword=keyword,
                risk=risk,
                quality=quality,
                lifecycle_stage=lifecycle_stage,
            )
        )
        payload["llm_usage"] = _llm_usage_dashboard(
            date_from=date_from,
            date_to=date_to,
        )
        payload["comparison"] = _project_dashboard_comparison(
            compare=compare,
            date_from=date_from,
            date_to=date_to,
            skill_id=skill_id,
            status=status,
            source_type=source_type,
            keyword=keyword,
            risk=risk,
            quality=quality,
            lifecycle_stage=lifecycle_stage,
        )
        return payload
    except ProjectLedgerError as exc:
        raise HTTPException(status_code=503, detail=f"项目台账暂不可用：{exc}") from exc


@app.get("/api/projects")
def list_projects(
    date_from: str = Query(default=""),
    date_to: str = Query(default=""),
    skill_id: str = Query(default=""),
    status: str = Query(default=""),
    source_type: str = Query(default=""),
    keyword: str = Query(default=""),
    risk: str = Query(default=""),
    quality: str = Query(default=""),
    lifecycle_stage: str = Query(default=""),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=100),
    sort_by: str = Query(default="updated_at"),
    sort_order: str = Query(default="desc"),
) -> dict[str, object]:
    try:
        return _project_ledger().list_projects(
            page=page,
            page_size=page_size,
            sort_by=sort_by,
            sort_order=sort_order,
            **_project_filters(
                date_from=date_from,
                date_to=date_to,
                skill_id=skill_id,
                status=status,
                source_type=source_type,
                keyword=keyword,
                risk=risk,
                quality=quality,
                lifecycle_stage=lifecycle_stage,
            ),
        )
    except ProjectLedgerError as exc:
        raise HTTPException(status_code=503, detail=f"项目台账暂不可用：{exc}") from exc


@app.post("/api/projects/backfill")
def backfill_projects(
    include_unclassified_legacy: bool = Query(default=False),
) -> dict[str, object]:
    try:
        return backfill_project_ledger(
            _project_ledger(),
            runtime_root=RUNTIME_DIR,
            collaboration_db_path=feishu_app_bot.DB_PATH,
            collaboration_runtime_root=feishu_app_bot.RUNTIME_ROOT,
            include_unclassified_legacy=include_unclassified_legacy,
        )
    except ProjectLedgerError as exc:
        raise HTTPException(status_code=503, detail=f"项目台账回填失败：{exc}") from exc


@app.get("/api/projects/{project_id}")
def get_project(project_id: str) -> dict[str, object]:
    try:
        return _project_ledger().project_detail(project_id)
    except ProjectNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


@app.get("/api/projects/{project_id}/tasks")
def get_project_tasks(project_id: str) -> dict[str, object]:
    try:
        items = _business_task_store().list_project_tasks(project_id)
    except BusinessTaskError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return {"items": items, "count": len(items)}


@app.get("/api/tasks/{task_id}")
def get_business_task(task_id: str) -> dict[str, object]:
    try:
        return _business_task_detail(task_id)
    except BusinessTaskNotFound as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except BusinessTaskError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except (OSError, sqlite3.Error, json.JSONDecodeError) as exc:
        raise HTTPException(status_code=503, detail="任务轨迹暂不可用，专业处理不受影响。") from exc


@app.get("/api/tasks/{task_id}/timeline")
def get_business_task_timeline(task_id: str) -> dict[str, object]:
    try:
        return _business_task_store().timeline(task_id)
    except BusinessTaskNotFound as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except BusinessTaskError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except (OSError, sqlite3.Error, json.JSONDecodeError) as exc:
        raise HTTPException(status_code=503, detail="任务轨迹暂不可用，专业处理不受影响。") from exc
    except ProjectLedgerError as exc:
        raise HTTPException(status_code=503, detail=f"项目台账暂不可用：{exc}") from exc


@app.get("/api/projects/{project_id}/runs")
def get_project_runs(project_id: str) -> dict[str, object]:
    try:
        return {"items": _project_ledger().list_runs(project_id)}
    except ProjectNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


@app.get("/api/projects/{project_id}/artifacts")
def get_project_artifacts(project_id: str) -> dict[str, object]:
    try:
        return {"items": _project_ledger().list_artifacts(project_id)}
    except ProjectNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


@app.get("/api/projects/{project_id}/runs/{run_id}/result")
def get_project_run_result(project_id: str, run_id: str) -> dict[str, object]:
    try:
        run = _project_ledger().get_run(project_id, run_id)
    except ProjectNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    job_id = str(run.get("job_id") or "")
    if not job_id:
        raise HTTPException(status_code=409, detail="该协同任务暂无可恢复的网页预览")
    job_dir = RUNTIME_DIR / job_id
    if not job_dir.is_dir():
        raise HTTPException(status_code=410, detail="历史任务运行文件已失效")
    state = _load_process_state(job_dir)
    summary = _summary_from_dict(state.get("summary", {}))
    output_excel = _state_path(job_dir, state, "output_excel", required=False)
    output_report = _state_path(job_dir, state, "output_report", required=False)
    task_tracking = _sync_business_task_from_job(job_dir, activity="restore")
    return _attach_job_skill(
        {
            "job_id": job_id,
            "summary": summary.to_dict(),
            "downloads": {
                "excel": f"/api/download/{job_id}/excel"
                if output_excel and output_excel.is_file()
                else "",
                "report": f"/api/download/{job_id}/report"
                if output_report and output_report.is_file()
                else "",
            },
            "project_tracking": {
                "status": "available",
                "project_id": project_id,
                "run_id": run_id,
            },
            "task_tracking": task_tracking,
        },
        job_dir,
        state,
    )


@app.get("/api/projects/{project_id}/artifacts/{artifact_id}/download")
def download_project_artifact(project_id: str, artifact_id: str) -> FileResponse:
    try:
        path = _project_ledger().get_artifact_path(project_id, artifact_id)
    except ProjectArtifactNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except ProjectNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    return FileResponse(path, filename=path.name)


@app.post("/api/inspect")
async def inspect_excel(
    file: UploadFile = File(...),
    header_row: int | None = Form(default=None),
    sheet_name: str | None = Form(default=None),
    field_preferences: str | None = Form(default=None),
) -> dict[str, object]:
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 文件")

    job_id = uuid4().hex
    job_dir = RUNTIME_DIR / "inspect" / job_id
    job_dir.mkdir(parents=True, exist_ok=True)
    input_path = job_dir / file.filename
    input_path.write_bytes(await file.read())

    detected_row, headers = _read_headers(input_path, header_row=header_row, sheet_name=sheet_name)
    columns = _build_column_options(headers)
    input_preferences = _parse_input_field_preferences_form(field_preferences)
    sheets = _inspect_candidate_sheets(input_path, preferences=input_preferences)
    return {
        "header_row": detected_row,
        "headers": headers,
        "columns": columns,
        "suggested_mapping": _suggest_column_mapping(headers, input_preferences),
        "sheets": sheets,
    }


@app.get("/api/project-default-settings")
async def get_project_default_settings() -> dict[str, object]:
    return _project_default_settings_payload()


@app.get("/api/professional-skills")
async def list_professional_skills() -> dict[str, object]:
    try:
        return PROFESSIONAL_SKILL_REGISTRY.list_public()
    except ProfessionalSkillError as exc:
        raise HTTPException(status_code=exc.status_code, detail=exc.detail()) from exc


@app.post("/api/professional-skills/recommend")
async def recommend_professional_skill(file: UploadFile = File(...)) -> dict[str, object]:
    try:
        content = await file.read(MAX_RECOMMENDATION_FILE_BYTES + 1)
        return PROFESSIONAL_SKILL_REGISTRY.recommend_for_file(file.filename or "", content)
    except ProfessionalSkillError as exc:
        raise HTTPException(status_code=exc.status_code, detail=exc.detail()) from exc


@app.get("/api/professional-skills/open-format")
async def get_professional_skill_open_format() -> dict[str, object]:
    return PROFESSIONAL_SKILL_REGISTRY.open_format()


@app.get("/api/professional-skills/management")
async def get_professional_skill_management() -> dict[str, object]:
    try:
        return PROFESSIONAL_SKILL_REGISTRY.management_overview()
    except ProfessionalSkillError as exc:
        raise HTTPException(status_code=exc.status_code, detail=exc.detail()) from exc


@app.post("/api/professional-skills/management/plan")
async def plan_professional_skill_lifecycle(payload: dict[str, object] = Body(...)) -> dict[str, object]:
    try:
        return PROFESSIONAL_SKILL_REGISTRY.plan_lifecycle(
            str(payload.get("skill_id") or ""),
            str(payload.get("action") or ""),
            str(payload.get("target_version") or "") or None,
        )
    except ProfessionalSkillError as exc:
        raise HTTPException(status_code=exc.status_code, detail=exc.detail()) from exc


@app.get("/api/professional-skills/{skill_id}/onboarding-evidence")
async def get_professional_skill_onboarding_evidence(skill_id: str) -> dict[str, object]:
    try:
        skill = PROFESSIONAL_SKILL_REGISTRY.get_public(skill_id)
    except ProfessionalSkillError as exc:
        raise HTTPException(status_code=exc.status_code, detail=exc.detail()) from exc

    warnings: list[str] = []
    task_details: list[dict[str, object]] = []
    try:
        task_summaries = _business_task_store().list_recent_tasks(skill_id=skill_id, limit=5)
    except Exception:
        task_summaries = []
        warnings.append("Task 聚合暂不可用；专业处理主流程不受影响。")
    for task in task_summaries:
        try:
            detail = _business_task_detail(str(task.get("task_id") or ""))
        except Exception:
            detail = task
            warnings.append("个别 Task 血缘暂不可用，已保留安全摘要。")
        lineage = detail.get("lineage") if isinstance(detail.get("lineage"), dict) else {}
        events = lineage.get("experience_events") if isinstance(lineage.get("experience_events"), list) else []
        for event in events:
            if not isinstance(event, dict) or not event.get("candidate_id") or not event.get("project_key"):
                continue
            try:
                candidate = _knowledge_memory_store().get_item(
                    str(event["candidate_id"]),
                    str(event["project_key"]),
                )
                event["governance_status"] = str(candidate.get("status") or "candidate")
            except Exception:
                event["governance_status"] = "candidate"
                warnings.append("个别 Experience 治理状态暂不可用，按候选状态保守展示。")
        task_details.append(detail)

    try:
        validation_evidence = PROFESSIONAL_SKILL_REGISTRY.onboarding_validation(skill_id)
    except ProfessionalSkillError:
        validation_evidence = {}
        warnings.append("正式验证事实暂不可用，已回退到 Manifest 验证状态。")
    try:
        experience_metrics = _trusted_experience_store().metrics()
    except Exception:
        experience_metrics = {}
        warnings.append("可信经验指标暂不可用；已有任务和成果证据仍可查看。")

    return build_onboarding_evidence(
        skill=skill,
        tasks=task_details,
        validation_evidence=validation_evidence,
        experience_metrics=experience_metrics,
        aggregation_warnings=warnings,
    )


@app.get("/api/professional-skills/{skill_id}")
async def get_professional_skill(skill_id: str) -> dict[str, object]:
    try:
        return PROFESSIONAL_SKILL_REGISTRY.get_public(skill_id)
    except ProfessionalSkillError as exc:
        raise HTTPException(status_code=exc.status_code, detail=exc.detail()) from exc


@app.get("/api/input/field-preferences")
async def get_input_field_preferences() -> dict[str, object]:
    return _input_field_preferences_payload()


@app.post("/api/input/field-preferences")
async def save_input_field_preferences(payload: dict[str, object] = Body(...)) -> dict[str, object]:
    raw_preferences = payload.get("preferences")
    if raw_preferences is None:
        raw_preferences = payload
    if not isinstance(raw_preferences, dict):
        raise HTTPException(status_code=400, detail="输入字段偏好必须是对象")
    preferences = _sanitize_input_field_preferences(raw_preferences)
    return _input_field_preferences_payload(preferences)


@app.get("/api/ui-preferences")
async def get_ui_preferences() -> dict[str, object]:
    return _ui_preferences_payload()


@app.post("/api/ui-preferences")
async def save_ui_preferences(payload: dict[str, object] = Body(...)) -> dict[str, object]:
    raw_preferences = payload.get("preferences")
    if raw_preferences is None:
        raw_preferences = payload
    if not isinstance(raw_preferences, dict):
        raise HTTPException(status_code=400, detail="页面用户设置必须是对象")
    preferences = _sanitize_ui_preferences(raw_preferences)
    _save_ui_preferences(preferences)
    return _ui_preferences_payload(preferences)


@app.get("/api/preview-column-preferences")
async def get_preview_column_preferences() -> dict[str, object]:
    return _preview_column_preferences_payload()


@app.post("/api/preview-column-preferences")
async def save_preview_column_preferences(payload: dict[str, object] = Body(...)) -> dict[str, object]:
    raw_preferences = payload.get("preferences")
    if raw_preferences is None:
        raw_preferences = payload
    if not isinstance(raw_preferences, dict):
        raise HTTPException(status_code=400, detail="预览列设置必须是对象")
    preferences = _sanitize_preview_column_preferences(raw_preferences)
    return _preview_column_preferences_payload(preferences)


@app.get("/api/collaboration/feishu-webhook/status")
def get_feishu_webhook_status() -> dict[str, object]:
    return feishu_webhook.get_status()


@app.post("/api/collaboration/feishu-webhook/settings")
def update_feishu_webhook_settings(payload: dict[str, object] = Body(...)) -> dict[str, object]:
    try:
        return feishu_webhook.save_settings(payload)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/api/collaboration/feishu-webhook/test")
def test_feishu_webhook() -> dict[str, object]:
    outcome = feishu_webhook.send_notification("test")
    if outcome.skipped:
        raise HTTPException(status_code=409, detail=outcome.error)
    if not outcome.success:
        raise HTTPException(status_code=502, detail=outcome.error or "飞书 Webhook 测试发送失败")
    return outcome.to_dict()


@app.get("/api/collaboration/feishu-webhook/history")
def get_feishu_webhook_history(limit: int = 50) -> dict[str, object]:
    return {"items": feishu_webhook.read_history(limit=limit)}


@app.get("/api/collaboration/feishu-app-bot/status")
def get_feishu_app_bot_status() -> dict[str, object]:
    return feishu_app_bot.bot_status()


@app.get("/api/collaboration/feishu-app-bot/tasks")
def get_feishu_app_bot_tasks(limit: int = 30) -> dict[str, object]:
    store = feishu_app_bot.TaskStore()
    return {"items": [feishu_app_bot.public_task(task) for task in store.list_tasks(limit=limit)]}


@app.get("/api/collaboration/feishu-app-bot/logs")
def get_feishu_app_bot_logs(limit: int = 200) -> dict[str, object]:
    return {"items": feishu_app_bot.read_console_events(limit=limit)}


@app.post("/api/collaboration/feishu-app-bot/settings")
def update_feishu_app_bot_settings(payload: dict[str, object] = Body(...)) -> dict[str, object]:
    if "enabled" not in payload or not isinstance(payload.get("enabled"), bool):
        raise HTTPException(status_code=400, detail="enabled 必须是布尔值")
    enabled = bool(payload["enabled"])
    profile_id = payload.get("profile_id")
    if profile_id is None:
        next_profile = feishu_app_bot.active_profile_id()
    elif not isinstance(profile_id, str) or not profile_id.strip():
        raise HTTPException(status_code=400, detail="profile_id 必须是非空字符串")
    else:
        next_profile = profile_id.strip()
    if next_profile not in {item["profile_id"] for item in feishu_app_bot.credential_profiles()}:
        raise HTTPException(status_code=400, detail="未找到指定的飞书机器人配置")
    feishu_app_bot.save_bot_enabled(enabled, next_profile)
    if enabled:
        configuration_issue = feishu_app_bot.credential_configuration_issue(next_profile)
        if configuration_issue:
            feishu_app_bot.save_bot_enabled(False, next_profile)
            raise HTTPException(status_code=409, detail=configuration_issue)
    try:
        start_results = feishu_app_bot.start_enabled_bot_processes()
    except RuntimeError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    if enabled and not start_results.get(
        next_profile,
        feishu_app_bot.bot_process_running(next_profile),
    ):
        feishu_app_bot.save_bot_enabled(False, next_profile)
        raise HTTPException(status_code=409, detail="该平台机器人凭证未配置完整或进程启动失败")
    return feishu_app_bot.bot_status()


@app.get("/api/collaboration/external-dispatch/options")
def get_external_dispatch_options(
    profile_id: str | None = Query(default=None),
    refresh_directory: bool = Query(default=False),
) -> dict[str, object]:
    try:
        service = external_task_dispatch.build_service(
            registry=PROFESSIONAL_SKILL_REGISTRY,
            profile_id=profile_id,
        )
        return {
            **service.options(refresh_directory=refresh_directory),
            "active_profile": service.profile_id,
            "platforms": external_task_dispatch.configured_platforms(),
            "skills": [
                item for item in PROFESSIONAL_SKILL_REGISTRY.list_public().get("items", [])
                if isinstance(item, dict) and item.get("can_create_task")
            ],
        }
    except external_task_dispatch.DispatchValidationError as exc:
        raise HTTPException(status_code=exc.status_code, detail=str(exc)) from exc
    except (httpx.HTTPError, RuntimeError, ValueError) as exc:
        raise HTTPException(status_code=502, detail=feishu_app_bot.sanitize_error(exc)) from exc


@app.get("/api/collaboration/external-dispatch/tasks")
def get_external_dispatch_tasks(limit: int = Query(default=30, ge=1, le=100)) -> dict[str, object]:
    store = external_task_dispatch.ExternalDispatchStore()
    return {
        "items": [external_task_dispatch.public_dispatch_task(task) for task in store.list_tasks(limit=limit)]
    }


@app.post("/api/collaboration/external-dispatch/tasks")
async def create_external_dispatch_task(
    file: UploadFile = File(...),
    event_id: str = Form(...),
    source_task_id: str = Form(default=""),
    task_name: str = Form(...),
    project_name: str = Form(default=""),
    skill_id: str = Form(...),
    skill_version: str = Form(default=""),
    delivery_mode: str = Form(default="group"),
    delivery_policy_json: str = Form(default="{}"),
    platform_profile_id: str = Form(...),
    target_group_ref: str = Form(default=""),
    assignee_ref: str = Form(...),
    reviewer_refs_json: str = Form(default="[]"),
    deadline: str = Form(...),
    instructions: str = Form(...),
    template_version: str = Form(default="v1.0"),
) -> dict[str, object]:
    try:
        service = external_task_dispatch.build_service(
            registry=PROFESSIONAL_SKILL_REGISTRY,
            profile_id=platform_profile_id,
        )
        file_bytes = await file.read()
        max_bytes = int(feishu_app_bot.load_bot_defaults().get("maxFileSizeMb") or 50) * 1024 * 1024
        if len(file_bytes) > max_bytes:
            raise external_task_dispatch.DispatchValidationError(
                f"待填模板超过 {max_bytes // 1024 // 1024} MB 上限"
            )
        resolved_source_task_id = source_task_id.strip() or external_task_dispatch.generate_dispatch_source_task_id()
        resolved_project_name = project_name.strip() or external_task_dispatch.generate_dispatch_project_name()
        try:
            reviewer_refs_value = json.loads(reviewer_refs_json)
        except json.JSONDecodeError as exc:
            raise external_task_dispatch.DispatchValidationError("复核人参数格式无效") from exc
        if not isinstance(reviewer_refs_value, list):
            raise external_task_dispatch.DispatchValidationError("复核人参数必须是列表")
        try:
            delivery_policy_value = json.loads(delivery_policy_json)
        except json.JSONDecodeError as exc:
            raise external_task_dispatch.DispatchValidationError("投递策略参数格式无效") from exc
        if not isinstance(delivery_policy_value, dict):
            raise external_task_dispatch.DispatchValidationError("投递策略参数必须是对象")
        envelope = external_task_dispatch.TaskEnvelope(
            event_id=event_id.strip(),
            event_type=external_task_dispatch.EVENT_TYPE,
            source_system=external_task_dispatch.SOURCE_SYSTEM,
            source_task_id=resolved_source_task_id,
            task_name=task_name.strip(),
            project_name=resolved_project_name,
            skill_id=skill_id.strip(),
            skill_version=skill_version.strip(),
            delivery_mode=delivery_mode.strip(),
            platform_profile_id=platform_profile_id.strip(),
            target_group_ref=target_group_ref.strip(),
            assignee_ref=assignee_ref.strip(),
            deadline=deadline.strip(),
            instructions=instructions.strip(),
            input_artifact=external_task_dispatch.TaskArtifact(
                template_asset_id=str(file.filename or "").strip(),
                template_version=template_version.strip(),
            ),
            reviewer_refs=tuple(str(item).strip() for item in reviewer_refs_value if str(item).strip()),
            delivery_policy=delivery_policy_value,
        )
        task, created = service.create_and_deliver(
            envelope,
            file_name=str(file.filename or ""),
            file_bytes=file_bytes,
        )
        task_tracking = _sync_business_task_from_dispatch(task)
        return {"created": created, "task": task, "task_tracking": task_tracking}
    except external_task_dispatch.DispatchValidationError as exc:
        raise HTTPException(status_code=exc.status_code, detail=str(exc)) from exc


@app.post("/api/collaboration/external-dispatch/web-review")
def create_web_result_review(payload: dict[str, object] = Body(...)) -> dict[str, object]:
    job_id = str(payload.get("job_id") or "").strip()
    if not re.fullmatch(r"[0-9a-fA-F]{32}", job_id):
        raise HTTPException(status_code=400, detail="网页填价任务编号格式无效")
    start_new_round = payload.get("start_new_round", False)
    if not isinstance(start_new_round, bool):
        raise HTTPException(status_code=400, detail="新一轮复核参数格式无效")
    existing_task_id = str(payload.get("existing_task_id") or "").strip()
    previous_review_round = payload.get("previous_review_round", 0)
    if isinstance(previous_review_round, bool) or not isinstance(previous_review_round, int):
        raise HTTPException(status_code=400, detail="上一轮复核轮次格式无效")
    if start_new_round and (not existing_task_id or previous_review_round < 1):
        raise HTTPException(status_code=400, detail="请先刷新并确认上一轮复核任务")
    job_dir = RUNTIME_DIR / job_id
    if not job_dir.is_dir():
        raise HTTPException(status_code=404, detail="未找到网页填价任务")
    state = _load_process_state(job_dir)
    summary = state.get("summary") if isinstance(state.get("summary"), dict) else {}
    if str(summary.get("matching_status") or "") == "pending":
        raise HTTPException(status_code=409, detail="请先完成批量匹配，再提交同事复核")
    output_path = _state_path(job_dir, state, "output_excel")
    if not output_path or not output_path.is_file():
        raise HTTPException(status_code=404, detail="网页填价成果不存在，请重新生成")
    max_bytes = int(feishu_app_bot.load_bot_defaults().get("maxFileSizeMb") or 50) * 1024 * 1024
    file_bytes = output_path.read_bytes()
    if len(file_bytes) > max_bytes:
        raise HTTPException(status_code=413, detail=f"网页填价成果超过 {max_bytes // 1024 // 1024} MB 上限")

    reviewer_refs = payload.get("reviewer_refs")
    if not isinstance(reviewer_refs, list):
        raise HTTPException(status_code=400, detail="复核人参数必须是列表")
    platform_profile_id = str(payload.get("platform_profile_id") or "").strip()
    if not platform_profile_id:
        raise HTTPException(status_code=400, detail="请选择飞书或 WeAct 投递平台")
    skill_snapshot = state.get("skill_snapshot") if isinstance(state.get("skill_snapshot"), dict) else {}
    skill_id = str(skill_snapshot.get("id") or "").strip()
    skill_version = str(skill_snapshot.get("version") or "").strip()
    if not skill_id:
        raise HTTPException(status_code=409, detail="当前网页任务缺少专业能力快照，不能发起复核")

    input_name = str(state.get("input_filename") or output_path.name).strip()
    project_name = (
        str(payload.get("project_name") or "").strip()
        or str(state.get("project_name") or "").strip()
        or Path(input_name).stem
        or "网页填价项目"
    )
    task_name = (
        str(payload.get("task_name") or "").strip()
        or f"{project_name}填价成果复核"
    )
    try:
        service = external_task_dispatch.build_service(
            registry=PROFESSIONAL_SKILL_REGISTRY,
            profile_id=platform_profile_id,
        )
        task, created = service.create_web_result_review(
            job_id=job_id,
            file_name=output_path.name,
            file_bytes=file_bytes,
            task_name=task_name,
            project_name=project_name,
            skill_id=skill_id,
            skill_version=skill_version,
            reviewer_refs=tuple(str(item).strip() for item in reviewer_refs if str(item).strip()),
            deadline=str(payload.get("deadline") or "").strip(),
            instructions=str(payload.get("instructions") or "").strip(),
            delivery_mode=str(payload.get("delivery_mode") or "").strip(),
            target_group_ref=str(payload.get("target_group_ref") or "").strip(),
            start_new_round=start_new_round,
            existing_task_id=existing_task_id,
            previous_review_round=previous_review_round,
        )
        task_tracking = _link_business_task_collaboration(job_dir, task)
        return {
            "created": created,
            "started_new_round": bool(start_new_round and created),
            "task": task,
            "task_tracking": task_tracking,
        }
    except external_task_dispatch.DispatchValidationError as exc:
        raise HTTPException(status_code=exc.status_code, detail=str(exc)) from exc
    except (httpx.HTTPError, RuntimeError, ValueError) as exc:
        raise HTTPException(status_code=502, detail=feishu_app_bot.sanitize_error(exc)) from exc


@app.post("/api/collaboration/external-dispatch/tasks/{task_id}/retry")
def retry_external_dispatch_task(task_id: str) -> dict[str, object]:
    store = external_task_dispatch.ExternalDispatchStore()
    task = store.get_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="未找到外部派发任务")
    try:
        service = external_task_dispatch.build_service(
            registry=PROFESSIONAL_SKILL_REGISTRY,
            profile_id=str(task.get("platform_profile_id") or ""),
            store=store,
        )
        return {"task": service.retry(task_id)}
    except external_task_dispatch.DispatchValidationError as exc:
        raise HTTPException(status_code=exc.status_code, detail=str(exc)) from exc


@app.post("/api/collaboration/feishu-webhook/notify")
def send_feishu_webhook_notification(payload: dict[str, object] = Body(...)) -> dict[str, object]:
    notification_type = str(payload.get("notification_type") or "").strip()
    if notification_type not in feishu_webhook.ALLOWED_NOTIFICATION_TYPES - {"test"}:
        raise HTTPException(status_code=400, detail="不支持的通知类型")
    context = payload.get("context") or {}
    if not isinstance(context, dict):
        raise HTTPException(status_code=400, detail="通知上下文必须是对象")
    return feishu_webhook.send_notification(notification_type, context).to_dict()


@app.post("/api/process")
async def process_excel(
    file: UploadFile = File(...),
    column_mapping: str | None = Form(default=None),
    sheet_configs: str | None = Form(default=None),
    header_row: int = Form(default=1),
    output_match_report: bool = Form(default=True),
    merge_vertical_cells: bool = Form(default=True),
    merge_horizontal_cells: bool = Form(default=True),
    only_match_rows_with_value: bool = Form(default=True),
    match_value_filter_field: str = Form(default=DEFAULT_WARNING_FILTER_FIELD),
    defer_matching: bool = Form(default=False),
    skill_id: str | None = Form(default=None),
    skill_version: str | None = Form(default=None),
    project_id: str | None = Form(default=None),
    project_name: str | None = Form(default=None),
    source_type: str = Form(default="web"),
    source_task_id: str | None = Form(default=None),
    task_name: str | None = Form(default=None),
    task_objective: str | None = Form(default=None),
    task_instructions: str | None = Form(default=None),
) -> dict[str, object]:
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 文件")
    skill_snapshot = _resolve_professional_skill_snapshot(skill_id, skill_version)
    runtime_context = _skill_runtime_context(skill_snapshot)
    allowed_extensions = {
        str(value).lower()
        for value in runtime_context.input_profile.get("extensions", [])
        if str(value).strip()
    }
    if Path(file.filename).suffix.lower() not in allowed_extensions:
        raise HTTPException(status_code=400, detail="上传文件不符合当前专业能力输入规范")

    job_id = uuid4().hex
    job_dir = RUNTIME_DIR / job_id
    job_dir.mkdir(parents=True, exist_ok=True)
    input_path = job_dir / file.filename
    input_path.write_bytes(await file.read())

    initial_state = {
        "input_filename": file.filename,
        "input_excel": input_path.name,
        "output_excel": "",
        "output_report": "",
        "summary": {"matching_status": "pending"},
        "skill_snapshot": skill_snapshot,
        "source_task_id": str(source_task_id or "").strip(),
        "business_task_seed": {
            "task_name": str(task_name or "").strip(),
            "objective": str(task_objective or "").strip(),
            "instructions": str(task_instructions or "").strip(),
        },
        "project_id": str(project_id or "").strip(),
        "project_name": str(project_name or "").strip(),
        "source_type": str(source_type or "web").strip(),
        "created_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "updated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
    }
    (job_dir / PROCESS_STATE_FILENAME).write_text(
        json.dumps(initial_state, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    _sync_business_task_from_job(job_dir, activity="process_received")

    output_timestamp = _output_timestamp()
    output_excel = job_dir / f"{OUTPUT_FILE_PREFIX}-控制价计算表-{output_timestamp}.xlsx"
    output_report = job_dir / f"{OUTPUT_FILE_PREFIX}-控制价报告-{output_timestamp}.docx"

    mapping = _parse_column_mapping(column_mapping)
    parsed_sheet_configs = _parse_sheet_configs(sheet_configs)
    parsed_match_value_filter_field = _parse_warning_filter_field(match_value_filter_field)
    if defer_matching:
        output_excel.write_bytes(input_path.read_bytes())
        try:
            summary = _build_pending_match_summary(
                output_excel,
                column_mapping=mapping,
                header_row=header_row,
                sheet_configs=parsed_sheet_configs,
                only_match_rows_with_value=only_match_rows_with_value,
                match_value_filter_field=parsed_match_value_filter_field,
            )
        except ValueError as exc:
            _record_business_task_failure(
                job_dir, event_type="structure_recognized", source_module="fill_engine",
                detail=str(exc), event_key="structure-failed",
            )
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        summary.output_excel = output_excel.name
        summary.output_report = ""
        summary.warning_summary = _warning_not_run_summary(
            runtime_context.risk_profile.get("warningSettings"),
            runtime_context.risk_profile.get("experiencePool"),
        )
        summary.warning_details = []
        _save_process_state(
            job_dir,
            file.filename,
            input_path,
            output_excel,
            None,
            summary,
            extra={
                "skill_snapshot": skill_snapshot,
                "deferred_matching": True,
                "source_task_id": str(source_task_id or "").strip(),
                "business_task_seed": {
                    "task_name": str(task_name or "").strip(),
                    "objective": str(task_objective or "").strip(),
                    "instructions": str(task_instructions or "").strip(),
                },
                "process_options": {
                    "column_mapping": mapping,
                    "sheet_configs": parsed_sheet_configs,
                    "header_row": header_row,
                    "output_match_report": output_match_report,
                    "merge_vertical_cells": merge_vertical_cells,
                    "merge_horizontal_cells": merge_horizontal_cells,
                    "only_match_rows_with_value": only_match_rows_with_value,
                    "match_value_filter_field": parsed_match_value_filter_field,
                },
            },
        )
        project_tracking = _sync_project_ledger(
            job_dir,
            project_id=project_id,
            project_name=project_name or "",
            source_type=source_type,
            create_project=bool(str(project_name or "").strip()),
        )
        task_tracking = _sync_business_task_from_job(job_dir, activity="process_pending")
        return _attach_job_skill({
            "job_id": job_id,
            "summary": summary.to_dict(),
            "downloads": {
                "excel": "",
                "report": "",
            },
            "project_tracking": project_tracking,
            "task_tracking": task_tracking,
        }, job_dir)
    try:
        summary = _build_skill_fill_engine(runtime_context).fill_workbook(
            input_path,
            output_excel,
            column_mapping=mapping,
            header_row=header_row,
            output_match_report=output_match_report,
            merge_vertical_cells=merge_vertical_cells,
            merge_horizontal_cells=merge_horizontal_cells,
            only_match_rows_with_value=only_match_rows_with_value,
            match_value_filter_field=parsed_match_value_filter_field,
            sheet_configs=parsed_sheet_configs,
        )
    except ValueError as exc:
        _record_business_task_failure(
            job_dir, event_type="rules_executed", source_module="fill_engine",
            detail=str(exc), event_key="rules-failed",
        )
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    recalculate_workbook(output_excel)
    summary.warning_summary = _warning_not_run_summary(
        runtime_context.risk_profile.get("warningSettings"),
        runtime_context.risk_profile.get("experiencePool"),
    )
    summary.warning_details = []
    summary.table_preview = _refresh_table_preview_from_output(summary.table_preview, output_excel)
    _require_skill_capability(runtime_context, "wordReport")
    output_report = write_report(
        output_report,
        file.filename,
        summary,
        output_excel_path=output_excel,
        input_excel_path=input_path,
        report_template_path=runtime_context.report_template_path,
    )
    summary.output_report = output_report.name
    _save_process_state(
        job_dir,
        file.filename,
        input_path,
        output_excel,
        output_report,
        summary,
        extra={
            "skill_snapshot": skill_snapshot,
            "source_task_id": str(source_task_id or "").strip(),
            "business_task_seed": {
                "task_name": str(task_name or "").strip(),
                "objective": str(task_objective or "").strip(),
                "instructions": str(task_instructions or "").strip(),
            },
        },
    )
    project_tracking = _sync_project_ledger(
        job_dir,
        project_id=project_id,
        project_name=project_name or "",
        source_type=source_type,
        create_project=bool(str(project_name or "").strip()),
    )
    task_tracking = _sync_business_task_from_job(job_dir, activity="process_completed")

    return _attach_job_skill({
        "job_id": job_id,
        "summary": summary.to_dict(),
        "downloads": {
            "excel": f"/api/download/{job_id}/excel",
            "report": f"/api/download/{job_id}/report",
        },
        "project_tracking": project_tracking,
        "task_tracking": task_tracking,
    }, job_dir)


@app.post("/api/process/batch-match")
async def batch_match_process(payload: dict[str, object] = Body(...)) -> dict[str, object]:
    job_id = str(payload.get("job_id") or "").strip()
    if not job_id:
        raise HTTPException(status_code=400, detail="缺少任务编号")
    job_dir = RUNTIME_DIR / job_id
    if not job_dir.exists():
        raise HTTPException(status_code=404, detail="任务不存在")
    state = _load_process_state(job_dir)
    runtime_context = _job_skill_runtime_context(job_dir, state)
    input_path = _state_path(job_dir, state, "input_excel", required=False)
    output_excel = _state_path(job_dir, state, "output_excel")
    if not output_excel or not output_excel.exists():
        raise HTTPException(status_code=404, detail="输出任务不存在，请重新上传")

    options = dict(state.get("process_options") or {})
    output_report = job_dir / f"{OUTPUT_FILE_PREFIX}-控制价报告-{_output_timestamp()}.docx"
    try:
        summary = _build_skill_fill_engine(runtime_context).fill_workbook(
            output_excel,
            output_excel,
            column_mapping=options.get("column_mapping"),
            header_row=int(options.get("header_row") or 1),
            output_match_report=bool(options.get("output_match_report", True)),
            merge_vertical_cells=bool(options.get("merge_vertical_cells", True)),
            merge_horizontal_cells=bool(options.get("merge_horizontal_cells", True)),
            only_match_rows_with_value=bool(options.get("only_match_rows_with_value", False)),
            match_value_filter_field=str(options.get("match_value_filter_field") or DEFAULT_WARNING_FILTER_FIELD),
            sheet_configs=options.get("sheet_configs"),
        )
    except ValueError as exc:
        _record_business_task_failure(
            job_dir, event_type="rules_executed", source_module="fill_engine",
            detail=str(exc), event_key="batch-rules-failed",
        )
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    recalculate_workbook(output_excel)
    summary.warning_summary = _warning_not_run_summary(
        runtime_context.risk_profile.get("warningSettings"),
        runtime_context.risk_profile.get("experiencePool"),
    )
    summary.warning_details = []
    summary.matching_status = "completed"
    summary.table_preview = _refresh_table_preview_from_output(
        summary.table_preview,
        output_excel,
        header_rows=_parse_preview_header_rows(payload.get("header_rows")),
    )
    _require_skill_capability(runtime_context, "wordReport")
    output_report = write_report(
        output_report,
        str(state.get("input_filename") or (input_path.name if input_path else output_excel.name)),
        summary,
        output_excel_path=output_excel,
        input_excel_path=input_path,
        report_template_path=runtime_context.report_template_path,
    )
    summary.output_excel = output_excel.name
    summary.output_report = output_report.name
    _save_process_state(
        job_dir,
        str(state.get("input_filename") or (input_path.name if input_path else output_excel.name)),
        input_path,
        output_excel,
        output_report,
        summary,
        extra={
            "deferred_matching": False,
            "process_options": options,
        },
    )
    project_tracking = _sync_project_ledger_from_job(job_dir)
    task_tracking = _sync_business_task_from_job(job_dir, activity="batch_match")
    return _attach_job_skill({
        "job_id": job_id,
        "summary": summary.to_dict(),
        "downloads": {
            "excel": f"/api/download/{job_id}/excel",
            "report": f"/api/download/{job_id}/report",
        },
        "project_tracking": project_tracking,
        "task_tracking": task_tracking,
    }, job_dir)


@app.post("/api/demo/load-sample")
async def load_demo_sample() -> dict[str, object]:
    sample_path = _find_demo_sample_path()
    if not sample_path:
        raise HTTPException(status_code=404, detail="未找到演示样例文件")
    return _process_existing_workbook(sample_path, demo_mode=True, sheet_configs=_demo_sample_sheet_configs(sample_path))


@app.get("/api/quality/experience-pool")
async def experience_pool_quality() -> dict[str, object]:
    report = build_experience_pool_governance_report(_resolve_experience_pool_path())
    report_path = RUNTIME_DIR / "experience-pool-governance-report.md"
    write_governance_markdown(report, report_path)
    return {**report, "report_path": str(report_path)}


@app.get("/api/risk/summary")
async def risk_summary(job_id: str) -> dict[str, object]:
    job_dir = RUNTIME_DIR / str(job_id).strip()
    if not job_dir.exists():
        raise HTTPException(status_code=404, detail="任务不存在")
    state = _load_process_state(job_dir)
    _job_skill_runtime_context(job_dir, state)
    summary = _summary_from_dict(state.get("summary", {}))
    items = build_structured_risk_items(summary)
    return _attach_job_skill({
        "job_id": job_id,
        "summary": summarize_risk_items(items),
        "items": items,
    }, job_dir, state)


@app.get("/api/standard-trace")
async def standard_trace(job_id: str, sheet_name: str, row_number: int) -> dict[str, object]:
    job_dir = RUNTIME_DIR / str(job_id).strip()
    if not job_dir.exists():
        raise HTTPException(status_code=404, detail="任务不存在")
    state = _load_process_state(job_dir)
    summary = _summary_from_dict(state.get("summary", {}))
    return _attach_job_skill({
        "job_id": job_id,
        "sheet_name": sheet_name,
        "row_number": row_number,
        "trace": build_standard_trace(summary, sheet_name, row_number),
    }, job_dir, state)


@app.post("/api/fill-assist/candidates")
async def fill_assist_candidates(payload: dict[str, object] = Body(...)) -> dict[str, object]:
    job_id = str(payload.get("job_id") or "").strip()
    sheet_name = str(payload.get("sheet_name") or "").strip()
    target_header = str(payload.get("target_header") or payload.get("field") or "").strip()
    if not job_id:
        raise HTTPException(status_code=400, detail="缺少任务编号")
    if not sheet_name:
        raise HTTPException(status_code=400, detail="缺少 sheet 名称")
    try:
        row_number = int(payload.get("row_number") or 0)
    except (TypeError, ValueError) as exc:
        raise HTTPException(status_code=400, detail="行号必须是数字") from exc
    if row_number < 1:
        raise HTTPException(status_code=400, detail="行号必须大于等于 1")

    job_dir = RUNTIME_DIR / job_id
    if not job_dir.exists():
        raise HTTPException(status_code=404, detail="任务不存在")
    state = _load_process_state(job_dir)
    runtime_context = _job_skill_runtime_context(job_dir, state)
    excel_path = _state_path(job_dir, state, "output_excel")
    if not excel_path or not excel_path.exists():
        raise HTTPException(status_code=404, detail="输出 Excel 不存在，请先完成转换")
    try:
        context = build_fill_assist_context(excel_path, sheet_name, row_number, target_header)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    summary = _summary_from_dict(state.get("summary", {}))
    candidates = build_fill_assist_candidates(
        dict(context.get("row") or {}),
        knowledge_base=KnowledgeBase.from_excel(runtime_context.knowledge_base_path),
        pool_path=runtime_context.risk_profile.get("experiencePool", _resolve_experience_pool_path()),
    )
    return _attach_job_skill({
        "job_id": job_id,
        "context": context,
        "candidates": candidates,
        "trace": build_standard_trace(summary, sheet_name, row_number),
    }, job_dir, state)


@app.post("/api/fill-assist/confirm")
async def confirm_fill_assist(payload: dict[str, object] = Body(...)) -> dict[str, object]:
    candidate = payload.get("candidate")
    note = str(payload.get("note") or "").strip()
    if not isinstance(candidate, dict):
        raise HTTPException(status_code=400, detail="缺少候选信息")
    if candidate.get("source") == "custom" and not note:
        raise HTTPException(status_code=400, detail="自定义值必须填写依据备注")
    payload = dict(payload)
    payload["value"] = candidate.get("value")
    payload["edit_source"] = "fill-assist"
    payload["edit_note"] = note
    payload["candidate_meta"] = candidate
    return await update_preview_cell(payload)


@app.post("/api/experience-warnings/run")
async def run_experience_warnings(
    job_id: str = Form(...),
    preview_header_rows: str | None = Form(default=None),
) -> dict[str, object]:
    job_dir = RUNTIME_DIR / job_id
    if not job_dir.exists():
        raise HTTPException(status_code=404, detail="任务不存在")

    state = _load_process_state(job_dir)
    runtime_context = _job_skill_runtime_context(job_dir, state)
    _require_skill_capability(runtime_context, "experienceWarning")
    excel_path = _state_path(job_dir, state, "output_excel")
    input_path = _state_path(job_dir, state, "input_excel", required=False)
    if not excel_path or not excel_path.exists():
        excel_matches = list(job_dir.glob("*-填价结果-【codex】.xlsx"))
        if not excel_matches:
            raise HTTPException(status_code=404, detail="输出 Excel 不存在，请先完成转换")
        excel_path = excel_matches[0]

    summary = _summary_from_dict(state.get("summary", {}))
    warning_settings = _load_experience_warning_settings(runtime_context.risk_profile.get("warningSettings"))
    _set_warning_progress(job_id, {"status": "running"})
    try:
        warning_result = analyze_workbook_warnings_with_progress(
            excel_path,
            runtime_context.risk_profile.get("experiencePool", _resolve_experience_pool_path()),
            progress_callback=lambda payload: _set_warning_progress(job_id, payload),
            low_risk_warning_ratio=float(warning_settings["low_risk_warning_ratio"]) / 100,
            high_risk_warning_ratio=float(warning_settings["high_risk_warning_ratio"]) / 100,
            only_check_rows_with_value=bool(warning_settings["only_check_rows_with_value"]),
            value_filter_field=str(warning_settings["value_filter_field"]),
        )
    except ValueError as exc:
        _set_warning_progress(job_id, {"status": "failed", "error": str(exc)})
        _record_business_task_failure(
            job_dir, event_type="risk_checked", source_module="experience_warning",
            detail=str(exc), event_key="risk-failed",
        )
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        _set_warning_progress(job_id, {"status": "failed", "error": str(exc)})
        _record_business_task_failure(
            job_dir, event_type="risk_checked", source_module="experience_warning",
            detail=str(exc), event_key="risk-failed-unexpected",
        )
        raise
    summary.warning_summary = warning_result["summary"]
    summary.warning_summary["executed"] = True
    summary.warning_details = warning_result["warnings"]
    write_warnings_to_workbook(excel_path, list(warning_result.get("row_results") or summary.warning_details))
    summary.table_preview = _refresh_table_preview_from_output(
        summary.table_preview,
        excel_path,
        header_rows=_parse_preview_header_rows(preview_header_rows),
    )

    input_name = str(state.get("input_filename") or (input_path.name if input_path else "input.xlsx"))
    report_path = _state_path(job_dir, state, "output_report", required=False)
    if not report_path:
        report_matches = _find_report_files(job_dir, ".docx")
        report_path = report_matches[0] if report_matches else job_dir / f"{OUTPUT_FILE_PREFIX}-控制价报告-{_output_timestamp()}.docx"
    _require_skill_capability(runtime_context, "wordReport")
    report_path = write_report(
        report_path,
        input_name,
        summary,
        output_excel_path=excel_path,
        input_excel_path=input_path,
        report_template_path=runtime_context.report_template_path,
    )
    summary.output_excel = excel_path.name
    summary.output_report = report_path.name
    _save_process_state(job_dir, input_name, input_path, excel_path, report_path, summary)
    project_tracking = _sync_project_ledger_from_job(job_dir)
    task_tracking = _sync_business_task_from_job(job_dir, activity="risk_checked")
    _set_warning_progress(
        job_id,
        {
            "status": "completed",
            "processed_rows": int(summary.warning_summary.get("candidate_rows") or 0),
            "total_rows": int(summary.warning_summary.get("total_candidate_rows") or summary.warning_summary.get("candidate_rows") or 0),
            "matched_rows": int(summary.warning_summary.get("checked_rows") or 0),
            "warning_rows": int(summary.warning_summary.get("warning_rows") or 0),
        },
    )

    return _attach_job_skill({
        "job_id": job_id,
        "summary": summary.to_dict(),
        "downloads": {
            "excel": f"/api/download/{job_id}/excel",
            "report": f"/api/download/{job_id}/report",
        },
        "project_tracking": project_tracking,
        "task_tracking": task_tracking,
    }, job_dir)


@app.post("/api/preview/refresh")
async def refresh_table_preview(payload: dict[str, object] = Body(...)) -> dict[str, object]:
    job_id = str(payload.get("job_id") or "").strip()
    if not job_id:
        raise HTTPException(status_code=400, detail="缺少任务编号")
    job_dir = RUNTIME_DIR / job_id
    if not job_dir.exists():
        raise HTTPException(status_code=404, detail="任务不存在")

    state = _load_process_state(job_dir)
    excel_path = _state_path(job_dir, state, "output_excel")
    if not excel_path or not excel_path.exists():
        raise HTTPException(status_code=404, detail="输出 Excel 不存在，请先完成转换")

    summary = _summary_from_dict(state.get("summary", {}))
    summary.table_preview = _refresh_table_preview_from_output(
        summary.table_preview,
        excel_path,
        header_rows=_parse_preview_header_rows(payload.get("header_rows")),
    )
    input_path = _state_path(job_dir, state, "input_excel", required=False)
    report_path = _state_path(job_dir, state, "output_report", required=False) or job_dir / str(state.get("output_report") or "")
    input_name = str(state.get("input_filename") or (input_path.name if input_path else "input.xlsx"))
    summary.output_excel = excel_path.name
    summary.output_report = report_path.name if report_path else str(state.get("output_report") or "")
    _save_process_state(job_dir, input_name, input_path, excel_path, report_path, summary)
    return _attach_job_skill({
        "job_id": job_id,
        "summary": summary.to_dict(),
        "downloads": {
            "excel": f"/api/download/{job_id}/excel",
            "report": f"/api/download/{job_id}/report",
        },
    }, job_dir)


@app.post("/api/preview/cell")
async def update_preview_cell(payload: dict[str, object] = Body(...)) -> dict[str, object]:
    job_id = str(payload.get("job_id") or "").strip()
    sheet_name = str(payload.get("sheet_name") or "").strip()
    if not job_id:
        raise HTTPException(status_code=400, detail="缺少任务编号")
    if not sheet_name:
        raise HTTPException(status_code=400, detail="缺少 sheet 名称")
    try:
        row_number = int(payload.get("row_number") or 0)
        column_number = int(payload.get("column_number") or 0)
    except (TypeError, ValueError) as exc:
        raise HTTPException(status_code=400, detail="行号和列号必须是数字") from exc
    if row_number < 1 or column_number < 1:
        raise HTTPException(status_code=400, detail="行号和列号必须大于等于 1")

    job_dir = RUNTIME_DIR / job_id
    if not job_dir.exists():
        raise HTTPException(status_code=404, detail="任务不存在")
    state = _load_process_state(job_dir)
    runtime_context = _job_skill_runtime_context(job_dir, state)
    excel_path = _state_path(job_dir, state, "output_excel")
    if not excel_path or not excel_path.exists():
        raise HTTPException(status_code=404, detail="输出 Excel 不存在，请先完成转换")

    summary = _summary_from_dict(state.get("summary", {}))
    header_rows = _parse_preview_header_rows(payload.get("header_rows"))
    should_recalculate = bool(payload.get("recalculate"))
    edit_record = _write_preview_cell_edit(
        excel_path,
        job_id=job_id,
        sheet_name=sheet_name,
        row_number=row_number,
        column_number=column_number,
        new_value=payload.get("value"),
        header_rows=header_rows,
        edit_source=str(payload.get("edit_source") or "manual"),
        edit_note=str(payload.get("edit_note") or "").strip(),
        edit_actor=str(payload.get("actor") or "本机试点用户").strip(),
        candidate_meta=payload.get("candidate_meta") if isinstance(payload.get("candidate_meta"), dict) else None,
    )
    _append_manual_edit_log(job_dir, edit_record)
    input_path = _state_path(job_dir, state, "input_excel", required=False)
    input_name = str(state.get("input_filename") or (input_path.name if input_path else "input.xlsx"))
    report_path = _state_path(job_dir, state, "output_report", required=False)
    recalculated = False
    if should_recalculate:
        recalculated = recalculate_workbook(excel_path)
        summary.table_preview = _refresh_table_preview_from_output(
            summary.table_preview,
            excel_path,
            header_rows=header_rows,
        )
        if not report_path:
            report_matches = _find_report_files(job_dir, ".docx")
            report_path = report_matches[0] if report_matches else job_dir / f"{OUTPUT_FILE_PREFIX}-控制价报告-{_output_timestamp()}.docx"
        report_path = write_report(
            report_path,
            input_name,
            summary,
            output_excel_path=excel_path,
            input_excel_path=input_path,
            report_template_path=runtime_context.report_template_path,
        )
    else:
        summary.table_preview = _apply_manual_edit_to_table_preview(summary.table_preview, edit_record, header_rows)
    summary.output_excel = excel_path.name
    summary.output_report = report_path.name if report_path else str(state.get("output_report") or "")
    _save_process_state(job_dir, input_name, input_path, excel_path, report_path, summary)
    trusted_experience = _capture_manual_edit_experience(
        job_dir,
        state=_load_process_state(job_dir),
        runtime_context=runtime_context,
        excel_path=excel_path,
        edit_record=edit_record,
    )
    task_tracking = _sync_business_task_from_job(
        job_dir,
        activity="manual_edit",
        activity_payload={"trusted_experience": trusted_experience},
    )

    return _attach_job_skill({
        "job_id": job_id,
        "summary": summary.to_dict(),
        "downloads": {
            "excel": f"/api/download/{job_id}/excel",
            "report": f"/api/download/{job_id}/report",
        },
        "manual_edit": edit_record,
        "manual_edits": _load_manual_edit_log(job_dir),
        "formula_recalculated": recalculated,
        "needs_recalculate": not should_recalculate,
        "trusted_experience": trusted_experience,
        "task_tracking": task_tracking,
    }, job_dir)


@app.post("/api/preview/recalculate")
async def recalculate_preview_workbook(payload: dict[str, object] = Body(...)) -> dict[str, object]:
    job_id = str(payload.get("job_id") or "").strip()
    if not job_id:
        raise HTTPException(status_code=400, detail="缺少任务编号")
    job_dir = RUNTIME_DIR / job_id
    if not job_dir.exists():
        raise HTTPException(status_code=404, detail="任务不存在")

    state = _load_process_state(job_dir)
    runtime_context = _job_skill_runtime_context(job_dir, state)
    excel_path = _state_path(job_dir, state, "output_excel")
    if not excel_path or not excel_path.exists():
        raise HTTPException(status_code=404, detail="输出 Excel 不存在，请先完成转换")

    summary = _summary_from_dict(state.get("summary", {}))
    header_rows = _parse_preview_header_rows(payload.get("header_rows"))
    recalculated = recalculate_workbook(excel_path)
    summary.table_preview = _refresh_table_preview_from_output(
        summary.table_preview,
        excel_path,
        header_rows=header_rows,
    )
    input_path = _state_path(job_dir, state, "input_excel", required=False)
    input_name = str(state.get("input_filename") or (input_path.name if input_path else "input.xlsx"))
    report_path = _state_path(job_dir, state, "output_report", required=False)
    if not report_path:
        report_matches = _find_report_files(job_dir, ".docx")
        report_path = report_matches[0] if report_matches else job_dir / f"{OUTPUT_FILE_PREFIX}-控制价报告-{_output_timestamp()}.docx"
    report_path = write_report(
        report_path,
        input_name,
        summary,
        output_excel_path=excel_path,
        input_excel_path=input_path,
        report_template_path=runtime_context.report_template_path,
    )
    summary.output_excel = excel_path.name
    summary.output_report = report_path.name
    _save_process_state(job_dir, input_name, input_path, excel_path, report_path, summary)
    task_tracking = _sync_business_task_from_job(job_dir, activity="recalculate")
    return _attach_job_skill({
        "job_id": job_id,
        "summary": summary.to_dict(),
        "downloads": {
            "excel": f"/api/download/{job_id}/excel",
            "report": f"/api/download/{job_id}/report",
        },
        "manual_edits": _load_manual_edit_log(job_dir),
        "formula_recalculated": recalculated,
        "needs_recalculate": False,
        "task_tracking": task_tracking,
    }, job_dir)


@app.get("/api/experience-warnings/progress/{job_id}")
async def get_experience_warning_progress(job_id: str) -> dict[str, object]:
    payload = _get_warning_progress(job_id)
    job_dir = RUNTIME_DIR / str(job_id).strip()
    if not job_dir.exists():
        return payload
    return _attach_job_skill(payload, job_dir)


@app.post("/api/experience-pool/inspect")
async def inspect_experience_pool_excel(
    file: UploadFile = File(...),
    header_row: int | None = Form(default=None),
    sheet_name: str | None = Form(default=None),
) -> dict[str, object]:
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 文件")

    job_id = uuid4().hex
    job_dir = RUNTIME_DIR / "experience-pool-inspect" / job_id
    job_dir.mkdir(parents=True, exist_ok=True)
    input_path = job_dir / file.filename
    input_path.write_bytes(await file.read())

    sheets = _inspect_experience_sheets(input_path, header_row=header_row, sheet_name=sheet_name)
    first = sheets[0] if sheets else {"header_row": 1, "headers": [], "columns": [], "suggested_mapping": {}}
    return {
        "header_row": first["header_row"],
        "headers": first["headers"],
        "columns": first["columns"],
        "suggested_mapping": first["suggested_mapping"],
        "sheets": sheets,
    }


@app.get("/api/experience-pool/field-preferences")
async def get_experience_field_preferences() -> dict[str, object]:
    return _experience_field_preferences_payload()


@app.post("/api/experience-pool/field-preferences")
async def save_experience_field_preferences(payload: dict[str, object] = Body(...)) -> dict[str, object]:
    raw_preferences = payload.get("preferences")
    if raw_preferences is None:
        raw_preferences = payload
    if not isinstance(raw_preferences, dict):
        raise HTTPException(status_code=400, detail="经验池字段偏好必须是对象")
    preferences = _sanitize_experience_field_preferences(raw_preferences)
    _save_experience_field_preferences(preferences)
    return _experience_field_preferences_payload(preferences)


@app.get("/api/experience-warnings/settings")
async def get_experience_warning_settings() -> dict[str, object]:
    return _experience_warning_settings_payload()


@app.post("/api/experience-warnings/settings")
async def save_experience_warning_settings(payload: dict[str, object] = Body(...)) -> dict[str, object]:
    raw_settings = payload.get("settings")
    if raw_settings is None:
        raw_settings = payload
    if not isinstance(raw_settings, dict):
        raise HTTPException(status_code=400, detail="预警设置必须是对象")
    settings = _sanitize_experience_warning_settings(raw_settings)
    _save_experience_warning_settings(settings)
    return _experience_warning_settings_payload(settings)


@app.post("/api/experience-pool/import")
async def import_experience_pool_endpoint(
    file: UploadFile = File(...),
    selected_fields: str | None = Form(default=None),
    sheet_configs: str | None = Form(default=None),
    only_import_rows_with_value: bool = Form(default=True),
    value_filter_field: str = Form(default="工程量"),
) -> dict[str, object]:
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 文件")

    selected = _parse_selected_experience_fields(selected_fields)
    job_id = uuid4().hex
    job_dir = RUNTIME_DIR / "experience-pool" / job_id
    job_dir.mkdir(parents=True, exist_ok=True)
    source_path = job_dir / file.filename
    source_path.write_bytes(await file.read())
    parsed_sheet_configs = _parse_experience_sheet_configs(sheet_configs)
    filter_field = _parse_experience_filter_field(value_filter_field) if only_import_rows_with_value else None
    try:
        summary = import_experience_pool(
            source_path,
            DEFAULT_EXPERIENCE_POOL_PATH,
            selected_fields=selected,
            sheet_configs=parsed_sheet_configs,
            template_path=DEFAULT_EXPERIENCE_POOL_TEMPLATE_PATH,
            filter_non_empty_field=filter_field,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return {
        "job_id": job_id,
        "summary": summary,
        "pool_file": str(DEFAULT_EXPERIENCE_POOL_PATH),
    }


@app.post("/api/workload-capture/inspect")
async def inspect_workload_capture_excel(
    file: UploadFile = File(...),
    role: str = Form(default="source"),
    header_row: int | None = Form(default=None),
    sheet_name: str | None = Form(default=None),
    field_preferences: str | None = Form(default=None),
    adjacent_fallback_enabled: str | None = Form(default=None),
    element_sequence_enabled: str | None = Form(default=None),
) -> dict[str, object]:
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 文件")
    clean_role = role if role in {"source", "target"} else "source"

    job_id = uuid4().hex
    job_dir = RUNTIME_DIR / "workload-capture-inspect" / job_id
    job_dir.mkdir(parents=True, exist_ok=True)
    input_path = job_dir / file.filename
    input_path.write_bytes(await file.read())

    sheets = _inspect_workload_sheets(
        input_path,
        clean_role,
        header_row=header_row,
        sheet_name=sheet_name,
        preferences=_parse_workload_field_preferences_form(field_preferences, clean_role),
        adjacent_fallback_enabled=_sanitize_optional_bool_setting(adjacent_fallback_enabled),
        element_sequence_enabled=_sanitize_optional_bool_setting(element_sequence_enabled),
    )
    first = sheets[0] if sheets else {"header_row": 1, "headers": [], "columns": [], "suggested_mapping": {}}
    return {
        "header_row": first["header_row"],
        "headers": first["headers"],
        "columns": first["columns"],
        "suggested_mapping": first["suggested_mapping"],
        "sheets": sheets,
    }


@app.post("/api/workload-capture/inspect-current-target")
async def inspect_current_workload_target(
    job_id: str = Form(...),
    header_row: int | None = Form(default=None),
    sheet_name: str | None = Form(default=None),
    field_preferences: str | None = Form(default=None),
    adjacent_fallback_enabled: str | None = Form(default=None),
    element_sequence_enabled: str | None = Form(default=None),
) -> dict[str, object]:
    clean_job_id = str(job_id or "").strip()
    if not clean_job_id:
        raise HTTPException(status_code=400, detail="缺少当前任务编号")
    job_dir = RUNTIME_DIR / clean_job_id
    if not job_dir.exists():
        raise HTTPException(status_code=404, detail="当前任务不存在，请先完成转换")
    state = _load_process_state(job_dir)
    runtime_context = _job_skill_runtime_context(job_dir, state)
    _require_skill_capability(runtime_context, "workloadCapture")
    excel_path = _state_path(job_dir, state, "output_excel")
    if not excel_path or not excel_path.exists():
        raise HTTPException(status_code=404, detail="当前预览控制价表不存在，请先完成转换")

    sheets = _inspect_workload_sheets(
        excel_path,
        "target",
        header_row=header_row,
        sheet_name=sheet_name,
        preferences=_parse_workload_field_preferences_form(field_preferences, "target"),
        adjacent_fallback_enabled=_sanitize_optional_bool_setting(adjacent_fallback_enabled),
        element_sequence_enabled=_sanitize_optional_bool_setting(element_sequence_enabled),
    )
    first = sheets[0] if sheets else {"header_row": 1, "headers": [], "columns": [], "suggested_mapping": {}}
    return {
        "header_row": first["header_row"],
        "headers": first["headers"],
        "columns": first["columns"],
        "suggested_mapping": first["suggested_mapping"],
        "sheets": sheets,
    }


@app.get("/api/workload-capture/field-preferences")
async def get_workload_field_preferences() -> dict[str, object]:
    return _workload_field_preferences_payload()


@app.post("/api/workload-capture/field-preferences")
async def save_workload_field_preferences(payload: dict[str, object] = Body(...)) -> dict[str, object]:
    raw_preferences = payload.get("preferences")
    if raw_preferences is None:
        raw_preferences = payload
    if not isinstance(raw_preferences, dict):
        raise HTTPException(status_code=400, detail="工作量字段偏好必须是对象")
    preferences = _sanitize_workload_field_preferences(raw_preferences)
    adjacent_fallback_enabled = _sanitize_bool_setting(payload.get("adjacent_fallback_enabled"), True)
    element_sequence_enabled = _sanitize_bool_setting(payload.get("element_sequence_enabled"), True)
    return _workload_field_preferences_payload(
        preferences,
        adjacent_fallback_enabled=adjacent_fallback_enabled,
        element_sequence_enabled=element_sequence_enabled,
    )


@app.get("/api/workload-capture/target-field-preferences")
async def get_workload_target_field_preferences() -> dict[str, object]:
    return _workload_target_field_preferences_payload()


@app.post("/api/workload-capture/target-field-preferences")
async def save_workload_target_field_preferences(payload: dict[str, object] = Body(...)) -> dict[str, object]:
    raw_preferences = payload.get("preferences")
    if raw_preferences is None:
        raw_preferences = payload
    if not isinstance(raw_preferences, dict):
        raise HTTPException(status_code=400, detail="控制价计算表字段偏好必须是对象")
    preferences = _sanitize_workload_target_field_preferences(raw_preferences)
    adjacent_fallback_enabled = _sanitize_bool_setting(payload.get("adjacent_fallback_enabled"), True)
    element_sequence_enabled = _sanitize_bool_setting(payload.get("element_sequence_enabled"), False)
    return _workload_target_field_preferences_payload(
        preferences,
        adjacent_fallback_enabled=adjacent_fallback_enabled,
        element_sequence_enabled=element_sequence_enabled,
    )


@app.post("/api/workload-capture/run")
async def run_workload_capture(
    workload_file: UploadFile = File(...),
    target_file: UploadFile = File(...),
    selected_fields: str | None = Form(default=None),
    source_sheet_configs: str | None = Form(default=None),
    target_sheet_configs: str | None = Form(default=None),
    only_capture_rows_with_value: bool = Form(default=True),
    value_filter_field: str = Form(default=SOURCE_QUANTITY_FIELD),
) -> dict[str, object]:
    if not workload_file.filename or not workload_file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 格式的工作量表格")
    if not target_file.filename or not target_file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 格式的控制价计算表")

    selected = _parse_workload_selected_fields(selected_fields)
    source_configs = _parse_workload_sheet_configs(source_sheet_configs, "source")
    target_configs = _parse_workload_sheet_configs(target_sheet_configs, "target")
    filter_field = _parse_workload_filter_field(value_filter_field) if only_capture_rows_with_value else None
    if not source_configs or not target_configs:
        raise HTTPException(status_code=400, detail="请先完成工作量表和控制价计算表的 sheet/列映射")

    job_id = uuid4().hex
    job_dir = RUNTIME_DIR / "workload-capture" / job_id
    job_dir.mkdir(parents=True, exist_ok=True)
    workload_path = job_dir / workload_file.filename
    target_path = job_dir / target_file.filename
    workload_path.write_bytes(await workload_file.read())
    target_path.write_bytes(await target_file.read())
    output_timestamp = _output_timestamp()
    output_workload = job_dir / f"{TEMP_FILE_PREFIX}-原表-(工作量信息抓取后标注符合用)-{output_timestamp}.xlsx"
    output_target = job_dir / f"{TEMP_FILE_PREFIX}-控制价计算表（填好数量后）-{output_timestamp}.xlsx"

    try:
        summary = capture_workload(
            workload_path,
            target_path,
            output_workload,
            output_target,
            source_configs,
            target_configs,
            selected_fields=selected,
            filter_non_empty_field=filter_field,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    return {
        "job_id": job_id,
        "summary": summary,
        "downloads": {
            "workload": f"/api/workload-capture/download/{job_id}/workload",
            "target": f"/api/workload-capture/download/{job_id}/target",
        },
    }


@app.post("/api/workload-capture/apply-to-current")
async def apply_workload_capture_to_current(
    workload_file: UploadFile = File(...),
    job_id: str = Form(...),
    selected_fields: str | None = Form(default=None),
    source_sheet_configs: str | None = Form(default=None),
    target_sheet_configs: str | None = Form(default=None),
    only_capture_rows_with_value: bool = Form(default=True),
    value_filter_field: str = Form(default=SOURCE_QUANTITY_FIELD),
    write_mode: str = Form(default="conservative"),
) -> dict[str, object]:
    if not workload_file.filename or not workload_file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 格式的工作量表格")
    clean_job_id = str(job_id or "").strip()
    if not clean_job_id:
        raise HTTPException(status_code=400, detail="缺少当前任务编号")
    job_dir = RUNTIME_DIR / clean_job_id
    if not job_dir.exists():
        raise HTTPException(status_code=404, detail="当前任务不存在，请先完成转换")

    selected = _parse_workload_selected_fields(selected_fields)
    source_configs = _parse_workload_sheet_configs(source_sheet_configs, "source")
    target_configs = _parse_workload_sheet_configs(target_sheet_configs, "target")
    filter_field = _parse_workload_filter_field(value_filter_field) if only_capture_rows_with_value else None
    mode = _parse_workload_write_mode(write_mode)
    if not source_configs or not target_configs:
        raise HTTPException(status_code=400, detail="请先完成工作量表和当前控制价表的 sheet/列映射")

    state = _load_process_state(job_dir)
    runtime_context = _job_skill_runtime_context(job_dir, state)
    _require_skill_capability(runtime_context, "workloadCapture")
    excel_path = _state_path(job_dir, state, "output_excel")
    if not excel_path or not excel_path.exists():
        raise HTTPException(status_code=404, detail="当前预览控制价表不存在，请先完成转换")
    input_path = _state_path(job_dir, state, "input_excel", required=False)
    input_name = str(state.get("input_filename") or (input_path.name if input_path else "input.xlsx"))

    workload_dir = job_dir / "workload-current"
    workload_dir.mkdir(parents=True, exist_ok=True)
    workload_path = workload_dir / workload_file.filename
    workload_path.write_bytes(await workload_file.read())
    marked_workload = workload_dir / f"{TEMP_FILE_PREFIX}-原表-(工作量信息抓取后标注符合用)-{_output_timestamp()}.xlsx"

    try:
        workload_summary = capture_workload(
            workload_path,
            excel_path,
            marked_workload,
            excel_path,
            source_configs,
            target_configs,
            selected_fields=selected,
            filter_non_empty_field=filter_field,
            write_mode=mode,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    recalculate_workbook(excel_path)
    summary = _summary_from_dict(state.get("summary", {}))
    summary.table_preview = _refresh_table_preview_from_output(summary.table_preview, excel_path)
    report_path = _state_path(job_dir, state, "output_report", required=False)
    if report_path:
        report_path = write_report(
            report_path,
            input_name,
            summary,
            output_excel_path=excel_path,
            input_excel_path=input_path,
            report_template_path=runtime_context.report_template_path,
        )
        summary.output_report = report_path.name
    else:
        summary.output_report = str(state.get("output_report") or "")
    summary.output_excel = excel_path.name
    _save_process_state(
        job_dir,
        input_name,
        input_path,
        excel_path,
        report_path,
        summary,
        extra={"workload_capture_summary": workload_summary},
    )
    return _attach_job_skill({
        "job_id": clean_job_id,
        "summary": summary.to_dict(),
        "downloads": {
            "excel": f"/api/download/{clean_job_id}/excel",
            "report": f"/api/download/{clean_job_id}/report",
        },
        "workload_summary": workload_summary,
        "workload_downloads": {
            "workload": f"/api/workload-capture/current-download/{clean_job_id}/workload",
        },
    }, job_dir)


@app.get("/api/workload-capture/current-download/{job_id}/{kind}")
def download_current_workload_capture(job_id: str, kind: str) -> FileResponse:
    if kind != "workload":
        raise HTTPException(status_code=400, detail="当前预览写入流程只提供标注工作量表下载")
    job_dir = RUNTIME_DIR / str(job_id).strip()
    if not job_dir.exists():
        raise HTTPException(status_code=404, detail="任务不存在")
    _job_skill_runtime_context(job_dir)
    workload_dir = job_dir / "workload-current"
    if not workload_dir.exists():
        raise HTTPException(status_code=404, detail="标注工作量表不存在")
    matches = sorted(workload_dir.glob(f"{TEMP_FILE_PREFIX}-原表-*.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True)
    if not matches:
        raise HTTPException(status_code=404, detail="标注工作量表不存在")
    path = matches[0]
    return FileResponse(path, filename=path.name, headers=_professional_skill_headers(job_dir))


@app.get("/api/workload-capture/download/{job_id}/{kind}")
def download_workload_capture(job_id: str, kind: str) -> FileResponse:
    job_dir = RUNTIME_DIR / "workload-capture" / job_id
    if not job_dir.exists():
        raise HTTPException(status_code=404, detail="任务不存在")
    if kind == "workload":
        matches = _find_workload_capture_files(job_dir, "workload")
    elif kind == "target":
        matches = _find_workload_capture_files(job_dir, "target")
    else:
        raise HTTPException(status_code=400, detail="下载类型只能是 workload 或 target")
    if not matches:
        raise HTTPException(status_code=404, detail="文件不存在")
    path = matches[0]
    return FileResponse(path, filename=path.name)


@app.get("/api/download/{job_id}/{kind}")
def download(
    job_id: str,
    kind: str,
    hide_empty_rows: bool = False,
    value_filter_field: str | None = None,
) -> FileResponse:
    job_dir = RUNTIME_DIR / job_id
    if not job_dir.exists():
        raise HTTPException(status_code=404, detail="任务不存在")
    _job_skill_runtime_context(job_dir)

    if kind == "excel":
        matches = _find_output_excel_files(job_dir)
    elif kind == "report":
        matches = _find_report_files(job_dir, ".docx")
    else:
        raise HTTPException(status_code=400, detail="下载类型只能是 excel 或 report")

    if not matches:
        raise HTTPException(status_code=404, detail="文件不存在")
    path = matches[0]
    if kind == "excel" and hide_empty_rows:
        filter_field = _parse_warning_filter_field(value_filter_field)
        path = _excel_with_hidden_empty_rows(path, filter_field)
    return FileResponse(path, filename=path.name, headers=_professional_skill_headers(job_dir))


@app.post("/api/risk-report")
async def generate_risk_report(
    job_id: str = Form(...),
    provider: str = Form(default=DEFAULT_PROVIDER),
    model: str = Form(default=DEFAULT_MODEL),
    base_url: str = Form(default=DEFAULT_BASE_URL),
) -> dict[str, object]:
    job_dir = RUNTIME_DIR / job_id
    if not job_dir.exists():
        raise HTTPException(status_code=404, detail="任务不存在")
    state = _load_process_state(job_dir)
    runtime_context = _job_skill_runtime_context(job_dir, state)
    _require_skill_capability(runtime_context, "knowledgeQa")
    _require_skill_capability(runtime_context, "wordReport")

    excel_matches = _find_output_excel_files(job_dir)
    report_matches = _find_report_files(job_dir, ".docx")
    markdown_matches = _find_report_files(job_dir, ".md")
    if not excel_matches or not report_matches or not markdown_matches:
        raise HTTPException(status_code=404, detail="任务文件不完整，请先完成转换")

    markdown_text = markdown_matches[0].read_text(encoding="utf-8")
    knowledge_evidence, knowledge_sources = _build_risk_report_knowledge_evidence(runtime_context, job_dir)
    config = LlmConfig(provider=provider, model=model, base_url=base_url)
    messages = build_risk_prompt(markdown_text, excel_matches[0], knowledge_evidence)
    prompt_path = _write_llm_prompt_markdown("风险报告", config, messages, job_dir)
    try:
        risk_text = _call_chat_completion_tracked(
            config,
            messages,
            source="风险报告",
            prompt_path=prompt_path,
        )
    except (ValueError, RuntimeError) as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    risk_text = ensure_risk_report_evidence(risk_text, markdown_text)
    risk_md_path = job_dir / "大模型风险报告-【codex】.md"
    risk_md_path.write_text(risk_text + "\n", encoding="utf-8")
    append_risk_report(report_matches[0], risk_text)
    task_tracking = _sync_business_task_from_job(job_dir, activity="risk_report")
    return _attach_job_skill({
        "job_id": job_id,
        "risk_report": risk_text,
        "knowledge_sources": knowledge_sources,
        "downloads": {"report": f"/api/download/{job_id}/report"},
        "debug": _build_llm_debug(config, messages, prompt_path),
        "task_tracking": task_tracking,
    }, job_dir)


GENERAL_MODEL_FALLBACK_NOTICE = "已自动转为大模型回答"
GENERAL_MODEL_FALLBACK_BOUNDARY = (
    "提示：本回答未检索到本地知识库明确依据，仅供一般参考；"
    "涉及价格、系数、正式标准或项目口径时，请以项目规则和人工复核为准。"
)


def _build_general_model_fallback_messages(question: str) -> list[dict[str, str]]:
    return [
        {
            "role": "system",
            "content": (
                "你是造价智算的普通大模型兜底助手。本次本地知识库检索没有提供可引用的明确依据，"
                "请基于通用知识直接回答用户问题。不得声称回答来自本地知识库或项目正式资料，"
                "不得把回答当作知识库依据；不得编造价格、系数、条款编号、标准出处或项目内部事实。"
                "如果问题必须依赖项目资料或权威文件，应明确说明无法核验的部分和建议补充的材料。"
                "不要输出固定句“当前知识库未找到明确依据，需要人工复核”。"
                f"{ASSISTANT_TABLE_FORMAT_RULE}"
            ),
        },
        {"role": "user", "content": question},
    ]


def _general_model_fallback_unavailable(reason: str) -> dict[str, object]:
    return {
        "answer": (
            f"{GENERAL_MODEL_FALLBACK_NOTICE}\n\n"
            "大模型回答暂时不可用，请稍后重试。\n\n"
            f"{GENERAL_MODEL_FALLBACK_BOUNDARY}"
        ),
        "answer_mode": "general_model_fallback_unavailable",
        "generated_by_model": False,
        "debug": None,
        "model_degradation_reason": reason,
    }


def _run_general_model_fallback(
    question: str,
    *,
    provider: str,
    model: str,
    base_url: str,
    source: str,
) -> dict[str, object]:
    config = LlmConfig(provider=provider, model=model, base_url=base_url)
    messages = _build_general_model_fallback_messages(question)
    prompt_path = _write_llm_prompt_markdown(source, config, messages)
    try:
        model_answer = _call_chat_completion_tracked(
            config,
            messages,
            source=source,
            prompt_path=prompt_path,
        ).strip()
    except (ValueError, RuntimeError) as exc:
        return _general_model_fallback_unavailable(
            f"general_model_unavailable:{type(exc).__name__}"
        )

    if not model_answer:
        return _general_model_fallback_unavailable("general_model_empty_answer")
    model_answer = model_answer.replace(
        "当前知识库未找到明确依据，需要人工复核。",
        "该问题缺少可核验的项目依据，以下仅作一般性说明。",
    )
    return {
        "answer": f"{GENERAL_MODEL_FALLBACK_NOTICE}\n\n{model_answer}\n\n{GENERAL_MODEL_FALLBACK_BOUNDARY}",
        "answer_mode": "general_model_fallback",
        "generated_by_model": True,
        "debug": _build_llm_debug(config, messages, prompt_path),
        "model_degradation_reason": None,
    }


@app.post("/api/llm-chat")
async def llm_chat(
    message: str = Form(...),
    provider: str = Form(default=DEFAULT_PROVIDER),
    model: str = Form(default=DEFAULT_MODEL),
    base_url: str = Form(default=DEFAULT_BASE_URL),
) -> dict[str, object]:
    clean_message = message.strip()
    if not clean_message:
        raise HTTPException(status_code=400, detail="请输入要发送给大模型的问题")

    config = LlmConfig(provider=provider, model=model, base_url=base_url)
    knowledge_message, force_knowledge = strip_force_knowledge_prefix(clean_message)
    if force_knowledge:
        config = LlmConfig(
            provider=provider,
            model=model,
            base_url=base_url,
            temperature=KNOWLEDGE_QA_TEMPERATURE,
        )
        if not knowledge_message:
            raise HTTPException(status_code=400, detail="请输入查库问题")
        results = search_knowledge(knowledge_message, limit=8)
        if not results:
            fallback = _run_general_model_fallback(
                knowledge_message,
                provider=provider,
                model=model,
                base_url=base_url,
                source="强制知识库问答-大模型兜底",
            )
            return {
                "provider": provider,
                "model": model,
                "forced_knowledge": True,
                "evidence_found": False,
                "sources": [],
                **fallback,
            }
        messages = build_knowledge_answer_prompt(knowledge_message, results)
        prompt_path = _write_llm_prompt_markdown("强制知识库问答", config, messages)
        try:
            answer = _call_chat_completion_tracked(
                config,
                messages,
                source="强制知识库问答",
                prompt_path=prompt_path,
            )
        except (ValueError, RuntimeError) as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        return {
            "provider": provider,
            "model": model,
            "answer": ensure_knowledge_answer(answer, knowledge_message, results),
            "forced_knowledge": True,
            "evidence_found": True,
            "sources": [result.__dict__ for result in results],
            "debug": _build_llm_debug(config, messages, prompt_path),
        }

    messages = [
        {
            "role": "system",
            "content": (
                "你是造价智算本地原型的大模型测试助手，回答应简洁、准确，避免编造未提供的事实。"
                f"{ASSISTANT_TABLE_FORMAT_RULE}"
            ),
        },
        {"role": "user", "content": clean_message},
    ]
    prompt_path = _write_llm_prompt_markdown("问答测试", config, messages)
    try:
        answer = _call_chat_completion_tracked(
            config,
            messages,
            source="问答测试",
            prompt_path=prompt_path,
        )
    except (ValueError, RuntimeError) as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    return {
        "provider": provider,
        "model": model,
        "answer": answer,
        "debug": _build_llm_debug(config, messages, prompt_path),
    }


def _knowledge_library_selection(
    payload: dict[str, Any],
    runtime_context: SkillRuntimeContext,
    job_dir: Path | None,
) -> KnowledgeLibrarySelection:
    base_kwargs = _skill_knowledge_search_kwargs(runtime_context, job_dir)
    try:
        requested_ids = parse_requested_library_ids(payload.get("library_ids"))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    has_row_context = isinstance(payload.get("row_context"), dict)
    if str(payload.get("context_type") or "").strip().lower() == "row" or has_row_context:
        requested_ids = (PROFESSIONAL_KNOWLEDGE_LIBRARY_ID,)
    project_root = Path(base_kwargs.get("project_root") or PROJECT_ROOT)
    sources_value = base_kwargs.get("sources")
    base_sources = (
        tuple(Path(source) for source in sources_value)
        if isinstance(sources_value, (list, tuple))
        else None
    )
    index_value = base_kwargs.get("index_path")
    base_index_path = Path(index_value) if index_value else None
    return resolve_knowledge_library_selection(
        requested_ids,
        project_root=project_root,
        base_sources=base_sources,
        base_index_path=base_index_path,
    )


def _search_selected_knowledge(
    question: str,
    row_context: dict[str, Any] | None,
    limit: int,
    selection: KnowledgeLibrarySelection,
):
    if not selection.sources:
        return []
    return search_knowledge(
        question,
        row_context=row_context,
        limit=limit,
        project_root=selection.project_root,
        index_path=selection.index_path,
        sources=list(selection.sources),
    )


def _parse_retrieval_mode(payload: dict[str, Any]) -> str:
    value = str(payload.get("retrieval_mode") or "classic").strip().lower()
    if value not in {"classic", "hybrid"}:
        raise HTTPException(status_code=400, detail="retrieval_mode 只支持 classic 或 hybrid")
    return value


def _classic_retrieval_trace(
    *,
    requested_mode: str,
    results: list[Any],
    fallback_reason: str | None = None,
) -> dict[str, Any]:
    return {
        "requested_mode": requested_mode,
        "retrieval_mode_used": "classic",
        "fallback_mode": "classic" if requested_mode == "hybrid" else None,
        "fallback_reason": fallback_reason,
        "degraded": bool(fallback_reason),
        "degradation_reasons": [fallback_reason] if fallback_reason else [],
        "channels": {
            "classic": {
                "available": True,
                "hits": len(results),
                "algorithm": "legacy weighted lexical retrieval",
            }
        },
        "fusion": None,
        "rerank": None,
        "hard_gate": None,
        "evidence_status": "sufficient" if results else "insufficient",
        "evidence_sufficient": bool(results),
    }


def _search_selected_knowledge_with_mode(
    question: str,
    row_context: dict[str, Any] | None,
    limit: int,
    selection: KnowledgeLibrarySelection,
    requested_mode: str,
) -> tuple[list[Any], dict[str, Any]]:
    if requested_mode == "classic":
        results = _search_selected_knowledge(question, row_context, limit, selection)
        return results, _classic_retrieval_trace(requested_mode=requested_mode, results=results)
    config = load_knowledge_retrieval_config()
    if not config.get("hybridEnabled", True):
        results = _search_selected_knowledge(question, row_context, limit, selection)
        return results, _classic_retrieval_trace(
            requested_mode=requested_mode,
            results=results,
            fallback_reason="hybrid_feature_disabled",
        )
    if not selection.sources:
        return [], _classic_retrieval_trace(
            requested_mode=requested_mode,
            results=[],
            fallback_reason="hybrid_no_static_sources",
        )
    try:
        response = hybrid_search_knowledge(
            question,
            row_context=row_context,
            limit=limit,
            project_root=selection.project_root,
            index_path=selection.index_path,
            sources=list(selection.sources),
            source_library_ids=selection.source_library_ids,
        )
    except HybridRetrievalError as exc:
        results = _search_selected_knowledge(question, row_context, limit, selection)
        return results, _classic_retrieval_trace(
            requested_mode=requested_mode,
            results=results,
            fallback_reason=str(exc),
        )
    trace = dict(response.trace)
    trace["requested_mode"] = requested_mode
    return response.results, trace


def _retrieval_response_fields(trace: dict[str, Any]) -> dict[str, object]:
    reasons = trace.get("degradation_reasons") or []
    return {
        "requested_retrieval_mode": trace.get("requested_mode", "classic"),
        "retrieval_mode_used": trace.get("retrieval_mode_used", "classic"),
        "actual_retrieval_mode": trace.get("retrieval_mode_used", "classic"),
        "retrieval_channels": trace.get("channels") or {},
        "evidence_status": trace.get("evidence_status", "insufficient"),
        "degradation_reason": reasons[0] if reasons else trace.get("fallback_reason"),
        "degradation_reasons": reasons,
        "retrieval_trace": trace,
    }


def _knowledge_source_payload(
    result: Any,
    selection: KnowledgeLibrarySelection,
) -> dict[str, object]:
    library = selection.library_for_source(result.source_file)
    return {
        **result.__dict__,
        "library_id": library.id if library else None,
        "library_name": library.name if library else None,
    }


def _selected_library_payload(selection: KnowledgeLibrarySelection) -> list[dict[str, str]]:
    selected = set(selection.selected_ids)
    return [
        {
            "id": library.id,
            "name": library.name,
            "kind": library.kind,
        }
        for library in selection.libraries
        if library.id in selected
    ]


@app.get("/api/knowledge/libraries")
async def knowledge_libraries(
    job_id: str = Query(default=""),
    skill_id: str = Query(default=""),
    skill_version: str = Query(default=""),
) -> dict[str, object]:
    request_payload: dict[str, Any] = {}
    if job_id.strip():
        request_payload["job_id"] = job_id.strip()
    if skill_id.strip():
        request_payload["skill_id"] = skill_id.strip()
    if skill_version.strip():
        request_payload["skill_version"] = skill_version.strip()
    runtime_context, job_dir, skill_snapshot = _knowledge_runtime_from_payload(request_payload)
    base_kwargs = _skill_knowledge_search_kwargs(runtime_context, job_dir)
    project_root = Path(base_kwargs.get("project_root") or PROJECT_ROOT)
    sources_value = base_kwargs.get("sources")
    base_sources = (
        tuple(Path(source) for source in sources_value)
        if isinstance(sources_value, (list, tuple))
        else None
    )
    default_selection = resolve_knowledge_library_selection(
        None,
        project_root=project_root,
        base_sources=base_sources,
    )
    retrieval_capabilities = knowledge_retrieval_capabilities(
        project_root=default_selection.project_root,
        index_path=default_selection.index_path,
        sources=list(default_selection.sources),
    )
    if retrieval_capabilities["index_status"] in {"not_built", "stale"} and retrieval_capabilities["available"]:
        try:
            load_or_build_hybrid_index(
                project_root=default_selection.project_root,
                index_path=default_selection.index_path,
                sources=list(default_selection.sources),
                source_library_ids=default_selection.source_library_ids,
            )
            retrieval_capabilities = knowledge_retrieval_capabilities(
                project_root=default_selection.project_root,
                index_path=default_selection.index_path,
                sources=list(default_selection.sources),
            )
        except HybridRetrievalError as exc:
            retrieval_capabilities = {
                **retrieval_capabilities,
                "index_status": "invalid",
                "index_ready": False,
                "degradation_reason": str(exc),
            }
    return {
        **knowledge_library_catalog(
            project_root=project_root,
            base_sources=base_sources,
        ),
        "retrieval_capabilities": retrieval_capabilities,
        "professional_skill": skill_snapshot,
    }


@app.post("/api/knowledge/search")
async def knowledge_search(payload: dict[str, Any] = Body(...)) -> dict[str, object]:
    question = str(payload.get("question") or "").strip()
    question, prefix_forced = strip_force_knowledge_prefix(question)
    if not question:
        raise HTTPException(status_code=400, detail="请输入要检索的问题")
    requested_retrieval_mode = _parse_retrieval_mode(payload)
    force_knowledge = bool(payload.get("force_knowledge")) or prefix_forced
    limit = _parse_knowledge_limit(payload.get("limit"))
    row_context = _parse_row_context(payload.get("row_context"))
    runtime_context, job_dir, skill_snapshot = _knowledge_runtime_from_payload(payload)
    selection = _knowledge_library_selection(payload, runtime_context, job_dir)
    results, retrieval_trace = _search_selected_knowledge_with_mode(
        question,
        row_context,
        limit,
        selection,
        requested_retrieval_mode,
    )
    project_key = normalize_project_key(payload.get("project_key"))
    if selection.memory_enabled:
        project_memories, memory_available = _safe_search_project_memories(
            question,
            project_key,
            limit=limit,
        )
    else:
        project_memories, memory_available = [], True
    memory_hit_audit = _record_trusted_memory_hits(project_memories, payload)
    if project_memories and retrieval_trace.get("evidence_status") == "insufficient":
        retrieval_trace = {**retrieval_trace, "evidence_status": "memory_only", "evidence_sufficient": True}
    return {
        "query": question,
        "project_key": project_key or None,
        "context_type": str(payload.get("context_type") or "general"),
        "knowledge_question": is_knowledge_question(question),
        "forced_knowledge": force_knowledge,
        "evidence_found": bool(results or project_memories),
        "results": [_knowledge_source_payload(result, selection) for result in results],
        "project_memories": project_memories,
        "memory_available": memory_available,
        "memory_enabled": selection.memory_enabled,
        "memory_hit_audit": memory_hit_audit,
        "selected_library_ids": list(selection.selected_ids),
        "selected_libraries": _selected_library_payload(selection),
        "professional_skill": skill_snapshot,
        **_retrieval_response_fields(retrieval_trace),
    }


@app.post("/api/knowledge/ask")
async def knowledge_ask(payload: dict[str, Any] = Body(...)) -> dict[str, object]:
    question = str(payload.get("question") or "").strip()
    question, prefix_forced = strip_force_knowledge_prefix(question)
    if not question:
        raise HTTPException(status_code=400, detail="请输入要询问的问题")
    requested_retrieval_mode = _parse_retrieval_mode(payload)
    force_knowledge = bool(payload.get("force_knowledge")) or prefix_forced
    row_context = _parse_row_context(payload.get("row_context"))

    demo_answer = get_row_demo_answer(question, row_context) or get_demo_answer(question)
    if demo_answer is not None:
        return {
            "answer": demo_answer["answer"],
            "sources": demo_answer["sources"],
            "project_memories": [],
            "project_key": normalize_project_key(payload.get("project_key")) or None,
            "memory_available": True,
            "memory_enabled": False,
            "evidence_found": True,
            "forced_knowledge": force_knowledge,
            "debug": None,
            "selected_library_ids": [],
            "selected_libraries": [],
            "professional_skill": None,
            "preset_answer": True,
            "answer_mode": "curated_demo",
            "generated_by_model": False,
            "preset_id": demo_answer["id"],
            "chart": demo_answer["chart"],
            "requested_retrieval_mode": requested_retrieval_mode,
            "retrieval_mode_used": "curated_demo",
            "actual_retrieval_mode": "curated_demo",
            "retrieval_channels": {},
            "evidence_status": "curated_demo",
            "degradation_reason": None,
            "degradation_reasons": [],
            "retrieval_trace": {
                "requested_mode": requested_retrieval_mode,
                "retrieval_mode_used": "curated_demo",
                "evidence_status": "curated_demo",
                "bypassed_retrieval": True,
                "bypass_reason": demo_answer.get("bypass_reason", "curated_demo_answer"),
            },
        }

    limit = _parse_knowledge_limit(payload.get("limit"))
    runtime_context, job_dir, skill_snapshot = _knowledge_runtime_from_payload(payload)
    selection = _knowledge_library_selection(payload, runtime_context, job_dir)
    results, retrieval_trace = _search_selected_knowledge_with_mode(
        question,
        row_context,
        limit,
        selection,
        requested_retrieval_mode,
    )
    project_key = normalize_project_key(payload.get("project_key"))
    if selection.memory_enabled:
        project_memories, memory_available = _safe_search_project_memories(
            question,
            project_key,
            limit=limit,
        )
    else:
        project_memories, memory_available = [], True
    memory_hit_audit = _record_trusted_memory_hits(project_memories, payload)
    if project_memories and retrieval_trace.get("evidence_status") == "insufficient":
        retrieval_trace = {**retrieval_trace, "evidence_status": "memory_only", "evidence_sufficient": True}
    has_ranked_candidates = bool(
        row_context
        and isinstance(row_context.get("candidate_recommendations"), list)
        and row_context.get("candidate_recommendations")
    )
    if not results and not project_memories and not has_ranked_candidates:
        fallback = _run_general_model_fallback(
            question,
            provider=str(payload.get("provider") or DEFAULT_PROVIDER),
            model=str(payload.get("model") or DEFAULT_MODEL),
            base_url=str(payload.get("base_url") or DEFAULT_BASE_URL),
            source="知识库问答-大模型兜底",
        )
        return {
            "sources": [],
            "project_memories": [],
            "project_key": project_key or None,
            "memory_available": memory_available,
            "memory_enabled": selection.memory_enabled,
            "memory_hit_audit": memory_hit_audit,
            "evidence_found": False,
            "forced_knowledge": force_knowledge,
            "debug": None,
            "selected_library_ids": list(selection.selected_ids),
            "selected_libraries": _selected_library_payload(selection),
            "professional_skill": skill_snapshot,
            **_retrieval_response_fields(retrieval_trace),
            **fallback,
        }

    provider = str(payload.get("provider") or DEFAULT_PROVIDER)
    model = str(payload.get("model") or DEFAULT_MODEL)
    base_url = str(payload.get("base_url") or DEFAULT_BASE_URL)
    config = LlmConfig(
        provider=provider,
        model=model,
        base_url=base_url,
        temperature=KNOWLEDGE_QA_TEMPERATURE,
    )
    messages = build_knowledge_answer_prompt(
        question,
        results,
        row_context=row_context,
        project_memories=project_memories,
    )
    prompt_path = _write_llm_prompt_markdown("知识库问答", config, messages)
    model_degradation_reason: str | None = None
    try:
        answer = _call_chat_completion_tracked(
            config,
            messages,
            source="知识库问答",
            prompt_path=prompt_path,
        )
    except (ValueError, RuntimeError) as exc:
        if requested_retrieval_mode != "hybrid":
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        model_degradation_reason = f"model_unavailable:{type(exc).__name__}"
        answer = ensure_knowledge_answer("", question, results)
        retrieval_trace = {
            **retrieval_trace,
            "degraded": True,
            "degradation_reasons": [*(retrieval_trace.get("degradation_reasons") or []), model_degradation_reason],
            "evidence_status": retrieval_trace.get("evidence_status") or "sufficient",
        }
    answer = ensure_knowledge_answer(answer, question, results)
    answer = prepend_ranked_candidate_recommendation(answer, row_context)

    return {
        "answer": answer,
        "sources": [_knowledge_source_payload(result, selection) for result in results],
        "project_memories": project_memories,
        "project_key": project_key or None,
        "memory_available": memory_available,
        "memory_enabled": selection.memory_enabled,
        "memory_hit_audit": memory_hit_audit,
        "selected_library_ids": list(selection.selected_ids),
        "selected_libraries": _selected_library_payload(selection),
        "professional_skill": skill_snapshot,
        "evidence_found": True,
        "forced_knowledge": force_knowledge,
        "debug": None if model_degradation_reason else _build_llm_debug(config, messages, prompt_path),
        "answer_mode": "evidence_fallback" if model_degradation_reason else "model_answer",
        **_retrieval_response_fields(retrieval_trace),
    }


@app.post("/api/knowledge-memory/candidates")
async def create_knowledge_memory_candidate(
    payload: dict[str, Any] = Body(...),
) -> dict[str, object]:
    formal_sources: list[dict[str, object]] = []
    if classify_knowledge_type(payload) in {"price_factor", "standard_policy", "project_rule"}:
        try:
            formal_sources = [
                {
                    "id": result.id,
                    "source_file": result.source_file,
                    "source_type": result.source_type,
                    "title_path": result.title_path,
                }
                for result in search_knowledge(str(payload.get("question") or ""), limit=3)
            ]
        except (OSError, ValueError):
            formal_sources = []
    try:
        item = _knowledge_memory_store().create_candidate(payload)
    except Exception as exc:
        raise _knowledge_memory_http_error(exc) from exc
    quality_warnings = list(item.get("quality_warnings") or [])
    if formal_sources and item.get("review_policy") == "manual_review":
        quality_warnings.append("已检索到正式资料候选，请在确认前核对是否与正式依据一致。")
    return {
        "item": item,
        "auto_approved": item["scope_type"] == "general" and item["status"] == "confirmed",
        "duplicate_reused": bool(item.get("duplicate_reused")),
        "quality_warnings": list(dict.fromkeys(quality_warnings)),
        "similar_items": list(item.get("similar_items") or []),
        "conflicts": list(item.get("conflicts") or []),
        "formal_sources": formal_sources,
        "identity_mode": "local_trial",
        "identity_notice": "当前仅提供本地试点操作人、确认角色和审计留痕，不等于企业级身份认证。",
    }


@app.get("/api/knowledge-memory/items")
async def list_knowledge_memory_items(
    project_key: str = Query(...),
    status: str = Query(default=""),
    query: str = Query(default=""),
) -> dict[str, object]:
    statuses = {item.strip() for item in status.split(",") if item.strip()} or None
    try:
        items = _knowledge_memory_store().list_items(
            project_key,
            statuses=statuses,
            query=query,
        )
    except Exception as exc:
        raise _knowledge_memory_http_error(exc) from exc
    return {"project_key": normalize_project_key(project_key), "items": items}


@app.get("/api/knowledge-memory/items/{item_id}")
async def get_knowledge_memory_item(
    item_id: str,
    project_key: str = Query(...),
) -> dict[str, object]:
    try:
        item = _knowledge_memory_store().get_item(item_id, project_key)
    except Exception as exc:
        raise _knowledge_memory_http_error(exc) from exc
    return {"item": item}


@app.patch("/api/knowledge-memory/items/{item_id}")
async def update_knowledge_memory_item(
    item_id: str,
    payload: dict[str, Any] = Body(...),
) -> dict[str, object]:
    try:
        item = _knowledge_memory_store().update_item(
            item_id,
            str(payload.get("project_key") or ""),
            payload,
        )
    except Exception as exc:
        raise _knowledge_memory_http_error(exc) from exc
    return {"item": item}


@app.post("/api/knowledge-memory/items/{item_id}/revise")
async def revise_knowledge_memory_item(
    item_id: str,
    payload: dict[str, Any] = Body(...),
) -> dict[str, object]:
    try:
        item = _knowledge_memory_store().revise_item(
            item_id,
            str(payload.get("project_key") or ""),
            payload,
        )
    except Exception as exc:
        raise _knowledge_memory_http_error(exc) from exc
    return {
        "item": item,
        "auto_approved": item["scope_type"] == "general" and item["status"] == "confirmed",
        "duplicate_reused": bool(item.get("duplicate_reused")),
        "quality_warnings": list(item.get("quality_warnings") or []),
    }


@app.post("/api/knowledge-memory/items/{item_id}/promote-general")
async def promote_knowledge_memory_item_to_general(
    item_id: str,
    payload: dict[str, Any] = Body(...),
) -> dict[str, object]:
    try:
        item = _knowledge_memory_store().promote_to_general(
            item_id,
            str(payload.get("project_key") or ""),
            actor=str(payload.get("actor") or ""),
            reason=str(payload.get("reason") or ""),
        )
    except Exception as exc:
        raise _knowledge_memory_http_error(exc) from exc
    return {
        "item": item,
        "auto_approved": item["scope_type"] == "general" and item["status"] == "confirmed",
        "duplicate_reused": bool(item.get("duplicate_reused")),
        "quality_warnings": list(item.get("quality_warnings") or []),
    }


@app.post("/api/knowledge-memory/items/{item_id}/submit")
async def submit_knowledge_memory_item(
    item_id: str,
    payload: dict[str, Any] = Body(...),
) -> dict[str, object]:
    return _knowledge_memory_transition(item_id, payload, "submit")


@app.post("/api/knowledge-memory/items/{item_id}/confirm")
async def confirm_knowledge_memory_item(
    item_id: str,
    payload: dict[str, Any] = Body(...),
) -> dict[str, object]:
    return _knowledge_memory_transition(item_id, payload, "confirm")


@app.post("/api/knowledge-memory/items/{item_id}/reject")
async def reject_knowledge_memory_item(
    item_id: str,
    payload: dict[str, Any] = Body(...),
) -> dict[str, object]:
    return _knowledge_memory_transition(item_id, payload, "reject")


@app.post("/api/knowledge-memory/items/{item_id}/revoke")
async def revoke_knowledge_memory_item(
    item_id: str,
    payload: dict[str, Any] = Body(...),
) -> dict[str, object]:
    return _knowledge_memory_transition(item_id, payload, "revoke")


@app.post("/api/knowledge-memory/items/{item_id}/mark-stale")
async def mark_knowledge_memory_item_stale(
    item_id: str,
    payload: dict[str, Any] = Body(...),
) -> dict[str, object]:
    return _knowledge_memory_transition(item_id, payload, "mark_stale")


@app.get("/api/knowledge-memory/items/{item_id}/audit")
async def get_knowledge_memory_audit(
    item_id: str,
    project_key: str = Query(...),
) -> dict[str, object]:
    try:
        audit = _knowledge_memory_store().audit(item_id, project_key)
    except Exception as exc:
        raise _knowledge_memory_http_error(exc) from exc
    return {"item_id": item_id, "audit": audit}


@app.get("/api/trusted-experience/events")
async def list_trusted_experience_events(
    project_id: str = Query(default=""),
    task_id: str = Query(default=""),
    limit: int = Query(default=100),
) -> dict[str, object]:
    try:
        items = _trusted_experience_store().list_events(
            project_id=project_id,
            task_id=task_id,
            limit=limit,
        )
    except Exception as exc:
        raise _trusted_experience_http_error(exc) from exc
    return {"items": items, "count": len(items)}


@app.get("/api/trusted-experience/events/{event_id}")
async def get_trusted_experience_event(event_id: str) -> dict[str, object]:
    try:
        return {"event": _trusted_experience_store().get_event(event_id)}
    except Exception as exc:
        raise _trusted_experience_http_error(exc) from exc


@app.post("/api/trusted-experience/events/{event_id}/retry")
async def retry_trusted_experience_event(
    event_id: str,
    payload: dict[str, Any] = Body(default={}),
) -> dict[str, object]:
    try:
        result = _trusted_experience_store().retry_event(
            event_id,
            memory_store=_knowledge_memory_store(),
            actor=str(payload.get("actor") or "事件审计重试"),
        )
    except Exception as exc:
        raise _trusted_experience_http_error(exc) from exc
    return result


@app.get("/api/trusted-experience/capsules/{project_id}")
async def get_trusted_experience_capsule(project_id: str) -> dict[str, object]:
    try:
        capsule = _trusted_experience_store().get_capsule(project_id)
    except Exception as exc:
        raise _trusted_experience_http_error(exc) from exc
    return {
        "capsule": capsule,
        "empty_state": capsule is None,
        "message": "该项目尚未生成记忆胶囊。" if capsule is None else "",
    }


@app.post("/api/trusted-experience/capsules/{project_id}/refresh")
async def refresh_trusted_experience_capsule(project_id: str) -> dict[str, object]:
    try:
        project = _project_ledger().project_detail(project_id)
        capsule = _trusted_experience_store().refresh_capsule(
            project,
            memory_store=_knowledge_memory_store(),
        )
    except (ProjectNotFoundError, ProjectLedgerError) as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except Exception as exc:
        raise _trusted_experience_http_error(exc) from exc
    return {"capsule": capsule, "refreshed": True}


@app.get("/api/trusted-experience/metrics")
async def get_trusted_experience_metrics(
    project_id: str = Query(default=""),
) -> dict[str, object]:
    try:
        return _trusted_experience_store().metrics(project_id=project_id)
    except Exception as exc:
        raise _trusted_experience_http_error(exc) from exc


def _knowledge_memory_store() -> KnowledgeMemoryStore:
    settings = _project_knowledge_memory_defaults()
    return KnowledgeMemoryStore(
        DEFAULT_KNOWLEDGE_MEMORY_DB_PATH,
        auto_approve_types=set(settings["autoApproveTypes"]),
        duplicate_threshold=float(settings["duplicateSimilarityThreshold"]),
    )


def _trusted_experience_store() -> TrustedExperienceStore:
    return TrustedExperienceStore(DEFAULT_KNOWLEDGE_MEMORY_DB_PATH)


def _trusted_experience_http_error(exc: Exception) -> HTTPException:
    if isinstance(exc, TrustedExperienceError):
        return HTTPException(status_code=400, detail=str(exc))
    if isinstance(exc, (OSError, sqlite3.Error)):
        return HTTPException(status_code=503, detail="可信经验审计暂不可用，专业业务结果不受影响")
    if isinstance(exc, KnowledgeMemoryError):
        return _knowledge_memory_http_error(exc)
    return HTTPException(status_code=500, detail="可信经验操作失败")


def _knowledge_memory_transition(
    item_id: str,
    payload: dict[str, Any],
    action: str,
) -> dict[str, object]:
    try:
        item = _knowledge_memory_store().transition(
            item_id,
            str(payload.get("project_key") or ""),
            action,
            actor=str(payload.get("actor") or ""),
            reason=str(payload.get("reason") or ""),
            actor_role=str(payload.get("actor_role") or ""),
        )
    except Exception as exc:
        raise _knowledge_memory_http_error(exc) from exc
    return {"item": item}


def _safe_search_project_memories(
    question: str,
    project_key: str,
    *,
    limit: int,
) -> tuple[list[dict[str, Any]], bool]:
    try:
        memories = search_confirmed_project_memory(
            question,
            project_key,
            limit=limit,
            db_path=DEFAULT_KNOWLEDGE_MEMORY_DB_PATH,
        )
        return _filter_project_memories_for_question(question, memories), True
    except (OSError, sqlite3.Error):
        return [], False


def _record_trusted_memory_hits(
    memories: list[dict[str, Any]],
    payload: dict[str, Any],
) -> dict[str, object]:
    if not memories:
        return {"recorded": 0, "available": True, "warning": ""}
    try:
        records = _trusted_experience_store().record_memory_hits(
            memories,
            target_project_id=str(payload.get("project_id") or "").strip(),
            target_project_key=str(payload.get("project_key") or "").strip(),
        )
        return {"recorded": len(records), "available": True, "warning": ""}
    except (OSError, sqlite3.Error, TrustedExperienceError) as exc:
        return {
            "recorded": 0,
            "available": False,
            "warning": "知识问答已返回，但可信经验命中审计暂不可用，可重试本次查询。",
            "error_type": type(exc).__name__,
        }


def _filter_project_memories_for_question(
    question: str,
    memories: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """避免无关的已确认记忆污染正式规则问答提示词。"""
    if not memories:
        return []
    clean_question = re.sub(r"\s+", "", str(question or ""))
    required_terms: tuple[str, ...] = ()
    if "技术工作费" in clean_question or "技术系数" in clean_question:
        required_terms = ("技术工作费", "技术系数", "工程测量技术工作费")
    elif "实物工作费" in clean_question or "实物系数" in clean_question:
        required_terms = ("实物工作费", "实物系数", "附加调整系数")
    if not required_terms:
        return memories
    filtered: list[dict[str, Any]] = []
    for memory in memories:
        searchable = "".join(
            str(memory.get(field) or "")
            for field in ("title", "question", "conclusion", "conditions", "exceptions")
        )
        if any(term in searchable for term in required_terms):
            filtered.append(memory)
    return filtered


def _knowledge_memory_http_error(exc: Exception) -> HTTPException:
    if isinstance(exc, KnowledgeMemoryNotFound):
        return HTTPException(status_code=404, detail=str(exc))
    if isinstance(exc, KnowledgeMemoryPermissionError):
        return HTTPException(status_code=403, detail=str(exc))
    if isinstance(exc, KnowledgeMemoryConflict):
        return HTTPException(status_code=409, detail=str(exc))
    if isinstance(exc, KnowledgeMemoryError):
        return HTTPException(status_code=400, detail=str(exc))
    if isinstance(exc, (OSError, sqlite3.Error)):
        return HTTPException(status_code=503, detail="项目知识记忆暂不可用，现有专业主流程不受影响")
    return HTTPException(status_code=500, detail="项目知识记忆操作失败")


def _parse_knowledge_limit(value: object) -> int:
    try:
        return max(1, min(int(value), 20))
    except (TypeError, ValueError):
        return 8


def _build_risk_report_knowledge_evidence(
    context: SkillRuntimeContext,
    job_dir: Path,
) -> tuple[str, list[dict[str, object]]]:
    seen_ids: set[str] = set()
    collected: list[tuple[str, dict[str, object]]] = []
    for query in RISK_REPORT_KNOWLEDGE_QUERIES:
        for result in search_knowledge(
            query,
            limit=4,
            **_skill_knowledge_search_kwargs(context, job_dir),
        ):
            if result.id in seen_ids:
                continue
            seen_ids.add(result.id)
            collected.append((query, result.__dict__))
            if len(collected) >= 12:
                break
        if len(collected) >= 12:
            break

    if not collected:
        return "当前知识库未检索到明确风险报告依据，报告中应提示建议人工复核。", []

    blocks: list[str] = []
    sources: list[dict[str, object]] = []
    for index, (query, source) in enumerate(collected, start=1):
        source_with_query = {"query": query, **source}
        sources.append(source_with_query)
        title = source.get("title_path") or "未标注"
        blocks.append(
            "\n".join(
                [
                    f"资料{index}：",
                    f"检索问题：{query}",
                    f"来源文件：{source.get('source_file', '')}",
                    f"标题路径：{title}",
                    f"正文片段：{source.get('snippet', '')}",
                ]
            )
        )
    return "\n\n".join(blocks), sources


def _parse_row_context(value: object) -> dict[str, Any] | None:
    if isinstance(value, dict):
        return value
    return None


def _build_llm_debug(config: LlmConfig, messages: list[dict[str, str]], prompt_path: Path | None = None) -> dict[str, object]:
    return {
        "provider": config.provider,
        "model": config.model,
        "base_url": config.base_url,
        "temperature": config.temperature,
        "max_tokens": config.max_tokens,
        "messages": messages,
        "prompt_markdown": str(prompt_path) if prompt_path else "",
    }


def _call_chat_completion_tracked(
    config: LlmConfig,
    messages: list[dict[str, str]],
    *,
    source: str,
    prompt_path: Path,
) -> str:
    requested_at = datetime.now().astimezone().isoformat(timespec="seconds")
    try:
        answer = call_chat_completion(config, messages)
    except Exception:
        _record_llm_request(
            config,
            source=source,
            prompt_path=prompt_path,
            status="failed",
            requested_at=requested_at,
        )
        raise
    _record_llm_request(
        config,
        source=source,
        prompt_path=prompt_path,
        status="success",
        requested_at=requested_at,
    )
    return answer


def _record_llm_request(
    config: LlmConfig,
    *,
    source: str,
    prompt_path: Path,
    status: str,
    requested_at: str,
) -> None:
    try:
        ledger = _llm_usage_ledger()
        ledger.record_request(
            provider=config.provider,
            model=config.model,
            source=source,
            status=status,
            requested_at=requested_at,
            event_key=ledger.prompt_event_key(prompt_path),
        )
    except (OSError, sqlite3.Error, LlmUsageError):
        # Usage telemetry is an auditable sidecar and must never block model answers.
        return


def _write_llm_prompt_markdown(
    source: str,
    config: LlmConfig,
    messages: list[dict[str, str]],
    directory: Path | None = None,
) -> Path:
    output_dir = directory or (RUNTIME_DIR / "llm-prompts")
    output_dir.mkdir(parents=True, exist_ok=True)
    safe_source = "".join(char for char in source if char.isalnum() or char in "-_一二三四五六七八九十风险报告问答测试行级AI复核")
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S-%f")[:-3]
    path = output_dir / f"{timestamp}-{safe_source or '大模型'}-提示词-【codex】.md"
    lines = [
        "# 大模型提示词调试文件",
        "",
        f"- 来源：{source}",
        f"- Provider：{config.provider}",
        f"- Model：{config.model}",
        f"- Base URL：{config.base_url}",
        f"- Temperature：{config.temperature}",
        f"- Max tokens：{config.max_tokens}",
        "",
        "> 本文件只记录发送给大模型的提示词，不包含 API Key。",
        "",
    ]
    for index, message in enumerate(messages, start=1):
        lines.extend([
            f"## Message {index} - {message.get('role', '')}",
            "",
            str(message.get("content", "")),
            "",
        ])
    path.write_text("\n".join(lines), encoding="utf-8")
    return path


def _find_report_files(job_dir: Path, suffix: str) -> list[Path]:
    current_named = list(job_dir.glob(f"{OUTPUT_FILE_PREFIX}-控制价报告-*{suffix}"))
    if current_named:
        return current_named
    current = list(job_dir.glob(f"*-处理报告-*-【codex】{suffix}"))
    if current:
        return current
    return list(job_dir.glob(f"*-处理报告-【codex】{suffix}"))


def _find_output_excel_files(job_dir: Path) -> list[Path]:
    current = list(job_dir.glob(f"{OUTPUT_FILE_PREFIX}-控制价计算表-*.xlsx"))
    if current:
        return current
    return list(job_dir.glob("*-填价结果-【codex】.xlsx"))


def _excel_with_hidden_empty_rows(path: Path, value_filter_field: str) -> Path:
    output_path = path.with_name(f"{path.stem}-隐藏空行{path.suffix}")
    value_workbook = load_workbook(path, data_only=True)
    workbook = load_workbook(path)
    try:
        for sheet_name in workbook.sheetnames:
            if not _is_core_output_sheet(sheet_name):
                continue
            if sheet_name not in value_workbook.sheetnames:
                continue
            value_sheet = value_workbook[sheet_name]
            sheet = workbook[sheet_name]
            header_values = next(
                value_sheet.iter_rows(min_row=4, max_row=4, values_only=True),
                (),
            )
            filter_column = _warning_filter_column_index(list(header_values), value_filter_field)
            if filter_column is None:
                raise HTTPException(
                    status_code=400,
                    detail=f"下载 Excel 隐藏空行失败：{sheet_name} 未找到指定列：{value_filter_field}",
                )
            for row_index in range(5, sheet.max_row + 1):
                sheet.row_dimensions[row_index].hidden = False
            merged_value_map = FillEngine._build_merged_value_map(value_sheet)
            for row_index in range(5, value_sheet.max_row + 1):
                first_column_value = FillEngine._read_mapped_value(value_sheet, row_index, 1, merged_value_map)
                filter_value = FillEngine._read_mapped_value(value_sheet, row_index, filter_column, merged_value_map)
                if _is_total_label(first_column_value):
                    continue
                if not _has_warning_filter_value(filter_value):
                    sheet.row_dimensions[row_index].hidden = True
        workbook.save(output_path)
    finally:
        value_workbook.close()
        workbook.close()
    return output_path


def _is_core_output_sheet(sheet_name: str) -> bool:
    normalized = str(sheet_name or "").replace(" ", "")
    return any(token in normalized for token in ("表2", "表3", "表4", "表二", "表三", "表四"))


def _is_total_label(value: object) -> bool:
    return str(value or "").strip().replace(" ", "").startswith("合计")


def _find_workload_capture_files(job_dir: Path, kind: str) -> list[Path]:
    if kind == "workload":
        current = list(job_dir.glob(f"{TEMP_FILE_PREFIX}-原表-(工作量信息抓取后标注符合用)-*.xlsx"))
        if current:
            return current
        return list(job_dir.glob("*-工作量抓取标注-【codex】.xlsx"))
    if kind == "target":
        current = list(job_dir.glob(f"{TEMP_FILE_PREFIX}-控制价计算表（填好数量后）-*.xlsx"))
        if current:
            return current
        return list(job_dir.glob("*-已抓取工作量-【codex】.xlsx"))
    return []


def _output_timestamp() -> str:
    return datetime.now().strftime("%Y%m%d-%H%M")


def _warning_not_run_summary(
    settings_path: Path | None = None,
    pool_path: Path | None = None,
) -> dict[str, object]:
    pool_path = pool_path or _resolve_experience_pool_path()
    warning_settings = _load_experience_warning_settings(settings_path)
    return {
        "pool_enabled": pool_path.exists(),
        "executed": False,
        "candidate_rows": 0,
        "checked_rows": 0,
        "no_comparable_rows": 0,
        "warning_rows": 0,
        "high_rows": 0,
        "low_rows": 0,
        "medium_rows": 0,
        "metric_counts": {},
        "match_mode_counts": {},
        "low_risk_threshold_percent": warning_settings["low_risk_warning_ratio"],
        "high_risk_threshold_percent": warning_settings["high_risk_warning_ratio"],
        "summary_text": "经验池预警尚未执行：点击“运行经验池预警分析”后，会与经验池比选并写入预警列。",
    }


def _set_warning_progress(job_id: str, payload: dict[str, Any]) -> None:
    with WARNING_PROGRESS_LOCK:
        current = dict(WARNING_PROGRESS.get(job_id, WARNING_PROGRESS_DEFAULT))
        current.update(payload)
        for key in ("processed_rows", "total_rows", "matched_rows", "warning_rows"):
            current[key] = int(current.get(key) or 0)
        WARNING_PROGRESS[job_id] = current


def _get_warning_progress(job_id: str) -> dict[str, object]:
    with WARNING_PROGRESS_LOCK:
        return dict(WARNING_PROGRESS.get(job_id, WARNING_PROGRESS_DEFAULT))


def _find_demo_sample_path() -> Path | None:
    data_dir = DEFAULT_KB_PATH.parent
    for path in data_dir.glob("*.xlsx"):
        if path.name.startswith("~$") or "答案" in path.name:
            continue
        if all(token in path.name for token in DEMO_SAMPLE_TOKENS):
            return path
    return None


def _process_existing_workbook(
    input_source: Path,
    *,
    demo_mode: bool = False,
    sheet_configs: list[dict[str, object]] | None = None,
) -> dict[str, object]:
    if not input_source.exists():
        raise HTTPException(status_code=404, detail=f"输入文件不存在：{input_source}")
    skill_snapshot = _resolve_professional_skill_snapshot(None, None)
    runtime_context = _skill_runtime_context(skill_snapshot)

    job_id = uuid4().hex
    job_dir = RUNTIME_DIR / job_id
    job_dir.mkdir(parents=True, exist_ok=True)
    input_path = job_dir / input_source.name
    input_path.write_bytes(input_source.read_bytes())
    output_timestamp = _output_timestamp()
    output_excel = job_dir / f"{OUTPUT_FILE_PREFIX}-控制价计算表-{output_timestamp}.xlsx"
    output_report = job_dir / f"{OUTPUT_FILE_PREFIX}-控制价报告-{output_timestamp}.docx"

    try:
        summary = _build_skill_fill_engine(runtime_context).fill_workbook(
            input_path,
            output_excel,
            only_match_rows_with_value=True,
            match_value_filter_field=DEFAULT_WARNING_FILTER_FIELD,
            sheet_configs=sheet_configs,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    recalculate_workbook(output_excel)
    summary.warning_summary = _warning_not_run_summary(
        runtime_context.risk_profile.get("warningSettings"),
        runtime_context.risk_profile.get("experiencePool"),
    )
    summary.warning_details = []
    summary.table_preview = _refresh_table_preview_from_output(summary.table_preview, output_excel)
    output_report = write_report(
        output_report,
        input_source.name,
        summary,
        output_excel_path=output_excel,
        input_excel_path=input_path,
        report_template_path=runtime_context.report_template_path,
    )
    summary.output_report = output_report.name
    _save_process_state(
        job_dir,
        input_source.name,
        input_path,
        output_excel,
        output_report,
        summary,
        extra={"skill_snapshot": skill_snapshot},
    )
    task_tracking = _sync_business_task_from_job(job_dir, activity="process_completed")
    return _attach_job_skill({
        "job_id": job_id,
        "demo_mode": demo_mode,
        "sample_file": input_source.name,
        "summary": summary.to_dict(),
        "downloads": {
            "excel": f"/api/download/{job_id}/excel",
            "report": f"/api/download/{job_id}/report",
        },
        "task_tracking": task_tracking,
    }, job_dir)


def _demo_sample_sheet_configs(sample_path: Path) -> list[dict[str, object]]:
    configs: list[dict[str, object]] = []
    for sheet in _inspect_candidate_sheets(sample_path):
        mapping = {
            str(key): str(value).strip()
            for key, value in dict(sheet.get("suggested_mapping") or {}).items()
            if value is not None
        }
        if not all(mapping.get(field) for field in ("要素1", "单位", "输出-价格列")):
            continue
        configs.append(
            {
                "sheet_name": str(sheet.get("sheet_name") or "").strip(),
                "enabled": True,
                "header_row": int(sheet.get("header_row") or 1),
                "column_mapping": mapping,
                "output_match_report": True,
                "merge_vertical_cells": True,
                "merge_horizontal_cells": True,
                "only_match_rows_with_value": True,
                "match_value_filter_field": DEFAULT_WARNING_FILTER_FIELD,
            }
        )
    if not configs:
        raise HTTPException(status_code=400, detail="演示样例未找到可转换的业务明细 sheet")
    return configs


def _build_pending_match_summary(
    workbook_path: Path,
    *,
    column_mapping: dict[str, str] | None,
    header_row: int,
    sheet_configs: list[dict[str, object]] | None,
    only_match_rows_with_value: bool,
    match_value_filter_field: str,
) -> FillSummary:
    workbook = load_workbook(workbook_path)
    value_workbook = load_workbook(workbook_path, data_only=True)
    try:
        configs = sheet_configs or [
            {
                "sheet_name": workbook.active.title,
                "enabled": True,
                "header_row": header_row,
                "column_mapping": column_mapping,
                "merge_vertical_cells": True,
                "merge_horizontal_cells": True,
                "only_match_rows_with_value": only_match_rows_with_value,
                "match_value_filter_field": match_value_filter_field,
            }
        ]
        preview_header_rows = {
            str(config.get("sheet_name") or workbook.active.title): int(config.get("header_row") or header_row)
            for config in configs
        }
        total_data_rows = 0
        price_column_name = ""
        for config in configs:
            if config.get("enabled") is False:
                continue
            sheet_name = str(config.get("sheet_name") or workbook.active.title)
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"输入表不存在候选 sheet：{sheet_name}")
            sheet = workbook[sheet_name]
            value_sheet = value_workbook[sheet_name]
            current_header_row = int(config.get("header_row") or header_row)
            headers = [cell.value for cell in value_sheet[current_header_row]]
            header_map = {str(name).strip(): idx for idx, name in enumerate(headers, start=1) if name}
            current_mapping = config.get("column_mapping") or column_mapping
            field_map = FillEngine._resolve_field_map(header_map, current_mapping)
            missing = [name for name in FIELD_COLUMNS if name not in field_map]
            if missing:
                raise ValueError(f"输入表缺少必要列：{', '.join(missing)}")
            current_price_column_name, _ = FillEngine._find_price_column(header_map, headers, current_mapping)
            if not price_column_name:
                price_column_name = current_price_column_name
            current_only_match_rows_with_value = bool(config.get("only_match_rows_with_value", only_match_rows_with_value))
            current_match_value_filter_field = str(config.get("match_value_filter_field") or match_value_filter_field)
            merged_value_map = FillEngine._build_merged_value_map(
                value_sheet,
                merge_vertical_cells=bool(config.get("merge_vertical_cells", True)),
                merge_horizontal_cells=bool(config.get("merge_horizontal_cells", True)),
            )
            filter_column_index = (
                FillEngine._find_value_filter_column(headers, current_match_value_filter_field)
                if current_only_match_rows_with_value
                else None
            )
            if current_only_match_rows_with_value and filter_column_index is None:
                raise ValueError(f"{sheet.title} 未找到指定列：{current_match_value_filter_field}")
            for excel_row in range(current_header_row + 1, sheet.max_row + 1):
                values = {
                    name: FillEngine._read_mapped_value(value_sheet, excel_row, field_map[name], merged_value_map)
                    for name in FIELD_COLUMNS
                }
                if FillEngine._is_ignored_row(values.get("要素1")):
                    continue
                if filter_column_index is not None:
                    filter_value = FillEngine._read_mapped_value(value_sheet, excel_row, filter_column_index, merged_value_map)
                    if not FillEngine._has_value_for_matching_filter(filter_value):
                        continue
                total_data_rows += 1
        table_preview = FillEngine._build_multi_sheet_table_preview(
            [
                (workbook[sheet_name], preview_header_rows.get(sheet_name, 1))
                for sheet_name in workbook.sheetnames
            ],
            max_rows=50,
        )
        return FillSummary(
            total_data_rows=total_data_rows,
            price_column=price_column_name,
            filled_rows=0,
            matched_rows=0,
            unchanged_rows=0,
            review_rows=0,
            conflict_rows=0,
            output_excel=workbook_path.name,
            output_report="",
            report_text=f"已读取{total_data_rows}行，等待批量匹配价格和两个系数。",
            table_preview=table_preview,
            matching_status="pending",
            warning_summary=_warning_not_run_summary(),
            warning_details=[],
        )
    finally:
        value_workbook.close()
        workbook.close()


def _save_process_state(
    job_dir: Path,
    input_filename: str,
    input_path: Path | None,
    output_excel: Path,
    output_report: Path | None,
    summary: FillSummary,
    extra: dict[str, object] | None = None,
) -> None:
    state_path = job_dir / PROCESS_STATE_FILENAME
    preserved: dict[str, object] = {}
    if state_path.exists():
        try:
            previous_state = json.loads(state_path.read_text(encoding="utf-8"))
            for key in (
                "deferred_matching",
                "process_options",
                "skill_snapshot",
                "source_task_id",
                "business_task_seed",
                "business_task",
                "business_task_definition",
                "project_relation",
                "project_id",
                "project_name",
                "source_type",
                "created_at",
            ):
                if key in previous_state:
                    preserved[key] = previous_state[key]
        except json.JSONDecodeError:
            preserved = {}
    state = {
        **preserved,
        "input_filename": input_filename,
        "input_excel": input_path.name if input_path else "",
        "output_excel": output_excel.name,
        "output_report": output_report.name if output_report else "",
        "summary": summary.to_dict(),
        "created_at": preserved.get("created_at") or datetime.now().astimezone().isoformat(timespec="seconds"),
        "updated_at": datetime.now().isoformat(timespec="seconds"),
    }
    if extra:
        state.update(extra)
    state_path.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")


def _project_ledger() -> ProjectLedger:
    return ProjectLedger(RUNTIME_DIR / DEFAULT_PROJECT_LEDGER_DB_PATH.name, RUNTIME_DIR)


def _business_task_store() -> BusinessTaskStore:
    return BusinessTaskStore(RUNTIME_DIR / DEFAULT_BUSINESS_TASK_DB_PATH.name)


def _safe_file_snapshot(path: Path | None, *, version: int = 1) -> dict[str, object]:
    if not path or not path.is_file():
        return {"reference": "", "type": "", "version": version, "sha256": ""}
    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    return {
        "reference": path.name,
        "type": path.suffix.lower().lstrip("."),
        "version": max(1, int(version)),
        "sha256": digest,
    }


def _task_skill_snapshot(state: dict[str, object]) -> dict[str, object]:
    snapshot = state.get("skill_snapshot") if isinstance(state.get("skill_snapshot"), dict) else {}
    runtime = snapshot.get("runtime_context") if isinstance(snapshot.get("runtime_context"), dict) else {}
    return {
        **ProfessionalSkillRegistry.public_snapshot(snapshot),
        "runtime_summary": {
            "processor_id": str(runtime.get("processor_id") or ""),
            "capabilities": {
                str(key): bool(value)
                for key, value in dict(runtime.get("capabilities") or {}).items()
            },
            "rule_asset_count": len(dict(runtime.get("rule_assets") or {})),
            "knowledge_source_count": len(list(runtime.get("knowledge_sources") or [])),
        },
    }


def _task_status_from_state(state: dict[str, object], activity: str) -> tuple[str, str]:
    summary = state.get("summary") if isinstance(state.get("summary"), dict) else {}
    if activity == "process_pending":
        return "processing", "structure_recognized"
    if activity == "risk_checked":
        review_rows = int(summary.get("review_rows") or 0)
        return ("pending_review", "risk_checked") if review_rows else ("completed", "risk_checked")
    if activity == "manual_edit":
        return "processing", "experience_governed"
    if activity == "risk_report":
        return "completed", "artifact_generated"
    if activity in {"process_completed", "batch_match", "recalculate"}:
        return "processing", "artifact_generated"
    return "processing", "task_defined"


def _task_artifact_event(
    store: BusinessTaskStore,
    task_id: str,
    job_dir: Path,
    state: dict[str, object],
    *,
    source_module: str,
) -> int:
    artifacts: list[dict[str, object]] = []
    for state_key, artifact_type in (("output_excel", "excel"), ("output_report", "word")):
        name = str(state.get(state_key) or "").strip()
        if not name:
            continue
        path = job_dir / Path(name).name
        if path.is_file():
            artifacts.append({"type": artifact_type, **_safe_file_snapshot(path)})
    if not artifacts:
        return 0
    bundle_hash = hashlib.sha256(
        json.dumps(artifacts, ensure_ascii=False, sort_keys=True).encode("utf-8")
    ).hexdigest()
    store.record_event(
        task_id,
        event_key=f"artifact:{bundle_hash}",
        event_type="artifact_generated",
        status="completed",
        source_module=source_module,
        detail="Excel / Word 成果已按当前专业处理状态生成。",
        reference_type="job_id",
        reference_id=job_dir.name,
        payload={"tool": "report.write_report", "artifacts": artifacts, "bundle_hash": bundle_hash},
    )
    return store.event_count(task_id, "artifact_generated")


def _sync_business_task_from_job(
    job_dir: Path,
    *,
    activity: str,
    activity_payload: dict[str, object] | None = None,
) -> dict[str, object]:
    """Best-effort Task identity/lineage sync; never owns professional state."""
    try:
        state = _load_process_state(job_dir)
        snapshot = _task_skill_snapshot(state)
        skill_id = str(snapshot.get("id") or "").strip()
        skill_version = str(snapshot.get("version") or "").strip()
        if not skill_id:
            raise BusinessTaskError("专业处理状态缺少 Skill 快照")
        project_relation = state.get("project_relation") if isinstance(state.get("project_relation"), dict) else {}
        project_id = str(state.get("project_id") or project_relation.get("project_id") or "").strip()
        source_type = str(state.get("source_type") or project_relation.get("source_type") or "web").strip()
        source_task_id = str(state.get("source_task_id") or "").strip()
        input_path = _state_path(job_dir, state, "input_excel", required=False)
        seed = state.get("business_task_seed") if isinstance(state.get("business_task_seed"), dict) else {}
        input_name = str(state.get("input_filename") or (input_path.name if input_path else "专业处理任务.xlsx"))
        task_name = str(seed.get("task_name") or state.get("project_name") or Path(input_name).stem or "专业处理任务").strip()
        objective = str(seed.get("objective") or f"完成“{task_name}”的结构识别、规则匹配、风险复核和成果输出").strip()
        definition = state.get("business_task_definition") if isinstance(state.get("business_task_definition"), dict) else {
            "expected_artifacts": ["excel", "word"],
            "success_criteria": ["规则处理完成", "不确定项明确待复核", "成果可下载并可追溯"],
            "human_gates": ["多候选、无候选或字段冲突由人工复核"],
            "collaboration_required": False,
        }
        input_snapshot = _safe_file_snapshot(input_path)
        store = _business_task_store()
        existing_task_id = ""
        task_relation = state.get("business_task") if isinstance(state.get("business_task"), dict) else {}
        if task_relation:
            existing_task_id = str(task_relation.get("task_id") or "")
        if not existing_task_id and source_task_id:
            source_scope = f"{source_type}:{skill_id}:{skill_version}"
            existing = store.find_by_link("source_task_id", source_task_id, source_system=source_scope)
            existing_task_id = str(existing.get("task_id") or "") if existing else ""
        if existing_task_id:
            task = store.get_task(existing_task_id)
            frozen_skill = task.get("skill_snapshot") if isinstance(task.get("skill_snapshot"), dict) else {}
            if (
                str(frozen_skill.get("id") or "") != skill_id
                or str(frozen_skill.get("version") or "") != skill_version
                or str(frozen_skill.get("manifest_hash") or "") != str(snapshot.get("manifest_hash") or "")
            ):
                existing_task_id = ""
            else:
                created = False
        if not existing_task_id:
            identity_key = (
                f"source:{source_type}:{source_task_id}:{skill_id}:{skill_version}"
                if source_task_id else f"job:{job_dir.name}:{skill_id}:{skill_version}"
            )
            task, created = store.ensure_task(
                identity_key=identity_key,
                project_id=project_id,
                source_type=source_type,
                source_reference=source_task_id,
                task_name=task_name,
                objective=objective,
                instructions=str(seed.get("instructions") or ""),
                definition=definition,
                skill_snapshot=snapshot,
                input_snapshot=input_snapshot,
                responsibility={},
                classification_status="classified" if project_id or source_task_id or job_dir.name else "pending_classification",
                created_at=str(state.get("created_at") or "") or None,
            )
        task_id = str(task["task_id"])
        store.link(task_id, "job_id", job_dir.name, source_system="professional")
        run_id = str(project_relation.get("run_id") or "").strip()
        if run_id:
            store.link(task_id, "run_id", run_id, source_system="project_ledger")
        if source_task_id:
            store.link(
                task_id,
                "source_task_id",
                source_task_id,
                source_system=f"{source_type}:{skill_id}:{skill_version}",
            )
        if created:
            store.record_event(
                task_id, event_key="defined", event_type="task_defined", status="completed",
                source_module="business_tasks", detail="业务目标与成功标准已冻结。",
                reference_type="job_id", reference_id=job_dir.name,
            )
            store.record_event(
                task_id, event_key=f"skill:{snapshot.get('manifest_hash')}", event_type="skill_frozen", status="completed",
                source_module="professional_skills", detail=f"已锁定 {snapshot.get('display_name') or skill_id} v{skill_version}。",
                payload={"tool": "ProfessionalSkillRegistry.create_snapshot", "manifest_hash": snapshot.get("manifest_hash")},
            )
            store.record_event(
                task_id, event_key=f"input:{input_snapshot.get('sha256')}", event_type="input_received", status="completed",
                source_module="professional_process", detail=f"已接收 {input_snapshot.get('reference') or '输入文件'}。",
                reference_type="job_id", reference_id=job_dir.name, payload={"tool": "UploadFile", "input": input_snapshot},
            )
        summary = state.get("summary") if isinstance(state.get("summary"), dict) else {}
        if activity in {"process_pending", "process_completed", "batch_match"}:
            store.record_event(
                task_id, event_key=f"structure:{job_dir.name}", event_type="structure_recognized", status="completed",
                source_module="fill_engine", detail=f"已识别 {int(summary.get('total_data_rows') or 0)} 条数据行。",
                reference_type="job_id", reference_id=job_dir.name, payload={"tool": "FillEngine.inspect_workbook"},
            )
        if activity in {"process_completed", "batch_match"}:
            store.record_event(
                task_id, event_key=f"rules:{job_dir.name}", event_type="rules_executed", status="completed",
                source_module="fill_engine", detail="结构化价格与系数规则已按冻结 Skill 执行。",
                reference_type="job_id", reference_id=job_dir.name, payload={"tool": "FillEngine.fill_workbook"},
            )
        if activity == "risk_checked":
            warning = summary.get("warning_summary") if isinstance(summary.get("warning_summary"), dict) else {}
            store.record_event(
                task_id, event_key=f"risk:{job_dir.name}", event_type="risk_checked", status="completed",
                source_module="experience_warning", detail=f"风险检查已完成，预警 {int(warning.get('warning_rows') or 0)} 行。",
                reference_type="job_id", reference_id=job_dir.name,
                payload={"tool": "analyze_workbook_warnings_with_progress", "summary": warning},
            )
        if activity == "manual_edit":
            experience = dict((activity_payload or {}).get("trusted_experience") or {})
            experience_status = str(experience.get("status") or "")
            event_status = "failed" if experience_status in {"failed", "capture_failed"} else (
                "no_candidate" if experience_status in {"no_change", "empty_opinion", ""} else "completed"
            )
            store.record_event(
                task_id, event_key=f"experience:{hashlib.sha256(json.dumps(experience, sort_keys=True, default=str).encode()).hexdigest()}",
                event_type="experience_governed", status=event_status, source_module="trusted_experience",
                detail=str(experience.get("message") or ("已形成可信经验候选。" if event_status == "completed" else "未形成候选。")),
                reference_type="job_id", reference_id=job_dir.name,
                payload={"tool": "TrustedExperienceStore.capture_event", "status": experience_status},
            )
        artifact_version = (
            _task_artifact_event(store, task_id, job_dir, state, source_module=activity)
            if activity in {"process_completed", "batch_match", "risk_checked", "manual_edit", "recalculate", "risk_report"}
            else 0
        )
        if activity == "restore":
            status = str(task.get("status") or "processing")
            stage = str(task.get("stage") or "task_defined")
        else:
            status, stage = _task_status_from_state(state, activity)
        task = store.update_progress(
            task_id,
            status=status,
            stage=stage,
            current_run_id=run_id,
            artifact_version=artifact_version or int(task.get("artifact_version") or 0),
            project_id=project_id or None,
            completed_at=now_iso() if status == "completed" else None,
        )
        state["business_task"] = {
            "task_id": task_id,
            "status": task["status"],
            "stage": task["stage"],
            "artifact_version": task["artifact_version"],
            "review_round": task["review_round"],
        }
        state["business_task_definition"] = definition
        (job_dir / PROCESS_STATE_FILENAME).write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"status": "available", "task": task, "created": created}
    except (OSError, sqlite3.Error, BusinessTaskError, ValueError, json.JSONDecodeError, HTTPException) as exc:
        return {
            "status": "unavailable",
            "message": "专业处理已继续，但任务轨迹暂不可用。",
            "error_type": type(exc).__name__,
        }


def _record_business_task_failure(
    job_dir: Path,
    *,
    event_type: str,
    source_module: str,
    detail: str,
    event_key: str,
) -> None:
    try:
        state = _load_process_state(job_dir)
        relation = state.get("business_task") if isinstance(state.get("business_task"), dict) else {}
        task_id = str(relation.get("task_id") or "")
        if not task_id:
            tracking = _sync_business_task_from_job(job_dir, activity="process_received")
            tracked = tracking.get("task") if isinstance(tracking.get("task"), dict) else {}
            task_id = str(tracked.get("task_id") or "")
        if not task_id:
            return
        store = _business_task_store()
        store.record_event(
            task_id,
            event_key=event_key,
            event_type=event_type,
            status="failed",
            source_module=source_module,
            detail=_clean_task_failure_detail(detail),
            reference_type="job_id",
            reference_id=job_dir.name,
        )
        task = store.update_progress(task_id, status="failed", stage=event_type)
        state["business_task"] = {
            **dict(relation), "task_id": task_id, "status": task["status"],
            "stage": task["stage"], "artifact_version": task["artifact_version"],
            "review_round": task["review_round"],
        }
        (job_dir / PROCESS_STATE_FILENAME).write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")
    except (OSError, sqlite3.Error, BusinessTaskError, ValueError, json.JSONDecodeError, HTTPException):
        return


def _clean_task_failure_detail(value: object) -> str:
    text = str(value or "").replace("\x00", "").strip()
    text = re.sub(r"[A-Za-z]:[\\/][^\s]+", "[本机路径已隐藏]", text)
    return text[:500] or "该阶段执行失败，可在修正输入或配置后重试。"


def _business_task_detail(task_id: str) -> dict[str, object]:
    store = _business_task_store()
    task = store.get_task(task_id)
    timeline = store.timeline(task_id)
    project = None
    artifacts: list[dict[str, object]] = []
    if task["project_id"]:
        try:
            project = _project_ledger().project_detail(str(task["project_id"]))
            artifacts = list(project.get("artifacts") or [])
        except ProjectLedgerError:
            project = None
    experience_events: list[dict[str, object]] = []
    task_ids = [str(task["task_id"])] + [
        str(link["external_id"]) for link in task["links"] if link["link_type"] == "job_id"
    ]
    experience_store = TrustedExperienceStore(RUNTIME_DIR / DEFAULT_KNOWLEDGE_MEMORY_DB_PATH.name)
    for lineage_task_id in task_ids:
        experience_events.extend(experience_store.list_events(task_id=lineage_task_id, limit=100))
    collaboration: list[dict[str, object]] = []
    collaboration_links = [link for link in task["links"] if link["link_type"] == "collaboration_task_id"]
    dispatch_store = external_task_dispatch.ExternalDispatchStore() if collaboration_links else None
    for link in collaboration_links:
        item = dispatch_store.get_task(str(link["external_id"])) if dispatch_store else None
        if item:
            public = external_task_dispatch.public_dispatch_task(item)
            collaboration.append({
                "task_id": public["task_id"], "task_name": public["task_name"],
                "status": public["status"], "status_label": public["status_label"],
                "review_round": public["review_round"], "participants": public["participants"],
                "completed_at": public["completed_at"], "trusted_experience": public["trusted_experience"],
            })
    tools = sorted({
        str(item.get("payload", {}).get("tool") or "")
        for item in timeline["items"] if isinstance(item.get("payload"), dict) and item.get("payload", {}).get("tool")
    })
    return {
        **task,
        "timeline": timeline,
        "lineage": {
            "project": project,
            "artifacts": artifacts,
            "collaboration": collaboration,
            "experience_events": experience_events,
            "tools": tools,
        },
    }


def _sync_business_task_from_dispatch(task: dict[str, object]) -> dict[str, object]:
    try:
        collaboration_task_id = str(task.get("task_id") or "").strip()
        raw = external_task_dispatch.ExternalDispatchStore().get_task(collaboration_task_id)
        if not raw:
            raise BusinessTaskError("协同任务不存在")
        try:
            frozen_snapshot = json.loads(str(raw.get("skill_snapshot_json") or "{}"))
        except json.JSONDecodeError as exc:
            raise BusinessTaskError("协同任务 Skill 快照无效") from exc
        skill_snapshot = {
            **ProfessionalSkillRegistry.public_snapshot(frozen_snapshot),
            "runtime_summary": {
                "processor_id": str((frozen_snapshot.get("runtime_context") or {}).get("processor_id") or ""),
                "capabilities": dict((frozen_snapshot.get("runtime_context") or {}).get("capabilities") or {}),
            },
        }
        source_task_id = str(task.get("source_task_id") or "").strip()
        source_system = str(task.get("source_system") or "external").strip()
        definition = {
            "expected_artifacts": ["excel", "word"],
            "success_criteria": ["专业成果完成", "指定复核人完成审核", "成果按原协同范围交付"],
            "human_gates": ["指定复核人通过后完成"],
            "collaboration_required": True,
            "deadline": str(task.get("deadline") or ""),
        }
        input_snapshot = {
            "reference": str(task.get("file_name") or ""),
            "type": "xlsx",
            "version": str(raw.get("template_version") or "v1.0"),
            "sha256": str(raw.get("template_hash") or ""),
        }
        store = _business_task_store()
        business_task, created = store.ensure_task(
            identity_key=str(raw.get("business_key") or f"dispatch:{source_system}:{source_task_id}"),
            source_type=str(task.get("platform") or source_system),
            source_reference=source_task_id,
            task_name=str(task.get("task_name") or "协同专业任务"),
            objective=str(raw.get("instructions") or task.get("task_name") or "完成协同专业任务"),
            instructions=str(raw.get("instructions") or ""),
            definition=definition,
            skill_snapshot=skill_snapshot,
            input_snapshot=input_snapshot,
            responsibility={
                "participants": task.get("participants") or [],
                "deadline": task.get("deadline") or "",
            },
            classification_status="pending_classification",
            created_at=str(task.get("created_at") or "") or None,
        )
        task_id = str(business_task["task_id"])
        store.link(task_id, "collaboration_task_id", collaboration_task_id, source_system="external_dispatch")
        if source_task_id:
            store.link(
                task_id, "source_task_id", source_task_id,
                source_system=f"{source_system}:{skill_snapshot.get('id')}:{skill_snapshot.get('version')}",
            )
        if created:
            store.record_event(
                task_id, event_key="defined", event_type="task_defined", status="completed",
                source_module="external_task_dispatch", detail="TaskEnvelope 业务目标已冻结。",
                reference_type="collaboration_task_id", reference_id=collaboration_task_id,
            )
            store.record_event(
                task_id, event_key=f"skill:{skill_snapshot.get('manifest_hash')}", event_type="skill_frozen", status="completed",
                source_module="professional_skills", detail=f"已锁定 {skill_snapshot.get('display_name') or skill_snapshot.get('id')} v{skill_snapshot.get('version')}。",
                payload={"tool": "ProfessionalSkillRegistry.create_snapshot", "manifest_hash": skill_snapshot.get("manifest_hash")},
            )
            store.record_event(
                task_id, event_key=f"input:{input_snapshot.get('sha256')}", event_type="input_received", status="completed",
                source_module="external_task_dispatch", detail="协同输入模板已接收并冻结哈希。",
                reference_type="collaboration_task_id", reference_id=collaboration_task_id,
                payload={"tool": "TaskEnvelope", "input": input_snapshot},
            )
        store.record_event(
            task_id, event_key=f"collaboration:{collaboration_task_id}:started", event_type="collaboration_completed",
            status="in_progress", source_module="external_task_dispatch", detail="协同任务已创建，等待领取、编制与复核。",
            reference_type="collaboration_task_id", reference_id=collaboration_task_id,
            payload={"tool": "ExternalTaskDispatchService.create_and_deliver"},
        )
        business_task = store.update_progress(
            task_id, status="processing", stage="input_received",
            review_round=int(task.get("review_round") or 0),
        )
        return {"status": "available", "task": business_task, "created": created}
    except (OSError, sqlite3.Error, BusinessTaskError, ValueError, json.JSONDecodeError) as exc:
        return {
            "status": "unavailable", "message": "协同任务已继续，但任务轨迹暂不可用。",
            "error_type": type(exc).__name__,
        }


def _link_business_task_collaboration(job_dir: Path, task: dict[str, object]) -> dict[str, object]:
    try:
        state = _load_process_state(job_dir)
        relation = state.get("business_task") if isinstance(state.get("business_task"), dict) else {}
        task_id = str(relation.get("task_id") or "")
        if not task_id:
            tracking = _sync_business_task_from_job(job_dir, activity="process_completed")
            tracked_task = tracking.get("task") if isinstance(tracking.get("task"), dict) else {}
            task_id = str(tracked_task.get("task_id") or "")
        if not task_id:
            raise BusinessTaskError("网页专业任务尚未形成业务 Task")
        collaboration_task_id = str(task.get("task_id") or "").strip()
        store = _business_task_store()
        store.link(task_id, "collaboration_task_id", collaboration_task_id, source_system="external_dispatch")
        source_task_id = str(task.get("source_task_id") or "").strip()
        if source_task_id:
            store.link(task_id, "source_task_id", source_task_id, source_system="web_review")
        review_round = int(task.get("review_round") or 0)
        store.record_event(
            task_id, event_key=f"review:{collaboration_task_id}:{review_round}", event_type="human_reviewed",
            status="pending_review", source_module="external_task_dispatch",
            detail=f"第 {review_round or 1} 轮人工复核已发起。",
            reference_type="collaboration_task_id", reference_id=collaboration_task_id,
            payload={"tool": "ExternalTaskDispatchService.create_web_result_review"},
        )
        store.record_event(
            task_id, event_key=f"collaboration:{collaboration_task_id}:started", event_type="collaboration_completed",
            status="in_progress", source_module="external_task_dispatch", detail="协同复核已进入原有投递与审核状态机。",
            reference_type="collaboration_task_id", reference_id=collaboration_task_id,
        )
        business_task = store.update_progress(
            task_id, status="pending_review", stage="human_reviewed", review_round=review_round,
        )
        state["business_task"] = {
            **dict(relation), "task_id": task_id, "status": business_task["status"],
            "stage": business_task["stage"], "artifact_version": business_task["artifact_version"],
            "review_round": business_task["review_round"],
        }
        (job_dir / PROCESS_STATE_FILENAME).write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"status": "available", "task": business_task, "created": False}
    except (OSError, sqlite3.Error, BusinessTaskError, ValueError, json.JSONDecodeError, HTTPException) as exc:
        return {
            "status": "unavailable", "message": "复核已继续，但任务轨迹暂不可用。",
            "error_type": type(exc).__name__,
        }


def _llm_usage_ledger() -> LlmUsageLedger:
    return LlmUsageLedger(RUNTIME_DIR / DEFAULT_LLM_USAGE_DB_PATH.name, RUNTIME_DIR)


def _llm_usage_dashboard(*, date_from: str, date_to: str) -> dict[str, object]:
    try:
        ledger = _llm_usage_ledger()
        database_key = str(ledger.db_path.resolve()).casefold()
        if database_key not in LLM_USAGE_BACKFILLED_DATABASES:
            with LLM_USAGE_BACKFILL_LOCK:
                if database_key not in LLM_USAGE_BACKFILLED_DATABASES:
                    ledger.backfill_prompt_logs()
                    LLM_USAGE_BACKFILLED_DATABASES.add(database_key)
        return ledger.dashboard(date_from=date_from.strip(), date_to=date_to.strip())
    except (OSError, sqlite3.Error, LlmUsageError) as exc:
        return {
            "available": False,
            "scope": "local_instance",
            "total_requests": 0,
            "successful_requests": 0,
            "failed_requests": 0,
            "historical_requests": 0,
            "model_count": 0,
            "trend_granularity": "day",
            "trend": [],
            "models": [],
            "tracked_from": "",
            "message": f"大模型调用统计暂不可用：{exc}",
        }


def _sync_project_ledger(
    job_dir: Path,
    *,
    project_id: str | None,
    project_name: str,
    source_type: str,
    create_project: bool,
) -> dict[str, object]:
    try:
        state = _load_process_state(job_dir)
        relation = _project_ledger().record_process_state(
            job_id=job_dir.name,
            state=state,
            project_id=str(project_id or "").strip() or None,
            project_name=str(project_name or "").strip(),
            source_type=str(source_type or "web").strip(),
            create_project=create_project,
        )
        project_relation = {
            "project_id": relation.get("project_id") or "",
            "project_name": str(project_name or "").strip(),
            "source_type": str(source_type or "web").strip(),
            "run_id": relation["run_id"],
        }
        state["project_relation"] = project_relation
        state["project_id"] = project_relation["project_id"]
        state["project_name"] = project_relation["project_name"]
        state["source_type"] = project_relation["source_type"]
        (job_dir / PROCESS_STATE_FILENAME).write_text(
            json.dumps(state, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        return {"status": "available", **relation}
    except (OSError, ProjectLedgerError, ValueError, json.JSONDecodeError) as exc:
        return {
            "status": "unavailable",
            "message": "专业成果已生成，但项目台账暂时不可用。",
            "error_type": type(exc).__name__,
        }


def _sync_project_ledger_from_job(job_dir: Path) -> dict[str, object]:
    state = _load_process_state(job_dir)
    relation = state.get("project_relation")
    relation_dict = relation if isinstance(relation, dict) else {}
    project_id = str(state.get("project_id") or relation_dict.get("project_id") or "").strip()
    project_name = str(state.get("project_name") or relation_dict.get("project_name") or "").strip()
    source_type = str(state.get("source_type") or relation_dict.get("source_type") or "web").strip()
    return _sync_project_ledger(
        job_dir,
        project_id=project_id or None,
        project_name=project_name,
        source_type=source_type,
        create_project=bool(project_id and project_name),
    )


def _project_dashboard_comparison(
    *,
    compare: bool,
    date_from: str,
    date_to: str,
    skill_id: str,
    status: str,
    source_type: str,
    keyword: str,
    risk: str,
    quality: str,
    lifecycle_stage: str,
) -> dict[str, object]:
    if not compare:
        return {"enabled": False, "available": False}
    try:
        start = datetime.strptime(date_from, "%Y-%m-%d").date()
        end = datetime.strptime(date_to, "%Y-%m-%d").date()
    except ValueError:
        return {
            "enabled": True,
            "available": False,
            "message": "选择明确起止日期后可比较上一周期。",
        }
    if end < start:
        return {
            "enabled": True,
            "available": False,
            "message": "时间范围无效。",
        }
    period_days = (end - start).days + 1
    previous_end = start - timedelta(days=1)
    previous_start = previous_end - timedelta(days=period_days - 1)
    common = {
        "skill_id": skill_id,
        "status": status,
        "source_type": source_type,
        "keyword": keyword,
        "risk": risk,
        "quality": quality,
        "lifecycle_stage": lifecycle_stage,
    }
    ledger = _project_ledger()
    current = ledger.dashboard(
        date_from=start.isoformat(),
        date_to=end.isoformat(),
        **common,
    )["kpis"]
    previous = ledger.dashboard(
        date_from=previous_start.isoformat(),
        date_to=previous_end.isoformat(),
        **common,
    )["kpis"]
    if int(previous["total_projects"]) == 0:
        return {
            "enabled": True,
            "available": False,
            "message": "上一周期没有可比项目，不计算下降。",
            "period": {
                "current": [start.isoformat(), end.isoformat()],
                "previous": [previous_start.isoformat(), previous_end.isoformat()],
            },
        }
    return {
        "enabled": True,
        "available": True,
        "period": {
            "current": [start.isoformat(), end.isoformat()],
            "previous": [previous_start.isoformat(), previous_end.isoformat()],
        },
        "current": {
            "new_projects": int(current["total_projects"]),
            "completed_projects": int(current["completed"]),
        },
        "previous": {
            "new_projects": int(previous["total_projects"]),
            "completed_projects": int(previous["completed"]),
        },
        "delta": {
            "new_projects": int(current["total_projects"]) - int(previous["total_projects"]),
            "completed_projects": int(current["completed"]) - int(previous["completed"]),
        },
    }


def _resolve_professional_skill_snapshot(skill_id: str | None, skill_version: str | None) -> dict[str, object]:
    try:
        return PROFESSIONAL_SKILL_REGISTRY.resolve_for_task(skill_id, skill_version)
    except ProfessionalSkillError as exc:
        raise HTTPException(status_code=exc.status_code, detail=exc.detail()) from exc


def _skill_runtime_context(snapshot: object) -> SkillRuntimeContext:
    try:
        return PROFESSIONAL_SKILL_REGISTRY.runtime_from_snapshot(snapshot)
    except ProfessionalSkillError as exc:
        if exc.code != "skill_runtime_missing":
            raise HTTPException(status_code=exc.status_code, detail=exc.detail()) from exc
    snapshot_dict = snapshot if isinstance(snapshot, dict) else {}
    return SkillRuntimeContext(
        skill_id=str(snapshot_dict.get("id") or "survey-measurement-limit-price"),
        skill_version=str(snapshot_dict.get("version") or "compatibility"),
        manifest_hash=str(snapshot_dict.get("manifest_hash") or "compatibility-default-survey-measurement"),
        input_profile={"extensions": [".xlsx"]},
        processor_id="survey-measurement-v1",
        knowledge_base_path=DEFAULT_KB_PATH,
        rule_assets={
            "technicalRules": (PROJECT_ROOT / "backend/app/rules/technical_fee_rules.xlsx",),
            "physicalRules": (PROJECT_ROOT / "backend/app/rules/physical_factor_rules.xlsx",),
            "physicalOverrides": (PROJECT_ROOT / "backend/app/rules/physical_factor_overrides.xlsx",),
        },
        risk_profile={
            "experiencePool": _resolve_experience_pool_path(),
            "warningSettings": DEFAULT_EXPERIENCE_WARNING_SETTINGS_PATH,
        },
        knowledge_sources=(),
        report_template_path=DEFAULT_REPORT_TEMPLATE_PATH,
        validation_profile={"status": "compatibility"},
        capabilities={
            "pricing": True,
            "workloadCapture": True,
            "experienceWarning": True,
            "knowledgeQa": True,
            "wordReport": True,
        },
    )


def _job_skill_runtime_context(job_dir: Path, state: dict[str, object] | None = None) -> SkillRuntimeContext:
    current_state = state if state is not None else _load_process_state(job_dir)
    return _skill_runtime_context(current_state.get("skill_snapshot"))


def _require_skill_capability(context: SkillRuntimeContext, capability: str) -> None:
    if not context.capabilities.get(capability, False):
        raise HTTPException(status_code=409, detail="当前专业能力不支持该操作")


def _first_rule_asset(context: SkillRuntimeContext, key: str, *, required: bool = True) -> Path | None:
    paths = context.rule_assets.get(key, ())
    if paths:
        return paths[0]
    if required:
        raise HTTPException(status_code=409, detail="任务专业能力缺少必要规则资产")
    return None


def _build_skill_fill_engine(context: SkillRuntimeContext) -> FillEngine:
    _require_skill_capability(context, "pricing")
    if context.processor_id != "survey-measurement-v1":
        raise HTTPException(status_code=503, detail="当前专业能力处理器不可用")
    if not context.knowledge_base_path.exists():
        raise HTTPException(status_code=409, detail="任务专业能力计价库已不可用")
    adjustment_engine = AdjustmentEngine.from_rule_assets(
        physical_rules_path=_first_rule_asset(context, "physicalRules", required=False),
        technical_rules_path=_first_rule_asset(context, "technicalRules"),
    )
    return FillEngine(KnowledgeBase.from_excel(context.knowledge_base_path), adjustment_engine=adjustment_engine)


def _skill_knowledge_search_kwargs(
    context: SkillRuntimeContext,
    job_dir: Path | None,
) -> dict[str, object]:
    if not context.knowledge_sources:
        return {}
    index_path = (
        job_dir / "skill-knowledge-index.json"
        if job_dir
        else RUNTIME_DIR / "knowledge" / f"{context.skill_id}-{context.manifest_hash[:12]}.json"
    )
    return {
        "project_root": PROFESSIONAL_SKILL_REGISTRY.project_root,
        "index_path": index_path,
        "sources": context.knowledge_sources,
    }


def _knowledge_runtime_from_payload(
    payload: dict[str, Any],
) -> tuple[SkillRuntimeContext, Path | None, dict[str, object]]:
    job_id = str(payload.get("job_id") or "").strip()
    if job_id:
        job_dir = RUNTIME_DIR / job_id
        if not job_dir.exists():
            raise HTTPException(status_code=404, detail="任务不存在")
        state = _load_process_state(job_dir)
        context = _job_skill_runtime_context(job_dir, state)
        snapshot = ProfessionalSkillRegistry.public_snapshot(state.get("skill_snapshot"))
    else:
        snapshot_state = _resolve_professional_skill_snapshot(
            str(payload.get("skill_id") or "").strip() or None,
            str(payload.get("skill_version") or "").strip() or None,
        )
        context = _skill_runtime_context(snapshot_state)
        snapshot = ProfessionalSkillRegistry.public_snapshot(snapshot_state)
        job_dir = None
    _require_skill_capability(context, "knowledgeQa")
    return context, job_dir, snapshot


def _attach_job_skill(
    payload: dict[str, object],
    job_dir: Path,
    state: dict[str, object] | None = None,
) -> dict[str, object]:
    current_state = state if state is not None else _load_process_state(job_dir)
    return {
        **payload,
        "professional_skill": ProfessionalSkillRegistry.public_snapshot(current_state.get("skill_snapshot")),
    }


def _professional_skill_headers(job_dir: Path) -> dict[str, str]:
    state = _load_process_state(job_dir)
    snapshot = ProfessionalSkillRegistry.public_snapshot(state.get("skill_snapshot"))
    return {
        "X-Professional-Skill-Id": str(snapshot.get("id") or ""),
        "X-Professional-Skill-Version": str(snapshot.get("version") or ""),
    }


def _load_process_state(job_dir: Path) -> dict[str, object]:
    state_path = job_dir / PROCESS_STATE_FILENAME
    if not state_path.exists():
        return {}
    try:
        return json.loads(state_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=500, detail="任务状态文件损坏，请重新转换") from exc


def _manual_edit_log_path(job_dir: Path) -> Path:
    return job_dir / MANUAL_EDIT_LOG_FILENAME


def _load_manual_edit_log(job_dir: Path) -> list[dict[str, object]]:
    path = _manual_edit_log_path(job_dir)
    if not path.exists():
        return []
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=500, detail="人工修改记录文件损坏，请重新转换") from exc
    if not isinstance(payload, list):
        return []
    return [item for item in payload if isinstance(item, dict)]


def _append_manual_edit_log(job_dir: Path, record: dict[str, object]) -> None:
    records = _load_manual_edit_log(job_dir)
    records.append(record)
    _manual_edit_log_path(job_dir).write_text(
        json.dumps(records, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _capture_manual_edit_experience(
    job_dir: Path,
    *,
    state: dict[str, object],
    runtime_context: SkillRuntimeContext,
    excel_path: Path,
    edit_record: dict[str, object],
) -> dict[str, object]:
    if edit_record.get("original_value") == edit_record.get("new_value"):
        return {
            "status": "no_change",
            "captured": False,
            "retryable": False,
            "message": "原值与新值相同，未生成知识候选。",
        }
    project_relation = state.get("project_relation") if isinstance(state.get("project_relation"), dict) else {}
    project_id = str(state.get("project_id") or project_relation.get("project_id") or "").strip()
    project_name = str(state.get("project_name") or project_relation.get("project_name") or "").strip()
    artifact_version = str(project_relation.get("run_id") or state.get("updated_at") or "").strip()
    event_basis = {
        "task_id": str((state.get("business_task") or {}).get("task_id") or job_dir.name)
        if isinstance(state.get("business_task"), dict)
        else job_dir.name,
        "sheet": edit_record.get("sheet"),
        "row": edit_record.get("row_number"),
        "column": edit_record.get("column_number"),
        "old": edit_record.get("original_value"),
        "new": edit_record.get("new_value"),
    }
    event_key = "cell-edit:" + hashlib.sha256(
        json.dumps(event_basis, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")
    ).hexdigest()
    field_name = str(edit_record.get("header") or "").strip()
    knowledge_type = (
        "price_factor"
        if any(token in field_name for token in ("基价", "单价", "价格", "系数", "费率", "金额"))
        else "project_rule"
    )
    try:
        result = TrustedExperienceStore(
            job_dir.parent / DEFAULT_KNOWLEDGE_MEMORY_DB_PATH.name
        ).capture_event(
            {
                "event_key": event_key,
                "event_type": "cell_edit",
                "project_id": project_id,
                "project_name": project_name,
                "task_id": str((state.get("business_task") or {}).get("task_id") or job_dir.name)
                if isinstance(state.get("business_task"), dict)
                else job_dir.name,
                "skill_id": runtime_context.skill_id,
                "skill_version": runtime_context.skill_version,
                "sheet_name": edit_record.get("sheet"),
                "row_number": edit_record.get("row_number"),
                "field_name": field_name,
                "old_value": edit_record.get("original_value"),
                "new_value": edit_record.get("new_value"),
                "reason": edit_record.get("note"),
                "artifact_version": artifact_version,
                "artifact_hash": hashlib.sha256(excel_path.read_bytes()).hexdigest(),
                "actor": edit_record.get("actor") or "本机试点用户",
                "knowledge_type": knowledge_type,
                "evidence_summary": "人工改单元格成功落盘后自动形成的待审核经验候选。",
            },
            memory_store=KnowledgeMemoryStore(
                job_dir.parent / DEFAULT_KNOWLEDGE_MEMORY_DB_PATH.name,
                auto_approve_types=set(_project_knowledge_memory_defaults()["autoApproveTypes"]),
                duplicate_threshold=float(_project_knowledge_memory_defaults()["duplicateSimilarityThreshold"]),
            ),
        )
        return {
            "status": result["status"],
            "captured": bool(result.get("event")),
            "retryable": False,
            "event_id": (result.get("event") or {}).get("id"),
            "candidate_id": (result.get("candidate") or {}).get("id"),
            "classification_status": (result.get("event") or {}).get("classification_status"),
        }
    except (OSError, sqlite3.Error, TrustedExperienceError, KnowledgeMemoryError) as exc:
        return {
            "status": "capture_failed",
            "captured": False,
            "retryable": True,
            "warning": "改单已成功保存，但可信经验候选捕获失败；可从事件审计重试。",
            "error_type": type(exc).__name__,
        }


def _apply_manual_edit_to_table_preview(
    table_preview: dict[str, object],
    record: dict[str, object],
    header_rows: dict[str, int],
) -> dict[str, object]:
    target_sheet = str(record.get("sheet") or "").strip()
    row_number = int(record.get("row_number") or 0)
    column_number = int(record.get("column_number") or 0)
    next_value = record.get("new_value")

    def apply_to_sheet(preview: dict[str, object]) -> dict[str, object]:
        sheet_name = str(preview.get("sheet_name") or "").strip()
        if sheet_name != target_sheet:
            return preview
        header_row = header_rows.get(sheet_name)
        if header_row is None:
            try:
                header_row = int(preview.get("header_row") or 1)
            except (TypeError, ValueError):
                header_row = 1
        row_numbers = preview.get("row_numbers")
        if isinstance(row_numbers, list):
            row_index = next(
                (
                    index
                    for index, value in enumerate(row_numbers)
                    if _safe_int(value) == row_number
                ),
                -1,
            )
        else:
            row_index = row_number - header_row - 1
        column_index = column_number - 1
        rows = preview.get("rows")
        if not isinstance(rows, list) or row_index < 0 or row_index >= len(rows):
            return preview
        next_rows: list[object] = []
        for index, row in enumerate(rows):
            if index != row_index or not isinstance(row, list):
                next_rows.append(row)
                continue
            next_row = list(row)
            if column_index >= len(next_row):
                next_row.extend([""] * (column_index - len(next_row) + 1))
            next_row[column_index] = next_value
            next_rows.append(next_row)
        return {**dict(preview), "rows": next_rows}

    preview_sheets = table_preview.get("sheets")
    if isinstance(preview_sheets, list) and preview_sheets:
        sheets = [
            apply_to_sheet(sheet)
            for sheet in preview_sheets
            if isinstance(sheet, dict)
        ]
        if not sheets:
            return table_preview
        first = sheets[0]
        return {
            **dict(table_preview),
            "sheet_name": first.get("sheet_name", ""),
            "header_row": first.get("header_row"),
            "headers": first.get("headers", []),
            "rows": first.get("rows", []),
            "row_numbers": first.get("row_numbers", []),
            "sheets": sheets,
        }
    return apply_to_sheet(table_preview)


def _write_preview_cell_edit(
    output_excel: Path,
    *,
    job_id: str,
    sheet_name: str,
    row_number: int,
    column_number: int,
    new_value: object,
    header_rows: dict[str, int],
    edit_source: str = "manual",
    edit_note: str = "",
    edit_actor: str = "本机试点用户",
    candidate_meta: dict[str, object] | None = None,
) -> dict[str, object]:
    workbook = load_workbook(output_excel)
    try:
        if sheet_name not in workbook.sheetnames:
            raise HTTPException(status_code=400, detail=f"输出 Excel 不存在 sheet：{sheet_name}")
        sheet = workbook[sheet_name]
        if row_number > sheet.max_row or column_number > sheet.max_column:
            raise HTTPException(status_code=400, detail="编辑位置超出当前输出 Excel 范围")
        header_row = header_rows.get(sheet_name) or _find_preview_header_row(sheet, [])
        header_text = _preview_edit_header(sheet, header_row, column_number)
        cell = sheet.cell(row=row_number, column=column_number)
        _ensure_preview_cell_editable(sheet, cell, header_row, header_text)

        original_value = cell.value
        saved_value = _coerce_manual_edit_value(new_value, original_value, header_text)
        updated_at = datetime.now().isoformat(timespec="seconds")
        cell.value = saved_value
        cell.fill = MANUAL_EDIT_FILL
        column_letter = get_column_letter(column_number)
        source_text = "辅助填价人工确认" if edit_source == "fill-assist" else "人工修改"
        extra_comment_parts = []
        if candidate_meta:
            extra_comment_parts.append(f"候选来源：{candidate_meta.get('source_label') or candidate_meta.get('source') or ''}")
            extra_comment_parts.append(f"候选依据：{candidate_meta.get('basis') or ''}")
        if edit_note:
            extra_comment_parts.append(f"备注：{edit_note}")
        cell.comment = Comment(
            (
                f"{source_text}\n"
                f"原值：{_manual_edit_value_text(original_value)}\n"
                f"新值：{_manual_edit_value_text(saved_value)}\n"
                f"时间：{updated_at}"
                + ("\n" + "\n".join(part for part in extra_comment_parts if part.strip()) if extra_comment_parts else "")
            ),
            MANUAL_EDIT_COMMENT_AUTHOR,
        )
        workbook.save(output_excel)
        record = {
            "job_id": job_id,
            "sheet": sheet_name,
            "row_number": row_number,
            "column_number": column_number,
            "column_letter": column_letter,
            "header": header_text,
            "original_value": _jsonable_cell_value(original_value),
            "new_value": _jsonable_cell_value(saved_value),
            "updated_at": updated_at,
            "source": edit_source,
            "note": edit_note,
            "actor": edit_actor or "本机试点用户",
        }
        if candidate_meta:
            record["candidate"] = {
                "source": candidate_meta.get("source"),
                "source_label": candidate_meta.get("source_label"),
                "basis": candidate_meta.get("basis"),
                "reason": candidate_meta.get("reason"),
                "confidence": candidate_meta.get("confidence"),
                "confidence_label": candidate_meta.get("confidence_label"),
            }
        return record
    finally:
        workbook.close()


def _preview_edit_header(sheet: object, header_row: int, column_number: int) -> str:
    if 1 <= header_row <= sheet.max_row:
        value = sheet.cell(row=header_row, column=column_number).value
        return str(value or "").strip() or f"列{column_number}"
    return f"列{column_number}"


def _ensure_preview_cell_editable(sheet: object, cell: object, header_row: int, header_text: str) -> None:
    if cell.row <= header_row:
        raise HTTPException(status_code=400, detail="表头和标题行暂不支持人工修改")
    if sheet.row_dimensions[cell.row].hidden:
        raise HTTPException(status_code=400, detail="隐藏行暂不支持人工修改")
    if _is_readonly_preview_header(header_text):
        raise HTTPException(status_code=400, detail=f"{header_text} 属于系统生成列，暂不支持人工修改")
    if sheet.protection.sheet and cell.protection.locked:
        raise HTTPException(status_code=400, detail="受保护单元格暂不支持人工修改")
    if isinstance(cell.value, str) and cell.value.startswith("="):
        raise HTTPException(status_code=400, detail="公式单元格暂不支持人工修改")
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate not in merged_range:
            continue
        if cell.coordinate != merged_range.start_cell.coordinate:
            raise HTTPException(status_code=400, detail="合并单元格非左上角暂不支持人工修改")
        return


def _is_readonly_preview_header(header_text: str) -> bool:
    compact = header_text.replace(" ", "")
    return compact in {item.replace(" ", "") for item in MANUAL_EDIT_READONLY_HEADERS}


def _coerce_manual_edit_value(value: object, original_value: object, header_text: str) -> object:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if text == "":
            return None
        if _should_coerce_manual_edit_to_number(original_value, header_text):
            numeric_text = text.replace(",", "")
            try:
                number = float(numeric_text)
            except ValueError:
                return value
            return int(number) if number.is_integer() else number
        return value
    return value


def _should_coerce_manual_edit_to_number(original_value: object, header_text: str) -> bool:
    if isinstance(original_value, (int, float)) and not isinstance(original_value, bool):
        return True
    compact = header_text.replace(" ", "")
    return any(token in compact for token in MANUAL_EDIT_NUMERIC_HEADER_TOKENS)


def _jsonable_cell_value(value: object) -> object:
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    return str(value)


def _manual_edit_value_text(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def _safe_int(value: object) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _state_path(job_dir: Path, state: dict[str, object], key: str, required: bool = True) -> Path | None:
    name = str(state.get(key) or "").strip()
    if not name:
        if required:
            raise HTTPException(status_code=404, detail="任务状态不完整，请重新转换")
        return None
    return job_dir / name


def _summary_from_dict(raw: dict[str, object]) -> FillSummary:
    if not raw:
        raise HTTPException(status_code=404, detail="任务状态不完整，请重新转换")
    review_details = [
        ReviewRow(
            excel_row=int(row.get("excel_row") or 0),
            status=str(row.get("status") or ""),
            message=str(row.get("message") or ""),
            values=dict(row.get("values") or {}),
        )
        for row in raw.get("review_details", [])
        if isinstance(row, dict)
    ]
    return FillSummary(
        total_data_rows=int(raw.get("total_data_rows") or 0),
        price_column=str(raw.get("price_column") or ""),
        filled_rows=int(raw.get("filled_rows") or 0),
        matched_rows=int(raw.get("matched_rows") or 0),
        unchanged_rows=int(raw.get("unchanged_rows") or 0),
        review_rows=int(raw.get("review_rows") or 0),
        conflict_rows=int(raw.get("conflict_rows") or 0),
        output_excel=str(raw.get("output_excel") or ""),
        output_report=str(raw.get("output_report") or ""),
        report_text=str(raw.get("report_text") or ""),
        table_preview=dict(raw.get("table_preview") or {}),
        review_details=review_details,
        price_logs=list(raw.get("price_logs") or []),
        physical_matched_rows=int(raw.get("physical_matched_rows") or 0),
        physical_experience_rows=int(raw.get("physical_experience_rows") or 0),
        physical_review_rows=int(raw.get("physical_review_rows") or 0),
        technical_matched_rows=int(raw.get("technical_matched_rows") or 0),
        technical_experience_rows=int(raw.get("technical_experience_rows") or 0),
        technical_review_rows=int(raw.get("technical_review_rows") or 0),
        warning_summary=dict(raw.get("warning_summary") or {}),
        warning_details=list(raw.get("warning_details") or []),
        matching_status=str(raw.get("matching_status") or "completed"),
    )


def _refresh_table_preview_from_output(
    table_preview: dict[str, object],
    output_excel: Path,
    header_rows: dict[str, int] | None = None,
) -> dict[str, object]:
    if not output_excel.exists():
        return table_preview
    table_preview = _with_preview_header_rows(table_preview, header_rows or {})
    try:
        resolver = WorkbookFormulaResolver(output_excel)
    except Exception:
        resolver = None
    if resolver is not None:
        try:
            preview_sheets = table_preview.get("sheets")
            sheets = preview_sheets if isinstance(preview_sheets, list) and preview_sheets else [table_preview]
            refreshed = [
                _refresh_one_preview_sheet_with_resolver(resolver, sheet)
                for sheet in sheets
                if isinstance(sheet, dict)
            ]
            if not refreshed:
                return table_preview
            first = refreshed[0]
            return {
                "sheet_name": first["sheet_name"],
                "header_row": first.get("header_row"),
                "headers": first["headers"],
                "rows": first["rows"],
                "row_numbers": first.get("row_numbers", []),
                "sheets": refreshed,
            }
        finally:
            resolver.close()

    workbook = load_workbook(output_excel, read_only=True, data_only=True)
    try:
        preview_sheets = table_preview.get("sheets")
        sheets = preview_sheets if isinstance(preview_sheets, list) and preview_sheets else [table_preview]
        refreshed = [
            _refresh_one_preview_sheet(workbook, sheet)
            for sheet in sheets
            if isinstance(sheet, dict)
        ]
        if not refreshed:
            return table_preview
        first = refreshed[0]
        return {
            "sheet_name": first["sheet_name"],
            "header_row": first.get("header_row"),
            "headers": first["headers"],
            "rows": first["rows"],
            "row_numbers": first.get("row_numbers", []),
            "sheets": refreshed,
        }
    finally:
        workbook.close()


def _parse_preview_header_rows(raw: object) -> dict[str, int]:
    if raw is None:
        return {}
    payload = raw
    if isinstance(raw, str):
        if not raw.strip():
            return {}
        try:
            payload = json.loads(raw)
        except json.JSONDecodeError as exc:
            raise HTTPException(status_code=400, detail="预览表头行设置不是合法 JSON") from exc
    if not isinstance(payload, dict):
        raise HTTPException(status_code=400, detail="预览表头行设置必须是对象")
    header_rows: dict[str, int] = {}
    for key, value in payload.items():
        sheet_name = str(key or "").strip()
        if not sheet_name:
            continue
        try:
            row_number = int(value)
        except (TypeError, ValueError) as exc:
            raise HTTPException(status_code=400, detail=f"{sheet_name} 的预览表头行不是有效数字") from exc
        if row_number < 1:
            raise HTTPException(status_code=400, detail=f"{sheet_name} 的预览表头行必须大于等于 1")
        header_rows[sheet_name] = row_number
    return header_rows


def _with_preview_header_rows(table_preview: dict[str, object], header_rows: dict[str, int]) -> dict[str, object]:
    if not header_rows:
        return table_preview

    def apply_header_row(preview: dict[str, object]) -> dict[str, object]:
        sheet_name = str(preview.get("sheet_name") or "").strip()
        header_row = header_rows.get(sheet_name)
        if header_row is None:
            return preview
        next_preview = dict(preview)
        next_preview["header_row"] = header_row
        return next_preview

    preview_sheets = table_preview.get("sheets")
    if isinstance(preview_sheets, list) and preview_sheets:
        sheets = [
            apply_header_row(sheet)
            for sheet in preview_sheets
            if isinstance(sheet, dict)
        ]
        if not sheets:
            return table_preview
        first = sheets[0]
        return {
            **dict(table_preview),
            "sheet_name": first.get("sheet_name", ""),
            "header_row": first.get("header_row"),
            "headers": first.get("headers", []),
            "rows": first.get("rows", []),
            "row_numbers": first.get("row_numbers", []),
            "sheets": sheets,
        }
    return apply_header_row(table_preview)


def _refresh_one_preview_sheet_with_resolver(
    resolver: WorkbookFormulaResolver,
    preview: dict[str, object],
) -> dict[str, object]:
    sheet_name = str(preview.get("sheet_name") or "")
    if sheet_name not in resolver.sheetnames:
        return preview
    sheet = resolver.value_workbook[sheet_name]
    merged_value_map = FillEngine._build_merged_value_map(sheet)
    header_row = _preview_header_row(preview, sheet, list(preview.get("headers") or []))
    preview_rows = preview.get("rows") or []
    max_rows = len(preview_rows) or 50
    column_count = _preview_column_count(
        list(preview.get("headers") or []),
        preview_rows,
        resolver.sheet_max_column(sheet_name),
    )
    column_count = _extend_preview_column_count_for_warning_columns(sheet, header_row, column_count)
    raw_headers = [
        value if value is not None else ""
        for value in _resolved_preview_row_values(resolver, sheet_name, header_row, column_count, merged_value_map)
    ]
    headers = FillEngine.preview_display_headers(sheet, header_row, raw_headers, column_count)
    rows = [
        [
            value if value is not None else ""
            for value in _resolved_preview_row_values(resolver, sheet_name, row_index, column_count, merged_value_map)
        ]
        for row_index in range(header_row + 1, min(resolver.sheet_max_row(sheet_name), header_row + max_rows) + 1)
    ]
    row_numbers = list(range(header_row + 1, header_row + 1 + len(rows)))
    return {"sheet_name": sheet.title, "header_row": header_row, "headers": headers, "rows": rows, "row_numbers": row_numbers}


def _resolved_preview_row_values(
    resolver: WorkbookFormulaResolver,
    sheet_name: str,
    row_index: int,
    column_count: int,
    merged_value_map: dict[tuple[int, int], Any],
) -> list[Any]:
    return [
        merged_value_map.get((row_index, column_index), resolver.cell_value(sheet_name, row_index, column_index))
        for column_index in range(1, column_count + 1)
    ]


def _refresh_one_preview_sheet(workbook: object, preview: dict[str, object]) -> dict[str, object]:
    sheet_name = str(preview.get("sheet_name") or "")
    if sheet_name not in workbook.sheetnames:
        return preview
    sheet = workbook[sheet_name]
    return _refresh_preview_from_sheet(sheet, preview)


def _refresh_preview_from_sheet(sheet: object, preview: dict[str, object]) -> dict[str, object]:
    headers = list(preview.get("headers") or [])
    header_row = _preview_header_row(preview, sheet, headers)
    preview_rows = preview.get("rows") or []
    max_rows = len(preview_rows) or 50
    column_count = _preview_column_count(headers, preview_rows, sheet.max_column)
    column_count = _extend_preview_column_count_for_warning_columns(sheet, header_row, column_count)
    raw_headers = [
        value if value is not None else ""
        for value in next(
            sheet.iter_rows(
                min_row=header_row,
                max_row=header_row,
                max_col=column_count,
                values_only=True,
            ),
            (),
        )
    ]
    headers = FillEngine.preview_display_headers(sheet, header_row, raw_headers, column_count)
    rows = [
        [value if value is not None else "" for value in row]
        for row in sheet.iter_rows(
            min_row=header_row + 1,
            max_row=min(sheet.max_row, header_row + max_rows),
            max_col=column_count,
            values_only=True,
        )
    ]
    row_numbers = list(range(header_row + 1, header_row + 1 + len(rows)))
    return {"sheet_name": sheet.title, "header_row": header_row, "headers": headers, "rows": rows, "row_numbers": row_numbers}


def _preview_header_row(preview: dict[str, object], sheet: object, headers: list[object]) -> int:
    try:
        header_row = int(preview.get("header_row") or 0)
    except (TypeError, ValueError):
        header_row = 0
    if 1 <= header_row <= sheet.max_row:
        return header_row
    return _find_preview_header_row(sheet, headers)


def _extend_preview_column_count_for_warning_columns(sheet: object, header_row: int, column_count: int) -> int:
    scan_limit = min(sheet.max_column, max(column_count + len(WARNING_OUTPUT_FIELDS) + 8, column_count))
    header_values = next(
        sheet.iter_rows(
            min_row=header_row,
            max_row=header_row,
            max_col=scan_limit,
            values_only=True,
        ),
        (),
    )
    for index, value in enumerate(header_values, start=1):
        if str(value or "").strip() in WARNING_OUTPUT_FIELDS:
            column_count = max(column_count, index)
    return column_count


def _preview_column_count(headers: list[object], rows: object, sheet_max_column: int) -> int:
    widths = [len(headers)]
    if isinstance(rows, list):
        widths.extend(len(row) for row in rows if isinstance(row, list))
    column_count = max(widths) if widths else 0
    if column_count <= 0:
        return min(sheet_max_column, 80)
    return min(column_count, sheet_max_column)


def _find_preview_header_row(sheet: object, headers: list[object]) -> int:
    compact_headers = [str(value or "").strip() for value in headers]
    non_empty_headers = [value for value in compact_headers if value]
    if not non_empty_headers:
        return 1
    for row_index in range(1, min(sheet.max_row, 8) + 1):
        values = [
            str(value or "").strip()
            for value in next(sheet.iter_rows(min_row=row_index, max_row=row_index, values_only=True), ())
        ]
        if values[: len(compact_headers)] == compact_headers:
            return row_index
        if len(non_empty_headers) >= 3 and sum(1 for value in non_empty_headers if value in values) >= 3:
            return row_index
    return 1


def _read_headers(path: Path, header_row: int | None = None, sheet_name: str | None = None) -> tuple[int, list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            workbook.close()
            raise HTTPException(status_code=400, detail=f"输入表不存在 sheet：{sheet_name}")
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active
    try:
        detected_row = header_row or _detect_header_row(sheet)
        row = next(sheet.iter_rows(min_row=detected_row, max_row=detected_row, values_only=True))
    except StopIteration as exc:
        raise HTTPException(status_code=400, detail="输入表没有表头行") from exc
    finally:
        workbook.close()
    return detected_row, [str(value).strip() if value is not None else "" for value in row]


def _inspect_candidate_sheets(
    path: Path,
    preferences: dict[str, list[str]] | None = None,
) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_names = [name for name in workbook.sheetnames if _is_candidate_sheet_name(name)]
        resolved_preferences = preferences if preferences is not None else _load_input_field_preferences()
        return [_inspect_sheet(workbook[name], resolved_preferences) for name in sheet_names]
    finally:
        workbook.close()


def _inspect_sheet(sheet: object, preferences: dict[str, list[str]] | None = None) -> dict[str, object]:
    detected_row = _detect_header_row(sheet)
    row = next(sheet.iter_rows(min_row=detected_row, max_row=detected_row, values_only=True), ())
    headers = [str(value).strip() if value is not None else "" for value in row]
    return {
        "sheet_name": sheet.title,
        "enabled": True,
        "header_row": detected_row,
        "headers": headers,
        "columns": _build_column_options(headers),
        "suggested_mapping": _suggest_column_mapping(headers, preferences),
    }


def _inspect_experience_sheets(path: Path, header_row: int | None = None, sheet_name: str | None = None) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise HTTPException(status_code=400, detail=f"输入表不存在 sheet：{sheet_name}")
            names = [sheet_name]
        else:
            names = list(workbook.sheetnames)
        candidate_names = [name for name in names if _is_candidate_sheet_name(name)]
        default_enabled = set(candidate_names) if candidate_names else {names[0]} if names else set()
        return [
            _inspect_experience_sheet(workbook[name], enabled=name in default_enabled, header_row=header_row)
            for name in names
        ]
    finally:
        workbook.close()


def _inspect_experience_sheet(sheet: object, enabled: bool, header_row: int | None = None) -> dict[str, object]:
    detected_row = header_row or _detect_header_row(sheet)
    row = next(sheet.iter_rows(min_row=detected_row, max_row=detected_row, values_only=True), ())
    headers = [str(value).strip() if value is not None else "" for value in row]
    return {
        "sheet_name": sheet.title,
        "enabled": enabled,
        "header_row": detected_row,
        "headers": headers,
        "columns": _build_column_options(headers),
        "suggested_mapping": _suggest_experience_column_mapping(headers),
    }


def _inspect_workload_sheets(
    path: Path,
    role: str,
    header_row: int | None = None,
    sheet_name: str | None = None,
    preferences: dict[str, list[str]] | None = None,
    adjacent_fallback_enabled: bool | None = None,
    element_sequence_enabled: bool | None = None,
) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise HTTPException(status_code=400, detail=f"输入表不存在 sheet：{sheet_name}")
            names = [sheet_name]
        else:
            names = list(workbook.sheetnames)
        if role == "source":
            default_enabled = {name for name in names if "工作量" in name and "范围" not in name}
            if not default_enabled and names:
                default_enabled = {names[0]}
        else:
            target_candidates = [name for name in names if _is_candidate_sheet_name(name)]
            default_enabled = set(target_candidates) if target_candidates else {names[0]} if names else set()
        return [
            _inspect_workload_sheet(
                workbook[name],
                role=role,
                enabled=name in default_enabled,
                header_row=header_row,
                preferences=preferences,
                adjacent_fallback_enabled=adjacent_fallback_enabled,
                element_sequence_enabled=element_sequence_enabled,
            )
            for name in names
        ]
    finally:
        workbook.close()


def _inspect_workload_sheet(
    sheet: object,
    role: str,
    enabled: bool,
    header_row: int | None = None,
    preferences: dict[str, list[str]] | None = None,
    adjacent_fallback_enabled: bool | None = None,
    element_sequence_enabled: bool | None = None,
) -> dict[str, object]:
    detected_row = header_row or _detect_workload_header_row(sheet, role)
    row = next(
        sheet.iter_rows(
            min_row=detected_row,
            max_row=detected_row,
            max_col=min(sheet.max_column, 300),
            values_only=True,
        ),
        (),
    )
    headers = [str(value).strip() if value is not None else "" for value in row]
    resolved_preferences = preferences if preferences is not None else (
        _load_workload_field_preferences() if role == "source" else _load_workload_target_field_preferences()
    )
    resolved_adjacent_fallback = adjacent_fallback_enabled if adjacent_fallback_enabled is not None else (
        _load_workload_adjacent_fallback_enabled() if role == "source" else _load_workload_target_adjacent_fallback_enabled()
    )
    resolved_element_sequence = element_sequence_enabled if element_sequence_enabled is not None else (
        _load_workload_element_sequence_enabled() if role == "source" else _load_workload_target_element_sequence_enabled()
    )
    return {
        "sheet_name": sheet.title,
        "enabled": enabled,
        "header_row": detected_row,
        "headers": headers,
        "columns": _build_column_options(headers),
        "suggested_mapping": suggest_workload_column_mapping(
            headers,
            role,
            resolved_preferences,
            resolved_adjacent_fallback,
            resolved_element_sequence,
        ),
    }


def _is_candidate_sheet_name(name: str) -> bool:
    return any(token in name for token in ["表2", "表3", "表4"])


def _detect_header_row(sheet: object) -> int:
    max_scan_row = min(sheet.max_row, 4)
    for row_index in range(1, max_scan_row + 1):
        values = next(sheet.iter_rows(min_row=row_index, max_row=row_index, values_only=True))
        if any(str(value).strip() == "要素1" for value in values if value is not None):
            return row_index
    return 1


def _detect_workload_header_row(sheet: object, role: str) -> int:
    if role == "target":
        return _detect_header_row(sheet)
    markers = ["项目", "工作任务", "内容", "类别", "单位", "数量", "工程量合计", "调整系数", "备注"]
    max_scan_row = min(sheet.max_row, 8)
    max_scan_col = min(sheet.max_column, 300)
    best_row = 1
    best_score = -1
    for row_index in range(1, max_scan_row + 1):
        values = [
            str(value or "").replace(" ", "")
            for value in next(
                sheet.iter_rows(
                    min_row=row_index,
                    max_row=row_index,
                    max_col=max_scan_col,
                    values_only=True,
                ),
                (),
            )
        ]
        score = sum(1 for marker in markers if any(marker in value for value in values))
        if score > best_score:
            best_score = score
            best_row = row_index
    return best_row


def _build_column_options(headers: list[str]) -> list[dict[str, str]]:
    columns: list[dict[str, str]] = []
    for index, header in enumerate(headers, start=1):
        letter = get_column_letter(index)
        columns.append(
            {
                "letter": letter,
                "header": header,
                "label": f"{letter}列 - {header}" if header else f"{letter}列",
            }
        )
    return columns


def _load_project_default_settings() -> dict[str, object]:
    if not PROJECT_DEFAULT_SETTINGS_PATH.exists():
        return {}
    try:
        raw = json.loads(PROJECT_DEFAULT_SETTINGS_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    return raw if isinstance(raw, dict) else {}


def _project_default_section(name: str) -> dict[str, object]:
    section = _load_project_default_settings().get(name, {})
    return section if isinstance(section, dict) else {}


def _project_default_bool(section: dict[str, object], key: str, default: bool) -> bool:
    return _sanitize_bool_setting(section.get(key), default)


def _project_default_int(section: dict[str, object], key: str, default: int, min_value: int = 1, max_value: int = 999) -> int:
    try:
        value = int(float(str(section.get(key, default)).strip()))
    except (TypeError, ValueError):
        value = default
    return max(min_value, min(max_value, value))


def _project_input_mapping_defaults() -> dict[str, object]:
    section = _project_default_section("inputMapping")
    match_value_filter_field = _parse_warning_filter_field(
        str(section.get("matchValueFilterField", DEFAULT_WARNING_FILTER_FIELD) or DEFAULT_WARNING_FILTER_FIELD)
    )
    return {
        "headerRow": _project_default_int(section, "headerRow", 4),
        "outputMatchReport": _project_default_bool(section, "outputMatchReport", True),
        "onlyMatchRowsWithValue": _project_default_bool(section, "onlyMatchRowsWithValue", True),
        "matchValueFilterField": match_value_filter_field,
        "mergeVerticalCells": _project_default_bool(section, "mergeVerticalCells", True),
        "mergeHorizontalCells": _project_default_bool(section, "mergeHorizontalCells", True),
        "fieldPreferences": _default_input_field_preferences(),
    }


def _project_workload_capture_defaults() -> dict[str, object]:
    section = _project_default_section("workloadCapture")
    raw_selected_fields = section.get("selectedFields", DEFAULT_SELECTED_WORKLOAD_FIELDS)
    if isinstance(raw_selected_fields, list):
        selected_fields = [
            str(field).strip()
            for field in raw_selected_fields
            if str(field).strip() in DEFAULT_SELECTED_WORKLOAD_FIELDS
        ]
    else:
        selected_fields = []
    write_mode = str(section.get("writeMode", WRITE_MODE_CONSERVATIVE) or WRITE_MODE_CONSERVATIVE)
    if write_mode not in {WRITE_MODE_CONSERVATIVE, WRITE_MODE_OVERWRITE}:
        write_mode = WRITE_MODE_CONSERVATIVE
    value_filter_field = _parse_workload_filter_field(
        str(section.get("valueFilterField", DEFAULT_WORKLOAD_FILTER_FIELD) or DEFAULT_WORKLOAD_FILTER_FIELD)
    )
    return {
        "selectedFields": selected_fields or list(DEFAULT_SELECTED_WORKLOAD_FIELDS),
        "writeMode": write_mode,
        "onlyCaptureRowsWithValue": _project_default_bool(section, "onlyCaptureRowsWithValue", True),
        "valueFilterField": value_filter_field,
        "source": {
            "adjacentFallbackEnabled": _load_workload_adjacent_fallback_enabled(),
            "elementSequenceEnabled": _load_workload_element_sequence_enabled(),
            "fieldPreferences": _default_workload_field_preferences(),
        },
        "target": {
            "adjacentFallbackEnabled": _load_workload_target_adjacent_fallback_enabled(),
            "elementSequenceEnabled": _load_workload_target_element_sequence_enabled(),
            "fieldPreferences": _default_workload_target_field_preferences(),
        },
    }


def _project_zhisuan_window_defaults() -> dict[str, object]:
    section = _project_default_section("zhisuanWindow")
    quick_settings = section.get("quickSettings", {})
    dock_visibility = section.get("dockVisibility", {})
    raw_common_questions = section.get("commonQuestions", [])
    common_questions = []
    if isinstance(raw_common_questions, list):
        for value in raw_common_questions:
            question = str(value).strip()
            if question and question not in common_questions:
                common_questions.append(question)
    return {
        "chatHeight": _project_default_int(section, "chatHeight", 430, 300, 720),
        "dockWidth": _project_default_int(section, "dockWidth", 400, 300, 560),
        "useViewportHeight": _project_default_bool(section, "useViewportHeight", False),
        "showAssistantAvatar": _project_default_bool(section, "showAssistantAvatar", False),
        "quickSettings": quick_settings if isinstance(quick_settings, dict) else {},
        "dockVisibility": dock_visibility if isinstance(dock_visibility, dict) else {},
        "welcomeMessage": str(section.get("welcomeMessage", "") or "").strip(),
        "dockStyle": str(section.get("dockStyle", "") or "").strip(),
        "commonQuestions": common_questions,
    }


def _project_knowledge_memory_defaults() -> dict[str, object]:
    section = _project_default_section("knowledgeMemory")
    raw_auto_types = section.get("autoApproveTypes", list(DEFAULT_AUTO_APPROVE_KNOWLEDGE_TYPES))
    if isinstance(raw_auto_types, list):
        auto_types = [
            str(value).strip()
            for value in raw_auto_types
            if str(value).strip() in KNOWLEDGE_MEMORY_TYPES
        ]
    else:
        auto_types = list(DEFAULT_AUTO_APPROVE_KNOWLEDGE_TYPES)
    try:
        duplicate_threshold = float(section.get("duplicateSimilarityThreshold", 0.92))
    except (TypeError, ValueError):
        duplicate_threshold = 0.92
    return {
        "autoApproveTypes": auto_types,
        "manualReviewTypes": sorted(KNOWLEDGE_MEMORY_TYPES - set(auto_types)),
        "duplicateSimilarityThreshold": max(0.8, min(duplicate_threshold, 1.0)),
    }


def _project_professional_skills_defaults() -> dict[str, object]:
    section = _project_default_section("professionalSkills")
    default_skill_id = str(section.get("defaultSkillId") or "survey-measurement-limit-price").strip()
    return {"defaultSkillId": default_skill_id or "survey-measurement-limit-price"}


def _project_default_settings_payload() -> dict[str, object]:
    return {
        "version": int(_load_project_default_settings().get("version", 1) or 1),
        "file_path": str(PROJECT_DEFAULT_SETTINGS_PATH),
        "previewColumns": _default_preview_column_preferences(),
        "zhisuanWindow": _project_zhisuan_window_defaults(),
        "inputMapping": _project_input_mapping_defaults(),
        "workloadCapture": _project_workload_capture_defaults(),
        "knowledgeMemory": _project_knowledge_memory_defaults(),
        "professionalSkills": _project_professional_skills_defaults(),
        "feishuAppBot": feishu_app_bot.load_bot_defaults(),
    }


def _suggest_column_mapping(headers: list[str], preferences: dict[str, list[str]] | None = None) -> dict[str, str]:
    mapping: dict[str, str] = {}
    defaults = _default_input_field_preferences()
    resolved_preferences = _sanitize_input_field_preferences(preferences or {})
    for name in FIELD_COLUMNS:
        mapping[name] = _find_preferred_input_column(headers, name, resolved_preferences, defaults)
        if not mapping[name] and name in {"要素2", "要素3", "要素4", "要素5"}:
            mapping[name] = EMPTY_ELEMENT_COLUMN

    price_column = _find_preferred_input_column(headers, "输出-价格列", resolved_preferences, defaults)
    mapping["输出-价格列"] = price_column
    mapping["价格列"] = price_column
    mapping[PHYSICAL_ADJUSTMENT_FIELD] = _find_preferred_input_column(
        headers,
        PHYSICAL_ADJUSTMENT_FIELD,
        resolved_preferences,
        defaults,
    )
    mapping[TECHNICAL_ADJUSTMENT_FIELD] = _find_preferred_input_column(
        headers,
        TECHNICAL_ADJUSTMENT_FIELD,
        resolved_preferences,
        defaults,
    )
    return mapping


def _builtin_input_field_preferences() -> dict[str, list[str]]:
    return {
        "要素1": ["要素1", "项目名称", "项目", "专业"],
        "要素2": ["要素2", "工作内容", "作业内容", "内容"],
        "要素3": ["要素3", "类别", "类别名称"],
        "要素4": ["要素4", "比例尺", "规格", "方法"],
        "要素5": ["要素5", "复杂程度", "等级"],
        "单位": ["单位", "计量单位"],
        "输出-价格列": ["单价匹配-测试", "基价测试列", "基价", "单价", "价格"],
        PHYSICAL_ADJUSTMENT_FIELD: ["实物工作费调整系数", "输出-实物工作费调整系数"],
        TECHNICAL_ADJUSTMENT_FIELD: ["技术工作费调整系数", "输出-技术工作费调整系数"],
    }


def _default_input_field_preferences() -> dict[str, list[str]]:
    defaults = _builtin_input_field_preferences()
    section = _project_default_section("inputMapping")
    raw_preferences = section.get("fieldPreferences", {})
    if not isinstance(raw_preferences, dict):
        return defaults
    return {**defaults, **_sanitize_input_field_preferences(raw_preferences)}


def _input_field_preferences_payload(preferences: dict[str, list[str]] | None = None) -> dict[str, object]:
    defaults = _default_input_field_preferences()
    return {
        "fields": INPUT_FIELD_PREFERENCE_FIELDS,
        "defaults": defaults,
        "preferences": preferences if preferences is not None else _load_input_field_preferences(),
        "mapping_defaults": _project_input_mapping_defaults(),
        "file_path": str(PROJECT_DEFAULT_SETTINGS_PATH),
    }


def _default_ui_preferences() -> dict[str, object]:
    return {
        "enabled": False,
        "styles": {},
        "text": {},
    }


def _ui_preferences_payload(preferences: dict[str, object] | None = None) -> dict[str, object]:
    return {
        "defaults": _default_ui_preferences(),
        "preferences": preferences if preferences is not None else _load_ui_preferences(),
        "file_path": str(DEFAULT_UI_PREFERENCES_PATH),
    }


def _builtin_preview_column_preferences() -> dict[str, object]:
    return {
        "defaultLabels": DEFAULT_CORE_PREVIEW_LABELS,
        "sheetOverrides": {},
        "headerRows": {},
        "maxDisplayChars": DEFAULT_PREVIEW_CELL_MAX_DISPLAY_CHARS,
        "columnWidths": {},
    }


def _default_preview_column_preferences() -> dict[str, object]:
    defaults = _builtin_preview_column_preferences()
    section = _project_default_section("previewColumns")
    if not section:
        return defaults
    return _sanitize_preview_column_preferences(section, fallback=defaults)


def _preview_column_preferences_payload(preferences: dict[str, object] | None = None) -> dict[str, object]:
    return {
        "defaults": _default_preview_column_preferences(),
        "preferences": preferences if preferences is not None else _load_preview_column_preferences(),
        "file_path": str(PROJECT_DEFAULT_SETTINGS_PATH),
    }


def _suggest_experience_column_mapping(headers: list[str]) -> dict[str, str]:
    mapping: dict[str, str] = {field: "" for field in EXPERIENCE_MAPPING_FIELDS}
    defaults = _default_experience_field_preferences()
    preferences = _load_experience_field_preferences()
    direct_fields = [field for field in EXPERIENCE_MAPPING_FIELDS if field not in DEFAULT_SELECTED_EXPERIENCE_FIELDS]
    for field in direct_fields:
        mapping[field] = _find_experience_field_column(headers, field, preferences, defaults)
    for metric in DEFAULT_SELECTED_EXPERIENCE_FIELDS:
        mapping[metric] = _find_experience_metric_column(headers, metric, preferences.get(metric, []))
    remark_columns = [
        (index, header)
        for index, header in enumerate(headers, start=1)
        if any(token in header.replace(" ", "") for token in ["备注", "批注", "说明"])
    ]
    for offset, field in enumerate(["原表备注1", "原表备注2", "原表备注3"]):
        if offset < len(remark_columns):
            mapping[field] = get_column_letter(remark_columns[offset][0])
    return mapping


def _default_experience_field_preferences() -> dict[str, list[str]]:
    return {
        "要素1": ["要素1", "项目名称", "项目", "专业"],
        "要素2": ["要素2", "工作内容", "作业内容", "内容"],
        "要素3": ["要素3", "类别", "类别名称"],
        "要素4": ["要素4", "比例尺", "规格", "方法"],
        "要素5": ["要素5", "复杂程度", "等级"],
        "单位": ["单位", "计量单位"],
        PRICE_METRIC: ["【经验数】单价", "【经验数】基价", "经验单价", "经验基价", "基价", "单价", "价格"],
        "工程量": ["工程量", "数量"],
        PHYSICAL_METRIC: ["【经验数】实物工作费调整系数", "经验实物工作费调整系数", "实物工作费调整系数"],
        TECHNICAL_METRIC: ["【经验数】技术工作费调整系数", "经验技术工作费调整系数", "技术工作费调整系数"],
        "其他参数1": ["其他参数1"],
        "其他参数2": ["其他参数2"],
        "原表备注1": ["原表备注1", "备注1", "备注", "批注", "说明"],
        "原表备注2": ["原表备注2", "备注2"],
        "原表备注3": ["原表备注3", "备注3"],
    }


def _experience_field_preferences_payload(preferences: dict[str, list[str]] | None = None) -> dict[str, object]:
    defaults = _default_experience_field_preferences()
    return {
        "fields": EXPERIENCE_MAPPING_FIELDS,
        "defaults": defaults,
        "preferences": preferences if preferences is not None else _load_experience_field_preferences(),
        "file_path": str(DEFAULT_EXPERIENCE_FIELD_PREFERENCES_PATH),
    }


def _default_experience_warning_settings() -> dict[str, float | bool | str]:
    return {
        "low_risk_warning_ratio": DEFAULT_LOW_RISK_WARNING_PERCENT,
        "high_risk_warning_ratio": DEFAULT_HIGH_RISK_WARNING_PERCENT,
        "only_check_rows_with_value": True,
        "value_filter_field": DEFAULT_WARNING_FILTER_FIELD,
    }


def _experience_warning_settings_payload(settings: dict[str, float | bool | str] | None = None) -> dict[str, object]:
    defaults = _default_experience_warning_settings()
    return {
        "defaults": defaults,
        "settings": settings if settings is not None else _load_experience_warning_settings(),
        "filter_fields": list(WARNING_FILTER_FIELDS),
        "file_path": str(DEFAULT_EXPERIENCE_WARNING_SETTINGS_PATH),
    }


def _default_workload_field_preferences() -> dict[str, list[str]]:
    defaults = default_workload_field_preferences()
    source = _project_default_section("workloadCapture").get("source", {})
    raw_preferences = source.get("fieldPreferences", {}) if isinstance(source, dict) else {}
    if not isinstance(raw_preferences, dict):
        return defaults
    return {**defaults, **_sanitize_workload_field_preferences(raw_preferences)}


def _default_workload_target_field_preferences() -> dict[str, list[str]]:
    defaults = default_workload_target_field_preferences()
    target = _project_default_section("workloadCapture").get("target", {})
    raw_preferences = target.get("fieldPreferences", {}) if isinstance(target, dict) else {}
    if not isinstance(raw_preferences, dict):
        return defaults
    return {**defaults, **_sanitize_workload_target_field_preferences(raw_preferences)}


def _workload_field_preferences_payload(
    preferences: dict[str, list[str]] | None = None,
    adjacent_fallback_enabled: bool | None = None,
    element_sequence_enabled: bool | None = None,
) -> dict[str, object]:
    defaults = _default_workload_field_preferences()
    return {
        "fields": WORKLOAD_FIELD_PREFERENCE_FIELDS,
        "defaults": defaults,
        "preferences": preferences if preferences is not None else _load_workload_field_preferences(),
        "adjacent_fallback_enabled": (
            _load_workload_adjacent_fallback_enabled()
            if adjacent_fallback_enabled is None
            else adjacent_fallback_enabled
        ),
        "element_sequence_enabled": (
            _load_workload_element_sequence_enabled()
            if element_sequence_enabled is None
            else element_sequence_enabled
        ),
        "file_path": str(PROJECT_DEFAULT_SETTINGS_PATH),
    }


def _workload_target_field_preferences_payload(
    preferences: dict[str, list[str]] | None = None,
    adjacent_fallback_enabled: bool | None = None,
    element_sequence_enabled: bool | None = None,
) -> dict[str, object]:
    defaults = _default_workload_target_field_preferences()
    return {
        "fields": WORKLOAD_TARGET_FIELD_PREFERENCE_FIELDS,
        "defaults": defaults,
        "preferences": preferences if preferences is not None else _load_workload_target_field_preferences(),
        "adjacent_fallback_enabled": (
            _load_workload_target_adjacent_fallback_enabled()
            if adjacent_fallback_enabled is None
            else adjacent_fallback_enabled
        ),
        "element_sequence_enabled": (
            _load_workload_target_element_sequence_enabled()
            if element_sequence_enabled is None
            else element_sequence_enabled
        ),
        "file_path": str(PROJECT_DEFAULT_SETTINGS_PATH),
    }


def _load_input_field_preferences() -> dict[str, list[str]]:
    return {}


def _load_ui_preferences() -> dict[str, object]:
    defaults = _default_ui_preferences()
    if not DEFAULT_UI_PREFERENCES_PATH.exists():
        return defaults
    try:
        raw = json.loads(DEFAULT_UI_PREFERENCES_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return defaults
    if isinstance(raw, dict) and isinstance(raw.get("preferences"), dict):
        raw = raw["preferences"]
    if not isinstance(raw, dict):
        return defaults
    return _sanitize_ui_preferences(raw)


def _load_preview_column_preferences() -> dict[str, object]:
    return _default_preview_column_preferences()


def _load_experience_field_preferences() -> dict[str, list[str]]:
    defaults = _default_experience_field_preferences()
    if not DEFAULT_EXPERIENCE_FIELD_PREFERENCES_PATH.exists():
        return {}
    try:
        raw = json.loads(DEFAULT_EXPERIENCE_FIELD_PREFERENCES_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    if isinstance(raw, dict) and isinstance(raw.get("preferences"), dict):
        raw = raw["preferences"]
    if not isinstance(raw, dict):
        return {}
    return {
        field: aliases
        for field, aliases in _sanitize_experience_field_preferences(raw).items()
        if aliases != defaults.get(field, [])
    }


def _load_workload_field_preferences() -> dict[str, list[str]]:
    return _default_workload_field_preferences()


def _load_workload_adjacent_fallback_enabled() -> bool:
    section = _project_default_section("workloadCapture").get("source", {})
    if isinstance(section, dict):
        return _sanitize_bool_setting(section.get("adjacentFallbackEnabled"), True)
    return True


def _load_workload_element_sequence_enabled() -> bool:
    section = _project_default_section("workloadCapture").get("source", {})
    if isinstance(section, dict):
        return _sanitize_bool_setting(section.get("elementSequenceEnabled"), True)
    return True


def _load_workload_target_field_preferences() -> dict[str, list[str]]:
    return _default_workload_target_field_preferences()


def _load_workload_target_adjacent_fallback_enabled() -> bool:
    section = _project_default_section("workloadCapture").get("target", {})
    if isinstance(section, dict):
        return _sanitize_bool_setting(section.get("adjacentFallbackEnabled"), True)
    return True


def _load_workload_target_element_sequence_enabled() -> bool:
    section = _project_default_section("workloadCapture").get("target", {})
    if isinstance(section, dict):
        return _sanitize_bool_setting(section.get("elementSequenceEnabled"), False)
    return False


def _load_adjacent_fallback_enabled(path: Path) -> bool:
    if not path.exists():
        return True
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return True
    if not isinstance(raw, dict):
        return True
    return _sanitize_bool_setting(raw.get("adjacent_fallback_enabled"), True)


def _load_element_sequence_enabled(path: Path, default: bool) -> bool:
    if not path.exists():
        return default
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return default
    if not isinstance(raw, dict):
        return default
    return _sanitize_bool_setting(raw.get("element_sequence_enabled"), default)


def _load_experience_warning_settings(
    settings_path: Path | None = None,
) -> dict[str, float | bool | str]:
    defaults = _default_experience_warning_settings()
    path = settings_path or DEFAULT_EXPERIENCE_WARNING_SETTINGS_PATH
    if not path.exists():
        return defaults
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return defaults
    if isinstance(raw, dict) and isinstance(raw.get("settings"), dict):
        raw = raw["settings"]
    if not isinstance(raw, dict):
        return defaults
    return _sanitize_experience_warning_settings(raw)


def _save_input_field_preferences(preferences: dict[str, list[str]]) -> None:
    DEFAULT_INPUT_FIELD_PREFERENCES_PATH.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "version": 1,
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "preferences": preferences,
    }
    DEFAULT_INPUT_FIELD_PREFERENCES_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _save_ui_preferences(preferences: dict[str, object]) -> None:
    DEFAULT_UI_PREFERENCES_PATH.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "version": 1,
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "preferences": preferences,
    }
    DEFAULT_UI_PREFERENCES_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _save_preview_column_preferences(preferences: dict[str, object]) -> None:
    DEFAULT_PREVIEW_COLUMN_PREFERENCES_PATH.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "version": 1,
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "preferences": preferences,
    }
    DEFAULT_PREVIEW_COLUMN_PREFERENCES_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _save_experience_field_preferences(preferences: dict[str, list[str]]) -> None:
    DEFAULT_EXPERIENCE_FIELD_PREFERENCES_PATH.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "version": 1,
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "preferences": preferences,
    }
    DEFAULT_EXPERIENCE_FIELD_PREFERENCES_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _save_workload_field_preferences(
    preferences: dict[str, list[str]],
    adjacent_fallback_enabled: bool = True,
    element_sequence_enabled: bool = True,
) -> None:
    DEFAULT_WORKLOAD_FIELD_PREFERENCES_PATH.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "version": 1,
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "adjacent_fallback_enabled": adjacent_fallback_enabled,
        "element_sequence_enabled": element_sequence_enabled,
        "preferences": preferences,
    }
    DEFAULT_WORKLOAD_FIELD_PREFERENCES_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _save_workload_target_field_preferences(
    preferences: dict[str, list[str]],
    adjacent_fallback_enabled: bool = True,
    element_sequence_enabled: bool = False,
) -> None:
    DEFAULT_WORKLOAD_TARGET_FIELD_PREFERENCES_PATH.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "version": 1,
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "adjacent_fallback_enabled": adjacent_fallback_enabled,
        "element_sequence_enabled": element_sequence_enabled,
        "preferences": preferences,
    }
    DEFAULT_WORKLOAD_TARGET_FIELD_PREFERENCES_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _save_experience_warning_settings(settings: dict[str, float | bool | str]) -> None:
    DEFAULT_EXPERIENCE_WARNING_SETTINGS_PATH.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "version": 1,
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "settings": settings,
    }
    DEFAULT_EXPERIENCE_WARNING_SETTINGS_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _sanitize_ui_preferences(raw_preferences: dict[object, object]) -> dict[str, object]:
    raw_styles = raw_preferences.get("styles", {})
    raw_text = raw_preferences.get("text", {})
    raw_common_questions = raw_preferences.get("commonQuestions")
    styles: dict[str, dict[str, float]] = {}
    text: dict[str, str] = {}
    common_questions: list[str] | None = None

    if isinstance(raw_styles, dict):
        for raw_key, raw_values in raw_styles.items():
            key = _clean_ui_key(raw_key)
            if not key or not isinstance(raw_values, dict):
                continue
            values = _sanitize_ui_style_values(raw_values)
            if values:
                styles[key] = values

    if isinstance(raw_text, dict):
        for raw_key, raw_value in raw_text.items():
            key = _clean_ui_key(raw_key)
            if not key:
                continue
            value = str(raw_value).replace("\r", "").strip()
            if len(value) > 200:
                value = value[:200]
            text[key] = value

    if isinstance(raw_common_questions, list):
        common_questions = []
        for raw_question in raw_common_questions:
            question = str(raw_question or "").replace("\r", "").strip()
            if not question or question in common_questions:
                continue
            common_questions.append(question[:500])
            if len(common_questions) >= 60:
                break

    sanitized: dict[str, object] = {
        "enabled": bool(raw_preferences.get("enabled", False)),
        "styles": styles,
        "text": text,
    }
    if common_questions is not None:
        sanitized["commonQuestions"] = common_questions
    return sanitized


def _sanitize_preview_column_preferences(
    raw_preferences: dict[object, object],
    fallback: dict[str, object] | None = None,
) -> dict[str, object]:
    defaults = fallback or _builtin_preview_column_preferences()
    raw_default_labels = raw_preferences.get("defaultLabels", defaults["defaultLabels"])
    raw_sheet_overrides = raw_preferences.get("sheetOverrides", {})
    raw_header_rows = raw_preferences.get("headerRows", {})
    raw_max_display_chars = raw_preferences.get("maxDisplayChars", defaults["maxDisplayChars"])
    raw_column_widths = raw_preferences.get("columnWidths", {})

    default_labels = _sanitize_text_list(raw_default_labels)
    if not default_labels:
        default_labels = list(defaults["defaultLabels"])

    sheet_overrides: dict[str, list[str]] = {}
    if isinstance(raw_sheet_overrides, dict):
        for raw_sheet_name, raw_labels in raw_sheet_overrides.items():
            sheet_name = str(raw_sheet_name or "").strip()
            labels = _sanitize_text_list(raw_labels)
            if sheet_name and labels:
                sheet_overrides[sheet_name] = labels

    header_rows: dict[str, int] = {}
    if isinstance(raw_header_rows, dict):
        for raw_sheet_name, raw_row in raw_header_rows.items():
            sheet_name = str(raw_sheet_name or "").strip()
            if not sheet_name:
                continue
            try:
                row_number = int(float(str(raw_row).strip()))
            except (TypeError, ValueError):
                continue
            if row_number >= 1:
                header_rows[sheet_name] = min(row_number, 999)

    try:
        max_display_chars = int(float(str(raw_max_display_chars).strip()))
    except (TypeError, ValueError):
        max_display_chars = DEFAULT_PREVIEW_CELL_MAX_DISPLAY_CHARS
    max_display_chars = max(4, min(40, max_display_chars))

    column_widths: dict[str, dict[str, int]] = {}
    if isinstance(raw_column_widths, dict):
        for raw_sheet_name, raw_widths in raw_column_widths.items():
            sheet_name = str(raw_sheet_name or "").strip()
            if not sheet_name or not isinstance(raw_widths, dict):
                continue
            widths: dict[str, int] = {}
            for raw_column_label, raw_width in raw_widths.items():
                column_label = str(raw_column_label or "").strip()
                if not column_label:
                    continue
                try:
                    width = int(round(float(str(raw_width).strip())))
                except (TypeError, ValueError):
                    continue
                widths[column_label] = max(MIN_PREVIEW_COLUMN_WIDTH_PX, min(MAX_PREVIEW_COLUMN_WIDTH_PX, width))
            if widths:
                column_widths[sheet_name] = widths

    return {
        "defaultLabels": default_labels,
        "sheetOverrides": sheet_overrides,
        "headerRows": header_rows,
        "maxDisplayChars": max_display_chars,
        "columnWidths": column_widths,
    }


def _sanitize_text_list(raw_values: object) -> list[str]:
    if not isinstance(raw_values, list):
        return []
    values: list[str] = []
    for raw_value in raw_values:
        value = str(raw_value or "").replace("\r", "").strip()
        if value and value not in values:
            values.append(value)
    return values


def _clean_ui_key(raw_key: object) -> str:
    key = str(raw_key).strip()
    if len(key) > 80:
        key = key[:80]
    return "".join(char for char in key if char.isalnum() or char in {"-", "_", "."})


def _sanitize_ui_style_values(raw_values: dict[object, object]) -> dict[str, float]:
    limits = {
        "paddingX": (0, 96),
        "paddingY": (0, 96),
        "fontSize": (10, 72),
        "radius": (0, 60),
        "gap": (0, 64),
        "marginTop": (-120, 120),
        "opacity": (20, 100),
    }
    sanitized: dict[str, float] = {}
    for key, raw_value in raw_values.items():
        name = str(key)
        if name not in limits:
            continue
        try:
            value = float(raw_value)
        except (TypeError, ValueError):
            continue
        lower, upper = limits[name]
        value = max(lower, min(upper, value))
        sanitized[name] = round(value, 2)
    return sanitized


def _parse_json_form_object(raw_value: str | None, label: str) -> dict[object, object] | None:
    if raw_value is None or not str(raw_value).strip():
        return None
    try:
        parsed = json.loads(raw_value)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail=f"{label}必须是 JSON 对象") from exc
    if not isinstance(parsed, dict):
        raise HTTPException(status_code=400, detail=f"{label}必须是 JSON 对象")
    return parsed


def _parse_input_field_preferences_form(raw_value: str | None) -> dict[str, list[str]] | None:
    parsed = _parse_json_form_object(raw_value, "输入字段偏好")
    return None if parsed is None else _sanitize_input_field_preferences(parsed)


def _parse_workload_field_preferences_form(raw_value: str | None, role: str) -> dict[str, list[str]] | None:
    parsed = _parse_json_form_object(raw_value, "工作量字段偏好")
    if parsed is None:
        return None
    if role == "target":
        return _sanitize_workload_target_field_preferences(parsed)
    return _sanitize_workload_field_preferences(parsed)


def _sanitize_bool_setting(value: object, default: bool) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in {"true", "1", "yes", "on", "是", "开启"}:
            return True
        if normalized in {"false", "0", "no", "off", "否", "关闭"}:
            return False
    return bool(value)


def _sanitize_optional_bool_setting(value: object) -> bool | None:
    if value is None:
        return None
    return _sanitize_bool_setting(value, False)


def _sanitize_input_field_preferences(raw_preferences: dict[object, object]) -> dict[str, list[str]]:
    allowed = set(INPUT_FIELD_PREFERENCE_FIELDS)
    sanitized: dict[str, list[str]] = {}
    for key, raw_aliases in raw_preferences.items():
        field = str(key).strip()
        if field not in allowed:
            continue
        if isinstance(raw_aliases, str):
            aliases = raw_aliases.replace(",", "\n").replace("，", "\n").splitlines()
        elif isinstance(raw_aliases, list):
            aliases = [str(alias) for alias in raw_aliases]
        else:
            aliases = []
        cleaned: list[str] = []
        seen: set[str] = set()
        for alias in aliases:
            value = str(alias).strip()
            if not value or value in seen:
                continue
            cleaned.append(value)
            seen.add(value)
        if cleaned:
            sanitized[field] = cleaned
    return sanitized


def _sanitize_experience_warning_settings(raw_settings: dict[object, object]) -> dict[str, float | bool | str]:
    defaults = _default_experience_warning_settings()
    low_raw = raw_settings.get("low_risk_warning_ratio", defaults["low_risk_warning_ratio"])
    high_raw = raw_settings.get("high_risk_warning_ratio", defaults["high_risk_warning_ratio"])
    only_check_rows_with_value = bool(raw_settings.get("only_check_rows_with_value", defaults["only_check_rows_with_value"]))
    value_filter_field = _parse_warning_filter_field(
        str(raw_settings.get("value_filter_field", defaults["value_filter_field"]) or defaults["value_filter_field"])
    )
    try:
        low = float(low_raw)
        high = float(high_raw)
    except (TypeError, ValueError) as exc:
        raise HTTPException(status_code=400, detail="预警比率必须是数字") from exc
    if low < 0 or high < 0:
        raise HTTPException(status_code=400, detail="预警比率必须大于等于 0")
    if high < low:
        raise HTTPException(status_code=400, detail="高风险预警比率必须大于等于低风险预警比率")
    return {
        "low_risk_warning_ratio": round(low, 6),
        "high_risk_warning_ratio": round(high, 6),
        "only_check_rows_with_value": only_check_rows_with_value,
        "value_filter_field": value_filter_field,
    }


def _sanitize_experience_field_preferences(raw_preferences: dict[object, object]) -> dict[str, list[str]]:
    allowed = set(EXPERIENCE_MAPPING_FIELDS)
    sanitized: dict[str, list[str]] = {}
    for key, raw_aliases in raw_preferences.items():
        field = str(key).strip()
        if field not in allowed:
            continue
        if isinstance(raw_aliases, str):
            aliases = raw_aliases.replace(",", "\n").replace("，", "\n").splitlines()
        elif isinstance(raw_aliases, list):
            aliases = [str(alias) for alias in raw_aliases]
        else:
            aliases = []
        cleaned: list[str] = []
        seen: set[str] = set()
        for alias in aliases:
            value = str(alias).strip()
            if not value or value in seen:
                continue
            cleaned.append(value)
            seen.add(value)
        if cleaned:
            sanitized[field] = cleaned
    return sanitized


def _sanitize_workload_field_preferences(raw_preferences: dict[object, object]) -> dict[str, list[str]]:
    allowed = set(WORKLOAD_FIELD_PREFERENCE_FIELDS)
    sanitized: dict[str, list[str]] = {}
    for key, raw_aliases in raw_preferences.items():
        field = str(key).strip()
        if field not in allowed:
            continue
        if isinstance(raw_aliases, str):
            aliases = raw_aliases.replace(",", "\n").replace("，", "\n").splitlines()
        elif isinstance(raw_aliases, list):
            aliases = [str(alias) for alias in raw_aliases]
        else:
            aliases = []
        cleaned: list[str] = []
        seen: set[str] = set()
        for alias in aliases:
            value = str(alias).strip()
            if not value or value in seen:
                continue
            cleaned.append(value)
            seen.add(value)
        if cleaned:
            sanitized[field] = cleaned
    return sanitized


def _sanitize_workload_target_field_preferences(raw_preferences: dict[object, object]) -> dict[str, list[str]]:
    allowed = set(WORKLOAD_TARGET_FIELD_PREFERENCE_FIELDS)
    sanitized: dict[str, list[str]] = {}
    for key, raw_aliases in raw_preferences.items():
        field = str(key).strip()
        if field not in allowed:
            continue
        if isinstance(raw_aliases, str):
            aliases = raw_aliases.replace(",", "\n").replace("，", "\n").splitlines()
        elif isinstance(raw_aliases, list):
            aliases = [str(alias) for alias in raw_aliases]
        else:
            aliases = []
        cleaned: list[str] = []
        seen: set[str] = set()
        for alias in aliases:
            value = str(alias).strip()
            if not value or value in seen:
                continue
            cleaned.append(value)
            seen.add(value)
        if cleaned:
            sanitized[field] = cleaned
    return sanitized


def _find_experience_field_column(
    headers: list[str],
    field: str,
    preferences: dict[str, list[str]],
    defaults: dict[str, list[str]],
) -> str:
    preferred_aliases = preferences.get(field, [])
    default_aliases = defaults.get(field, [])
    for alias in preferred_aliases:
        found = _find_column_letter(headers, [alias])
        if found:
            return found
    for alias in preferred_aliases:
        if alias in default_aliases:
            continue
        found = _find_column_by_token(headers, alias)
        if found:
            return found
    found = _find_column_letter(headers, default_aliases)
    if found:
        return found
    return ""


def _find_preferred_input_column(
    headers: list[str],
    field: str,
    preferences: dict[str, list[str]],
    defaults: dict[str, list[str]],
) -> str:
    exact_field = _find_column_letter(headers, [field])
    if exact_field:
        return exact_field
    preferred_aliases = preferences.get(field, [])
    default_aliases = defaults.get(field, [])
    for alias in preferred_aliases:
        found = _find_column_letter(headers, [alias])
        if found:
            return found
    for alias in preferred_aliases:
        if alias in default_aliases:
            continue
        found = _find_column_by_token(headers, alias)
        if found:
            return found
    found = _find_column_letter(headers, default_aliases)
    if found:
        return found
    for alias in default_aliases:
        found = _find_column_by_token(headers, alias)
        if found:
            return found
    return ""


def _find_experience_metric_column(headers: list[str], metric: str, preferred_aliases: list[str] | None = None) -> str:
    compact_metric = metric.replace(" ", "")
    preferred_tokens = {
        PRICE_METRIC: ["【经验数】单价", "【经验数】基价", "经验单价", "经验基价", "基价", "单价", "价格"],
        PHYSICAL_METRIC: ["【经验数】实物工作费调整系数", "经验实物工作费调整系数", "实物工作费调整系数"],
        TECHNICAL_METRIC: ["【经验数】技术工作费调整系数", "经验技术工作费调整系数", "技术工作费调整系数"],
    }
    tokens = [*(preferred_aliases or []), *preferred_tokens.get(metric, [compact_metric])]
    for token in tokens:
        found = _find_column_by_token(headers, token)
        if found:
            return found
    return ""


def _find_column_letter(headers: list[str], names: list[str]) -> str:
    for index, header in enumerate(headers, start=1):
        if header in names:
            return get_column_letter(index)
    return ""


def _find_column_by_token(headers: list[str], token: str) -> str:
    compact_token = token.replace(" ", "")
    for index, header in enumerate(headers, start=1):
        if compact_token in header.replace(" ", ""):
            return get_column_letter(index)
    return ""


def _parse_column_mapping(raw_mapping: str | None) -> dict[str, str] | None:
    if not raw_mapping:
        return None
    try:
        payload = json.loads(raw_mapping)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="列映射不是合法 JSON") from exc
    if not isinstance(payload, dict):
        raise HTTPException(status_code=400, detail="列映射必须是对象")
    return {str(key): str(value).strip() for key, value in payload.items() if value is not None}


def _parse_sheet_configs(raw_configs: str | None) -> list[dict[str, object]] | None:
    if not raw_configs:
        return None
    try:
        payload = json.loads(raw_configs)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="sheet 配置不是合法 JSON") from exc
    if not isinstance(payload, list):
        raise HTTPException(status_code=400, detail="sheet 配置必须是数组")
    configs: list[dict[str, object]] = []
    for item in payload:
        if not isinstance(item, dict):
            raise HTTPException(status_code=400, detail="sheet 配置项必须是对象")
        column_mapping = item.get("column_mapping")
        if column_mapping is not None and not isinstance(column_mapping, dict):
            raise HTTPException(status_code=400, detail="sheet 列映射必须是对象")
        config = {
            "sheet_name": str(item.get("sheet_name", "")).strip(),
            "enabled": bool(item.get("enabled", True)),
            "header_row": int(item.get("header_row") or 1),
            "column_mapping": {
                str(key): str(value).strip()
                for key, value in (column_mapping or {}).items()
                if value is not None
            },
            "output_match_report": bool(item.get("output_match_report", True)),
            "merge_vertical_cells": bool(item.get("merge_vertical_cells", True)),
            "merge_horizontal_cells": bool(item.get("merge_horizontal_cells", True)),
        }
        if "only_match_rows_with_value" in item:
            config["only_match_rows_with_value"] = bool(item.get("only_match_rows_with_value"))
        if "match_value_filter_field" in item:
            config["match_value_filter_field"] = _parse_warning_filter_field(
                str(item.get("match_value_filter_field") or DEFAULT_WARNING_FILTER_FIELD)
            )
        configs.append(config)
    return configs


def _parse_experience_sheet_configs(raw_configs: str | None) -> list[dict[str, object]] | None:
    if not raw_configs:
        return None
    try:
        payload = json.loads(raw_configs)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="经验池 sheet 配置不是合法 JSON") from exc
    if not isinstance(payload, list):
        raise HTTPException(status_code=400, detail="经验池 sheet 配置必须是数组")
    allowed = set(EXPERIENCE_MAPPING_FIELDS)
    configs: list[dict[str, object]] = []
    for item in payload:
        if not isinstance(item, dict):
            raise HTTPException(status_code=400, detail="经验池 sheet 配置项必须是对象")
        column_mapping = item.get("column_mapping")
        if column_mapping is not None and not isinstance(column_mapping, dict):
            raise HTTPException(status_code=400, detail="经验池列映射必须是对象")
        configs.append(
            {
                "sheet_name": str(item.get("sheet_name", "")).strip(),
                "enabled": bool(item.get("enabled", True)),
                "header_row": int(item.get("header_row") or 1),
                "column_mapping": {
                    str(key): str(value).strip()
                    for key, value in (column_mapping or {}).items()
                    if str(key) in allowed and value is not None
                },
            }
        )
    return configs


def _parse_selected_experience_fields(raw_fields: str | None) -> list[str]:
    if not raw_fields:
        return DEFAULT_SELECTED_EXPERIENCE_FIELDS
    try:
        payload = json.loads(raw_fields)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="经验字段不是合法 JSON") from exc
    if not isinstance(payload, list):
        raise HTTPException(status_code=400, detail="经验字段必须是数组")
    allowed = set(DEFAULT_SELECTED_EXPERIENCE_FIELDS)
    selected = [str(item) for item in payload if str(item) in allowed]
    if not selected:
        raise HTTPException(status_code=400, detail="至少选择一种经验字段")
    return selected


def _parse_experience_filter_field(raw_field: str | None) -> str:
    field = str(raw_field or "工程量").strip() or "工程量"
    if field not in EXPERIENCE_MAPPING_FIELDS:
        raise HTTPException(status_code=400, detail=f"经验池导入过滤字段不支持：{field}")
    return field


def _parse_warning_filter_field(raw_field: str | None) -> str:
    field = str(raw_field or DEFAULT_WARNING_FILTER_FIELD).strip() or DEFAULT_WARNING_FILTER_FIELD
    if field not in WARNING_FILTER_FIELDS:
        raise HTTPException(status_code=400, detail=f"预警过滤字段不支持：{field}")
    return field


def _parse_workload_selected_fields(raw_fields: str | None) -> list[str]:
    if not raw_fields:
        return DEFAULT_SELECTED_WORKLOAD_FIELDS
    try:
        payload = json.loads(raw_fields)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="工作量抓取字段不是合法 JSON") from exc
    if not isinstance(payload, list):
        raise HTTPException(status_code=400, detail="工作量抓取字段必须是数组")
    allowed = set(DEFAULT_SELECTED_WORKLOAD_FIELDS)
    selected = [str(item) for item in payload if str(item) in allowed]
    if not selected:
        raise HTTPException(status_code=400, detail="至少选择一个工作量抓取字段")
    return selected


def _parse_workload_filter_field(raw_field: str | None) -> str:
    field = str(raw_field or DEFAULT_WORKLOAD_FILTER_FIELD).strip() or DEFAULT_WORKLOAD_FILTER_FIELD
    if field not in SOURCE_MAPPING_FIELDS:
        raise HTTPException(status_code=400, detail=f"工作量抓取过滤字段不支持：{field}")
    return field


def _parse_workload_write_mode(raw_mode: str | None) -> str:
    mode = str(raw_mode or WRITE_MODE_CONSERVATIVE).strip().lower()
    if mode in {WRITE_MODE_CONSERVATIVE, "safe", "保守", "保守模式"}:
        return WRITE_MODE_CONSERVATIVE
    if mode in {WRITE_MODE_OVERWRITE, "cover", "覆盖", "覆盖模式"}:
        return WRITE_MODE_OVERWRITE
    raise HTTPException(status_code=400, detail="工作量抓取写入模式只能是保守模式或覆盖模式")


def _parse_workload_sheet_configs(raw_configs: str | None, role: str) -> list[dict[str, object]] | None:
    if not raw_configs:
        return None
    try:
        payload = json.loads(raw_configs)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="工作量抓取 sheet 配置不是合法 JSON") from exc
    if not isinstance(payload, list):
        raise HTTPException(status_code=400, detail="工作量抓取 sheet 配置必须是数组")
    allowed = set(SOURCE_MAPPING_FIELDS if role == "source" else TARGET_MAPPING_FIELDS)
    configs: list[dict[str, object]] = []
    for item in payload:
        if not isinstance(item, dict):
            raise HTTPException(status_code=400, detail="工作量抓取 sheet 配置项必须是对象")
        column_mapping = item.get("column_mapping")
        if column_mapping is not None and not isinstance(column_mapping, dict):
            raise HTTPException(status_code=400, detail="工作量抓取列映射必须是对象")
        configs.append(
            {
                "sheet_name": str(item.get("sheet_name", "")).strip(),
                "enabled": bool(item.get("enabled", True)),
                "header_row": int(item.get("header_row") or 1),
                "column_mapping": {
                    str(key): str(value).strip()
                    for key, value in (column_mapping or {}).items()
                    if str(key) in allowed and value is not None
                },
            }
        )
    return configs


def _resolve_experience_pool_path() -> Path:
    if DEFAULT_EXPERIENCE_POOL_PATH.exists():
        return DEFAULT_EXPERIENCE_POOL_PATH
    if LEGACY_EXPERIENCE_POOL_PATH.exists():
        return LEGACY_EXPERIENCE_POOL_PATH
    return DEFAULT_EXPERIENCE_POOL_PATH


def _resolve_frontend_static_dir() -> Path | None:
    configured = os.getenv("GUANKAN_FRONTEND_DIR", "").strip()
    candidates = [
        Path(configured) if configured else None,
        PROJECT_ROOT / "web",
        PROJECT_ROOT / "frontend" / "dist",
    ]
    for candidate in candidates:
        if candidate and (candidate / "index.html").exists():
            return candidate
    return None


def _mount_frontend_static_files() -> None:
    static_dir = _resolve_frontend_static_dir()
    if static_dir is None:
        return
    app.mount("/", StaticFiles(directory=static_dir, html=True), name="frontend")


_mount_frontend_static_files()



