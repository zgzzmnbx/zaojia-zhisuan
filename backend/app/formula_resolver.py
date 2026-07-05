from __future__ import annotations

import ast
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string


SHEET_PREFIX_PATTERN = r"(?:(?P<sheet_q>'(?:[^']|'')+')|(?P<sheet_u>[^'!+\-*/(),:&=<>]+))!"
CELL_REF_PATTERN = re.compile(
    rf"(?:{SHEET_PREFIX_PATTERN})?\$?(?P<col>[A-Z]{{1,3}})\$?(?P<row>\d+)",
    re.IGNORECASE,
)
RANGE_REF_PATTERN = re.compile(
    rf"(?:{SHEET_PREFIX_PATTERN})?\$?(?P<start_col>[A-Z]{{1,3}})\$?(?P<start_row>\d+):"
    r"\$?(?P<end_col>[A-Z]{1,3})\$?(?P<end_row>\d+)",
    re.IGNORECASE,
)
PERCENT_PATTERN = re.compile(r"(?<![A-Za-z0-9_])(\d+(?:\.\d+)?)%")
FUNCTION_PATTERN = re.compile(r"\b(SUM|ROUND|MIN|MAX)\s*\(", re.IGNORECASE)


class WorkbookFormulaResolver:
    """Resolve simple Excel formulas when openpyxl has no cached formula value."""

    def __init__(self, path: str | Path):
        self.path = Path(path)
        self.value_workbook = load_workbook(self.path, data_only=True, read_only=False)
        self.formula_workbook = load_workbook(self.path, data_only=False, read_only=False)
        self._cache: dict[tuple[str, int, int], Any] = {}
        self._visiting: set[tuple[str, int, int]] = set()

    def close(self) -> None:
        self.value_workbook.close()
        self.formula_workbook.close()

    def __enter__(self) -> "WorkbookFormulaResolver":
        return self

    def __exit__(self, *_args: object) -> None:
        self.close()

    @property
    def sheetnames(self) -> list[str]:
        return list(self.formula_workbook.sheetnames)

    def sheet_max_row(self, sheet_name: str) -> int:
        return self.formula_workbook[sheet_name].max_row

    def sheet_max_column(self, sheet_name: str) -> int:
        return self.formula_workbook[sheet_name].max_column

    def cell_value(self, sheet_name: str, row: int, column: int) -> Any:
        key = (sheet_name, row, column)
        if key in self._cache:
            return self._cache[key]
        if key in self._visiting:
            return None
        if sheet_name not in self.formula_workbook.sheetnames:
            return None

        self._visiting.add(key)
        try:
            value_sheet = self.value_workbook[sheet_name]
            formula_sheet = self.formula_workbook[sheet_name]
            cached_value = value_sheet.cell(row=row, column=column).value
            formula = formula_sheet.cell(row=row, column=column).value
            if isinstance(formula, str) and formula.startswith("="):
                resolved = self.evaluate_formula(sheet_name, formula)
                self._cache[key] = resolved if resolved is not None else cached_value
                return self._cache[key]

            if cached_value not in (None, ""):
                self._cache[key] = cached_value
                return cached_value

            self._cache[key] = formula if formula is not None else cached_value
            return self._cache[key]
        finally:
            self._visiting.discard(key)

    def row_values(self, sheet_name: str, row: int, max_col: int) -> list[Any]:
        return [self.cell_value(sheet_name, row, column) for column in range(1, max_col + 1)]

    def evaluate_formula(self, sheet_name: str, formula: str) -> Any:
        expression = formula.strip()
        if expression.startswith("="):
            expression = expression[1:]
        expression = expression.replace("$", "")
        expression = PERCENT_PATTERN.sub(r"(\1/100)", expression)
        expression = RANGE_REF_PATTERN.sub(
            lambda match: str(self._range_sum(self._resolve_sheet_name(sheet_name, match), match)),
            expression,
        )
        expression = CELL_REF_PATTERN.sub(
            lambda match: str(self._as_formula_number(self._cell_ref_value(sheet_name, match))),
            expression,
        )
        expression = FUNCTION_PATTERN.sub(lambda match: f"{match.group(1).upper()}(", expression)

        try:
            tree = ast.parse(expression, mode="eval")
            _validate_ast(tree)
            value = eval(
                compile(tree, "<excel-formula>", "eval"),
                {"__builtins__": {}},
                {
                    "SUM": lambda *args: sum(_to_number(arg) for arg in args),
                    "ROUND": lambda value, digits=0: round(_to_number(value), int(_to_number(digits))),
                    "MIN": lambda *args: min((_to_number(arg) for arg in args), default=0),
                    "MAX": lambda *args: max((_to_number(arg) for arg in args), default=0),
                },
            )
        except Exception:
            return None
        return value

    def _cell_ref_value(self, current_sheet: str, match: re.Match[str]) -> Any:
        target_sheet = self._resolve_sheet_name(current_sheet, match)
        column = column_index_from_string(match.group("col").upper())
        row = int(match.group("row"))
        return self.cell_value(target_sheet, row, column)

    def _range_sum(self, target_sheet: str, match: re.Match[str]) -> float:
        start_column = column_index_from_string(match.group("start_col").upper())
        end_column = column_index_from_string(match.group("end_col").upper())
        start_row = int(match.group("start_row"))
        end_row = int(match.group("end_row"))
        if start_column > end_column:
            start_column, end_column = end_column, start_column
        if start_row > end_row:
            start_row, end_row = end_row, start_row
        total = 0.0
        for row in range(start_row, end_row + 1):
            for column in range(start_column, end_column + 1):
                total += _to_number(self.cell_value(target_sheet, row, column))
        return total

    @staticmethod
    def _resolve_sheet_name(current_sheet: str, match: re.Match[str]) -> str:
        quoted = match.groupdict().get("sheet_q")
        if quoted:
            return quoted[1:-1].replace("''", "'")
        unquoted = match.groupdict().get("sheet_u")
        if unquoted:
            return unquoted.strip()
        return current_sheet

    @staticmethod
    def _as_formula_number(value: Any) -> float | int:
        return _to_number(value)


def _to_number(value: Any) -> float | int:
    if value in (None, ""):
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip().replace(",", "")
    if text.endswith("%"):
        try:
            return float(text[:-1]) / 100
        except ValueError:
            return 0
    try:
        return float(text)
    except ValueError:
        return 0


def _validate_ast(tree: ast.AST) -> None:
    allowed_nodes = (
        ast.Expression,
        ast.BinOp,
        ast.UnaryOp,
        ast.Constant,
        ast.Call,
        ast.Name,
        ast.Load,
        ast.Add,
        ast.Sub,
        ast.Mult,
        ast.Div,
        ast.Pow,
        ast.Mod,
        ast.USub,
        ast.UAdd,
    )
    allowed_names = {"SUM", "ROUND", "MIN", "MAX"}
    for node in ast.walk(tree):
        if not isinstance(node, allowed_nodes):
            raise ValueError(f"Unsupported formula node: {type(node).__name__}")
        if isinstance(node, ast.Name) and node.id not in allowed_names:
            raise ValueError(f"Unsupported formula name: {node.id}")
        if isinstance(node, ast.Call):
            if not isinstance(node.func, ast.Name) or node.func.id not in allowed_names:
                raise ValueError("Unsupported formula function")
