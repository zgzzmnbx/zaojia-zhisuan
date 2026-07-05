from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from pathlib import Path
from statistics import mean
from typing import Any

from openpyxl import load_workbook

from .experience_warning import EXPERIENCE_POOL_HEADERS, PHYSICAL_METRIC, PRICE_METRIC, TECHNICAL_METRIC
from .knowledge_base import KnowledgeBase
from .normalization import normalize_key_part, normalize_price
from .schemas import FIELD_COLUMNS


GOVERNANCE_METRICS = [PRICE_METRIC, PHYSICAL_METRIC, TECHNICAL_METRIC]
CRITICAL_FIELDS = [*FIELD_COLUMNS, PRICE_METRIC]


def build_experience_pool_governance_report(pool_path: Path, *, max_items: int = 100) -> dict[str, Any]:
    if not pool_path.exists():
        return {
            "pool_path": str(pool_path),
            "exists": False,
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "summary": {
                "total_rows": 0,
                "valid_key_rows": 0,
                "issue_count": 1,
                "categories": {"missing_pool": 1},
            },
            "issues": [
                {
                    "category": "missing_pool",
                    "severity": "high",
                    "message": f"经验池文件不存在：{pool_path}",
                    "suggestion": "先通过经验池导入模块生成独立经验池，再运行治理检查。",
                }
            ],
        }

    records = _read_pool_records(pool_path)
    issues: list[dict[str, Any]] = []
    issues.extend(_empty_field_issues(records))
    issues.extend(_unparseable_metric_issues(records))
    issues.extend(_duplicate_issues(records))
    issues.extend(_unit_inconsistency_issues(records))
    issues.extend(_sample_shortage_issues(records))
    issues.extend(_dispersion_issues(records))

    categories: dict[str, int] = {}
    for issue in issues:
        category = str(issue.get("category") or "other")
        categories[category] = categories.get(category, 0) + 1

    return {
        "pool_path": str(pool_path),
        "exists": True,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "summary": {
            "total_rows": len(records),
            "valid_key_rows": sum(1 for record in records if _has_key(record)),
            "issue_count": len(issues),
            "categories": categories,
        },
        "issues": issues[:max_items],
        "truncated": len(issues) > max_items,
    }


def write_governance_markdown(report: dict[str, Any], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    summary = dict(report.get("summary") or {})
    lines = [
        "# 经验池治理报告",
        "",
        f"- 生成时间：{report.get('generated_at', '')}",
        f"- 经验池文件：{report.get('pool_path', '')}",
        f"- 总记录数：{summary.get('total_rows', 0)}",
        f"- 有效关键字段记录数：{summary.get('valid_key_rows', 0)}",
        f"- 问题数：{summary.get('issue_count', 0)}",
        "",
        "## 问题分类",
        "",
    ]
    categories = dict(summary.get("categories") or {})
    if categories:
        lines.extend(f"- {category}：{count}" for category, count in sorted(categories.items()))
    else:
        lines.append("- 未发现治理问题。")
    lines.extend(["", "## 问题明细", ""])
    for index, issue in enumerate(report.get("issues") or [], start=1):
        location = _issue_location(issue)
        lines.append(f"### {index}. {issue.get('title') or issue.get('category')}")
        lines.append("")
        if location:
            lines.append(f"- 位置：{location}")
        lines.append(f"- 等级：{issue.get('severity', 'info')}")
        lines.append(f"- 说明：{issue.get('message', '')}")
        if issue.get("key_text"):
            lines.append(f"- 关键要素：{issue.get('key_text')}")
        lines.append(f"- 建议：{issue.get('suggestion', '人工复核后处理。')}")
        lines.append("")
    if report.get("truncated"):
        lines.append("> 问题较多，报告仅展示前 100 条。")
    output_path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")
    return output_path


def _read_pool_records(pool_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(pool_path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        raw_headers = [str(value or "").strip() for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        headers = raw_headers if any(raw_headers) else EXPERIENCE_POOL_HEADERS
        records: list[dict[str, Any]] = []
        for excel_row, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            record = {
                header: row[index] if index < len(row) else None
                for index, header in enumerate(headers)
                if header
            }
            if not any(value not in (None, "") for value in record.values()):
                continue
            record["_sheet_name"] = sheet.title
            record["_excel_row"] = excel_row
            records.append(record)
        return records
    finally:
        workbook.close()


def _empty_field_issues(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    for record in records:
        missing = [field for field in CRITICAL_FIELDS if _blank(record.get(field))]
        if missing:
            issues.append(
                _issue(
                    "empty_field",
                    "high" if "单位" in missing or PRICE_METRIC in missing else "low",
                    record,
                    f"关键字段为空：{'、'.join(missing)}",
                    "补齐字段后再参与预警比选；无法确认的记录应从经验池中人工剔除或标注。",
                )
            )
    return issues


def _unparseable_metric_issues(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    for record in records:
        for metric in GOVERNANCE_METRICS:
            value = record.get(metric)
            if _blank(value):
                continue
            if _number(value) is None:
                issues.append(
                    _issue(
                        "unparseable_metric",
                        "high",
                        record,
                        f"{metric} 不可解析为数字：{value}",
                        "检查原始控制价文件或经验池导入映射，必要时人工修正后重新导入。",
                    )
                )
    return issues


def _duplicate_issues(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, ...], list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        if _has_key(record):
            grouped[KnowledgeBase.make_key(record)].append(record)
    issues: list[dict[str, Any]] = []
    for group in grouped.values():
        if len(group) <= 1:
            continue
        prices = sorted({_number(record.get(PRICE_METRIC)) for record in group if _number(record.get(PRICE_METRIC)) is not None})
        severity = "high" if len(prices) > 1 else "low"
        first = group[0]
        rows = "、".join(str(item.get("_excel_row")) for item in group[:10])
        issues.append(
            _issue(
                "duplicate_record",
                severity,
                first,
                f"同一要素组合出现 {len(group)} 条经验记录；行号：{rows}；基价集合：{prices or '无有效数字'}",
                "人工判断是否为真实多项目样本；如是重复导入，应重新整理经验池。",
                related_rows=[int(item.get("_excel_row") or 0) for item in group],
            )
        )
    return issues


def _unit_inconsistency_issues(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, ...], list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        element_key = tuple(normalize_key_part(record.get(field)) for field in FIELD_COLUMNS[:5])
        if any(element_key):
            grouped[element_key].append(record)
    issues: list[dict[str, Any]] = []
    for group in grouped.values():
        units = sorted({normalize_key_part(record.get("单位")) for record in group if normalize_key_part(record.get("单位"))})
        if len(units) <= 1:
            continue
        first = group[0]
        rows = "、".join(str(item.get("_excel_row")) for item in group[:10])
        issues.append(
            _issue(
                "unit_inconsistent",
                "high",
                first,
                f"同一要素1-5存在多个单位：{'、'.join(units)}；行号：{rows}",
                "核对计量单位是否为真实不同子项；若只是写法不一致，应人工统一后重新导入。",
                related_rows=[int(item.get("_excel_row") or 0) for item in group],
            )
        )
    return issues


def _sample_shortage_issues(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, ...], list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        if _has_key(record):
            grouped[KnowledgeBase.make_key(record)].append(record)
    return [
        _issue(
            "sample_shortage",
            "low",
            group[0],
            f"同类样本数仅 {len(group)} 条，经验均值代表性有限。",
            "风险预警可作为线索，但不宜作为强结论；建议继续导入同类历史样本。",
            related_rows=[int(item.get("_excel_row") or 0) for item in group],
        )
        for group in grouped.values()
        if 0 < len(group) < 3
    ]


def _dispersion_issues(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, ...], list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        if _has_key(record):
            grouped[KnowledgeBase.make_key(record)].append(record)
    issues: list[dict[str, Any]] = []
    for group in grouped.values():
        for metric in GOVERNANCE_METRICS:
            values = [_number(record.get(metric)) for record in group]
            numbers = [float(value) for value in values if value is not None]
            if len(numbers) < 3:
                continue
            avg = mean(numbers)
            if avg == 0:
                continue
            spread = (max(numbers) - min(numbers)) / abs(avg)
            if spread >= 0.4:
                issues.append(
                    _issue(
                        "extreme_dispersion",
                        "high" if spread >= 0.8 else "low",
                        group[0],
                        f"{metric} 样本离散度 {round(spread * 100, 2)}%，范围 {min(numbers)} - {max(numbers)}。",
                        "逐条核对历史项目口径，确认是否混入不同工作内容或单位。",
                        related_rows=[int(item.get("_excel_row") or 0) for item in group],
                    )
                )
    return issues


def _issue(
    category: str,
    severity: str,
    record: dict[str, Any],
    message: str,
    suggestion: str,
    *,
    related_rows: list[int] | None = None,
) -> dict[str, Any]:
    return {
        "category": category,
        "title": _category_title(category),
        "severity": severity,
        "sheet": record.get("_sheet_name", ""),
        "row": int(record.get("_excel_row") or 0),
        "related_rows": related_rows or [int(record.get("_excel_row") or 0)],
        "key": {field: record.get(field) for field in FIELD_COLUMNS},
        "key_text": _key_text(record),
        "message": message,
        "suggestion": suggestion,
    }


def _category_title(category: str) -> str:
    return {
        "missing_pool": "经验池文件缺失",
        "empty_field": "关键字段空值",
        "unparseable_metric": "经验数不可解析",
        "duplicate_record": "重复或冲突记录",
        "unit_inconsistent": "单位不一致",
        "sample_shortage": "同类样本不足",
        "extreme_dispersion": "极端值或离散值",
    }.get(category, category)


def _issue_location(issue: dict[str, Any]) -> str:
    sheet = str(issue.get("sheet") or "")
    row = issue.get("row")
    if sheet and row:
        return f"{sheet} 第 {row} 行"
    return ""


def _key_text(record: dict[str, Any]) -> str:
    return " / ".join(str(record.get(field) or "").strip() for field in FIELD_COLUMNS if str(record.get(field) or "").strip())


def _blank(value: Any) -> bool:
    return normalize_key_part(value) == ""


def _has_key(record: dict[str, Any]) -> bool:
    return bool(normalize_key_part(record.get("要素1")) and normalize_key_part(record.get("单位")))


def _number(value: Any) -> int | float | None:
    parsed = normalize_price(value)
    if isinstance(parsed, (int, float)) and not isinstance(parsed, bool):
        return parsed
    return None
