from __future__ import annotations

import hashlib
import json
import logging
import re
from io import BytesIO
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


LOGGER = logging.getLogger(__name__)
ALLOWED_STATUSES = {"active", "beta", "planned", "disabled"}
TASK_ENABLED_STATUSES = {"active"}
SUB_SKILL_TYPES = {"professional", "shared"}
SUB_SKILL_STATUSES = {"available", "planned"}
REQUIRED_FIELDS = {
    "id",
    "displayName",
    "version",
    "status",
    "domain",
    "description",
    "inputProfile",
    "capabilities",
    "assets",
    "validation",
}
PROHIBITED_KEYS = {
    "script",
    "scripts",
    "command",
    "commands",
    "executable",
    "entrypoint",
    "module",
    "pythonmodule",
    "shell",
}
SECRET_KEYS = {
    "apikey",
    "appsecret",
    "password",
    "secret",
    "token",
    "accesstoken",
    "refreshtoken",
}
EXECUTABLE_SUFFIXES = {".bat", ".cmd", ".com", ".exe", ".ps1", ".py", ".sh"}
SEMVER_PATTERN = re.compile(r"^\d+\.\d+\.\d+$")
SKILL_ID_PATTERN = re.compile(r"^[a-z0-9]+(?:-[a-z0-9]+)*$")
ASSET_LABELS = {
    "knowledgeBase": "结构化计价库",
    "reportTemplate": "Word 报告模板",
    "ruleDocuments": "核心规则说明",
    "ruleWorkbooks": "结构化规则表",
    "validationSample": "回归验证样例",
}
TRUSTED_PROCESSOR_IDS = {"survey-measurement-v1"}
KNOWLEDGE_SOURCE_SUFFIXES = {".csv", ".md", ".xlsx"}
MAX_RECOMMENDATION_FILE_BYTES = 25 * 1024 * 1024
MANAGEMENT_ACTIONS = {"install", "enable", "disable", "upgrade", "rollback", "uninstall", "export"}


@dataclass(frozen=True)
class SkillRuntimeContext:
    skill_id: str
    skill_version: str
    manifest_hash: str
    input_profile: dict[str, object]
    processor_id: str
    knowledge_base_path: Path
    rule_assets: dict[str, tuple[Path, ...]]
    risk_profile: dict[str, Path]
    knowledge_sources: tuple[Path, ...]
    report_template_path: Path
    validation_profile: dict[str, object]
    capabilities: dict[str, bool]

    def to_state(self, project_root: Path) -> dict[str, object]:
        root = project_root.resolve()

        def relative(path: Path) -> str:
            return path.resolve().relative_to(root).as_posix()

        return {
            "skill_id": self.skill_id,
            "skill_version": self.skill_version,
            "manifest_hash": self.manifest_hash,
            "input_profile": self.input_profile,
            "processor_id": self.processor_id,
            "knowledge_base": relative(self.knowledge_base_path),
            "rule_assets": {
                key: [relative(path) for path in paths]
                for key, paths in self.rule_assets.items()
            },
            "risk_profile": {
                key: relative(path)
                for key, path in self.risk_profile.items()
            },
            "knowledge_sources": [relative(path) for path in self.knowledge_sources],
            "report_template": relative(self.report_template_path),
            "validation_profile": self.validation_profile,
            "capabilities": self.capabilities,
        }


class ProfessionalSkillError(ValueError):
    def __init__(self, code: str, message: str, *, status_code: int = 400) -> None:
        super().__init__(message)
        self.code = code
        self.message = message
        self.status_code = status_code

    def detail(self) -> dict[str, str]:
        return {"code": self.code, "message": self.message}


class ProfessionalSkillRegistry:
    def __init__(
        self,
        project_root: Path,
        skills_root: Path,
        settings_path: Path,
        *,
        fallback_default_skill_id: str = "survey-measurement-limit-price",
    ) -> None:
        self.project_root = project_root.resolve()
        self.skills_root = skills_root.resolve()
        self.settings_path = settings_path
        self.fallback_default_skill_id = fallback_default_skill_id

    def default_skill_id(self) -> str:
        try:
            payload = json.loads(self.settings_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            payload = {}
        section = payload.get("professionalSkills", {}) if isinstance(payload, dict) else {}
        configured = str(section.get("defaultSkillId") or "").strip() if isinstance(section, dict) else ""
        return configured or self.fallback_default_skill_id

    def list_public(self) -> dict[str, object]:
        manifests = self._load_all()
        default_id = self.default_skill_id()
        return {
            "default_skill_id": default_id,
            "items": [self.public_summary(manifest, default_id=default_id) for manifest in manifests],
        }

    def get_public(self, skill_id: str) -> dict[str, object]:
        manifest = self.load(skill_id)
        return self.public_detail(manifest, default_id=self.default_skill_id())

    def recommend_for_file(self, filename: str, content: bytes) -> dict[str, object]:
        clean_name = Path(str(filename or "")).name
        extension = Path(clean_name).suffix.lower()
        if extension != ".xlsx":
            raise ProfessionalSkillError("skill_recommendation_input_invalid", "专业能力推荐当前只分析 .xlsx 文件")
        if not content:
            raise ProfessionalSkillError("skill_recommendation_input_invalid", "待分析文件为空")
        if len(content) > MAX_RECOMMENDATION_FILE_BYTES:
            raise ProfessionalSkillError(
                "skill_recommendation_input_too_large",
                "待分析文件超过 25 MB，未执行专业能力推荐",
                status_code=413,
            )
        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=False)
        except Exception as exc:
            raise ProfessionalSkillError("skill_recommendation_input_invalid", "无法读取待分析 Excel 文件") from exc
        try:
            sheet_count = len(workbook.sheetnames)
            sheet_names = workbook.sheetnames[:12]
            headers: list[str] = []
            for sheet_name in sheet_names:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 12), max_col=min(sheet.max_column, 80), values_only=True):
                    headers.extend(str(value).strip() for value in row if value not in (None, ""))
        finally:
            workbook.close()

        manifests = self._load_all()
        items = [
            self._recommendation_item(
                manifest,
                filename=clean_name,
                extension=extension,
                sheet_names=sheet_names,
                headers=headers,
            )
            for manifest in manifests
        ]
        items.sort(key=lambda item: (-int(item["score"]), int(item["display_order"]), str(item["display_name"])))
        top = items[0] if items else None
        second_score = int(items[1]["score"]) if len(items) > 1 else -1
        recommended = top if top and int(top["score"]) >= 35 and int(top["score"]) > second_score else None
        return {
            "file": {
                "name": clean_name,
                "extension": extension,
                "sheet_count": sheet_count,
                "observed_sheet_names": sheet_names,
                "observed_header_count": len(set(headers)),
            },
            "recommended_skill_id": recommended["id"] if recommended else None,
            "requires_confirmation": recommended is not None,
            "items": [{key: value for key, value in item.items() if key != "display_order"} for item in items],
            "notice": "推荐仅依据文件结构特征，必须由用户确认；不会自动切换专业能力或创建任务。",
        }

    def open_format(self) -> dict[str, object]:
        return {
            "format": "zaojiazhisuan-professional-skill",
            "format_version": "1.0",
            "descriptor": "manifest.json",
            "documentation": "SKILL.md",
            "required_manifest_fields": sorted(REQUIRED_FIELDS),
            "supported_statuses": sorted(ALLOWED_STATUSES),
            "asset_policy": "资产只允许项目内相对路径，禁止可执行文件、越界路径和运行秘密。",
            "runtime_policy": "只有进入代码白名单的 processorId 才能运行；SKILL.md 不参与价格、系数或风险裁决。",
            "lifecycle_policy": "安装、启停、升级、回滚和卸载当前只提供审核计划接口，不直接改写能力包。",
        }

    def management_overview(self) -> dict[str, object]:
        manifests = self._load_all()
        default_id = self.default_skill_id()
        return {
            "mode": "governed-interface",
            "changes_enabled": False,
            "items": [self._management_item(manifest, default_id=default_id) for manifest in manifests],
            "governance": {
                "publisher": "local-project",
                "signature_required_for_external_packages": True,
                "approval_required": True,
                "arbitrary_script_execution": False,
                "audit_fields": ["operator", "action", "skill_id", "source_version", "target_version", "checks", "decision", "created_at"],
            },
        }

    def plan_lifecycle(self, skill_id: str, action: str, target_version: str | None = None) -> dict[str, object]:
        clean_action = str(action or "").strip().lower()
        if clean_action not in MANAGEMENT_ACTIONS:
            raise ProfessionalSkillError("skill_management_action_invalid", "不支持的专业能力管理动作")
        clean_skill_id = str(skill_id or "").strip()
        if not SKILL_ID_PATTERN.fullmatch(clean_skill_id):
            raise ProfessionalSkillError("skill_not_found", "未找到指定的专业能力", status_code=404)
        manifest = None if clean_action == "install" else self.load(clean_skill_id)
        clean_target = str(target_version or "").strip()
        if clean_target and not SEMVER_PATTERN.fullmatch(clean_target):
            raise ProfessionalSkillError("skill_management_version_invalid", "目标版本必须使用 x.y.z")
        checks = [
            {"id": "manifest", "label": "Manifest 结构与安全校验", "passed": True},
            {"id": "paths", "label": "项目内相对资产路径", "passed": True},
            {"id": "scripts", "label": "无任意脚本或命令入口", "passed": True},
            {"id": "signature", "label": "可信发布签名", "passed": False},
            {"id": "approval", "label": "人工发布审批", "passed": False},
        ]
        blockers = ["当前为治理接口模式，未启用能力包文件写入。", "尚未接入可信签名与发布审批。"]
        if clean_action == "enable" and manifest and str(manifest["status"]) == "planned":
            blockers.append("规划中能力缺少独立业务资产和已验证样例，不能启用。")
        if clean_action in {"upgrade", "rollback"} and not clean_target:
            blockers.append("升级或回滚必须指定目标版本。")
        if clean_action == "disable" and manifest and manifest["id"] == self.default_skill_id():
            blockers.append("默认专业能力停用前必须先指定可用替代能力。")
        if clean_action == "install":
            blockers.append("安装前必须提供已签名能力包、发布者身份、兼容性声明和人工审批记录。")
        return {
            "skill_id": clean_skill_id,
            "display_name": manifest["displayName"] if manifest else clean_skill_id,
            "action": clean_action,
            "source_version": manifest["version"] if manifest else None,
            "target_version": clean_target or None,
            "status": "review_required",
            "changes_applied": False,
            "checks": checks,
            "blockers": blockers,
            "created_at": datetime.now().isoformat(timespec="seconds"),
        }

    def load(self, skill_id: str) -> dict[str, Any]:
        clean_id = str(skill_id or "").strip()
        if not SKILL_ID_PATTERN.fullmatch(clean_id):
            raise ProfessionalSkillError("skill_not_found", "未找到指定的专业能力", status_code=404)
        manifest_path = self.skills_root / clean_id / "manifest.json"
        if not manifest_path.is_file():
            raise ProfessionalSkillError("skill_not_found", "未找到指定的专业能力", status_code=404)
        return self._load_manifest(manifest_path)

    def resolve_for_task(self, skill_id: str | None, skill_version: str | None) -> dict[str, object]:
        explicit = bool(str(skill_id or "").strip() or str(skill_version or "").strip())
        resolved_id = str(skill_id or "").strip() or self.default_skill_id()
        try:
            manifest = self.load(resolved_id)
        except ProfessionalSkillError:
            if explicit:
                raise
            LOGGER.warning("默认专业能力 Manifest 不可用，启用勘察测量安全兼容快照")
            return self._compatibility_snapshot(resolved_id)

        status = str(manifest["status"])
        if status not in TASK_ENABLED_STATUSES:
            label = {"planned": "规划中", "beta": "内测中", "disabled": "已停用"}.get(status, "不可用")
            raise ProfessionalSkillError(
                "skill_not_available",
                f"专业能力“{manifest['displayName']}”当前为{label}，不能创建真实任务",
                status_code=409,
            )
        requested_version = str(skill_version or "").strip()
        if requested_version and requested_version != manifest["version"]:
            raise ProfessionalSkillError(
                "skill_version_mismatch",
                "专业能力版本已更新，请刷新能力清单后重试",
                status_code=409,
            )
        return self.create_snapshot(manifest)

    def create_snapshot(self, manifest: dict[str, Any]) -> dict[str, object]:
        runtime_context = self.create_runtime_context(manifest)
        return {
            "id": manifest["id"],
            "display_name": manifest["displayName"],
            "version": manifest["version"],
            "manifest_hash": manifest["_manifest_hash"],
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "asset_ids": sorted(manifest["assets"].keys()),
            "runtime_context": runtime_context.to_state(self.project_root),
            "compatibility_fallback": False,
        }

    def create_runtime_context(self, manifest: dict[str, Any]) -> SkillRuntimeContext:
        runtime = manifest.get("runtime")
        if not isinstance(runtime, dict):
            raise ProfessionalSkillError(
                "skill_runtime_invalid",
                "上线专业能力缺少运行时绑定",
                status_code=503,
            )
        processor_id = str(runtime.get("processorId") or "").strip()
        if processor_id not in TRUSTED_PROCESSOR_IDS:
            raise ProfessionalSkillError(
                "skill_processor_not_allowed",
                "专业能力处理器未进入可信代码白名单",
                status_code=503,
            )

        knowledge_base = self._single_bound_asset(manifest, runtime, "knowledgeBaseAsset")
        report_template = self._single_bound_asset(manifest, runtime, "reportTemplateAsset")
        rule_assets = self._named_bound_assets(manifest, runtime.get("ruleAssets"), "规则资产")
        risk_profile = {
            key: paths[0]
            for key, paths in self._named_bound_assets(
                manifest,
                runtime.get("riskProfile"),
                "风险配置",
                require_single=True,
            ).items()
        }
        knowledge_sources = self._knowledge_sources(
            self._bound_asset_paths(manifest, runtime.get("knowledgeSourceAssets"), "知识源")
        )
        validation_profile = dict(manifest["validation"])
        validation_asset_id = str(runtime.get("validationAsset") or "").strip()
        if validation_asset_id:
            validation_profile["asset"] = self._relative_asset_path(
                self._asset_paths(manifest, validation_asset_id, "验证资产")[0]
            )
        return SkillRuntimeContext(
            skill_id=str(manifest["id"]),
            skill_version=str(manifest["version"]),
            manifest_hash=str(manifest["_manifest_hash"]),
            input_profile=dict(manifest["inputProfile"]),
            processor_id=processor_id,
            knowledge_base_path=knowledge_base,
            rule_assets=rule_assets,
            risk_profile=risk_profile,
            knowledge_sources=knowledge_sources,
            report_template_path=report_template,
            validation_profile=validation_profile,
            capabilities={key: bool(value) for key, value in manifest["capabilities"].items()},
        )

    def runtime_from_snapshot(self, snapshot: object) -> SkillRuntimeContext:
        if not isinstance(snapshot, dict) or not isinstance(snapshot.get("runtime_context"), dict):
            raise ProfessionalSkillError(
                "skill_runtime_missing",
                "任务缺少专业能力运行时快照",
                status_code=409,
            )
        payload = snapshot["runtime_context"]
        processor_id = str(payload.get("processor_id") or "").strip()
        if processor_id not in TRUSTED_PROCESSOR_IDS:
            raise ProfessionalSkillError(
                "skill_processor_not_allowed",
                "任务处理器未进入可信代码白名单",
                status_code=503,
            )
        skill_id = str(payload.get("skill_id") or "")
        skill_version = str(payload.get("skill_version") or "")
        manifest_hash = str(payload.get("manifest_hash") or "")
        if (
            skill_id != str(snapshot.get("id") or "")
            or skill_version != str(snapshot.get("version") or "")
            or manifest_hash != str(snapshot.get("manifest_hash") or "")
        ):
            raise ProfessionalSkillError(
                "skill_runtime_invalid",
                "任务专业能力快照身份不一致",
                status_code=409,
            )
        rule_assets_raw = payload.get("rule_assets")
        risk_profile_raw = payload.get("risk_profile")
        knowledge_sources_raw = payload.get("knowledge_sources")
        if not isinstance(rule_assets_raw, dict) or not isinstance(risk_profile_raw, dict) or not isinstance(knowledge_sources_raw, list):
            raise ProfessionalSkillError("skill_runtime_invalid", "任务专业能力运行时快照无效", status_code=409)
        return SkillRuntimeContext(
            skill_id=skill_id,
            skill_version=skill_version,
            manifest_hash=manifest_hash,
            input_profile=dict(payload.get("input_profile") or {}),
            processor_id=processor_id,
            knowledge_base_path=self._resolve_frozen_path(payload.get("knowledge_base")),
            rule_assets={
                str(key): tuple(self._resolve_frozen_path(path) for path in paths)
                for key, paths in rule_assets_raw.items()
                if isinstance(paths, list)
            },
            risk_profile={
                str(key): self._resolve_frozen_path(path)
                for key, path in risk_profile_raw.items()
            },
            knowledge_sources=tuple(self._resolve_frozen_path(path) for path in knowledge_sources_raw),
            report_template_path=self._resolve_frozen_path(payload.get("report_template")),
            validation_profile=dict(payload.get("validation_profile") or {}),
            capabilities={key: bool(value) for key, value in dict(payload.get("capabilities") or {}).items()},
        )

    @staticmethod
    def public_snapshot(snapshot: object) -> dict[str, object]:
        if not isinstance(snapshot, dict):
            return {}
        return {
            "id": str(snapshot.get("id") or ""),
            "display_name": str(snapshot.get("display_name") or ""),
            "version": str(snapshot.get("version") or ""),
            "manifest_hash": str(snapshot.get("manifest_hash") or ""),
            "created_at": str(snapshot.get("created_at") or ""),
            "compatibility_fallback": bool(snapshot.get("compatibility_fallback", False)),
        }

    def public_summary(self, manifest: dict[str, Any], *, default_id: str) -> dict[str, object]:
        status = str(manifest["status"])
        default_status_label = {"active": "已上线", "beta": "内测中", "planned": "规划中", "disabled": "已停用"}[status]
        return {
            "id": manifest["id"],
            "display_name": manifest["displayName"],
            "version": manifest["version"],
            "status": status,
            "status_label": str(manifest.get("statusLabel") or default_status_label),
            "domain": manifest["domain"],
            "description": manifest["description"],
            "capabilities": self._capability_names(manifest["capabilities"]),
            "asset_count": self._asset_count(manifest["assets"]),
            "validation_status": str(manifest["validation"].get("status") or "unverified"),
            "is_default": manifest["id"] == default_id,
            "can_create_task": status in TASK_ENABLED_STATUSES,
        }

    def public_detail(self, manifest: dict[str, Any], *, default_id: str) -> dict[str, object]:
        summary = self.public_summary(manifest, default_id=default_id)
        return {
            **summary,
            "input_profile": manifest["inputProfile"],
            "applicability": manifest.get("applicability", {}),
            "sub_skills": manifest.get("subSkills", []),
            "asset_summary": self._asset_summary(manifest["assets"]),
            "validation": manifest["validation"],
            "boundary": "规则裁决、模型解释、人工兜底；专业能力说明不得覆盖结构化计价规则。",
        }

    def _recommendation_item(
        self,
        manifest: dict[str, Any],
        *,
        filename: str,
        extension: str,
        sheet_names: list[str],
        headers: list[str],
    ) -> dict[str, object]:
        input_profile = manifest["inputProfile"]
        recommendation = manifest.get("recommendationProfile")
        profile = recommendation if isinstance(recommendation, dict) else {}
        reasons: list[str] = []
        score = 0
        extensions = {str(value).lower() for value in input_profile.get("extensions", [])}
        if extension in extensions:
            score += 20
            reasons.append(f"支持 {extension} 输入")
        score += self._keyword_score(filename, profile.get("fileNameKeywords"), 10, 20, "文件名", reasons)
        score += self._keyword_score(" ".join(sheet_names), profile.get("sheetKeywords"), 8, 24, "工作表", reasons)
        score += self._keyword_score(" ".join(headers), profile.get("headerKeywords"), 6, 36, "表头", reasons)
        confidence = "high" if score >= 70 else "medium" if score >= 35 else "low"
        return {
            "id": manifest["id"],
            "display_name": manifest["displayName"],
            "version": manifest["version"],
            "status": manifest["status"],
            "status_label": str(manifest.get("statusLabel") or {"active": "已上线", "beta": "内测中", "planned": "规划中", "disabled": "已停用"}[manifest["status"]]),
            "can_create_task": manifest["status"] in TASK_ENABLED_STATUSES,
            "score": score,
            "confidence": confidence,
            "reasons": reasons or ["未发现足够的专业结构特征"],
            "display_order": int(manifest.get("displayOrder", 1000)),
        }

    @staticmethod
    def _keyword_score(text: str, raw_keywords: object, points: int, maximum: int, label: str, reasons: list[str]) -> int:
        if not isinstance(raw_keywords, list):
            return 0
        haystack = text.casefold()
        matches = [str(keyword).strip() for keyword in raw_keywords if str(keyword).strip() and str(keyword).strip().casefold() in haystack]
        if matches:
            reasons.append(f"{label}命中：{'、'.join(matches[:4])}")
        return min(len(matches) * points, maximum)

    def _management_item(self, manifest: dict[str, Any], *, default_id: str) -> dict[str, object]:
        skill_dir = self.skills_root / str(manifest["id"])
        return {
            **self.public_summary(manifest, default_id=default_id),
            "manifest_valid": True,
            "skill_md_present": (skill_dir / "SKILL.md").is_file(),
            "runtime_ready": manifest["status"] in TASK_ENABLED_STATUSES,
            "manifest_hash": manifest["_manifest_hash"],
            "signature_status": "not_configured",
            "available_operations": ["inspect", "validate", "recommend", "plan_lifecycle"],
            "write_operations_enabled": False,
        }

    def _single_bound_asset(
        self,
        manifest: dict[str, Any],
        runtime: dict[str, Any],
        binding_name: str,
    ) -> Path:
        asset_id = str(runtime.get(binding_name) or "").strip()
        paths = self._asset_paths(manifest, asset_id, binding_name)
        if len(paths) != 1:
            raise ProfessionalSkillError("skill_runtime_invalid", f"{binding_name} 必须绑定一个资产", status_code=503)
        return paths[0]

    def _named_bound_assets(
        self,
        manifest: dict[str, Any],
        bindings: object,
        label: str,
        *,
        require_single: bool = False,
    ) -> dict[str, tuple[Path, ...]]:
        if not isinstance(bindings, dict) or not bindings:
            raise ProfessionalSkillError("skill_runtime_invalid", f"{label}绑定不能为空", status_code=503)
        resolved: dict[str, tuple[Path, ...]] = {}
        for key, asset_id in bindings.items():
            clean_key = str(key or "").strip()
            if not clean_key:
                raise ProfessionalSkillError("skill_runtime_invalid", f"{label}名称不能为空", status_code=503)
            paths = self._asset_paths(manifest, str(asset_id or "").strip(), label)
            if require_single and len(paths) != 1:
                raise ProfessionalSkillError("skill_runtime_invalid", f"{label}必须逐项绑定单个资产", status_code=503)
            resolved[clean_key] = tuple(paths)
        return resolved

    def _bound_asset_paths(
        self,
        manifest: dict[str, Any],
        asset_ids: object,
        label: str,
    ) -> list[Path]:
        if not isinstance(asset_ids, list) or not asset_ids:
            raise ProfessionalSkillError("skill_runtime_invalid", f"{label}绑定不能为空", status_code=503)
        paths: list[Path] = []
        for asset_id in asset_ids:
            paths.extend(self._asset_paths(manifest, str(asset_id or "").strip(), label))
        return paths

    def _asset_paths(self, manifest: dict[str, Any], asset_id: str, label: str) -> list[Path]:
        assets = manifest["assets"]
        if not asset_id or asset_id not in assets:
            raise ProfessionalSkillError("skill_runtime_invalid", f"{label}引用了未登记资产", status_code=503)
        value = assets[asset_id]
        values = value if isinstance(value, list) else [value]
        return [(self.project_root / str(path)).resolve() for path in values]

    def _knowledge_sources(self, roots: list[Path]) -> tuple[Path, ...]:
        sources: list[Path] = []
        for root in roots:
            if root.is_dir():
                sources.extend(
                    path.resolve()
                    for path in root.rglob("*")
                    if path.is_file() and path.suffix.lower() in KNOWLEDGE_SOURCE_SUFFIXES and not path.name.startswith("~$")
                )
            elif root.suffix.lower() in KNOWLEDGE_SOURCE_SUFFIXES:
                sources.append(root.resolve())
        unique = tuple(dict.fromkeys(sources))
        if not unique:
            raise ProfessionalSkillError("skill_runtime_invalid", "专业能力知识源为空", status_code=503)
        return unique

    def _relative_asset_path(self, path: Path) -> str:
        return path.resolve().relative_to(self.project_root).as_posix()

    def _resolve_frozen_path(self, value: object) -> Path:
        relative = Path(str(value or ""))
        if not str(value or "").strip() or relative.is_absolute() or ".." in relative.parts:
            raise ProfessionalSkillError("skill_runtime_invalid", "任务专业能力资产引用不安全", status_code=409)
        try:
            resolved = (self.project_root / relative).resolve(strict=True)
            resolved.relative_to(self.project_root)
        except (OSError, ValueError) as exc:
            raise ProfessionalSkillError(
                "skill_runtime_unavailable",
                "任务专业能力资产已不可用",
                status_code=409,
            ) from exc
        if resolved.suffix.lower() in EXECUTABLE_SUFFIXES:
            raise ProfessionalSkillError("skill_runtime_invalid", "任务专业能力资产引用不安全", status_code=409)
        return resolved

    def _load_all(self) -> list[dict[str, Any]]:
        if not self.skills_root.is_dir():
            raise ProfessionalSkillError("skill_registry_unavailable", "专业能力清单暂不可用", status_code=503)
        manifests = [self._load_manifest(path) for path in sorted(self.skills_root.glob("*/manifest.json"))]
        if not manifests:
            raise ProfessionalSkillError("skill_registry_empty", "当前没有可展示的专业能力", status_code=503)
        ids = [str(item["id"]) for item in manifests]
        if len(ids) != len(set(ids)):
            raise ProfessionalSkillError("skill_manifest_invalid", "专业能力清单存在重复标识", status_code=503)
        return sorted(manifests, key=lambda item: (int(item.get("displayOrder", 1000)), item["displayName"]))

    def _load_manifest(self, manifest_path: Path) -> dict[str, Any]:
        try:
            raw_text = manifest_path.read_text(encoding="utf-8")
            manifest = json.loads(raw_text)
        except (OSError, json.JSONDecodeError) as exc:
            raise ProfessionalSkillError("skill_manifest_invalid", "专业能力清单格式无效", status_code=503) from exc
        if not isinstance(manifest, dict):
            raise ProfessionalSkillError("skill_manifest_invalid", "专业能力清单必须是对象", status_code=503)
        missing = sorted(REQUIRED_FIELDS - manifest.keys())
        if missing:
            raise ProfessionalSkillError("skill_manifest_invalid", f"专业能力清单缺少字段：{', '.join(missing)}", status_code=503)
        self._validate_metadata(manifest, manifest_path)
        self._reject_unsafe_entries(manifest)
        self._validate_assets(manifest)
        canonical = json.dumps(manifest, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        loaded = {**manifest, "_manifest_hash": hashlib.sha256(canonical.encode("utf-8")).hexdigest()}
        if manifest["status"] in {"active", "beta"}:
            self.create_runtime_context(loaded)
        return loaded

    def _validate_metadata(self, manifest: dict[str, Any], manifest_path: Path) -> None:
        skill_id = str(manifest.get("id") or "").strip()
        if not SKILL_ID_PATTERN.fullmatch(skill_id) or manifest_path.parent.name != skill_id:
            raise ProfessionalSkillError("skill_manifest_invalid", "专业能力 ID 与目录不一致", status_code=503)
        if not str(manifest.get("displayName") or "").strip():
            raise ProfessionalSkillError("skill_manifest_invalid", "专业能力名称不能为空", status_code=503)
        if not SEMVER_PATTERN.fullmatch(str(manifest.get("version") or "")):
            raise ProfessionalSkillError("skill_manifest_invalid", "专业能力版本必须使用 x.y.z", status_code=503)
        if manifest.get("status") not in ALLOWED_STATUSES:
            raise ProfessionalSkillError("skill_manifest_invalid", "专业能力状态无效", status_code=503)
        if "statusLabel" in manifest and not str(manifest.get("statusLabel") or "").strip():
            raise ProfessionalSkillError("skill_manifest_invalid", "专业能力状态名称不能为空", status_code=503)
        if "displayOrder" in manifest and not isinstance(manifest.get("displayOrder"), int):
            raise ProfessionalSkillError("skill_manifest_invalid", "专业能力展示顺序必须是整数", status_code=503)
        for field in ("inputProfile", "capabilities", "assets", "validation"):
            if not isinstance(manifest.get(field), dict):
                raise ProfessionalSkillError("skill_manifest_invalid", f"专业能力字段 {field} 必须是对象", status_code=503)
        recommendation = manifest.get("recommendationProfile", {})
        if not isinstance(recommendation, dict):
            raise ProfessionalSkillError("skill_manifest_invalid", "专业能力推荐特征必须是对象", status_code=503)
        for field in ("fileNameKeywords", "sheetKeywords", "headerKeywords"):
            values = recommendation.get(field, [])
            if not isinstance(values, list) or any(not isinstance(value, str) or not value.strip() for value in values):
                raise ProfessionalSkillError("skill_manifest_invalid", f"专业能力推荐字段 {field} 必须是非空字符串数组", status_code=503)
        self._validate_sub_skills(manifest.get("subSkills", []))

    @staticmethod
    def _validate_sub_skills(sub_skills: object) -> None:
        if not isinstance(sub_skills, list):
            raise ProfessionalSkillError("skill_manifest_invalid", "专业能力子 Skill 必须是数组", status_code=503)
        names: set[str] = set()
        for item in sub_skills:
            if not isinstance(item, dict):
                raise ProfessionalSkillError("skill_manifest_invalid", "专业能力子 Skill 格式无效", status_code=503)
            name = str(item.get("name") or "").strip()
            description = str(item.get("description") or "").strip()
            if not name or not description:
                raise ProfessionalSkillError("skill_manifest_invalid", "专业能力子 Skill 名称和说明不能为空", status_code=503)
            if name in names:
                raise ProfessionalSkillError("skill_manifest_invalid", "专业能力子 Skill 名称不得重复", status_code=503)
            if item.get("type") not in SUB_SKILL_TYPES or item.get("status") not in SUB_SKILL_STATUSES:
                raise ProfessionalSkillError("skill_manifest_invalid", "专业能力子 Skill 类型或状态无效", status_code=503)
            names.add(name)

    def _reject_unsafe_entries(self, value: object, *, key: str = "") -> None:
        normalized_key = re.sub(r"[^a-z0-9]", "", key.lower())
        if normalized_key in PROHIBITED_KEYS:
            raise ProfessionalSkillError("skill_manifest_unsafe", "专业能力清单不得声明脚本或可执行入口", status_code=503)
        if normalized_key in SECRET_KEYS:
            raise ProfessionalSkillError("skill_manifest_unsafe", "专业能力清单不得保存运行秘密", status_code=503)
        if isinstance(value, dict):
            for child_key, child_value in value.items():
                self._reject_unsafe_entries(child_value, key=str(child_key))
        elif isinstance(value, list):
            for child_value in value:
                self._reject_unsafe_entries(child_value, key=key)

    def _validate_assets(self, manifest: dict[str, Any]) -> None:
        assets = manifest["assets"]
        if manifest["status"] in {"active", "beta"} and not assets:
            raise ProfessionalSkillError("skill_manifest_invalid", "上线专业能力必须声明资产", status_code=503)
        for value in assets.values():
            paths = value if isinstance(value, list) else [value]
            if not paths or any(not isinstance(path, str) or not path.strip() for path in paths):
                raise ProfessionalSkillError("skill_manifest_invalid", "专业能力资产引用格式无效", status_code=503)
            for path_text in paths:
                self._validate_asset_path(path_text)

    def _validate_asset_path(self, path_text: str) -> None:
        relative = Path(path_text)
        if relative.is_absolute() or ".." in relative.parts or relative.suffix.lower() in EXECUTABLE_SUFFIXES:
            raise ProfessionalSkillError("skill_asset_unsafe", "专业能力资产引用不安全", status_code=503)
        try:
            resolved = (self.project_root / relative).resolve(strict=True)
            resolved.relative_to(self.project_root)
        except (OSError, ValueError) as exc:
            raise ProfessionalSkillError("skill_asset_unavailable", "专业能力所需资产缺失或超出允许范围", status_code=503) from exc

    @staticmethod
    def _asset_count(assets: dict[str, object]) -> int:
        return sum(len(value) if isinstance(value, list) else 1 for value in assets.values())

    @staticmethod
    def _capability_names(capabilities: dict[str, object]) -> list[str]:
        labels = {
            "pricing": "三数字匹配",
            "workloadCapture": "工作量抓取",
            "experienceWarning": "经验池预警",
            "knowledgeQa": "知识库问答",
            "knowledgeMemory": "知识记忆",
            "wordReport": "Word 报告",
            "collaboration": "智能协同",
        }
        return [labels.get(key, key) for key, enabled in capabilities.items() if enabled]

    @staticmethod
    def _asset_summary(assets: dict[str, object]) -> list[dict[str, object]]:
        return [
            {
                "id": key,
                "name": ASSET_LABELS.get(key, key),
                "count": len(value) if isinstance(value, list) else 1,
            }
            for key, value in assets.items()
        ]

    @staticmethod
    def _compatibility_snapshot(skill_id: str) -> dict[str, object]:
        return {
            "id": skill_id,
            "display_name": "勘察测量最高投标限价编制",
            "version": "compatibility",
            "manifest_hash": "compatibility-default-survey-measurement",
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "asset_ids": [],
            "compatibility_fallback": True,
        }
