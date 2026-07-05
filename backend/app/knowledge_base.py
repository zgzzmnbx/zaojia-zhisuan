from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .normalization import normalize_key_part, normalize_price
from .schemas import FIELD_COLUMNS, MatchResult

ELEMENT_COLUMNS = FIELD_COLUMNS[:5]
EXPERIENCE_PHYSICAL_FIELD = "【经验数】实物工作费调整系数"
EXPERIENCE_PHYSICAL_NOTE_FIELD = "【经验数解释】-实物工作费调整系数"
EXPERIENCE_TECHNICAL_FIELD = "【经验数】技术工作费调整系数"
EXPERIENCE_TECHNICAL_NOTE_FIELD = "【经验数解释】-技术工作费调整系数"
EXPERIENCE_FIELDS = (
    EXPERIENCE_PHYSICAL_FIELD,
    EXPERIENCE_PHYSICAL_NOTE_FIELD,
    EXPERIENCE_TECHNICAL_FIELD,
    EXPERIENCE_TECHNICAL_NOTE_FIELD,
)


class KnowledgeBase:
    def __init__(self, rows: list[dict[str, Any]]) -> None:
        self.rows = rows
        self._index: dict[tuple[str, ...], list[dict[str, Any]]] = defaultdict(list)
        self._ordered_index: dict[tuple[str, tuple[str, ...]], list[dict[str, Any]]] = defaultdict(list)
        for row in rows:
            self._index[self.make_key(row)].append(row)
            self._ordered_index[self.make_ordered_key(row)].append(row)

    @classmethod
    def from_excel(cls, path: str | Path) -> "KnowledgeBase":
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.worksheets[0]
        row_iter = sheet.iter_rows(max_col=13, values_only=True)
        headers = list(next(row_iter))
        header_map: dict[str, int] = {}
        for idx, name in enumerate(headers, start=1):
            text = str(name).strip() if name else ""
            if text and text not in header_map:
                header_map[text] = idx

        missing = [name for name in [*FIELD_COLUMNS, "基价"] if name not in header_map]
        if missing:
            raise ValueError(f"知识库缺少必要列：{', '.join(missing)}")

        rows: list[dict[str, Any]] = []
        for excel_row, values in enumerate(row_iter, start=2):
            item = {
                name: values[header_map[name] - 1] if header_map[name] - 1 < len(values) else None
                for name in FIELD_COLUMNS
            }
            price_index = header_map["基价"] - 1
            item["基价"] = normalize_price(values[price_index] if price_index < len(values) else None)
            for field_name in EXPERIENCE_FIELDS:
                field_index = header_map.get(field_name)
                item[field_name] = (
                    values[field_index - 1]
                    if field_index is not None and field_index - 1 < len(values)
                    else None
                )
            item["_excel_row"] = excel_row
            if item["基价"] is not None:
                rows.append(item)
        workbook.close()
        return cls(rows)

    @staticmethod
    def make_key(row: dict[str, Any]) -> tuple[str, ...]:
        return tuple(normalize_key_part(row.get(name)) for name in FIELD_COLUMNS)

    @classmethod
    def make_ordered_key(cls, row: dict[str, Any]) -> tuple[str, tuple[str, ...]]:
        unit = normalize_key_part(row.get("单位"))
        elements = tuple(
            cls.canonical_key_part(row.get(name))
            for name in ELEMENT_COLUMNS
            if cls.canonical_key_part(row.get(name))
        )
        return unit, elements

    @staticmethod
    def canonical_key_part(value: Any) -> str:
        text = normalize_key_part(value)
        if text.startswith("级别-"):
            return text.removeprefix("级别-")
        if text.startswith("比例-"):
            return text.removeprefix("比例-")
        return text

    def lookup(self, row: dict[str, Any]) -> MatchResult:
        exact_candidates = self._index.get(self.make_key(row), [])
        ordered_candidates = self._ordered_index.get(self.make_ordered_key(row), [])
        candidates = exact_candidates or ordered_candidates
        if not candidates:
            return MatchResult("not_found", None, self._build_not_found_message(row))

        prices = {candidate["基价"] for candidate in candidates}
        if len(prices) == 1:
            candidate = candidates[0]
            match_type = "字段完全匹配" if exact_candidates else "要素顺序匹配"
            return MatchResult(
                "matched",
                candidate["基价"],
                f"{match_type}，匹配知识库第 {candidate['_excel_row']} 行",
                candidates,
            )

        return MatchResult("conflict", None, f"匹配到 {len(candidates)} 条记录但价格不一致", candidates)

    def _build_not_found_message(self, row: dict[str, Any]) -> str:
        unit, elements = self.make_ordered_key(row)
        same_unit_rows = [
            candidate
            for candidate in self.rows
            if normalize_key_part(candidate.get("单位")) == unit
        ]
        if not same_unit_rows:
            return "未找到单位匹配的知识库记录"

        best_prefix = 0
        for candidate in same_unit_rows:
            _, candidate_elements = self.make_ordered_key(candidate)
            prefix = 0
            for input_part, candidate_part in zip(elements, candidate_elements):
                if input_part != candidate_part:
                    break
                prefix += 1
            best_prefix = max(best_prefix, prefix)

        if best_prefix == 0:
            return "要素1开始没有匹配"
        return f"要素1至要素{best_prefix}匹配上，但是要素{best_prefix + 1}开始没有匹配"
