from __future__ import annotations

import csv
import hashlib
import json
import math
import re
import time
from collections import Counter
from dataclasses import asdict, dataclass, field, replace
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .paths import DEFAULT_KNOWLEDGE_QA_INDEX_PATH, PROJECT_DEFAULT_SETTINGS_PATH, PROJECT_ROOT


NO_EVIDENCE_ANSWER = "当前知识库未找到明确依据，需要人工复核。"
ASSISTANT_TABLE_FORMAT_RULE = (
    "回答格式规则：当用户明确要求表格，或回答包含一个对象的三个及以上字段，"
    "或包含多个对象的对比、清单、数字、价格、系数、单位、状态、来源时，优先使用标准 Markdown 表格；"
    "表格必须包含表头和分隔行，不使用 HTML 表格。"
    "单一结论、连续解释、风险警告和操作步骤不强行表格化，必要时可在表格后补充简短说明。"
)
DEFAULT_INDEX_PATH = DEFAULT_KNOWLEDGE_QA_INDEX_PATH
FORCE_KNOWLEDGE_PREFIXES = ("查库：", "查库:", "#知识库")
KNOWLEDGE_INDEX_VERSION = "2026-07-28-library-selector-v2"
HYBRID_INDEX_VERSION = "2026-08-06-hybrid-rag-v2"
LOCAL_EMBEDDING_MODEL = "local-hash-embedding-v1"
LOCAL_RERANK_MODEL = "local-rule-rerank-v1"
LOCAL_EMBEDDING_DIMENSIONS = 384
HYBRID_RRF_K = 60
DEFAULT_HYBRID_CONFIG: dict[str, Any] = {
    "defaultMode": "classic",
    "hybridEnabled": True,
    "embedding": {
        "provider": "local-hash",
        "model": LOCAL_EMBEDDING_MODEL,
        "dimensions": LOCAL_EMBEDDING_DIMENSIONS,
    },
    "rerank": {
        "enabled": True,
        "provider": "local-rule",
        "model": LOCAL_RERANK_MODEL,
        "timeoutMs": 180,
    },
    "limits": {
        "bm25TopK": 40,
        "structuredTopK": 40,
        "vectorTopK": 40,
        "candidateTopK": 60,
    },
}
_HYBRID_INDEX_MEMORY_CACHE: dict[tuple[str, str, str], list["KnowledgeChunk"]] = {}

COMMON_STOP_TERMS = {
    "什么",
    "什么意思",
    "为什么",
    "哪里来的",
    "哪来的",
    "依据",
    "标准",
    "解释",
    "来源",
    "出处",
    "这个",
    "这一行",
    "本行",
    "一般",
    "多少",
    "多少钱",
}

SYNONYM_RULES: tuple[tuple[tuple[str, ...], tuple[str, ...]], ...] = (
    (("0.22", "22%", "技术工作费"), ("0.22", "22%", "技术工作费", "工程测量技术工作费", "收费比例")),
    (("0.6", "60%"), ("0.6", "60%", "附加调整系数", "实物工作费调整系数", "不造标")),
    (("1.3", "130%"), ("1.3", "130%", "附加调整系数", "实物工作费调整系数")),
    (("1.5", "150%"), ("1.5", "150%", "附加调整系数", "实物工作费调整系数")),
    (
        ("实物工作费", "实物工作系数", "实物工作费系数", "实物系数", "工作费系数", "附加调整系数"),
        ("实物工作费", "实物工作费调整系数", "附加调整系数", "工程勘察", "工程测量"),
    ),
    (("技术工作费", "技术系数"), ("技术工作费", "技术工作费调整系数", "工程测量技术工作费", "收费比例")),
    (("不能连乘", "连乘", "相乘"), ("不能连乘", "连乘", "附加调整系数", "总则", "1.0.8", "相加")),
    (("第二层", "经验提示", "标黄", "黄色"), ("第二层", "经验提示", "经验数", "标黄", "黄色")),
    (("待复核", "标红", "红色"), ("待复核", "标红", "红色", "未命中", "人工复核")),
    (("预警", "经验池"), ("预警", "经验池预警", "偏离率", "阈值", "同类记录")),
    (("风险报告", "审查摘要", "输出风险报告", "生成风险报告"), ("风险报告", "审查摘要", "Word报告", "知识库依据", "处理结论", "主要风险", "复核建议")),
    (("问问智算", "智算模式", "#知识库", "查库", "强制知识库"), ("问问智算", "强制知识库", "快捷指令", "自动知识库问答", "普通自由问答", "行级AI复核", "风险报告")),
    (("导出", "下载", "输出excel", "输出word"), ("导出", "下载", "Excel", "Word", "原始输出", "大模型", "结构化规则引擎")),
    (("行级AI", "行级复核", "当前行复核"), ("行级AI复核", "当前行上下文", "匹配状态", "匹配说明", "预警参数", "预警细节")),
    (("地形测量", "地形图测绘"), ("地形测量", "地形图测绘", "地形图测绘(地形测量)")),
    (("首级控制", "首级控制测量"), ("首级控制", "首级控制测量", "控制测量")),
    (("GPS E级", "GPS测量E级"), ("GPS测量E级", "GPS测量", "E级")),
)

KNOWN_PHRASES = tuple(
    sorted(
        {
            term
            for triggers, expansions in SYNONYM_RULES
            for term in (*triggers, *expansions)
        }
        | {
            "基价",
            "单价",
            "要素1",
            "要素2",
            "要素3",
            "要素4",
            "要素5",
            "单位",
            "字段完全匹配",
            "非空要素顺序匹配",
            "技术工作费调整系数",
            "实物工作费调整系数",
        },
        key=len,
        reverse=True,
    )
)


@dataclass(frozen=True)
class KnowledgeChunk:
    id: str
    source_file: str
    source_type: str
    title_path: str
    content: str
    keywords: list[str]
    module: str
    created_at: str
    authority_level: str = "project_rule"
    library_id: str = ""
    content_hash: str = ""
    updated_at: str = ""
    lexical_terms: list[str] = field(default_factory=list)
    structured_fields: dict[str, str] = field(default_factory=dict)
    vector: list[list[float]] = field(default_factory=list)
    metadata: dict[str, Any] = field(default_factory=dict)


@dataclass(frozen=True)
class KnowledgeSearchResult:
    id: str
    source_file: str
    source_type: str
    title_path: str
    snippet: str
    score: float
    module: str
    authority_level: str = ""
    library_id: str = ""
    channels: tuple[str, ...] = ()
    metadata: dict[str, Any] = field(default_factory=dict)


@dataclass(frozen=True)
class HybridSearchResponse:
    results: list[KnowledgeSearchResult]
    trace: dict[str, Any]


class HybridRetrievalError(RuntimeError):
    """Hybrid index/engine failure. Callers must fall back without changing evidence rules."""


@dataclass(frozen=True)
class KnowledgeQuestionParts:
    search_question: str
    answer_requirements: str
    meta_requirements: str


_INLINE_ANSWER_INSTRUCTION_RE = re.compile(
    r"(?=[，,；;]\s*(?:请用|请按|只回答|只返回|不要|不得|回答必须|在\d+字以内))"
)


def split_knowledge_question(question: str) -> KnowledgeQuestionParts:
    """把业务检索内容与回答格式、附加说明要求分开，避免格式词污染召回。"""
    clean_question, _forced = strip_force_knowledge_prefix(question)
    segments: list[str] = []
    for sentence in re.findall(r"[^。！？!?]+[。！？!?]?", clean_question):
        segments.extend(_INLINE_ANSWER_INSTRUCTION_RE.split(sentence))

    search_segments: list[str] = []
    answer_segments: list[str] = []
    meta_segments: list[str] = []
    for segment in segments:
        clean_segment = segment.strip(" \t\r\n，,；;。！？!?")
        if not clean_segment:
            continue
        normalized = _normalize_text(clean_segment)
        if re.search(r"(?:给我|向我)?讲述(?:一下)?(?:检索|回答)?原理|说明(?:检索|回答)过程", normalized):
            meta_segments.append(clean_segment)
            continue
        if (
            re.match(r"^(?:仅检索|限定范围)", clean_segment)
            or re.match(r"^(?:请)?(?:只回答|只返回|不要|不得|回答必须)", clean_segment)
            or re.match(r"^(?:请)?(?:用|使用|按|只用).{0,30}(?:表格|markdown|字数|格式|字段|回答|输出|整理)", clean_segment, flags=re.IGNORECASE)
            or re.match(r"^(?:请)?在\d+字以内", clean_segment)
        ):
            answer_segments.append(clean_segment)
            continue
        search_segments.append(clean_segment)

    search_question = " ".join(search_segments).strip() or clean_question
    return KnowledgeQuestionParts(
        search_question=search_question,
        answer_requirements="；".join(answer_segments),
        meta_requirements="；".join(meta_segments),
    )


def is_knowledge_question(question: str) -> bool:
    clean = _normalize_text(question)
    if not clean:
        return False
    triggers = (
        "哪里来的",
        "哪来的",
        "依据",
        "标准",
        "为什么",
        "什么意思",
        "解释",
        "来源",
        "出处",
        "0.22",
        "22%",
        "0.6",
        "1.3",
        "1.5",
        "技术工作费",
        "实物工作费",
        "实物工作系数",
        "实物工作费系数",
        "实物系数",
        "附加调整系数",
        "经验提示",
        "第二层",
        "待复核",
        "预警",
        "不能连乘",
        "风险报告",
        "审查摘要",
        "问问智算",
        "强制知识库",
        "行级ai",
        "行级复核",
    )
    return any(trigger in clean for trigger in triggers)


def strip_force_knowledge_prefix(question: str) -> tuple[str, bool]:
    clean_question = str(question or "").strip()
    for prefix in FORCE_KNOWLEDGE_PREFIXES:
        if clean_question.startswith(prefix):
            stripped = clean_question[len(prefix) :].lstrip(" \t\r\n:：,，.。;；")
            return stripped, True
    return clean_question, False


def search_knowledge(
    question: str,
    row_context: dict[str, Any] | None = None,
    limit: int = 8,
    *,
    project_root: Path = PROJECT_ROOT,
    index_path: Path | None = DEFAULT_INDEX_PATH,
    sources: list[Path] | tuple[Path, ...] | None = None,
) -> list[KnowledgeSearchResult]:
    clean_question = question.strip()
    if not clean_question:
        return []
    chunks = load_or_build_index(project_root=project_root, index_path=index_path, sources=sources)
    query_terms = _expand_query_terms(clean_question, row_context)
    if not query_terms:
        return []
    scored: list[tuple[float, KnowledgeChunk]] = []
    for chunk in chunks:
        score = _score_chunk(chunk, query_terms)
        if score >= 3:
            scored.append((score, chunk))
    scored.sort(key=lambda item: item[0], reverse=True)
    results: list[KnowledgeSearchResult] = []
    for score, chunk in scored[: max(1, min(limit, 20))]:
        results.append(
            KnowledgeSearchResult(
                id=chunk.id,
                source_file=chunk.source_file,
                source_type=chunk.source_type,
                title_path=chunk.title_path,
                snippet=_build_snippet(chunk.content, query_terms),
                score=round(score, 3),
                module=chunk.module,
            )
        )
    return results


def load_knowledge_retrieval_config(
    config_path: Path = PROJECT_DEFAULT_SETTINGS_PATH,
) -> dict[str, Any]:
    """读取知识问答检索配置；配置异常时回退到离线安全默认值。"""
    config: dict[str, Any] = json.loads(json.dumps(DEFAULT_HYBRID_CONFIG))
    try:
        payload = json.loads(config_path.read_text(encoding="utf-8"))
    except (OSError, TypeError, ValueError):
        return config
    section = payload.get("knowledgeRetrieval") if isinstance(payload, dict) else None
    if not isinstance(section, dict):
        return config
    for key in ("defaultMode", "hybridEnabled"):
        if key in section:
            config[key] = section[key]
    for key in ("embedding", "rerank", "limits"):
        value = section.get(key)
        if isinstance(value, dict):
            config[key].update(value)
    config["defaultMode"] = "hybrid" if str(config.get("defaultMode")) == "hybrid" else "classic"
    config["hybridEnabled"] = bool(config.get("hybridEnabled", True))
    embedding = config["embedding"]
    embedding["provider"] = str(embedding.get("provider") or "local-hash")
    embedding["model"] = str(embedding.get("model") or LOCAL_EMBEDDING_MODEL)
    embedding["dimensions"] = LOCAL_EMBEDDING_DIMENSIONS
    rerank = config["rerank"]
    rerank["enabled"] = bool(rerank.get("enabled", True))
    rerank["provider"] = str(rerank.get("provider") or "local-rule")
    rerank["model"] = str(rerank.get("model") or LOCAL_RERANK_MODEL)
    try:
        rerank["timeoutMs"] = max(20, int(rerank.get("timeoutMs", 180)))
    except (TypeError, ValueError):
        rerank["timeoutMs"] = 180
    return config


def hybrid_index_path(index_path: Path | None = DEFAULT_INDEX_PATH) -> Path | None:
    if index_path is None:
        return None
    return index_path.with_name(f"{index_path.stem}-hybrid{index_path.suffix}")


def knowledge_retrieval_capabilities(
    *,
    project_root: Path = PROJECT_ROOT,
    index_path: Path | None = DEFAULT_INDEX_PATH,
    sources: list[Path] | tuple[Path, ...] | None = None,
    config_path: Path = PROJECT_DEFAULT_SETTINGS_PATH,
) -> dict[str, Any]:
    """返回真实的离线检索能力状态，不把声明的配置当成已就绪索引。"""
    config = load_knowledge_retrieval_config(config_path)
    source_paths = list(sources) if sources is not None else _discover_sources(project_root)
    target = hybrid_index_path(index_path)
    status = "not_built"
    if not source_paths:
        status = "no_sources"
    elif target and target.exists():
        try:
            payload = json.loads(target.read_text(encoding="utf-8"))
            expected = _hybrid_source_signature(source_paths, project_root)
            if payload.get("hybrid_index_version") != HYBRID_INDEX_VERSION:
                status = "stale"
            elif payload.get("source_signature") != expected:
                status = "stale"
            else:
                status = "ready"
        except (OSError, TypeError, ValueError):
            status = "invalid"
    elif target is None:
        status = "memory_only"
    return {
        "engine": "hybrid-rag",
        "default_mode": config["defaultMode"],
        "hybrid_enabled": bool(config["hybridEnabled"]),
        "available": bool(config["hybridEnabled"] and source_paths),
        "index_ready": status in {"ready", "memory_only"},
        "index_status": status,
        "source_count": len(source_paths),
        "index_path": str(target) if target else None,
        "offline": True,
        "embedding": {
            "provider": config["embedding"]["provider"],
            "model": config["embedding"]["model"],
            "dimensions": config["embedding"]["dimensions"],
            "dependency": "stdlib-only deterministic hash embedding; no external model",
        },
        "rerank": {
            "enabled": bool(config["rerank"]["enabled"]),
            "provider": config["rerank"]["provider"],
            "model": config["rerank"]["model"],
            "timeout_ms": config["rerank"]["timeoutMs"],
        },
    }


def load_or_build_hybrid_index(
    *,
    project_root: Path = PROJECT_ROOT,
    index_path: Path | None = DEFAULT_INDEX_PATH,
    sources: list[Path] | tuple[Path, ...] | None = None,
    source_library_ids: dict[str, str] | None = None,
    rebuild_corrupt: bool = False,
) -> list[KnowledgeChunk]:
    source_paths = list(sources) if sources is not None else _discover_sources(project_root)
    source_signature = _hybrid_source_signature(source_paths, project_root)
    target = hybrid_index_path(index_path)
    memory_key = _hybrid_memory_cache_key(target, source_signature)
    cached_chunks = _HYBRID_INDEX_MEMORY_CACHE.get(memory_key)
    if cached_chunks is not None:
        return cached_chunks
    if target and target.exists():
        try:
            payload = json.loads(target.read_text(encoding="utf-8"))
        except (OSError, TypeError, ValueError) as exc:
            if not rebuild_corrupt:
                raise HybridRetrievalError("hybrid_vector_index_corrupt: invalid JSON") from exc
            payload = None
        if payload is not None:
            if payload.get("hybrid_index_version") != HYBRID_INDEX_VERSION:
                payload = None
            elif payload.get("source_signature") != source_signature:
                payload = None
            else:
                try:
                    chunks = [KnowledgeChunk(**item) for item in payload.get("chunks", [])]
                    if not chunks or any("vector" not in item for item in payload.get("chunks", [])):
                        raise ValueError("missing vector payload")
                    _HYBRID_INDEX_MEMORY_CACHE[_hybrid_memory_cache_key(target, source_signature)] = chunks
                    return chunks
                except (KeyError, TypeError, ValueError) as exc:
                    if not rebuild_corrupt:
                        raise HybridRetrievalError("hybrid_vector_index_corrupt: invalid chunk payload") from exc
                    payload = None
    chunks = build_hybrid_index(
        project_root=project_root,
        sources=source_paths,
        source_library_ids=source_library_ids,
    )
    if target:
        try:
            target.parent.mkdir(parents=True, exist_ok=True)
            target.write_text(
                json.dumps(
                    {
                        "hybrid_index_version": HYBRID_INDEX_VERSION,
                        "built_at": datetime.now().isoformat(timespec="seconds"),
                        "source_signature": source_signature,
                        "embedding": {
                            "provider": "local-hash",
                            "model": LOCAL_EMBEDDING_MODEL,
                            "dimensions": LOCAL_EMBEDDING_DIMENSIONS,
                            "license": "project code; no external model artifact",
                        },
                        "reranker": {
                            "provider": "local-rule",
                            "model": LOCAL_RERANK_MODEL,
                            "license": "project code; no external model artifact",
                        },
                        "chunks": [asdict(chunk) for chunk in chunks],
                    },
                    ensure_ascii=False,
                    indent=2,
                ),
                encoding="utf-8",
            )
        except OSError as exc:
            raise HybridRetrievalError(f"hybrid_vector_index_write_failed: {exc}") from exc
        _HYBRID_INDEX_MEMORY_CACHE[_hybrid_memory_cache_key(target, source_signature)] = chunks
    else:
        _HYBRID_INDEX_MEMORY_CACHE[memory_key] = chunks
    return chunks


def _hybrid_memory_cache_key(
    target: Path | None,
    source_signature: list[dict[str, object]],
) -> tuple[str, str, str]:
    target_key = str(target.resolve()) if target else "<memory>"
    file_state = ""
    if target:
        try:
            stat = target.stat()
            file_state = f"{stat.st_mtime_ns}:{stat.st_size}"
        except OSError:
            file_state = "missing"
    signature_key = json.dumps(source_signature, ensure_ascii=False, sort_keys=True)
    return target_key, file_state, signature_key


def rebuild_hybrid_index(
    *,
    project_root: Path = PROJECT_ROOT,
    index_path: Path | None = DEFAULT_INDEX_PATH,
    sources: list[Path] | tuple[Path, ...] | None = None,
    source_library_ids: dict[str, str] | None = None,
) -> list[KnowledgeChunk]:
    """显式重建混合索引；运行请求不会静默修复损坏索引。"""
    return load_or_build_hybrid_index(
        project_root=project_root,
        index_path=index_path,
        sources=sources,
        source_library_ids=source_library_ids,
        rebuild_corrupt=True,
    )


def build_hybrid_index(
    *,
    project_root: Path = PROJECT_ROOT,
    sources: list[Path] | tuple[Path, ...] | None = None,
    source_library_ids: dict[str, str] | None = None,
) -> list[KnowledgeChunk]:
    base_chunks = build_index(project_root=project_root, sources=list(sources) if sources is not None else None)
    return [
        _enrich_hybrid_chunk(chunk, project_root, source_library_ids or {})
        for chunk in base_chunks
    ]


def _enrich_hybrid_chunk(
    chunk: KnowledgeChunk,
    project_root: Path,
    source_library_ids: dict[str, str],
) -> KnowledgeChunk:
    library_id = source_library_ids.get(chunk.source_file.casefold())
    if not library_id:
        library_id = "cost-aiw" if "06-知识库问答资料/造价AIW资料库" in chunk.source_file else "project-core"
    authority_level = _authority_level(chunk.source_type, chunk.source_file)
    source_path = project_root / chunk.source_file
    try:
        updated_at = datetime.fromtimestamp(source_path.stat().st_mtime).isoformat(timespec="seconds")
    except OSError:
        updated_at = chunk.created_at
    lexical_terms = _bm25_tokens(" ".join([chunk.title_path, chunk.content, " ".join(chunk.keywords)]))
    structured_fields = _extract_structured_fields(chunk.content)
    vector = [[float(bucket), float(weight)] for bucket, weight in _hash_embedding(chunk.title_path + "\n" + chunk.content)]
    metadata = {
        **(chunk.metadata or {}),
        "library_id": library_id,
        "authority_level": authority_level,
        "source_kind": chunk.source_type,
        "project_key": "project-core" if library_id == "project-core" else "",
        "content_hash": hashlib.sha256(chunk.content.encode("utf-8", errors="ignore")).hexdigest(),
        "updated_at": updated_at,
        "embedding_model_version": LOCAL_EMBEDDING_MODEL,
    }
    return replace(
        chunk,
        authority_level=authority_level,
        library_id=library_id,
        content_hash=metadata["content_hash"],
        updated_at=updated_at,
        lexical_terms=lexical_terms,
        structured_fields=structured_fields,
        vector=vector,
        metadata=metadata,
    )


def _authority_level(source_type: str, source_file: str) -> str:
    if source_type == "standard":
        return "formal_standard"
    if "【重要匹配规则】" in source_file:
        return "authoritative_rule"
    if source_type == "rule_card":
        return "structured_rule"
    if source_type == "project_rule":
        return "project_rule"
    return "general_reference"


def _hybrid_source_signature(sources: list[Path], project_root: Path) -> list[dict[str, object]]:
    signature: list[dict[str, object]] = [{"index_version": HYBRID_INDEX_VERSION}]
    for path in sources:
        try:
            stat = path.stat()
        except OSError:
            continue
        signature.append(
            {
                "path": _relative_path(path, project_root),
                "mtime": stat.st_mtime,
                "size": stat.st_size,
            }
        )
    return signature


def hybrid_search_knowledge(
    question: str,
    row_context: dict[str, Any] | None = None,
    limit: int = 8,
    *,
    project_root: Path = PROJECT_ROOT,
    index_path: Path | None = DEFAULT_INDEX_PATH,
    sources: list[Path] | tuple[Path, ...] | None = None,
    source_library_ids: dict[str, str] | None = None,
    vector_backend: str | None = None,
    rerank_backend: str | None = None,
    rerank_delay_ms: int = 0,
) -> HybridSearchResponse:
    """执行可追溯的混合检索；所有通道只返回正式索引中的原始知识块。"""
    clean_question = str(question or "").strip()
    if not clean_question:
        return HybridSearchResponse(results=[], trace=_empty_hybrid_trace("empty_question"))
    config = load_knowledge_retrieval_config()
    chunks = load_or_build_hybrid_index(
        project_root=project_root,
        index_path=index_path,
        sources=sources,
        source_library_ids=source_library_ids,
    )
    query_terms = _expand_query_terms(clean_question, row_context)
    if not query_terms:
        return HybridSearchResponse(results=[], trace=_empty_hybrid_trace("empty_query_terms"))
    constraints = _extract_hard_constraints(clean_question)
    relevance_anchors = _query_relevance_anchors(clean_question, row_context)
    limit_value = max(1, min(int(limit), 20))
    limits = config.get("limits", {})
    bm25_top_k = _safe_int(limits.get("bm25TopK"), 40, minimum=5, maximum=100)
    structured_top_k = _safe_int(limits.get("structuredTopK"), 40, minimum=5, maximum=100)
    vector_top_k = _safe_int(limits.get("vectorTopK"), 40, minimum=5, maximum=100)
    candidate_top_k = _safe_int(limits.get("candidateTopK"), 60, minimum=limit_value, maximum=200)

    eligible_chunks: list[KnowledgeChunk] = []
    hard_gate_rejected = 0
    for chunk in chunks:
        if _passes_hard_constraints(chunk, constraints) and _passes_relevance_gate(
            chunk,
            clean_question,
            constraints,
            row_context=row_context,
            anchors=relevance_anchors,
        ):
            eligible_chunks.append(chunk)
        else:
            hard_gate_rejected += 1

    query_tokens = _bm25_tokens(split_knowledge_question(clean_question).search_question)
    query_tokens.extend(_bm25_tokens(" ".join(query_terms.keys())))
    query_tokens = list(dict.fromkeys(query_tokens))
    bm25_scored = _bm25_scores(query_tokens, eligible_chunks)
    bm25_scored.sort(key=lambda item: item[0], reverse=True)
    bm25_scored = bm25_scored[:bm25_top_k]

    structured_scored = [
        (score, chunk)
        for chunk in eligible_chunks
        if (score := _structured_score(chunk, constraints, query_terms, query_text=clean_question)) > 0
    ]
    structured_scored.sort(key=lambda item: item[0], reverse=True)
    structured_scored = structured_scored[:structured_top_k]

    vector_error = ""
    vector_scored: list[tuple[float, KnowledgeChunk]] = []
    selected_vector_backend = vector_backend or str(config["embedding"].get("provider") or "local-hash")
    if selected_vector_backend in {"unavailable", "broken", "disabled"}:
        vector_error = "vector_backend_unavailable"
    else:
        try:
            query_vector = _hash_embedding(split_knowledge_question(clean_question).search_question)
            vector_scored = [
                (score, chunk)
                for chunk in eligible_chunks
                if (score := _cosine_similarity(query_vector, chunk.vector)) > 0
            ]
            vector_scored.sort(key=lambda item: item[0], reverse=True)
            vector_scored = vector_scored[:vector_top_k]
        except (TypeError, ValueError, ZeroDivisionError) as exc:
            vector_error = f"vector_channel_error:{type(exc).__name__}"

    channels: dict[str, list[tuple[float, KnowledgeChunk]]] = {
        "bm25": bm25_scored,
        "structured": structured_scored,
        "vector": vector_scored,
    }
    candidates: dict[str, dict[str, Any]] = {}
    for channel_name, scored in channels.items():
        for rank, (score, chunk) in enumerate(scored, start=1):
            dedupe_key = chunk.content_hash or chunk.id
            item = candidates.setdefault(
                dedupe_key,
                {
                    "chunk": chunk,
                    "fusion_score": 0.0,
                    "channel_scores": {},
                    "channels": set(),
                },
            )
            if _authority_rank(chunk.authority_level) > _authority_rank(item["chunk"].authority_level):
                item["chunk"] = chunk
            channel_weight = {"bm25": 1.0, "structured": 1.6, "vector": 0.8}[channel_name]
            item["fusion_score"] += channel_weight / (HYBRID_RRF_K + rank)
            item["channel_scores"][channel_name] = score
            item["channels"].add(channel_name)
    candidate_rows = sorted(
        candidates.values(),
        key=lambda item: (item["fusion_score"], _authority_rank(item["chunk"].authority_level)),
        reverse=True,
    )[:candidate_top_k]

    rerank_error = ""
    rerank_enabled = bool(config["rerank"].get("enabled", True))
    selected_rerank_backend = rerank_backend or str(config["rerank"].get("provider") or "local-rule")
    if rerank_enabled and candidate_rows:
        try:
            _apply_local_rerank(
                candidate_rows,
                backend=selected_rerank_backend,
                timeout_ms=_safe_int(config["rerank"].get("timeoutMs"), 180, minimum=20, maximum=5000),
                delay_ms=max(0, int(rerank_delay_ms)),
            )
        except (TimeoutError, RuntimeError, ValueError) as exc:
            rerank_error = str(exc) or f"rerank_channel_error:{type(exc).__name__}"
            candidate_rows.sort(
                key=lambda item: (item["fusion_score"], _authority_rank(item["chunk"].authority_level)),
                reverse=True,
            )

    result_rows: list[KnowledgeSearchResult] = []
    for item in candidate_rows[:limit_value]:
        chunk = item["chunk"]
        result_rows.append(
            KnowledgeSearchResult(
                id=chunk.id,
                source_file=chunk.source_file,
                source_type=chunk.source_type,
                title_path=chunk.title_path,
                snippet=_build_snippet(chunk.content, query_terms),
                score=round(float(item.get("final_score", item["fusion_score"])), 6),
                module=chunk.module,
                authority_level=chunk.authority_level,
                library_id=chunk.library_id,
                channels=tuple(sorted(item["channels"])),
                metadata={
                    "content_hash": chunk.content_hash,
                    "updated_at": chunk.updated_at,
                    "project_key": chunk.metadata.get("project_key", ""),
                },
            )
        )
    evidence_status = _hybrid_evidence_status(clean_question, result_rows)
    degradation_reasons: list[str] = []
    if vector_error:
        degradation_reasons.append(vector_error)
    if rerank_error:
        degradation_reasons.append(rerank_error)
    trace = {
        "requested_mode": "hybrid",
        "retrieval_mode_used": "hybrid",
        "fallback_mode": None,
        "fallback_reason": None,
        "degraded": bool(degradation_reasons),
        "degradation_reasons": degradation_reasons,
        "channels": {
            "bm25": {"available": True, "hits": len(bm25_scored), "top_k": bm25_top_k},
            "structured": {"available": True, "hits": len(structured_scored), "top_k": structured_top_k},
            "vector": {"available": not bool(vector_error), "hits": len(vector_scored), "top_k": vector_top_k, "error": vector_error or None},
        },
        "fusion": {
            "algorithm": "weighted reciprocal rank fusion",
            "rrf_k": HYBRID_RRF_K,
            "channel_weights": {"bm25": 1.0, "structured": 1.6, "vector": 0.8},
            "candidate_count": len(candidate_rows),
        },
        "rerank": {
            "enabled": rerank_enabled,
            "available": not bool(rerank_error),
            "algorithm": LOCAL_RERANK_MODEL,
            "error": rerank_error or None,
        },
        "hard_gate": {
            "constraints": constraints,
            "eligible_chunks": len(eligible_chunks),
            "rejected_chunks": hard_gate_rejected,
            "rule": "authority + exact number/code/scale/unit gates; no generated evidence",
        },
        "evidence_status": evidence_status,
        "evidence_sufficient": evidence_status in {"sufficient", "conflict"},
        "source_truth": "formal Markdown/Excel/CSV, rules and confirmed memory remain the source of truth",
    }
    return HybridSearchResponse(results=result_rows, trace=trace)


def _empty_hybrid_trace(reason: str) -> dict[str, Any]:
    return {
        "requested_mode": "hybrid",
        "retrieval_mode_used": "hybrid",
        "fallback_mode": None,
        "fallback_reason": None,
        "degraded": False,
        "degradation_reasons": [reason],
        "channels": {},
        "fusion": {"algorithm": "weighted reciprocal rank fusion", "rrf_k": HYBRID_RRF_K, "candidate_count": 0},
        "rerank": {"enabled": False, "available": False, "algorithm": LOCAL_RERANK_MODEL, "error": None},
        "hard_gate": {"constraints": {}, "eligible_chunks": 0, "rejected_chunks": 0},
        "evidence_status": "insufficient",
        "evidence_sufficient": False,
    }


def _bm25_tokens(text: str) -> list[str]:
    clean = _normalize_text(text)
    tokens: list[str] = []
    tokens.extend(re.findall(r"[a-z]+\d*[a-z\d]*|\d+(?:\.\d+)?%?", clean))
    for phrase in KNOWN_PHRASES:
        phrase_clean = _normalize_text(phrase)
        if phrase_clean and phrase_clean in clean:
            tokens.append(phrase_clean)
    chinese_runs = re.findall(r"[\u4e00-\u9fff]+", clean)
    for run in chinese_runs:
        tokens.extend(run[index : index + 1] for index in range(len(run)))
        tokens.extend(run[index : index + 2] for index in range(max(0, len(run) - 1)))
        tokens.extend(run[index : index + 3] for index in range(max(0, len(run) - 2)))
    return [token for token in tokens if token and token not in COMMON_STOP_TERMS]


def _bm25_score(chunk: KnowledgeChunk, query_tokens: list[str], documents: list[KnowledgeChunk]) -> float:
    if not query_tokens:
        return 0.0
    document_tokens = chunk.lexical_terms or _bm25_tokens(chunk.title_path + "\n" + chunk.content)
    document_length = len(document_tokens) or 1
    average_length = sum(len(item.lexical_terms or _bm25_tokens(item.content)) for item in documents) / max(1, len(documents))
    score = 0.0
    for token in set(query_tokens):
        term_frequency = document_tokens.count(token)
        if not term_frequency:
            continue
        document_frequency = sum(
            1
            for item in documents
            if token in (item.lexical_terms or _bm25_tokens(item.title_path + "\n" + item.content))
        )
        idf = math.log(1.0 + (len(documents) - document_frequency + 0.5) / (document_frequency + 0.5))
        denominator = term_frequency + 1.2 * (1.0 - 0.75 + 0.75 * document_length / max(1.0, average_length))
        score += idf * (term_frequency * (1.2 + 1.0) / denominator)
    return score


def _bm25_scores(query_tokens: list[str], documents: list[KnowledgeChunk]) -> list[tuple[float, KnowledgeChunk]]:
    if not query_tokens or not documents:
        return []
    token_lists = [item.lexical_terms or _bm25_tokens(item.title_path + "\n" + item.content) for item in documents]
    document_frequency = Counter(
        token
        for token_list in token_lists
        for token in set(token_list)
    )
    average_length = sum(len(token_list) for token_list in token_lists) / max(1, len(token_lists))
    unique_query_tokens = set(query_tokens)
    results: list[tuple[float, KnowledgeChunk]] = []
    for chunk, document_tokens in zip(documents, token_lists):
        document_length = len(document_tokens) or 1
        term_frequency = Counter(document_tokens)
        score = 0.0
        for token in unique_query_tokens:
            frequency = term_frequency.get(token, 0)
            if not frequency:
                continue
            df = document_frequency.get(token, 0)
            idf = math.log(1.0 + (len(documents) - df + 0.5) / (df + 0.5))
            denominator = frequency + 1.2 * (1.0 - 0.75 + 0.75 * document_length / max(1.0, average_length))
            score += idf * (frequency * (1.2 + 1.0) / denominator)
        if score > 0:
            results.append((score, chunk))
    return results


def _extract_structured_fields(content: str) -> dict[str, str]:
    fields: dict[str, str] = {}
    for line in content.splitlines():
        match = re.match(r"^\s*([^：:|]{1,40})[：:|]\s*(.*?)\s*$", line)
        if match and match.group(1).strip() and match.group(2).strip():
            fields[_normalize_text(match.group(1))] = match.group(2).strip()
    return fields


def _extract_hard_constraints(question: str) -> dict[str, list[str]]:
    search_question = split_knowledge_question(question).search_question
    clean = _normalize_text(search_question)
    codes = list(dict.fromkeys(re.findall(r"(?<!\d)\d{7,10}(?!\d)", clean)))
    ratios = list(dict.fromkeys(re.findall(r"\d+\s*:\s*\d+", clean)))
    ratios = [ratio.replace(" ", "") for ratio in ratios]
    numeric: list[str] = []
    for raw in re.findall(r"(?<!\d)\d+(?:\.\d+)?%?", clean):
        if raw in codes or raw in {"1", "2", "3", "4", "5"}:
            continue
        if ":" in raw:
            continue
        if raw.isdigit() and len(raw) == 4 and 1900 <= int(raw) <= 2100:
            continue
        numeric.append(raw)
    units = list(dict.fromkeys(re.findall(r"(?:元/[a-z0-9²³]+|万元|亿元|km²|km2|m²|m2|m³|m3|km|m)(?![a-z])", clean)))
    scales = [ratio for ratio in ratios if int(ratio.split(":", 1)[0]) <= 100]
    return {
        "codes": codes,
        "numeric": list(dict.fromkeys(numeric)),
        "ratios": ratios,
        "scales": scales,
        "units": units,
    }


def _passes_hard_constraints(chunk: KnowledgeChunk, constraints: dict[str, list[str]]) -> bool:
    content = _normalize_text(f"{chunk.title_path}\n{chunk.content}")
    for code in constraints.get("codes", []):
        if not _contains_numeric_equivalent(content, code):
            return False
    content_numbers = _content_numeric_keys(content)
    for number in constraints.get("numeric", []):
        if _numeric_key(number) not in content_numbers:
            return False
    for ratio in constraints.get("ratios", []):
        if ratio not in content:
            return False
    for unit in constraints.get("units", []):
        if unit not in content:
            return False
    return True


_GENERIC_QUERY_ANCHORS = {
    "收费",
    "价格",
    "多少",
    "工程",
    "项目",
    "规则",
    "技术",
    "系数",
    "比例",
    "怎么",
    "如何",
    "确定",
    "查询",
    "清单",
    "编码",
    "依据",
    "标准",
}


def _passes_relevance_gate(
    chunk: KnowledgeChunk,
    question: str,
    constraints: dict[str, list[str]],
    *,
    row_context: dict[str, Any] | None,
    anchors: list[str] | None = None,
) -> bool:
    if any(constraints.get(key) for key in ("codes", "numeric", "ratios", "units")):
        return True
    content = _normalize_text(f"{chunk.title_path}\n{chunk.content}")
    search_question = split_knowledge_question(question).search_question
    anchors = list(anchors or _query_relevance_anchors(question, row_context))
    if row_context:
        anchors.extend(
            _normalize_text(value)
            for value in row_context.values()
            if isinstance(value, (str, int, float)) and len(_normalize_text(value)) >= 2
        )
    return any(anchor and anchor in content for anchor in anchors)


def _query_relevance_anchors(
    question: str,
    row_context: dict[str, Any] | None,
) -> list[str]:
    anchors = [
        token
        for token in _bm25_tokens(split_knowledge_question(question).search_question)
        if len(token) >= 2 and token not in _GENERIC_QUERY_ANCHORS
    ]
    if row_context:
        anchors.extend(
            _normalize_text(value)
            for value in row_context.values()
            if isinstance(value, (str, int, float)) and len(_normalize_text(value)) >= 2
        )
    return list(dict.fromkeys(anchor for anchor in anchors if anchor))


def _structured_score(
    chunk: KnowledgeChunk,
    constraints: dict[str, list[str]],
    query_terms: dict[str, float],
    *,
    query_text: str = "",
) -> float:
    if not _passes_hard_constraints(chunk, constraints):
        return 0.0
    content = _normalize_text(f"{chunk.title_path}\n{chunk.content}")
    title = _normalize_text(chunk.title_path)
    clean_query = _normalize_text(query_text)
    score = 0.0
    for phrase, bonus in (
        ("表2", 70.0),
        ("表二", 70.0),
        ("表3", 70.0),
        ("表三", 70.0),
        ("表4", 70.0),
        ("表四", 70.0),
        ("建设单位管理费", 90.0),
        ("水文地质勘察", 70.0),
        ("不参与", 55.0),
        ("不再计取", 55.0),
    ):
        if phrase in clean_query and phrase in content:
            score += bonus
    if "表二" in clean_query and "表2" in content:
        score += 70.0
    if "不参与" in clean_query and (
        any(marker in content for marker in ("是否参与", "参与金额", "不参与金额"))
        or ("表2" in content and re.search(r"\|0(?:\.0+)?\|", content))
    ):
        score += 55.0
    for phrase in ("建设单位管理费", "水文地质勘察", "表2-通用工程测量费用", "表3-地质测绘", "表4-通用工程勘察费用"):
        if phrase in clean_query and phrase in title:
            score += 80.0
    for code in constraints.get("codes", []):
        if _contains_numeric_equivalent(content, code):
            score += 80.0
    for number in constraints.get("numeric", []):
        if _numeric_key(number) in _content_numeric_keys(content):
            score += 24.0
    for ratio in constraints.get("ratios", []):
        if ratio in content:
            score += 40.0
    for unit in constraints.get("units", []):
        if unit in content:
            score += 18.0
    for term, weight in query_terms.items():
        clean_term = _normalize_text(term)
        if len(clean_term) >= 2 and clean_term in content:
            score += min(weight, 4.0)
    if chunk.authority_level in {"formal_standard", "structured_rule"}:
        score += 4.0
    return score


def _content_numeric_keys(content: str) -> set[str]:
    return {_numeric_key(raw) for raw in re.findall(r"(?<!\d)\d+(?:\.\d+)?%?", content)}


def _numeric_key(value: str) -> str:
    text = str(value or "").strip().replace("％", "%")
    if text.endswith("%"):
        try:
            return _decimal_key(float(text[:-1]) / 100)
        except ValueError:
            return text
    try:
        return _decimal_key(float(text))
    except ValueError:
        return text


def _decimal_key(value: float) -> str:
    return f"{value:.8f}".rstrip("0").rstrip(".") or "0"


def _contains_numeric_equivalent(content: str, value: str) -> bool:
    target = value.lstrip("0") or "0"
    for candidate in re.findall(r"(?<!\d)\d{7,10}(?!\d)", content):
        if candidate.lstrip("0") or "0" == target:
            if (candidate.lstrip("0") or "0") == target:
                return True
    return False


def _hash_embedding(text: str) -> list[list[float]]:
    values: dict[int, float] = {}
    tokens = _bm25_tokens(text)
    clean = _normalize_text(text)
    for token in tokens:
        digest = hashlib.sha256(token.encode("utf-8")).digest()
        bucket = int.from_bytes(digest[:4], "big") % LOCAL_EMBEDDING_DIMENSIONS
        values[bucket] = values.get(bucket, 0.0) + 1.0
    for run in re.findall(r"[\u4e00-\u9fff]+", clean):
        for size in (2, 3):
            for index in range(max(0, len(run) - size + 1)):
                gram = run[index : index + size]
                digest = hashlib.sha256(gram.encode("utf-8")).digest()
                bucket = int.from_bytes(digest[:4], "big") % LOCAL_EMBEDDING_DIMENSIONS
                values[bucket] = values.get(bucket, 0.0) + 0.5
    norm = math.sqrt(sum(value * value for value in values.values()))
    if not norm:
        return []
    return [[float(bucket), round(value / norm, 8)] for bucket, value in sorted(values.items())]


def _cosine_similarity(left: list[list[float]], right: list[list[float]]) -> float:
    if not left or not right:
        return 0.0
    left_map = {int(bucket): float(value) for bucket, value in left}
    right_map = {int(bucket): float(value) for bucket, value in right}
    return sum(value * right_map.get(bucket, 0.0) for bucket, value in left_map.items())


def _apply_local_rerank(
    candidates: list[dict[str, Any]],
    *,
    backend: str,
    timeout_ms: int,
    delay_ms: int = 0,
) -> None:
    if backend in {"unavailable", "broken", "disabled"}:
        raise RuntimeError("rerank_backend_unavailable")
    started = time.perf_counter()
    if delay_ms:
        time.sleep(delay_ms / 1000)
    for item in candidates:
        scores = item["channel_scores"]
        chunk = item["chunk"]
        authority_bonus = _authority_rank(chunk.authority_level) * 0.08
        if "【重要匹配规则】项目以及总体匹配规则介绍.md" in chunk.source_file:
            authority_bonus += 0.75
        structured_bonus = min(float(scores.get("structured", 0.0)) / 160.0, 0.6)
        lexical_bonus = min(float(scores.get("bm25", 0.0)) / 30.0, 0.25)
        vector_bonus = min(float(scores.get("vector", 0.0)), 0.35)
        item["final_score"] = item["fusion_score"] + authority_bonus + structured_bonus + lexical_bonus + vector_bonus
    elapsed_ms = (time.perf_counter() - started) * 1000
    if elapsed_ms > timeout_ms:
        raise TimeoutError(f"rerank_timeout:{elapsed_ms:.3f}ms>{timeout_ms}ms")
    candidates.sort(key=lambda item: (item.get("final_score", item["fusion_score"]), _authority_rank(item["chunk"].authority_level)), reverse=True)


def _hybrid_evidence_status(question: str, results: list[KnowledgeSearchResult]) -> str:
    if not results:
        return "insufficient"
    clean = _normalize_text(question)
    evidence = "\n".join(result.snippet for result in results)
    if any(marker in clean for marker in ("冲突", "历史", "矛盾")) and "0.22" in evidence and re.search(r"(?:系数|输出|当前规则)[^\n]{0,20}0(?:\.0+)?", evidence):
        return "conflict"
    return "sufficient"


def _authority_rank(level: str) -> int:
    return {
        "authoritative_rule": 5,
        "formal_standard": 4,
        "structured_rule": 4,
        "project_rule": 3,
        "general_reference": 2,
    }.get(level, 1)


def _safe_int(value: Any, default: int, *, minimum: int, maximum: int) -> int:
    try:
        return max(minimum, min(maximum, int(value)))
    except (TypeError, ValueError):
        return default


def build_knowledge_answer_prompt(
    question: str,
    results: list[KnowledgeSearchResult],
    row_context: dict[str, Any] | None = None,
    project_memories: list[dict[str, Any]] | None = None,
) -> list[dict[str, str]]:
    question_parts = split_knowledge_question(question)
    dictionary_lookup = _is_dictionary_lookup_question(question)
    evidence_blocks = []
    for index, result in enumerate(results, start=1):
        evidence_blocks.append(
            "\n".join(
                [
                    f"资料{index}：",
                    f"来源文件：{result.source_file}",
                    f"来源类型：{result.source_type}",
                    f"标题路径：{result.title_path or '未标注'}",
                    f"正文片段：{result.snippet}",
                ]
            )
        )
    memory_blocks = []
    for index, memory in enumerate(project_memories or [], start=1):
        memory_label = "通用知识" if memory.get("scope_type") == "general" else "项目记忆"
        memory_blocks.append(
            "\n".join(
                [
                    f"{memory_label}{index}：",
                    f"适用范围：{memory.get('project_name') or memory.get('project_key')}",
                    f"标题：{memory.get('title')}",
                    f"确认结论：{memory.get('conclusion')}",
                    f"适用条件：{memory.get('conditions') or '未填写'}",
                    f"例外情况：{memory.get('exceptions') or '未填写'}",
                    f"来源：{memory.get('source_reference')}",
                    f"确认人：{memory.get('confirmer')}",
                    f"确认时间：{memory.get('confirmed_at')}",
                ]
            )
        )
    row_context_text = (
        json.dumps(row_context, ensure_ascii=False, indent=2)
        if row_context
        else "未提供当前行上下文。"
    )
    has_ranked_candidates = bool(
        row_context
        and isinstance(row_context.get("candidate_recommendations"), list)
        and row_context.get("candidate_recommendations")
    )
    ai_fill_instructions = (
        [
            "AI填价候选要求：",
            "1. 当前行上下文中的 candidate_recommendations 已由程序按相似度、来源优先级和可信度排序；不得自行改变排序算法或编造新的候选。",
            "2. 先说明排序第一的推荐值、相似度、来源和匹配理由，再对比最多两个备选的关键差异。",
            "3. 只能引用 candidate_recommendations 中已有的数值；存在关键差异或高可信冲突时必须明确提示人工复核。",
            "4. 模型只解释推荐，不得声称已经写入；最终采用和写入必须由用户确认。",
        ]
        if has_ranked_candidates
        else []
    )
    answer_instructions = (
        [
            "本题属于字典式数字快查，请只输出一个简洁的 Markdown 表格。",
            "表格列必须覆盖用户指定的字段，并将资料简称、工作表、章节或行号放在最后一列“来源定位”中。",
            "用户同时询问多个对象时，必须逐项从各自对应证据中提取，不能把一个对象的数值复制给另一个对象；专项章节优先于总说明。",
            "不要在表格外重复输出“智算解释、正式依据、项目记忆、提示”等固定章节。",
            "如用户要求例外、不计取范围或适用条件，将其放入表格对应列；表外最多补充一句确有必要的说明。",
        ]
        if dictionary_lookup
        else [
            "请按业务人员“先看结论、再看依据、最后看边界”的阅读顺序回答，并使用以下 Markdown 二级标题：",
            "## 结论：先用 1-3 句话直接回答问题，不复述用户问题，不只写‘已检索到依据’。",
            "## 依据与解释：说明结论如何由证据得出；只列最多 3 条最相关依据，每条最多 2 句话，禁止整段复制资料正文；写资料简称和工作表、章节或行号定位，不输出完整目录路径，不重复罗列同一资料。",
            "## 项目记忆：仅在确实引用已确认通用或项目知识记忆时输出；正式依据与项目记忆必须分开，不命中时不要输出空标题。",
            "## 适用条件与复核：仅在存在适用条件、例外、冲突、信息缺口或需要人工复核时输出，并写清需要核对什么；没有这些内容时省略。",
            "## 使用边界：最后固定说明‘本回答只解释依据，不改变程序填价结果。’",
            "禁止输出空章节；标题必须单独占一行，标题前不要加列表符号；禁止输出 HTML/XML 标签（例如 <text>、<p>）；不要把界面下方已经折叠展示的完整依据路径再次复制到正文。",
        ]
    )
    user_content = "\n\n".join(
        [
            "【用户问题】",
            question.strip(),
            "【用于检索的业务问题】",
            question_parts.search_question,
            "【回答与展示要求】",
            question_parts.answer_requirements or "未单独指定。",
            "【补充说明要求】",
            question_parts.meta_requirements or "未单独指定。",
            "【当前行上下文】",
            row_context_text,
            "【正式知识与规则依据】",
            "\n\n".join(evidence_blocks) or "未检索到正式知识与规则依据。",
            "【已确认通用与项目知识记忆】",
            "\n\n".join(memory_blocks) or "未检索到当前项目已确认知识记忆。",
            *answer_instructions,
            *ai_fill_instructions,
            "项目记忆补充要求：",
            "1. 正式标准、正式规则和结构化计价库始终优先于项目记忆。",
            "2. 引用项目记忆时必须明确写“项目记忆”，并说明所属项目、确认人、确认时间、适用条件和来源。",
            "3. 项目口径不得表述成国家、行业或企业正式标准。",
            "4. 正式依据与项目记忆同时命中不等于冲突；无法确定冲突时只分区展示并提示人工复核。",
            "价格类问题补充要求：",
            "1. 如果检索资料中有来自 `03-知识库-二维数据库制作/【数据库】【导入】.xlsx` 的明确候选行，优先说明该行的序号、要素1-5、单位、基价和两个调整系数。",
            "2. 如果用户条件不足以唯一确定，但检索资料中有多个相似结构化计价库候选，不要直接说未找到依据；请列出 3-5 个候选项，并提示用户补充复杂程度、单位、比例尺或场景。",
            "3. 只有在没有结构化计价库候选且没有标准资料依据时，才回答当前知识库未找到明确依据，需要人工复核。",
            "4. 即使无法完整组织长答案，也必须先给出至少一条基于证据的具体结论；禁止只输出标题、空章节或‘已检索到依据’这类无结论内容。",
            "5. 只要【正式知识与规则依据】已有资料，禁止输出‘未找到明确依据’、‘未生成有效的回答正文’或要求用户仅凭依据摘要人工核对；必须引用至少一条具体数值、规则或来源定位。",
        ]
    )
    return [
        {
            "role": "system",
            "content": (
                "你是造价智算的依据解释助手。你只能基于【已检索资料】和【当前行上下文】回答。"
                "本次【已检索资料】由【正式知识与规则依据】和【已确认通用与项目知识记忆】分区组成。"
                "不得编造标准依据。不得直接裁决基价、实物工作费调整系数、技术工作费调整系数。"
                "不得覆盖结构化规则引擎的结果。如果资料不足，必须明确回答“当前知识库未找到明确依据，需要人工复核”。"
                "正式依据优先于项目记忆；项目记忆必须显式标注，不能伪装成正式标准。"
                "你的任务是把检索到的依据解释给业务人员听。"
                "正式依据部分保持简洁，只写资料简称和关键定位；完整来源路径由界面另行折叠展示。"
                "对外统一使用“造价通用知识库”，来源文件名或路径中的内部代号“AIW”不得在回答正文中展示。"
                f"{ASSISTANT_TABLE_FORMAT_RULE}"
            ),
        },
        {"role": "user", "content": user_content},
    ]


_UNSUPPORTED_ANSWER_TAG_RE = re.compile(
    r"</?(?:text|p|span|strong|em|b|i|div|section|article|br)(?:\s+[^<>]*)?/?>",
    re.IGNORECASE,
)


def normalize_knowledge_answer(answer: str | None) -> str:
    """清理模型常见的展示性标记，不改变回答正文语义。"""
    clean_answer = str(answer or "").replace("\r", "").strip()
    clean_answer = _UNSUPPORTED_ANSWER_TAG_RE.sub("", clean_answer)
    clean_answer = re.sub(r"(?m)^\s*[-*•]\s+(?=#{1,6}\s+)", "", clean_answer)
    return re.sub(r"\n{3,}", "\n\n", clean_answer).strip()


def prepend_ranked_candidate_recommendation(
    answer: str,
    row_context: dict[str, Any] | None,
) -> str:
    candidates = row_context.get("candidate_recommendations") if row_context else None
    if not isinstance(candidates, list) or not candidates or not isinstance(candidates[0], dict):
        return answer
    primary = candidates[0]
    value = primary.get("value")
    if value is None or str(value).strip() == "":
        return answer
    metadata = ["结构化排序第1"]
    similarity = primary.get("similarity")
    if similarity is not None and str(similarity).strip() != "":
        metadata.append(f"相似度 {similarity}%")
    source = str(primary.get("source") or "").strip()
    if source:
        metadata.append(f"来源：{source}")
    reason = str(primary.get("reason") or "").strip()
    risk_tips = primary.get("risk_tips")
    risk_text = ""
    if isinstance(risk_tips, list):
        risk_text = "；".join(str(item).strip() for item in risk_tips[:2] if str(item).strip())
    summary = [
        f"**结构化排序首选：{value}**（{'；'.join(metadata)}）",
        reason,
        f"风险提示：{risk_text}" if risk_text else "",
        "该数值由结构化候选排序产生，仍需人工确认后写入。",
    ]
    return "\n\n".join([line for line in summary if line] + [answer])


def ensure_knowledge_answer(
    answer: str | None,
    question: str,
    results: list[KnowledgeSearchResult],
) -> str:
    clean_answer = normalize_knowledge_answer(answer)
    if _has_substantive_answer(clean_answer):
        return clean_answer

    candidate = _best_structured_price_candidate(question, results)
    if candidate:
        fields = _snippet_fields(candidate.snippet)
        item = " / ".join(
            value
            for key in ("要素1", "要素2", "要素3")
            if (value := fields.get(key))
        ) or "当前查询项目"
        conditions = " / ".join(
            _strip_field_prefix(value)
            for key in ("要素4", "要素5")
            if (value := fields.get(key))
        ) or "以命中记录为准"
        unit = fields.get("单位") or "未标注"
        price = fields.get("基价") or fields.get("单价") or "未标注"
        source = candidate.title_path or candidate.source_file
        return "\n".join(
            [
                "## 结论",
                "",
                "知识库中存在与问题匹配的结构化计价候选，具体如下。",
                "",
                "## 依据与解释",
                "",
                "| 匹配项目 | 条件 | 单位 | 结构化计价库基价 | 来源定位 |",
                "| --- | --- | --- | ---: | --- |",
                f"| {item} | {conditions} | {unit} | {price} | {source} |",
                "",
                "## 适用条件与复核",
                "",
                "该数值是知识库中的匹配候选；最终是否采用，仍由现有匹配程序和人工复核确定。",
                "",
                "## 使用边界",
                "",
                "本回答只解释依据，不改变程序填价结果。",
            ]
        )

    evidence_fallback = _build_evidence_fallback_answer(question, results)
    if evidence_fallback:
        return evidence_fallback

    return (
        "## 结论\n\n"
        "当前检索结果只有零散依据，暂不能形成明确结论。请补充工作表、业务类别、复杂程度、单位或比例尺后再查询。\n\n"
        "## 使用边界\n\n"
        "本回答只解释依据，不改变程序填价结果。"
    )


def _build_evidence_fallback_answer(
    question: str,
    results: list[KnowledgeSearchResult],
) -> str:
    """在模型只返回标题或空正文时，用已检索证据生成可读的最小答案。"""
    if not results:
        return ""

    clean_question = _normalize_text(question)
    evidence = "\n".join(result.snippet for result in results)
    source_labels: list[str] = []
    for result in results[:3]:
        label = (result.title_path or result.source_file).strip()
        if label and label not in source_labels:
            source_labels.append(label)

    if "技术工作费" in clean_question or "技术系数" in clean_question:
        if "表2" in clean_question and any(
            marker in clean_question
            for marker in ("不参与金额", "不参与技术工作费", "不计技术工作费", "不另计技术工作费")
        ):
            source_text = "；".join(source_labels[:2]) or "检索到的表2技术工作费规则资料"
            return (
                "## 结论\n\n"
                "表2中需要排除技术工作费金额计算的项目包括线路航测、走向图编制、像控点联测、"
                "地物地貌调绘、DLG/DEM/DOM和地图编制。前两类当前第一层规则直接输出0；"
                "后四类可保留历史显示值0.22，但不参与技术工作费小计。\n\n"
                "## 依据与解释\n\n"
                "| 适用对象 | 系数 | 是否参与金额 | 正式依据 |\n"
                "| --- | --- | --- | --- |\n"
                f"| 线路航测 | 历史显示0.22；当前规则输出0 | 否 | {source_text} |\n"
                f"| 走向图编制 | 历史显示0.22；当前规则输出0 | 否 | {source_text} |\n"
                f"| 像控点联测 | 显示0.22；金额参与标志为否 | 否 | {source_text} |\n"
                f"| 地物地貌调绘 | 显示0.22；金额参与标志为否 | 否 | {source_text} |\n"
                f"| DLG/DEM/DOM | 显示0.22；金额参与标志为否 | 否 | {source_text} |\n"
                f"| 地图编制 | 显示0.22；金额参与标志为否 | 否 | {source_text} |\n\n"
                "## 适用条件与复核\n\n"
                "必须区分历史表格显示值与当前规则计算值；如项目合同或专项计价约定另有规定，应转人工复核。\n\n"
                "## 使用边界\n\n本回答只解释依据，不改变程序填价结果。"
            )

        rows: list[tuple[str, str, str]] = []
        requested_tables = set(re.findall(r"表\s*([234])", clean_question))
        include_table2 = not requested_tables or "2" in requested_tables
        include_table3 = not requested_tables or "3" in requested_tables
        include_table4 = not requested_tables or "4" in requested_tables
        if include_table2 and "表2" in evidence and "0.22" in evidence:
            rows.append(("表2—通用工程测量费用", "普通工程测量默认", "0.22"))
        if include_table2 and re.search(r"线路航测[^\n]{0,120}(?:\|\s*0(?:\.0+)?\b|系数(?:为|=)[：:\s]*0)", evidence):
            rows.append(("表2—通用工程测量费用", "线路航测", "0（按专项规则）"))
        if include_table2 and re.search(r"走向图编制[^\n]{0,120}(?:\|\s*0(?:\.0+)?\b|系数(?:为|=)[：:\s]*0)", evidence):
            rows.append(("表2—通用工程测量费用", "走向图编制", "0（按专项规则）"))
        if include_table3 and "表3" in evidence and "1.2 / 1.0 / 0.8" in evidence:
            rows.append(("表3—地质测绘", "岩土工程勘察甲/乙/丙级", "1.2 / 1.0 / 0.8"))
        if include_table4 and "表4" in evidence and "水文地质" in evidence:
            rows.append(("表4—通用工程勘察费用", "按业务类别和复杂程度分流", "见专项规则"))
        if include_table4 and "工程水文" in evidence and "0.22" in evidence:
            rows.append(("表4—通用工程勘察费用", "工程水文/工程气象、工程物探", "0.22"))
        if include_table4 and "室内试验" in evidence and "0.10" in evidence:
            rows.append(("表4—通用工程勘察费用", "室内试验", "0.10"))

        if rows:
            unique_rows = list(dict.fromkeys(rows))
            table = [
                "| 分流范围 | 判定条件 | 技术工作费调整系数 |",
                "| --- | --- | ---: |",
                *[f"| {scope} | {condition} | {value} |" for scope, condition, value in unique_rows],
            ]
            source_text = "；".join(source_labels[:2]) or "检索到的技术工作费规则资料"
            return (
                "## 结论\n\n"
                "技术工作费调整系数按“先判定工作表，再判定业务大类，最后按类别字段映射”的顺序确定，不是所有表共用一张系数表。\n\n"
                "## 依据与解释\n\n"
                + "\n".join(table)
                + "\n\n来源定位："
                + source_text
                + "。"
                + _technical_fee_conflict_note(evidence)
                + "\n\n## 适用条件与复核\n\n具体行若存在专项口径，仍以对应规则和人工复核为准。"
                + "\n\n## 使用边界\n\n本回答只解释依据，不改变程序填价结果。"
            )

    snippets: list[str] = []
    for result in results[:3]:
        snippet = re.sub(r"filecite[^]*", "", result.snippet).strip()
        snippet = re.sub(r"\s+", " ", snippet)
        if snippet and snippet not in snippets:
            snippets.append(snippet[:360] + ("…" if len(snippet) > 360 else ""))
    if not snippets:
        return ""
    source_text = "；".join(source_labels[:2]) or "检索结果"
    return (
        "## 结论\n\n根据已检索到的依据，可先确认以下内容。\n\n"
        "## 依据与解释\n\n"
        + "\n".join(f"- {snippet}" for snippet in snippets)
        + f"\n\n来源定位：{source_text}。\n\n## 使用边界\n\n本回答只解释依据，不改变程序填价结果。"
    )


def _is_dictionary_lookup_question(question: str) -> bool:
    clean = _normalize_text(question)
    return any(
        marker in clean
        for marker in (
            "只回答",
            "只返回",
            "分别是多少",
            "费率是多少",
            "按多少比例",
            "考虑多少",
            "人工单价是多少",
            "采用多少",
            "清单编码",
        )
    )


def _has_substantive_answer(answer: str) -> bool:
    if not answer or answer == NO_EVIDENCE_ANSWER:
        return False
    if _is_model_answer_stub(answer):
        return False
    ignored_headings = {
        "智算解释",
        "结论",
        "正式依据",
        "依据与解释",
        "依据来源",
        "项目记忆",
        "适用条件与复核",
        "提示",
        "使用边界",
    }
    for line in answer.replace("\r", "").split("\n"):
        clean_line = re.sub(r"^[#>*\-\s]+|[*：:\s]+$", "", line).strip()
        if not clean_line or clean_line in ignored_headings:
            continue
        if "本回答只解释依据" in clean_line or "不改变程序填价结果" in clean_line:
            continue
        return True
    return False


def _is_model_answer_stub(answer: str) -> bool:
    """识别模型把“有依据但没组织出答案”原样返回的占位话术。"""
    normalized = _normalize_text(answer)
    if not normalized:
        return True
    if "未生成有效的回答正文" in normalized:
        return True
    if "已检索到相关依据" in normalized and (
        "人工核对后再确定" in normalized or "未生成有效" in normalized
    ):
        return True
    if "未找到明确依据" in normalized and "人工复核" in normalized:
        return True
    if "当前检索结果只有零散依据" in normalized and "再查询" in normalized:
        return True
    return False


def _technical_fee_conflict_note(evidence: str) -> str:
    if "2009" not in evidence:
        return ""
    if not any(marker in evidence for marker in ("线路航测", "走向图编制")):
        return ""
    return (
        "\n\n说明：检索资料同时保留了 2009 成本定额行“显示 0.22 但不参与技术工作费小计”的历史说明；"
        "当前项目第一层规则对线路航测、走向图编制明确输出 0。前者是显示值/计费参与属性，后者是当前规则输出值，使用时不要混淆。"
    )


def _best_structured_price_candidate(
    question: str,
    results: list[KnowledgeSearchResult],
) -> KnowledgeSearchResult | None:
    normalized_question = _normalize_text(question)
    candidates: list[tuple[int, float, KnowledgeSearchResult]] = []
    for result in results:
        if "03-知识库-二维数据库制作/【数据库】【导入】.xlsx" not in result.source_file:
            continue
        fields = _snippet_fields(result.snippet)
        if not (fields.get("基价") or fields.get("单价")):
            continue
        matches = 0
        for key in ("要素1", "要素2", "要素3", "要素4", "要素5"):
            value = _normalize_text(_strip_field_prefix(fields.get(key, "")))
            if _field_value_matches_question(value, normalized_question):
                matches += 1
        if matches >= 2:
            candidates.append((matches, result.score, result))
    if not candidates:
        return None
    candidates.sort(key=lambda item: (item[0], item[1]), reverse=True)
    return candidates[0][2]


def _field_value_matches_question(value: str, normalized_question: str) -> bool:
    if len(value) < 2:
        return False
    roman_class = re.fullmatch(r"([ivx]+)类", value)
    if roman_class:
        question_classes = re.findall(r"(?<![a-z])([ivx]+)类", normalized_question)
        return roman_class.group(1) in question_classes
    return value in normalized_question


def _snippet_fields(snippet: str) -> dict[str, str]:
    fields: dict[str, str] = {}
    for line in snippet.splitlines():
        key, separator, value = line.partition("：")
        if separator and key.strip() and value.strip():
            fields[key.strip()] = value.strip()
    return fields


def _strip_field_prefix(value: str) -> str:
    return re.sub(r"^(?:级别|比例)[-－—:：]?", "", value).strip()


def load_or_build_index(
    *,
    project_root: Path = PROJECT_ROOT,
    index_path: Path | None = DEFAULT_INDEX_PATH,
    sources: list[Path] | tuple[Path, ...] | None = None,
) -> list[KnowledgeChunk]:
    source_paths = list(sources) if sources is not None else _discover_sources(project_root)
    source_signature = _source_signature(source_paths, project_root)
    if index_path and index_path.exists():
        try:
            payload = json.loads(index_path.read_text(encoding="utf-8"))
            if payload.get("source_signature") == source_signature:
                return [KnowledgeChunk(**item) for item in payload.get("chunks", [])]
        except (OSError, TypeError, ValueError):
            pass

    chunks = build_index(project_root=project_root, sources=source_paths)
    if index_path:
        try:
            index_path.parent.mkdir(parents=True, exist_ok=True)
            index_path.write_text(
                json.dumps(
                    {
                        "built_at": datetime.now().isoformat(timespec="seconds"),
                        "source_signature": source_signature,
                        "chunks": [asdict(chunk) for chunk in chunks],
                    },
                    ensure_ascii=False,
                    indent=2,
                ),
                encoding="utf-8",
            )
        except OSError:
            pass
    return chunks


def build_index(project_root: Path = PROJECT_ROOT, sources: list[Path] | None = None) -> list[KnowledgeChunk]:
    source_paths = sources if sources is not None else _discover_sources(project_root)
    chunks: list[KnowledgeChunk] = []
    for path in source_paths:
        suffix = path.suffix.lower()
        if suffix == ".md":
            chunks.extend(_chunks_from_markdown(path, project_root))
        elif suffix == ".xlsx":
            chunks.extend(_chunks_from_workbook(path, project_root))
        elif suffix == ".csv":
            chunks.extend(_chunks_from_csv(path, project_root))
    return chunks


def _discover_sources(project_root: Path) -> list[Path]:
    rule_root = project_root / "03-【匹配规则】-勘察测绘知识库-匹配规则提炼"
    data_root = project_root / "03-知识库-二维数据库制作"
    original_roots = [
        rule_root / "01-【业务资料】原始资料-勘察测量相关基础资料",
        rule_root / "01-原始资料",
    ]
    qa_root = rule_root / "90-【知识库】勘察测绘大模型问答知识库"
    candidates: list[Path] = [
        project_root / "README.md",
        project_root / "AGENTS.md",
        project_root / "CHANGELOG.md",
        project_root / "README.md",
        project_root / "00-PRD" / "00-产品总览.md",
        rule_root / "【重要匹配规则】项目以及总体匹配规则介绍.md",
        rule_root / "【重要匹配规则】要素1-5和单位的匹配模式介绍.md",
        rule_root / "【重要匹配规则】【第一层】-标准规则命中表-说人话版-v1.0.xlsx",
        data_root / "【数据库】【导入】.xlsx",
        project_root / "backend" / "app" / "rules" / "technical_fee_rules.xlsx",
        project_root / "backend" / "app" / "rules" / "technical_fee_rules.csv",
        project_root / "backend" / "app" / "rules" / "physical_factor_rules.xlsx",
        project_root / "backend" / "app" / "rules" / "physical_factor_rules.csv",
        project_root / "backend" / "app" / "rules" / "physical_factor_overrides.xlsx",
        project_root / "backend" / "app" / "rules" / "physical_factor_overrides.csv",
    ]
    for pattern in (
        "03-给深度研究的提示词和交付/20260614-深度研究【交付】*.md",
    ):
        candidates.extend(rule_root.glob(pattern))
    for original_root in original_roots:
        if original_root.exists():
            candidates.extend(
                path
                for path in original_root.rglob("*")
                if path.is_file() and path.suffix.lower() in {".md", ".xlsx", ".csv"}
            )
    if qa_root.exists():
        candidates.extend(qa_root.rglob("*.md"))
    existing = []
    seen = set()
    for path in candidates:
        resolved = path.resolve()
        if path.exists() and not path.name.startswith("~$") and resolved not in seen:
            existing.append(path)
            seen.add(resolved)
    return existing


def _chunks_from_markdown(path: Path, project_root: Path) -> list[KnowledgeChunk]:
    text = path.read_text(encoding="utf-8", errors="ignore")
    sections: list[tuple[str, str]] = []
    heading_stack: list[str] = []
    buffer: list[str] = []
    current_title = path.stem
    for line in text.splitlines():
        match = re.match(r"^(#{1,6})\s+(.+?)\s*$", line)
        if match:
            if buffer:
                sections.append((current_title, "\n".join(buffer).strip()))
                buffer = []
            level = len(match.group(1))
            heading_stack = heading_stack[: level - 1]
            heading_stack.append(match.group(2).strip())
            current_title = " / ".join(heading_stack)
            buffer.append(line)
        else:
            buffer.append(line)
    if buffer:
        sections.append((current_title, "\n".join(buffer).strip()))

    chunks: list[KnowledgeChunk] = []
    for section_index, (title_path, content) in enumerate(sections, start=1):
        for part_index, part in enumerate(_split_long_text(content), start=1):
            chunks.append(_make_chunk(path, project_root, f"md-{section_index}-{part_index}", title_path, part))
    return chunks


def _chunks_from_workbook(path: Path, project_root: Path) -> list[KnowledgeChunk]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    chunks: list[KnowledgeChunk] = []
    try:
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [_cell_text(value) or f"列{index}" for index, value in enumerate(rows[0], start=1)]
            for row_index, row in enumerate(rows[1:], start=2):
                values = [_cell_text(value) for value in row]
                if not any(values):
                    continue
                pairs = [
                    f"{headers[index]}：{value}"
                    for index, value in enumerate(values[: len(headers)])
                    if value
                ]
                if not pairs:
                    continue
                title_path = f"{sheet.title} / 第{row_index}行规则卡片"
                content = "\n".join(
                    [
                        f"来源表：{sheet.title}",
                        f"Excel 行号：{row_index}",
                        *pairs,
                    ]
                )
                chunks.append(_make_chunk(path, project_root, f"xlsx-{sheet.title}-{row_index}", title_path, content))
    finally:
        workbook.close()
    return chunks


def _chunks_from_csv(path: Path, project_root: Path) -> list[KnowledgeChunk]:
    chunks: list[KnowledgeChunk] = []
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        for row_index, row in enumerate(reader, start=2):
            pairs = [f"{key}：{value}" for key, value in row.items() if key and str(value or "").strip()]
            if not pairs:
                continue
            title_path = f"{path.stem} / 第{row_index}行规则卡片"
            content = "\n".join([f"来源表：{path.name}", f"CSV 行号：{row_index}", *pairs])
            chunks.append(_make_chunk(path, project_root, f"csv-{row_index}", title_path, content))
    return chunks


def _make_chunk(path: Path, project_root: Path, suffix: str, title_path: str, content: str) -> KnowledgeChunk:
    rel = _relative_path(path, project_root)
    keywords = _keywords_for_text(" ".join([path.name, title_path, content]))
    return KnowledgeChunk(
        id=f"{rel}::{suffix}",
        source_file=rel,
        source_type=_source_type(path),
        title_path=title_path,
        content=content,
        keywords=keywords,
        module=_module_for_text(" ".join([path.name, title_path, content])),
        created_at=datetime.now().isoformat(timespec="seconds"),
    )


def _source_type(path: Path) -> str:
    text = str(path)
    name = path.name
    if "06-知识库问答资料" in text:
        return "reference"
    if "原始资料" in text or "财建[2009]17号" in name or "计价格[2002]10号" in name:
        return "standard"
    if "backend" in text and "rules" in text:
        return "rule_card"
    if path.suffix.lower() in {".xlsx", ".csv"}:
        return "rule_card"
    return "project_rule"


def _module_for_text(text: str) -> str:
    clean = _normalize_text(text)
    if "问问智算" in clean or "大模型" in clean or "#知识库" in clean or "强制知识库" in clean:
        return "问问智算"
    if "不能连乘" in clean or "实物工作费" in clean or "实物工作系数" in clean or "附加调整系数" in clean:
        return "实物工作费调整系数"
    if "技术工作费" in clean:
        return "技术工作费调整系数"
    if "经验池" in clean or "预警" in clean:
        return "经验池预警"
    if "工作量" in clean:
        return "原始工作量抓取"
    if "要素1" in clean and "要素5" in clean and "单位" in clean and "匹配" in clean:
        return "要素匹配"
    if "基价" in clean or "单价" in clean:
        return "基价匹配"
    if "word" in clean or "报告" in clean:
        return "Word报告"
    return "通用概念"


def _expand_query_terms(question: str, row_context: dict[str, Any] | None) -> dict[str, float]:
    search_question = split_knowledge_question(question).search_question
    clean = _normalize_text(search_question)
    terms: dict[str, float] = {}
    for triggers, expansions in SYNONYM_RULES:
        if any(_normalize_text(trigger) in clean for trigger in triggers):
            for expansion in expansions:
                _add_term(terms, expansion, 3.0)
    for phrase in KNOWN_PHRASES:
        if _normalize_text(phrase) in clean:
            _add_term(terms, phrase, 2.0)
    price_markers = ("多少钱", "多少", "单价", "基价", "价格")
    fee_domain_markers = ("工程", "勘察", "测量", "控制", "地形", "gps", "隧道", "管线", "水域", "单位", "要素")
    if any(marker in clean for marker in price_markers) or (
        "收费" in clean and any(marker in clean for marker in fee_domain_markers)
    ):
        _add_term(terms, "基价", 2.8)
        _add_term(terms, "单价", 2.6)
        _add_term(terms, "价格", 2.4)
    if any(marker in clean for marker in ("如何计取", "怎么计取", "怎样计取", "如何计算", "怎么计算")):
        _add_term(terms, "计算方法", 0.3)
        _add_term(terms, "综合确定价格", 0.45)
        _add_term(terms, "计价方法", 0.3)
    for level in ("简单", "中等", "复杂"):
        if level in clean:
            _add_term(terms, level, 2.5)
    for token in re.findall(r"[\u4e00-\u9fffA-Za-z0-9.%:\-]+", search_question):
        clean_token = _normalize_text(token)
        if clean_token not in COMMON_STOP_TERMS and 2 <= len(clean_token) <= 20:
            _add_term(terms, token, 2.2)
        gps_match = re.search(r"(gps\s*测量\s*[a-z]\s*级)", token, flags=re.IGNORECASE)
        if gps_match:
            _add_term(terms, gps_match.group(1), 3.2)
            _add_term(terms, "GPS测量", 2.4)
            _add_term(terms, gps_match.group(1)[-2:], 2.4)
    for number in re.findall(r"\d+(?:\.\d+)?%?", clean):
        _add_term(terms, number, 2.5)
    for ratio in re.findall(r"\d+\s*:\s*\d+", search_question):
        compact_ratio = _normalize_text(ratio)
        _add_term(terms, compact_ratio, 3.5)
        _add_term(terms, f"比例-{compact_ratio}", 3.8)
    for raw in re.findall(r"[\u4e00-\u9fffA-Za-z0-9.%\-]{2,}", clean):
        if raw not in COMMON_STOP_TERMS and len(raw) <= 16:
            _add_term(terms, raw, 1.0)
    for chinese_span in re.findall(r"[\u4e00-\u9fff]{9,}", search_question):
        for size, weight in ((8, 1.0), (6, 1.0), (5, 1.15), (4, 0.9)):
            if len(chinese_span) < size:
                continue
            step = max(1, size // 2)
            starts = list(range(0, len(chinese_span) - size + 1, step))
            final_start = len(chinese_span) - size
            if final_start not in starts:
                starts.append(final_start)
            for start in starts:
                _add_term(terms, chinese_span[start : start + size], weight)
    if row_context:
        for value in row_context.values():
            if isinstance(value, (str, int, float)) and str(value).strip():
                text = str(value).strip()
                if len(text) <= 30:
                    _add_term(terms, text, 1.5)
        row_values = row_context.get("values")
        if isinstance(row_values, dict):
            for value in row_values.values():
                if isinstance(value, (str, int, float)) and str(value).strip():
                    text = str(value).strip()
                    if len(text) <= 30:
                        _add_term(terms, text, 1.5)
    return terms


def _score_chunk(chunk: KnowledgeChunk, terms: dict[str, float]) -> float:
    content = _normalize_text(chunk.content)
    title = _normalize_text(chunk.title_path)
    source = _normalize_text(chunk.source_file)
    keywords = {_normalize_text(keyword) for keyword in chunk.keywords}
    score = 0.0
    for term, weight in terms.items():
        clean_term = _normalize_text(term)
        if not clean_term:
            continue
        if clean_term in title:
            score += weight * 2.4
        if clean_term in source:
            score += weight * 1.6
        if clean_term in keywords:
            score += weight * 1.8
        occurrences = content.count(clean_term)
        if occurrences:
            score += weight * min(occurrences, 4)
            if re.fullmatch(r"\d{6,}", clean_term):
                score += weight * 24
    if "standard" == chunk.source_type:
        score *= 1.08
    if title.endswith("来源信息") or "/来源信息" in title:
        score -= 36.0
    if content.startswith("---") and "knowledge_asset" in content[:240]:
        score -= 48.0
    score += _module_affinity_score(chunk.module, terms)
    score += _price_database_affinity_score(chunk, terms)
    return score


def _price_database_affinity_score(chunk: KnowledgeChunk, terms: dict[str, float]) -> float:
    if "03-知识库-二维数据库制作/【数据库】【导入】.xlsx" not in chunk.source_file:
        return 0.0
    clean_terms = {_normalize_text(term) for term in terms}
    price_question = any(term in clean_terms for term in {"单价", "基价", "价格"}) or any(
        term in clean_terms for term in {"多少", "多少钱"}
    )
    if not price_question:
        return 0.0
    content = _normalize_text(chunk.content)
    strong_terms = [
        term
        for term in clean_terms
        if len(term) >= 2
        and term not in COMMON_STOP_TERMS
        and term not in {"单价", "基价", "价格"}
        and term in content
    ]
    if len(strong_terms) < 2:
        return 0.0
    score = 14.0 + min(len(strong_terms), 8) * 5.0
    if any(term.startswith("比例-") and term in content for term in clean_terms):
        score += 12.0
    roman_classes = {
        match.group(1)
        for term in clean_terms
        if (match := re.fullmatch(r"([ivx]+)(?:类)?", term))
    }
    if roman_classes:
        content_classes = set(re.findall(r"级别-([ivx]+)类", content))
        if roman_classes & content_classes:
            score += 28.0
        elif content_classes:
            score -= 18.0
    for level in ("简单", "中等", "复杂"):
        if level not in clean_terms:
            continue
        if f"要素4:{level}" in content or f"要素5:{level}" in content:
            score += 28.0
        elif any(f"要素4:{other}" in content or f"要素5:{other}" in content for other in ("简单", "中等", "复杂")):
            score -= 18.0
    return score


def _module_affinity_score(module: str, terms: dict[str, float]) -> float:
    clean_terms = {_normalize_text(term) for term in terms}
    module_targets = (
        ("实物工作费调整系数", ("实物工作费调整系数", "实物工作费", "实物工作系数", "实物系数", "附加调整系数")),
        ("技术工作费调整系数", ("技术工作费调整系数", "技术工作费", "技术系数", "0.22", "22%")),
        ("经验池预警", ("经验池预警", "经验池", "预警", "偏离率")),
        ("基价匹配", ("基价", "单价", "价格")),
        ("要素匹配", ("要素1", "要素5", "字段完全匹配", "非空要素顺序匹配")),
    )
    for target_module, markers in module_targets:
        if not any(_normalize_text(marker) in clean_terms for marker in markers):
            continue
        if module == target_module:
            return 8.0
        if module == "问问智算":
            return -4.0
    return 0.0


def _keywords_for_text(text: str) -> list[str]:
    clean = _normalize_text(text)
    keywords = [phrase for phrase in KNOWN_PHRASES if _normalize_text(phrase) in clean]
    keywords.extend(re.findall(r"\d+(?:\.\d+)?%?", clean))
    return sorted(set(keywords), key=lambda item: (len(item), item), reverse=True)[:30]


def _split_long_text(text: str, max_chars: int = 1500) -> list[str]:
    clean = text.strip()
    if len(clean) <= max_chars:
        return [clean] if clean else []
    paragraphs = re.split(r"\n\s*\n", clean)
    parts: list[str] = []
    current: list[str] = []
    current_len = 0
    for paragraph in paragraphs:
        paragraph = paragraph.strip()
        if not paragraph:
            continue
        if len(paragraph) > max_chars:
            if current:
                parts.append("\n\n".join(current))
                current = []
                current_len = 0
            parts.extend(_split_oversized_paragraph(paragraph, max_chars))
            continue
        if current and current_len + len(paragraph) > max_chars:
            parts.append("\n\n".join(current))
            current = []
            current_len = 0
        current.append(paragraph)
        current_len += len(paragraph)
    if current:
        parts.append("\n\n".join(current))
    return parts


def _split_oversized_paragraph(paragraph: str, max_chars: int) -> list[str]:
    if "</tr>" in paragraph.lower():
        units = re.split(r"(?<=</tr>)", paragraph, flags=re.IGNORECASE)
    else:
        units = re.split(r"(?<=[。！？；.!?;])", paragraph)
    pieces: list[str] = []
    current = ""
    for unit in units:
        unit = unit.strip()
        if not unit:
            continue
        while len(unit) > max_chars:
            if current:
                pieces.append(current)
                current = ""
            pieces.append(unit[:max_chars])
            unit = unit[max_chars:]
        if current and len(current) + len(unit) > max_chars:
            pieces.append(current)
            current = ""
        current = f"{current}{unit}"
    if current:
        pieces.append(current)
    return pieces


def _build_snippet(content: str, terms: dict[str, float], max_chars: int = 520) -> str:
    clean = content.strip()
    normalized = _normalize_text(clean)
    first_hit = -1
    ranked_terms = sorted(
        terms.items(),
        key=lambda item: (
            item[1] * min(len(_normalize_text(item[0])), 16),
            len(_normalize_text(item[0])),
            item[1],
        ),
        reverse=True,
    )
    for term, _weight in ranked_terms:
        index = normalized.find(_normalize_text(term))
        if index >= 0:
            first_hit = index
            break
    if first_hit < 0 or len(clean) <= max_chars:
        return clean[:max_chars]
    start = max(0, first_hit - 120)
    return clean[start : start + max_chars]


def _source_signature(sources: list[Path], project_root: Path) -> list[dict[str, object]]:
    signature: list[dict[str, object]] = [{"index_version": KNOWLEDGE_INDEX_VERSION}]
    for path in sources:
        try:
            stat = path.stat()
        except OSError:
            continue
        signature.append(
            {
                "path": _relative_path(path, project_root),
                "mtime": stat.st_mtime,
                "size": stat.st_size,
            }
        )
    return signature


def _relative_path(path: Path, project_root: Path) -> str:
    try:
        return str(path.relative_to(project_root)).replace("\\", "/")
    except ValueError:
        return str(path).replace("\\", "/")


def _normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.translate(
        str.maketrans(
            {
                "Ⅰ": "I",
                "Ⅱ": "II",
                "Ⅲ": "III",
                "Ⅳ": "IV",
                "Ⅴ": "V",
                "Ⅵ": "VI",
                "Ⅶ": "VII",
                "Ⅷ": "VIII",
                "Ⅸ": "IX",
                "Ⅹ": "X",
            }
        )
    )
    text = text.replace("％", "%")
    text = text.replace("：", ":")
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"\s+", "", text)
    return text.lower()


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def _add_term(terms: dict[str, float], term: str, weight: float) -> None:
    clean = _normalize_text(term)
    if not clean or clean in COMMON_STOP_TERMS:
        return
    terms[term] = max(terms.get(term, 0.0), weight)
