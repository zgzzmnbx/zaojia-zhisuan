from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

import sys

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT / "backend"))

from app.fill_engine import FillEngine  # noqa: E402
from app.knowledge_base import KnowledgeBase  # noqa: E402
from app.normalization import normalize_key_part  # noqa: E402
from app.schemas import FIELD_COLUMNS  # noqa: E402

DATA_DIR = PROJECT_ROOT / "03-知识库-二维数据库制作"
SOURCE_DB = DATA_DIR / "【数据库】【导入】.xlsx"
PROJECT_EXAMPLE = DATA_DIR / "【项目例子】【测试输入】铜梁江津-遵义-贵阳-可行性研究勘察测量控制价 -v2.3【批注】.xlsx"
OUTPUT_DB = SOURCE_DB

EXPERIENCE_FIELDS = [
    "【经验数】实物工作费调整系数",
    "【经验数解释】-实物工作费调整系数",
    "【经验数】技术工作费调整系数",
    "【经验数解释】-技术工作费调整系数",
]

SHEET_CONFIGS = {
    "表2-通用工程测量费用": {
        "key_columns": {
            "要素1": "B",
            "要素2": "C",
            "要素3": None,
            "要素4": "D",
            "要素5": "E",
            "单位": "F",
        },
        "experience_columns": {
            EXPERIENCE_FIELDS[0]: "S",
            EXPERIENCE_FIELDS[1]: "T",
            EXPERIENCE_FIELDS[2]: "U",
            EXPERIENCE_FIELDS[3]: "V",
        },
    },
    "表3-地质测绘": {
        "key_columns": {
            "要素1": "B",
            "要素2": "C",
            "要素3": "D",
            "要素4": "E",
            "要素5": "F",
            "单位": "G",
        },
        "experience_columns": {
            EXPERIENCE_FIELDS[0]: "AN",
            EXPERIENCE_FIELDS[1]: "AO",
            EXPERIENCE_FIELDS[2]: "AP",
            EXPERIENCE_FIELDS[3]: "AQ",
        },
    },
    "表4-通用工程勘察费用": {
        "key_columns": {
            "要素1": "B",
            "要素2": "C",
            "要素3": "D",
            "要素4": "E",
            "要素5": "F",
            "单位": "G",
        },
        "experience_columns": {
            EXPERIENCE_FIELDS[0]: "X",
            EXPERIENCE_FIELDS[1]: "Y",
            EXPERIENCE_FIELDS[2]: "Z",
            EXPERIENCE_FIELDS[3]: "AA",
        },
    },
}


def main() -> None:
    if not SOURCE_DB.exists():
        raise FileNotFoundError(SOURCE_DB)
    if not PROJECT_EXAMPLE.exists():
        raise FileNotFoundError(PROJECT_EXAMPLE)

    db_wb = load_workbook(SOURCE_DB)
    db_ws = db_wb.worksheets[0]
    project_wb = load_workbook(PROJECT_EXAMPLE, data_only=True)
    try:
        header_map = _header_map(db_ws)
        _ensure_headers(header_map)
        db_index = _build_database_index(db_ws, header_map)
        filled = 0
        unmatched: list[str] = []

        for sheet_name, config in SHEET_CONFIGS.items():
            ws = project_wb[sheet_name]
            merged_value_map = FillEngine._build_merged_value_map(ws)
            for excel_row in range(5, ws.max_row + 1):
                key = _read_project_key(ws, excel_row, config["key_columns"], merged_value_map)
                if not _has_key_content(key):
                    continue
                experience = _read_experience(ws, excel_row, config["experience_columns"], merged_value_map)
                if not any(not _is_empty_experience(value) for value in experience.values()):
                    continue
                db_row = _find_database_row(db_index, key)
                if db_row is None:
                    unmatched.append(f"{sheet_name}!{excel_row}: {key}")
                    continue
                for field_name, value in experience.items():
                    target_cell = db_ws.cell(row=db_row, column=header_map[field_name])
                    target_cell.value = value
                filled += 1

        db_wb.save(OUTPUT_DB)
    finally:
        db_wb.close()
        project_wb.close()

    print(f"output={OUTPUT_DB}")
    print(f"filled_rows={filled}")
    print(f"unmatched_rows={len(unmatched)}")
    for line in unmatched[:20]:
        print(f"unmatched={line}")


def _header_map(sheet: Any) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in sheet[1]
        if cell.value is not None and str(cell.value).strip()
    }


def _ensure_headers(header_map: dict[str, int]) -> None:
    missing = [name for name in [*FIELD_COLUMNS, "基价", *EXPERIENCE_FIELDS] if name not in header_map]
    if missing:
        raise ValueError(f"母表缺少字段：{', '.join(missing)}")


def _build_database_index(sheet: Any, header_map: dict[str, int]) -> dict[tuple[str, tuple[str, ...]], list[int]]:
    index: dict[tuple[str, tuple[str, ...]], list[int]] = {}
    for row_idx in range(2, sheet.max_row + 1):
        row = {
            name: sheet.cell(row=row_idx, column=header_map[name]).value
            for name in FIELD_COLUMNS
        }
        key = KnowledgeBase.make_ordered_key(row)
        index.setdefault(key, []).append(row_idx)
    return index


def _read_project_key(
    sheet: Any,
    excel_row: int,
    key_columns: dict[str, str | None],
    merged_value_map: dict[tuple[int, int], Any],
) -> dict[str, Any]:
    return {
        field_name: _read_cell(sheet, excel_row, column, merged_value_map) if column else None
        for field_name, column in key_columns.items()
    }


def _read_experience(
    sheet: Any,
    excel_row: int,
    columns: dict[str, str],
    merged_value_map: dict[tuple[int, int], Any],
) -> dict[str, Any]:
    return {
        field_name: _read_cell(sheet, excel_row, column, merged_value_map)
        for field_name, column in columns.items()
    }


def _read_cell(
    sheet: Any,
    excel_row: int,
    column: str,
    merged_value_map: dict[tuple[int, int], Any],
) -> Any:
    return FillEngine._read_mapped_value(
        sheet,
        excel_row,
        column_index_from_string(column),
        merged_value_map,
    )


def _has_key_content(row: dict[str, Any]) -> bool:
    return bool(normalize_key_part(row.get("要素1"))) and bool(normalize_key_part(row.get("单位")))


def _is_empty_experience(value: Any) -> bool:
    return value is None or str(value).strip() in {"", "/"}


def _find_database_row(
    db_index: dict[tuple[str, tuple[str, ...]], list[int]],
    project_key: dict[str, Any],
) -> int | None:
    rows = db_index.get(KnowledgeBase.make_ordered_key(project_key), [])
    return rows[0] if len(rows) == 1 else None


if __name__ == "__main__":
    main()
