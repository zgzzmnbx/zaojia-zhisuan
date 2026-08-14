from __future__ import annotations

import hashlib
import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font, PatternFill


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = (
    PROJECT_ROOT
    / "03-知识库-二维数据库制作"
    / "05-260729-【结算】【前辈经验】结算和投标限价相关资料"
)
TEMPLATE_PATH = SOURCE_DIR / "【结算模板】260723-勘察测量结算统一报价模板-v1.0.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "00-PRD" / "01-模块PRD" / "10-结算审核助手模块" / "evals"
OUTPUT_PATH = OUTPUT_DIR / "结算审核演示样例-v0.1.xlsx"
EXPECTED_PATH = OUTPUT_DIR / "结算审核演示样例-预期结果.json"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _clear_range(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.value = None
            cell.comment = None


def build_sample(output_path: Path = OUTPUT_PATH) -> Path:
    if not TEMPLATE_PATH.is_file():
        raise FileNotFoundError(f"统一结算模板不存在：{TEMPLATE_PATH}")
    template_hash_before = sha256(TEMPLATE_PATH)
    workbook = load_workbook(TEMPLATE_PATH, data_only=False, keep_links=True)

    summary = workbook["汇总表"]
    measure = workbook["表1 测量费，注所有单体都放一张表中"]
    survey = workbook["表2-1 XX勘察费，注每个单体一张表"]
    other = workbook["表3 其他费用(一次报送)"]

    summary["A1"] = (
        "长输管道示范工程可行性研究阶段勘察测量结算费用汇总表\n"
        "                                                                                                                               单位：万元"
    )
    for row in range(5, 24):
        for column in range(3, 12):
            cell = summary.cell(row, column)
            if not isinstance(cell, MergedCell) and cell.data_type != "f":
                cell.value = None
    summary["M4"] = 0.95  # 演示风险：2024 口径应按模板参考值 0.90 核验。
    summary["N4"] = 0.05
    summary["O4"] = 0.06
    summary["C5"] = f"=SUM('{measure.title}'!M5:M80)/10000"
    summary["D6"] = f"=SUM('{survey.title}'!O5:O210)/10000"
    summary["C21"] = f"=SUM('{other.title}'!L4:L10)/10000"

    # 清空模板自带演示工程量和审核栏，只保留模板结构、公式与样式。
    for row in range(5, measure.max_row + 1):
        cell = measure.cell(row, 7)
        if not isinstance(cell, MergedCell):
            cell.value = None
    _clear_range(measure, 5, measure.max_row, 16, 29)
    for row in range(5, survey.max_row + 1):
        cell = survey.cell(row, 8)
        if not isinstance(cell, MergedCell):
            cell.value = None
    _clear_range(survey, 5, survey.max_row, 18, 31)
    _clear_range(other, 4, other.max_row, 3, 25)

    # 测量费：两条参数偏离、一条算术错误。
    measure["G11"] = 6
    measure["H11"] = 3500
    measure["I11"] = 0.6
    measure["K11"] = 0.22
    measure["M11"] = "=G11*I11*H11*(1+K11)"

    measure["G24"] = 0.25
    measure["I24"] = 2.0
    measure["K24"] = 0.22
    measure["M24"] = "=G24*I24*H24*(1+K24)"

    measure["G37"] = 1.5
    measure["I37"] = 1.0
    measure["K37"] = 0.22
    measure["M37"] = 5000

    # 勘察费：一条正确记录、深孔额外 1.2、室内试验 0.1。
    survey["H13"] = 13
    survey["J13"] = 1.0
    survey["M13"] = 1.2
    survey["O13"] = "=H13*J13*I13*(1+M13)"

    survey["H40"] = 20
    survey["J40"] = 1.0
    survey["M40"] = 1.2
    survey["O40"] = "=H40*J40*I40*(1+M40)"

    survey["H200"] = 1
    survey["J200"] = 1.0
    survey["M200"] = 0.1
    survey["O200"] = "=H200*J200*I200*(1+M200)"

    # 其他费用：K 列“其他”被人工合计遗漏；第二条缺少费用文件编号。
    other["A4"] = 1
    other["B4"] = "示范穿越"
    other["C4"] = 400
    other["D4"] = 400
    other["K4"] = 500
    other["L4"] = 800
    other["M4"] = "附件001-1~附件001-3"
    other["N4"] = "许可、资料及临时设施费用"

    other["A5"] = 2
    other["B5"] = "示范隧道"
    other["J5"] = 1800
    other["L5"] = 1800
    other["M5"] = None
    other["N5"] = "青苗补偿费用，费用文件编号待补"

    if "演示样例说明" in workbook.sheetnames:
        workbook.remove(workbook["演示样例说明"])
    notes = workbook.create_sheet("演示样例说明")
    notes.sheet_view.showGridLines = False
    notes.column_dimensions["A"].width = 16
    notes.column_dimensions["B"].width = 30
    notes.column_dimensions["C"].width = 68
    notes.merge_cells("A1:C1")
    notes["A1"] = "结算审核助手演示样例"
    notes["A1"].fill = PatternFill("solid", fgColor="163B65")
    notes["A1"].font = Font(name="等线", size=18, bold=True, color="FFFFFF")
    notes["A1"].alignment = Alignment(vertical="center")
    notes.row_dimensions[1].height = 34
    notes.merge_cells("A2:C2")
    notes["A2"] = "本文件由统一报价模板生成，仅用于竞赛演示与自动化回归，不代表真实项目结算结论。"
    notes["A2"].fill = PatternFill("solid", fgColor="EAF2FF")
    notes["A2"].font = Font(name="等线", size=10, color="163B65", bold=True)
    rows = [
        ("场景", "正确记录", "勘察费第 13 行参数与算术正确，用于验证系统不会把所有有值行都判成风险。"),
        ("已知风险", "模板参数", "测量费第 11 行基价、第 24 行实物系数偏离统一模板。"),
        ("已知风险", "金额算术", "测量费第 37 行金额被手工写为 5000 元。"),
        ("已知风险", "深孔专项", "勘察费第 40 行深度大于 300m，模板公式额外乘以 1.2。"),
        ("已知风险", "室内试验", "勘察费第 200 行技术工作费系数使用 0.1。"),
        ("已知风险", "其他费用", "第 4 行合计遗漏 K 列，第 5 行缺少费用文件编号。"),
        ("已知风险", "合同参数", "汇总表 M4 使用 0.95，偏离统一模板参考值 0.90，必须结合合同确认。"),
    ]
    for column, header in enumerate(("类型", "主题", "说明"), start=1):
        cell = notes.cell(4, column)
        cell.value = header
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.font = Font(name="等线", size=10, bold=True, color="FFFFFF")
    for row_index, values in enumerate(rows, start=5):
        for column, value in enumerate(values, start=1):
            cell = notes.cell(row_index, column)
            cell.value = value
            cell.font = Font(name="等线", size=10, color="172033")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor="F8FAFC" if row_index % 2 else "FFFFFF")
        notes.row_dimensions[row_index].height = 34
    notes.freeze_panes = "A5"

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()

    expected = {
        "sample": output_path.name,
        "reference_template": TEMPLATE_PATH.name,
        "reference_template_sha256": template_hash_before,
        "expected_rules": [
            {"rule_id": "JS-001", "coordinate": f"{measure.title}!H11"},
            {"rule_id": "JS-001", "coordinate": f"{measure.title}!I24"},
            {"rule_id": "JS-002", "coordinate": f"{measure.title}!M37"},
            {"rule_id": "JS-003", "coordinate": f"{survey.title}!I40"},
            {"rule_id": "JS-004", "coordinate": f"{survey.title}!M200"},
            {"rule_id": "JS-006", "coordinate": f"{other.title}!L4"},
            {"rule_id": "JS-007", "coordinate": f"{other.title}!M5"},
            {"rule_id": "JS-008", "coordinate": "汇总表!M4"},
        ],
        "expected_passing_row": f"{survey.title}!13",
        "boundary": "演示样例只验证结构化辅助审核，不代表真实项目结算审定。",
    }
    expected_path = (
        EXPECTED_PATH
        if output_path.resolve() == OUTPUT_PATH.resolve()
        else output_path.with_name(f"{output_path.stem}-预期结果.json")
    )
    expected_path.write_text(json.dumps(expected, ensure_ascii=False, indent=2), encoding="utf-8")

    template_hash_after = sha256(TEMPLATE_PATH)
    if template_hash_before != template_hash_after:
        raise RuntimeError("统一结算模板在样例生成过程中发生变化，已停止交付。")
    print(f"sample={output_path}")
    print(f"expected={expected_path}")
    print(f"template_sha256={template_hash_after}")
    return output_path


if __name__ == "__main__":
    build_sample()
