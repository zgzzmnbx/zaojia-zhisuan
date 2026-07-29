from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
AIW_MARKER = "【AIW】"
LIBRARY_ID = "cost-aiw"
SUPPORTED_SUFFIXES = {".doc", ".docx", ".pdf", ".xlsx"}
DEFAULT_PROJECT_OUTPUT = PROJECT_ROOT / "06-知识库问答资料" / "造价AIW资料库"
DEFAULT_EXTERNAL_OUTPUT_NAME = "00-造价智算问答知识资产"
ASSET_VERSION = "1.0"

HEADER_TERMS = {
    "序号",
    "编码",
    "项目编码",
    "项目名称",
    "项目名称及特征",
    "项目特征",
    "计量单位",
    "单位",
    "全费用单价",
    "人工费",
    "材料费",
    "机械费",
    "合计",
    "费用名称",
    "计算基数",
    "费率",
    "调整系数",
}


@dataclass(frozen=True)
class AssetSpec:
    title: str
    output_name: str
    converted_name: str | None = None


@dataclass(frozen=True)
class AssetRecord:
    asset_id: str
    title: str
    source_relative_path: str
    source_type: str
    source_sha256: str
    source_size: int
    source_modified_at: str
    output_name: str
    conversion_method: str
    status: str
    generated_size: int


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build reusable AIW-marked cost knowledge assets for @知识库.",
    )
    parser.add_argument(
        "--source-root",
        type=Path,
        default=Path(os.environ["AIW_KNOWLEDGE_ROOT"]) if os.environ.get("AIW_KNOWLEDGE_ROOT") else None,
        help="Root directory to scan. Only files whose names contain the complete marker 【AIW】 are included.",
    )
    parser.add_argument(
        "--project-output",
        type=Path,
        default=DEFAULT_PROJECT_OUTPUT,
        help="Project-local publishable snapshot directory.",
    )
    parser.add_argument(
        "--external-output",
        type=Path,
        help="Reusable external output directory. Defaults to <source-root>/00-造价智算问答知识资产.",
    )
    parser.add_argument(
        "--converted-dir",
        type=Path,
        help="Directory containing high-fidelity Markdown converted from PDF/DOC sources.",
    )
    parser.add_argument(
        "--no-external-sync",
        action="store_true",
        help="Only build the project-local snapshot.",
    )
    return parser.parse_args()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def source_modified_at(path: Path) -> str:
    return datetime.fromtimestamp(path.stat().st_mtime).astimezone().isoformat(timespec="seconds")


def json_string(value: str) -> str:
    return json.dumps(value, ensure_ascii=False)


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, ".15g")
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    return re.sub(r"\n+", " / ", text).strip()


def is_number_like(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def is_header_row(raw_values: list[Any], values: list[str]) -> bool:
    nonempty = [(raw, text) for raw, text in zip(raw_values, values) if text]
    if len(nonempty) < 2:
        return False
    joined = " ".join(text for _, text in nonempty)
    term_hits = sum(1 for term in HEADER_TERMS if term in joined)
    numeric_ratio = sum(1 for raw, _ in nonempty if is_number_like(raw)) / len(nonempty)
    return term_hits >= 1 and numeric_ratio <= 0.4


def heading_level(text: str) -> int:
    if re.match(r"^(第[一二三四五六七八九十百0-9]+[册篇章部分]|总说明|总则|目录)", text):
        return 3
    return 4


def is_heading_text(text: str) -> bool:
    if not text or len(text) > 90 or re.fullmatch(r"[-+]?\d+(?:\.\d+)?", text):
        return False
    if re.match(
        r"^(第[一二三四五六七八九十百0-9]+[册篇章部分条节]|总\s*说\s*明|总则|目录|"
        r"表\s*\d+|附件\s*\d*[:：]?|注[:：]?$)",
        text,
    ):
        return True
    if re.match(r"^[一二三四五六七八九十0-9]+[、.．]", text):
        return False
    if text.endswith(("。", "；", "，", ".", ";", ",")):
        return False
    return len(text) <= 30


def workbook_to_markdown(path: Path) -> tuple[str, dict[str, int]]:
    value_workbook = load_workbook(path, read_only=True, data_only=True)
    formula_workbook = load_workbook(path, read_only=True, data_only=False)
    lines: list[str] = []
    sheet_count = 0
    active_rows = 0
    try:
        for worksheet, formula_worksheet in zip(value_workbook.worksheets, formula_workbook.worksheets):
            sheet_count += 1
            lines.extend([f"## 工作表：{worksheet.title}", ""])
            header_labels: dict[int, str] = {}
            sheet_active_rows = 0
            value_rows = worksheet.iter_rows(values_only=True)
            formula_rows = formula_worksheet.iter_rows(values_only=True)
            for row_index, (value_row, formula_row) in enumerate(zip(value_rows, formula_rows), start=1):
                raw_values = [
                    value if value is not None else formula
                    for value, formula in zip(value_row, formula_row)
                ]
                values = [normalize_text(value) for value in raw_values]
                populated = [(index, text) for index, text in enumerate(values, start=1) if text]
                if not populated:
                    continue
                active_rows += 1
                sheet_active_rows += 1
                if len(populated) == 1:
                    column_index, text = populated[0]
                    if is_heading_text(text):
                        level = heading_level(text)
                        lines.extend(
                            [
                                f"{'#' * level} {text}",
                                "",
                                f"> 来源定位：工作表 `{worksheet.title}`，Excel 第 {row_index} 行，{get_column_letter(column_index)} 列。",
                                "",
                            ]
                        )
                    else:
                        lines.extend(
                            [
                                f"- Excel 第 {row_index} 行｜{get_column_letter(column_index)}列：{text}",
                                "",
                            ]
                        )
                    continue

                header = is_header_row(raw_values, values)
                if header:
                    for column_index, text in populated:
                        if len(text) <= 80:
                            header_labels[column_index] = text
                    fields = "｜".join(
                        f"{get_column_letter(column_index)}列={text}" for column_index, text in populated
                    )
                    lines.extend([f"**表头定位：Excel 第 {row_index} 行**｜{fields}", ""])
                    continue

                fields: list[str] = []
                for column_index, text in populated:
                    column = get_column_letter(column_index)
                    header_label = header_labels.get(column_index, "")
                    label = f"{column}列（{header_label}）" if header_label else f"{column}列"
                    fields.append(f"{label}={text}")
                lines.extend([f"- Excel 第 {row_index} 行｜" + "｜".join(fields), ""])
            if not sheet_active_rows:
                lines.extend(["> 本工作表没有非空数据。", ""])
    finally:
        value_workbook.close()
        formula_workbook.close()
    return "\n".join(lines).strip() + "\n", {"sheet_count": sheet_count, "active_rows": active_rows}


def asset_spec(path: Path) -> AssetSpec:
    name = path.name
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        workbook_rules = (
            ("第1册《油气输送管道工程》", "2024年建设项目全费用工程量清单单价：第1册 油气输送管道工程"),
            ("第2册《储罐制作安装工程》", "2024年建设项目全费用工程量清单单价：第2册 储罐制作安装工程"),
            ("第3册《通用安装工程》", "2024年建设项目全费用工程量清单单价：第3册 通用安装工程"),
            ("第4册《其他工程》", "2024年建设项目全费用工程量清单单价：第4册 其他工程"),
            ("第5册《参考计价依据》", "2024年建设项目全费用工程量清单单价：第5册 参考计价依据"),
        )
        for needle, title in workbook_rules:
            if needle in name:
                book = re.search(r"第([1-5])册", needle)
                book_number = book.group(1) if book else "0"
                short_title = title.split("：", maxsplit=1)[1].replace(" ", "")
                return AssetSpec(
                    title=title,
                    output_name=f"2024全费用清单单价-第{book_number}册{short_title[3:]}-Excel问答版.md",
                )

    binary_rules = (
        (
            "附件1-1：",
            "2024年建设项目全费用工程量清单单价：第1册 油气输送管道工程（正式正文）",
            "2024全费用清单单价-第1册油气输送管道工程-正文问答版.md",
            "2024全费用清单单价-第1册油气输送管道工程-正文版.md",
        ),
        (
            "附件3：",
            "2024年建设项目其他费用及相关费用计价依据",
            "2024建设项目其他费用及相关费用计价依据-正文问答版.md",
            "2024建设项目其他费用及相关费用计价依据-正文版.md",
        ),
        (
            "附件4：",
            "2024年数字与信息化项目投资计价依据",
            "2024数字与信息化项目投资计价依据-正文问答版.md",
            "2024数字与信息化项目投资计价依据-正文版.md",
        ),
        (
            "050000-【概算指标】",
            "石油建设安装工程概算指标（2005年）说明",
            "2005石油建设安装工程概算指标说明-正文问答版.md",
            "2005石油建设安装工程概算指标说明-正文版.md",
        ),
        (
            "120000-【其他费用】",
            "中国石油天然气集团公司建设项目其他费用和相关费用规定（2012年）",
            "2012石油建设项目其他费用和相关费用规定-正文问答版.md",
            "2012石油建设项目其他费用和相关费用规定-正文版.md",
        ),
        (
            "050000-【费用定额】",
            "石油建设安装工程费用定额（2005年）",
            "2005石油建设安装工程费用定额-正文问答版.md",
            "2005石油建设安装工程费用定额-正文版.md",
        ),
        (
            "150000-【费用定额】",
            "石油建设安装工程费用定额（2015年）",
            "2015石油建设安装工程费用定额-正文问答版.md",
            "2015石油建设安装工程费用定额-正文版.md",
        ),
        (
            "【标准-行业】《2025年中国软件行业基准数据》",
            "2025年中国软件行业基准数据",
            "2025中国软件行业基准数据-正文问答版.md",
            "2025中国软件行业基准数据-正文版.md",
        ),
    )
    for needle, title, output_name, converted_name in binary_rules:
        if needle in name:
            return AssetSpec(title=title, output_name=output_name, converted_name=converted_name)

    clean_stem = path.stem.replace(AIW_MARKER, "").strip(" -")
    return AssetSpec(
        title=clean_stem,
        output_name=f"{clean_stem}-问答版.md",
        converted_name=f"{clean_stem}-正文版.md" if suffix != ".xlsx" else None,
    )


def discover_sources(source_root: Path) -> list[Path]:
    return sorted(
        (
            path
            for path in source_root.rglob("*")
            if path.is_file()
            and AIW_MARKER in path.name
            and path.suffix.lower() in SUPPORTED_SUFFIXES
            and not path.name.startswith("~$")
        ),
        key=lambda item: item.relative_to(source_root).as_posix().lower(),
    )


def frontmatter(
    *,
    asset_id: str,
    spec: AssetSpec,
    source: Path,
    source_root: Path,
    source_sha256: str,
    conversion_method: str,
    generated_at: str,
) -> str:
    relative_path = source.relative_to(source_root).as_posix()
    fields = [
        "---",
        "knowledge_asset: true",
        f"asset_version: {json_string(ASSET_VERSION)}",
        f"asset_id: {json_string(asset_id)}",
        f"library_id: {json_string(LIBRARY_ID)}",
        f"title: {json_string(spec.title)}",
        f"source_marker: {json_string(AIW_MARKER)}",
        f"source_type: {json_string(source.suffix.lower().lstrip('.'))}",
        f"source_relative_path: {json_string(relative_path)}",
        f"source_sha256: {json_string(source_sha256)}",
        f"source_size: {source.stat().st_size}",
        f"source_modified_at: {json_string(source_modified_at(source))}",
        f"conversion_method: {json_string(conversion_method)}",
        f"generated_at: {json_string(generated_at)}",
        f"usage_scope: {json_string('仅用于知识问答检索与依据解释，不参与基价、单价或调整系数裁决。')}",
        "---",
        "",
        f"# {spec.title}",
        "",
        "> 本文是面向 AI 检索的派生知识资产。原始文件保持不变；回答时应同时给出来源文件与定位信息。",
        "> 本资料只进入 `@知识库` 解释链路，不能反向修改结构化计价、匹配规则、经验池或报告结果。",
        "",
        "## 来源信息",
        "",
        f"- 原始资料相对路径：`{relative_path}`",
        f"- 原始资料 SHA256：`{source_sha256}`",
        f"- 转换方式：{conversion_method}",
        "",
    ]
    return "\n".join(fields)


def build_asset(
    source: Path,
    source_root: Path,
    converted_dir: Path | None,
    generated_at: str,
) -> tuple[AssetRecord, str | None]:
    spec = asset_spec(source)
    relative_path = source.relative_to(source_root).as_posix()
    source_hash = sha256_file(source)
    asset_id = hashlib.sha256(relative_path.encode("utf-8")).hexdigest()[:16]
    body: str | None
    conversion_method: str
    if source.suffix.lower() == ".xlsx":
        body, workbook_stats = workbook_to_markdown(source)
        conversion_method = (
            "openpyxl 逐行扁平化；保留工作表、Excel 行号、列号、识别到的表头和全部非空单元格"
            f"；共 {workbook_stats['sheet_count']} 个工作表、{workbook_stats['active_rows']} 条非空行"
        )
    else:
        converted_path = converted_dir / spec.converted_name if converted_dir and spec.converted_name else None
        if not converted_path or not converted_path.exists():
            return (
                AssetRecord(
                    asset_id=asset_id,
                    title=spec.title,
                    source_relative_path=relative_path,
                    source_type=source.suffix.lower().lstrip("."),
                    source_sha256=source_hash,
                    source_size=source.stat().st_size,
                    source_modified_at=source_modified_at(source),
                    output_name=spec.output_name,
                    conversion_method="等待高保真 Markdown 转换",
                    status="pending_conversion",
                    generated_size=0,
                ),
                None,
            )
        body = converted_path.read_text(encoding="utf-8", errors="replace").strip() + "\n"
        conversion_method = "05-sm-xx2md 高保真版面解析并清理为 Markdown"

    content = (
        frontmatter(
            asset_id=asset_id,
            spec=spec,
            source=source,
            source_root=source_root,
            source_sha256=source_hash,
            conversion_method=conversion_method,
            generated_at=generated_at,
        )
        + body
    )
    encoded = content.encode("utf-8")
    return (
        AssetRecord(
            asset_id=asset_id,
            title=spec.title,
            source_relative_path=relative_path,
            source_type=source.suffix.lower().lstrip("."),
            source_sha256=source_hash,
            source_size=source.stat().st_size,
            source_modified_at=source_modified_at(source),
            output_name=spec.output_name,
            conversion_method=conversion_method,
            status="ready",
            generated_size=len(encoded),
        ),
        content,
    )


def inventory_markdown(records: list[AssetRecord], generated_at: str) -> str:
    ready_count = sum(record.status == "ready" for record in records)
    pending_count = len(records) - ready_count
    lines = [
        "# 造价 AIW 问答知识资产清单",
        "",
        f"- 生成时间：{generated_at}",
        f"- 源文件总数：{len(records)}",
        f"- 已就绪：{ready_count}",
        f"- 待转换：{pending_count}",
        f"- 纳入条件：文件名必须包含完整人工标记 `{AIW_MARKER}`",
        "- 使用边界：仅用于 `@知识库` 检索和依据解释，不参与价格、系数、匹配规则、经验池或报告裁决。",
        "",
        "| 状态 | 知识资产 | 源类型 | 原始资料 | 输出文件 | SHA256（前12位） |",
        "|---|---|---|---|---|---|",
    ]
    for record in records:
        status = "已就绪" if record.status == "ready" else "待转换"
        lines.append(
            f"| {status} | {record.title} | {record.source_type} | "
            f"`{record.source_relative_path}` | `{record.output_name}` | `{record.source_sha256[:12]}` |"
        )
    lines.extend(
        [
            "",
            "## 复用说明",
            "",
            "其他智能体可优先读取本清单，再按问题主题打开对应 Markdown。每份资产均包含原始资料相对路径、SHA256、转换方式和定位信息。",
            "若原文件更新，应重新运行生成脚本；不得直接修改派生资产来替代原始资料修订。",
            "",
        ]
    )
    return "\n".join(lines)


def write_outputs(
    output_dirs: list[Path],
    records: list[AssetRecord],
    contents: dict[str, str],
    generated_at: str,
) -> None:
    manifest = {
        "schema_version": 1,
        "library_id": LIBRARY_ID,
        "generated_at": generated_at,
        "source_marker": AIW_MARKER,
        "usage_scope": "仅用于@知识库检索与依据解释，不参与价格、系数或其他业务裁决。",
        "asset_count": len(records),
        "ready_count": sum(record.status == "ready" for record in records),
        "assets": [asdict(record) for record in records],
    }
    inventory = inventory_markdown(records, generated_at)
    for output_dir in output_dirs:
        output_dir.mkdir(parents=True, exist_ok=True)
        for output_name, content in contents.items():
            (output_dir / output_name).write_text(content, encoding="utf-8")
        (output_dir / "manifest.json").write_text(
            json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        (output_dir / "知识资产清单.md").write_text(inventory, encoding="utf-8")


def main() -> int:
    args = parse_args()
    if args.source_root is None:
        raise SystemExit("请通过 --source-root 或 AIW_KNOWLEDGE_ROOT 指定造价知识库目录。")
    source_root = args.source_root.resolve()
    if not source_root.exists():
        raise SystemExit(f"源目录不存在：{source_root}")
    converted_dir = args.converted_dir.resolve() if args.converted_dir else None
    project_output = args.project_output.resolve()
    external_output = (
        args.external_output.resolve()
        if args.external_output
        else source_root / DEFAULT_EXTERNAL_OUTPUT_NAME
    )

    sources = discover_sources(source_root)
    if not sources:
        raise SystemExit(f"未找到文件名含完整标记 {AIW_MARKER} 的受支持文件。")

    generated_at = datetime.now().astimezone().isoformat(timespec="seconds")
    records: list[AssetRecord] = []
    contents: dict[str, str] = {}
    for source in sources:
        record, content = build_asset(source, source_root, converted_dir, generated_at)
        records.append(record)
        if content is not None:
            if record.output_name in contents:
                raise RuntimeError(f"多个源文件映射到同一知识资产：{record.output_name}")
            contents[record.output_name] = content
        print(f"[{record.status}] {record.source_relative_path} -> {record.output_name}")

    output_dirs = [project_output]
    if not args.no_external_sync:
        output_dirs.append(external_output)
    write_outputs(output_dirs, records, contents, generated_at)

    pending = [record for record in records if record.status != "ready"]
    print(
        json.dumps(
            {
                "sources": len(records),
                "ready": len(records) - len(pending),
                "pending": len(pending),
                "project_output": str(project_output),
                "external_output": None if args.no_external_sync else str(external_output),
            },
            ensure_ascii=False,
        )
    )
    return 2 if pending else 0


if __name__ == "__main__":
    raise SystemExit(main())
